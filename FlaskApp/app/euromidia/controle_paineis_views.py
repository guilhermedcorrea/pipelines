from flask_sqlalchemy import SQLAlchemy
#from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify,abort,send_file,session
from flask import Blueprint, current_app, render_template, request, redirect, url_for, flash, jsonify, abort, send_file, session
from ..extensions import db,limiter
from ..models.euromidia_models import (DimPaineisEuromidia,DimFacesPaineis,Vendedores
                                       ,FatoControleContratosEuromidia
                                       ,FatoControleContratosItensEuromidia,DimEmpresas,DimCustoMensalPainel,
                                       DimMargemPaineisEuromidia,FatoOcupacaoPaineisEuromidia,DimCnaes,
                                       DimClassificacacaoClientes,DimCustoPainel,DimCheckingHistorico)

from ..autenticacao.autenticacao_views import requer_permissao
from ..models.autenticacao import (DimUsuarios,DimPerfilUsuario, DimPermissoes, PermissoesUsuario,)
from ..forms.euromidia_forms import (FormCadastroCliente,CadastroContratoManualForm,FormUsuarioNovo,
FormUsuarioEditar, FormTrocarSenha, FormPermissaoExtraUpsert, FormPermissaoExtraRemover,ReservaOcupacaoForm)
from ..models.admin_models import DimCalendario,DimEmpresaProprietaria,DimRecorrencia,DimPublicoAlvo
from sqlalchemy import case, String,cast, or_, and_,func,text,  select
from datetime import date, datetime, timedelta
from sqlalchemy.exc import OperationalError
from functools import wraps
import time,random,calendar,re,requests
from decimal import Decimal, InvalidOperation
from io import BytesIO
from openpyxl import Workbook
from werkzeug.security import generate_password_hash
from flask_login import login_required
from flask_wtf.csrf import CSRFError,validate_csrf
import json
import math
from flask_login import current_user
import threading
import os

    
    

from PIL import Image, ImageOps
from pathlib import Path
from werkzeug.utils import secure_filename
import uuid




from ..autenticacao.acl_menu_paineis import pode_acessar_menu_paineis, requer_item_menu_paineis



paineis_bp = Blueprint("Paineis", __name__)



@paineis_bp.app_context_processor
def injetar_acl_menu_paineis():
    return {
        "pode_acessar_menu_paineis": pode_acessar_menu_paineis,
    }



def exige_admin_ou_usuarios_gerenciar(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not current_user.is_authenticated:
            return abort(401)

        pode = (
            current_user.has_permission("ADMIN_TUDO") or
            current_user.has_permission("USUARIOS_GERENCIAR")
        )
        if not pode:
            return abort(403)

        return view_func(*args, **kwargs)
    return wrapper









_TRANSIENT_SQLSTATE = ("08S01",)           
_TRANSIENT_MSG = ("10054", "10060", "timeout", "communication link failure", "reset by peer")

def _is_disconnect_error(exc: OperationalError) -> bool:
    s = str(getattr(exc, "orig", exc)).lower()
    return any(code.lower() in s for code in _TRANSIENT_SQLSTATE) or any(m in s for m in _TRANSIENT_MSG)

def _cleanup_session_and_pool(db):
    try:
        db.session.rollback()
    except Exception:
        pass
    try:
        db.session.close()
    except Exception:
        pass
    
    try:
        db.engine.dispose()
    except Exception:
        pass


def retry_get_view(db, attempts: int = 6, base_delay: float = 0.2, max_delay: float = 1.5):

    def _decorator(fn):
        @wraps(fn)
        def _wrapped(*args, **kwargs):
            if request.method != "GET":
                return fn(*args, **kwargs)

            last_exc = None
            for i in range(attempts):
                try:
                    return fn(*args, **kwargs)
                except OperationalError as e:
                    if not _is_disconnect_error(e):
                        raise
                    last_exc = e
                    _cleanup_session_and_pool(db)
                  
                    sleep_s = min(max_delay, base_delay * (2 ** i)) + random.uniform(0, 0.15)
                    time.sleep(sleep_s)
                except Exception:
  
                    raise
     
            raise last_exc
        return _wrapped
    return _decorator




@paineis_bp.route("/ocupacao/novo", methods=["GET", "POST"])
@login_required
@limiter.limit("25 per minute", methods=["POST"])
@retry_get_view(db, attempts=6, base_delay=0.2, max_delay=1.5)
def cadastrar_ocupacao():

    LOOPS_PERMITIDOS = ["1MIN", "2MIN"]


    clientes = (
        db.session.query(DimClientes)
        .order_by(
            case((DimClientes.NomeFantasia == None, 1), else_=0),
            DimClientes.NomeFantasia.asc(),
            case((DimClientes.RazaoSocial == None, 1), else_=0),
            DimClientes.RazaoSocial.asc(),
        )
        .all()
    )

    vendedores = (
        db.session.query(Vendedores)
        .order_by(
            case((Vendedores.NomeVendedor == None, 1), else_=0),
            Vendedores.NomeVendedor.asc(),
        )
        .all()
    )

    rows_cp = (
        db.session.query(DimFacesPaineis.CodPonto)
        .filter(DimFacesPaineis.CodPonto != None)
        .group_by(DimFacesPaineis.CodPonto)
        .order_by(DimFacesPaineis.CodPonto.asc())
        .all()
    )

    codpontos = []
    for (cp,) in rows_cp:
        if cp is None:
            continue
        try:
            codpontos.append(int(cp))
        except:
            continue

    rows_faces = (
        db.session.query(DimFacesPaineis.CodPonto, DimFacesPaineis.CodFace)
        .filter(
            DimFacesPaineis.CodPonto != None,
            DimFacesPaineis.CodFace != None,
            DimFacesPaineis.CodFace != "",
        )
        .group_by(DimFacesPaineis.CodPonto, DimFacesPaineis.CodFace)
        .all()
    )

    faces_por_ponto: dict[int, list[str]] = {}
    for cp, cf in rows_faces:
        try:
            cp_int = int(cp)
        except:
            continue
        if not cf:
            continue
        faces_por_ponto.setdefault(cp_int, []).append(str(cf))

    for cp_int in list(faces_por_ponto.keys()):
        faces_por_ponto[cp_int] = sorted(set(faces_por_ponto[cp_int]))


    form = CadastroContratoManualForm()

  
    form.cliente_id.choices = [
        (
            int(c.IDCliente),
            f"{(c.NomeFantasia or c.RazaoSocial or 'Cliente sem nome').strip()} {(('— ' + (c.CNPJ or '').strip()) if (c.CNPJ or '').strip() else '')}".strip()
        )
        for c in clientes
    ]
    form.vendedor_id.choices = [
        (int(v.IDVendedor), (v.NomeVendedor or "Vendedor sem nome").strip())
        for v in vendedores
    ]
    form.cod_ponto.choices = [(int(cp), str(cp)) for cp in codpontos]
    form.loop_tipo.choices = [(lp, lp) for lp in LOOPS_PERMITIDOS]

   
    default_cod_ponto = 118
    default_cod_face = "118AD"

    cod_ponto_selecionado = None
    if form.cod_ponto.data:
        try:
            cod_ponto_selecionado = int(form.cod_ponto.data)
        except:
            cod_ponto_selecionado = None

    if cod_ponto_selecionado is None:
        cod_ponto_selecionado = (
            default_cod_ponto if default_cod_ponto in codpontos else (codpontos[0] if codpontos else None)
        )

    faces_lista = faces_por_ponto.get(cod_ponto_selecionado, []) if cod_ponto_selecionado else []
    form.cod_face.choices = [(cf, cf) for cf in faces_lista]



    if request.method == "GET":
        hoje = datetime.now().date()

        if cod_ponto_selecionado is not None:
            form.cod_ponto.data = cod_ponto_selecionado

        if faces_lista:
            form.cod_face.data = default_cod_face if default_cod_face in faces_lista else faces_lista[0]

        form.loop_tipo.data = "2MIN" if "2MIN" in LOOPS_PERMITIDOS else LOOPS_PERMITIDOS[0]

        if vendedores:
            form.vendedor_id.data = int(vendedores[0].IDVendedor)

     
        form.origem.data = "MANUAL"
        form.tipo_documento.data = "MANUAL"
        form.data_lancamento.data = hoje

      
        form.data_inicio_previsto.data = hoje
        form.data_termino_previsto.data = hoje
        form.cota.data = 1
        form.permuta.data = "0"

        return render_template(
            "euromidia/ocupacao_form.html",
            form=form,
            clientes=clientes,
            vendedores=vendedores,
            codpontos=codpontos,
            faces_por_ponto=faces_por_ponto,
            loops=LOOPS_PERMITIDOS,
        )



    if not form.validate_on_submit():
        erros = []
        for campo, msgs in form.errors.items():
            for msg in msgs:
                erros.append(f"{campo}: {msg}")
        flash("Form inválido: " + " | ".join(erros), "danger")
        return redirect(url_for("Paineis.cadastrar_ocupacao"))

    try:
     
        cliente_id = int(form.cliente_id.data)
        vendedor_id = int(form.vendedor_id.data)

        numero_contrato = (form.numero_contrato.data or "").strip() or None
        numero_previa = (form.numero_previa.data or "").strip() or None
        data_lancamento = form.data_lancamento.data
        cidade_exibicao = (form.cidade_exibicao.data or "").strip() or None
        origem = (form.origem.data or "MANUAL").strip()
        tipo_documento = (form.tipo_documento.data or "MANUAL").strip()

  
        item_1 = {
            "cod_ponto": int(form.cod_ponto.data),
            "cod_face": (form.cod_face.data or "").strip(),
            "data_inicio": form.data_inicio_previsto.data,
            "data_fim": form.data_termino_previsto.data,
            "loop": (form.loop_tipo.data or "").strip(),
            "cota": int(form.cota.data or 1),
            "qtd_parcelas": form.quantidade_parcelas.data,
            "bruto_mensal": form.faturamento_bruto_mensal.data,
            "liquido_mensal": form.faturamento_liquido_mensal.data,
            "permuta_flag": 1 if (form.permuta.data or "0") == "1" else 0,
            "valor_permuta": form.valor_permuta.data,
        }

   
        extras_cod_ponto = request.form.getlist("itens_cod_ponto[]")
        extras_cod_face = request.form.getlist("itens_cod_face[]")
        extras_data_inicio = request.form.getlist("itens_data_inicio[]")
        extras_data_fim = request.form.getlist("itens_data_fim[]")
        extras_loop = request.form.getlist("itens_loop[]")

        extras_cota = request.form.getlist("itens_cota[]")
        extras_parcelas = request.form.getlist("itens_quantidade_parcelas[]")
        extras_bruto = request.form.getlist("itens_faturamento_bruto_mensal[]")
        extras_liquido = request.form.getlist("itens_faturamento_liquido_mensal[]")
        extras_permuta = request.form.getlist("itens_permuta[]")
        extras_valor_permuta = request.form.getlist("itens_valor_permuta[]")

      
        itens = [item_1]

      
        n = len(extras_cod_ponto)
        if not (len(extras_cod_face) == n == len(extras_data_inicio) == len(extras_data_fim) == len(extras_loop)):
            raise ValueError("Itens adicionais inconsistentes (listas de campos com tamanhos diferentes).")

        def _to_int(v, padrao=None):
            s = (v or "").strip()
            if not s:
                return padrao
            return int(s)

        def _to_date(v):
            s = (v or "").strip()
            if not s:
                return None
            return datetime.strptime(s, "%Y-%m-%d").date()

        def _to_decimal_str(v):
    
            s = (v or "").strip()
            return s if s else None

        for i in range(n):
            cp = _to_int(extras_cod_ponto[i])
            cf = (extras_cod_face[i] or "").strip()
            di = _to_date(extras_data_inicio[i])
            df = _to_date(extras_data_fim[i])
            lp = (extras_loop[i] or "").strip()

            if not cp:
                raise ValueError(f"Item adicional #{i+2}: informe o CodPonto.")
            if not cf:
                raise ValueError(f"Item adicional #{i+2}: informe o CodFace.")
            if not di or not df:
                raise ValueError(f"Item adicional #{i+2}: informe Data Início e Data Fim.")
            if df < di:
                raise ValueError(f"Item adicional #{i+2}: Data Fim não pode ser menor que Data Início.")

            cota_i = _to_int(extras_cota[i] if i < len(extras_cota) else "", 1) or 1
            parcelas_i = _to_int(extras_parcelas[i] if i < len(extras_parcelas) else "", None)
            bruto_i = _to_decimal_str(extras_bruto[i] if i < len(extras_bruto) else "")
            liquido_i = _to_decimal_str(extras_liquido[i] if i < len(extras_liquido) else "")
            permuta_i = 1 if (extras_permuta[i] if i < len(extras_permuta) else "0") == "1" else 0
            valor_permuta_i = _to_decimal_str(extras_valor_permuta[i] if i < len(extras_valor_permuta) else "")

            itens.append(
                {
                    "cod_ponto": cp,
                    "cod_face": cf,
                    "data_inicio": di,
                    "data_fim": df,
                    "loop": lp,
                    "cota": cota_i,
                    "qtd_parcelas": parcelas_i,
                    "bruto_mensal": bruto_i,
                    "liquido_mensal": liquido_i,
                    "permuta_flag": permuta_i,
                    "valor_permuta": valor_permuta_i,
                }
            )

    
        if not numero_previa:
          
            numero_previa = f"MANUAL-{itens[0]['cod_ponto']}-{itens[0]['cod_face']}-{datetime.now().strftime('%Y%m%d%H%M%S')}"

     
   
        cliente = db.session.query(DimClientes).filter(DimClientes.IDCliente == cliente_id).first()
        if not cliente:
            raise ValueError("Cliente inválido (não encontrado).")

        vendedor_obj = db.session.query(Vendedores).filter(Vendedores.IDVendedor == vendedor_id).first()
        if not vendedor_obj:
            raise ValueError("Vendedor inválido (não encontrado).")

        marca_exibida = (
            (getattr(cliente, "NomeFantasia", None) or "").strip()
            or (getattr(cliente, "RazaoSocial", None) or "").strip()
            or "Cliente sem nome"
        )
        razao_social = (getattr(cliente, "RazaoSocial", None) or "").strip() or None
        cnpj = (getattr(cliente, "CNPJ", None) or "").strip() or None
        cpf = (getattr(cliente, "CPF", None) or "").strip() or None
        nome_vendedor = (getattr(vendedor_obj, "NomeVendedor", None) or "").strip() or "Vendedor"

 
        sql_insert_contrato = text("""
            INSERT INTO [Integracao].[Silver].[FatoControleContratos] (
                DataAtualizacao,
                IDEmpresaProprietaria,
                NumeroContrato,
                NumeroPrevia,
                DataLancamento,
                CidadeExibicao,
                Origem,
                TipoDocumento,
                RazaoSocial,
                CNPJ,
                CPF,
                MarcaExibida,
                IDDimClientesEuromidia,
                QtdeCodFace,
                QtdeCodPonto,
                TotalBrutoContrato,
                TotalLiquidoContratoAGBRCTACORDO,
                TotalLiquidoContratoAGBRVENDGERCOOR,
                TotalValorMensalAgencia,
                TotalValorBureauMensal,
                TotalValorAcordoMensal,
                TotalOutrasComissoes,
                TotalFaturamentoLiquidoMensal,
                TotalFaturamentoLiquidoMensalFinal,
                TotalValorVendedor,
                ValorVendedorTotal,
                TotalValorCoordenador,
                ValorCoordenadorTotal,
                ValorGerenciaTotal,
                DimVendedor
            )
            OUTPUT INSERTED.IDFatoControleContratosEuromidia
            VALUES (
                GETDATE(),
                NULL,
                :NumeroContrato,
                :NumeroPrevia,
                :DataLancamento,
                :CidadeExibicao,
                :Origem,
                :TipoDocumento,
                :RazaoSocial,
                :CNPJ,
                :CPF,
                :MarcaExibida,
                :IDDimClientesEuromidia,
                :QtdeCodFace,
                :QtdeCodPonto,
                NULL,
                NULL,
                NULL,
                NULL,
                NULL,
                NULL,
                NULL,
                NULL,
                NULL,
                NULL,
                NULL,
                NULL,
                NULL,
                NULL,
                NULL,
                :DimVendedor
            );
        """)

        id_fato_contrato = db.session.execute(
            sql_insert_contrato,
            {
                "NumeroContrato": numero_contrato,
                "NumeroPrevia": numero_previa,
                "DataLancamento": data_lancamento,
                "CidadeExibicao": cidade_exibicao,
                "Origem": origem,
                "TipoDocumento": tipo_documento,
                "RazaoSocial": razao_social,
                "CNPJ": cnpj,
                "CPF": cpf,
                "MarcaExibida": marca_exibida,
                "IDDimClientesEuromidia": cliente_id,
                "QtdeCodFace": int(len(itens)),
                "QtdeCodPonto": int(len(itens)),
                "DimVendedor": nome_vendedor,
            },
        ).scalar_one()

 
        sql_insert_item = text("""
            INSERT INTO [Integracao].[Silver].[FatoControleContratosItens] (
                DataAtualizacao,
                NumeroContrato,
                NumeroPrevia,
                DataLancamento,
                DataAssinaturaRenovacao,
                DataInicioPrevisto,
                DataTerminoPrevisto,
                DataCancelamento,
                CidadeExibicao,
                Tipo,
                Origem,
                TipoDocumento,
                RazaoSocial,
                CNPJ,
                CPF,
                MarcaExibida,
                Vendedor,
                SDR,
                InicioRenovacao,
                AtivoCancelamento,
                Cota,
                CodPonto,
                CodFace,
                TempoExibicaoDias,
                QuantidadeParcelas,
                FaturamentoBrutoMensal,
                ValorPermuta,
                FaturamentoLiquidoPermuta,
                TotalBrutoContrato,
                TotalLiquidoContratoAGBRCTACORDO,
                TotalLiquidoContratoAGBRVENDGERCOOR,
                ValorMensalAgencia,
                ValorBureauMensal,
                ValorAcordoMensal,
                OutrasComissoes,
                FaturamentoLiquidoMensal,
                FaturamentoLiquidoMensalFinal,
                ValorVendedor,
                ValorVendedorTotal,
                ValorCoordenador,
                ValorCoordenadorTotal,
                ValorGerencia,
                ValorGerenciaTotal,
                Permuta,
                PercentualAgencia,
                PercentualBureau,
                PercentualCartaAcordo,
                PercentualComissaoVendedor,
                ComissaoCoordenacao,
                PercentualComissaoGerencia,
                IDFatoControleContratosEuromidia,
                IDVendedor
            )
            VALUES (
                GETDATE(),
                :NumeroContrato,
                :NumeroPrevia,
                :DataLancamento,
                NULL,
                :DataInicioPrevisto,
                :DataTerminoPrevisto,
                NULL,
                :CidadeExibicao,
                :Tipo,
                :Origem,
                :TipoDocumento,
                :RazaoSocial,
                :CNPJ,
                :CPF,
                :MarcaExibida,
                :Vendedor,
                NULL,
                NULL,
                NULL,
                :Cota,
                :CodPonto,
                :CodFace,
                NULL,
                :QuantidadeParcelas,
                :FaturamentoBrutoMensal,
                :ValorPermuta,
                NULL,
                NULL,
                NULL,
                NULL,
                NULL,
                NULL,
                NULL,
                NULL,
                :FaturamentoLiquidoMensal,
                NULL,
                NULL,
                NULL,
                NULL,
                NULL,
                NULL,
                NULL,
                NULL,
                :Permuta,
                NULL,
                NULL,
                NULL,
                NULL,
                NULL,
                NULL,
                :IDFatoControleContratosEuromidia,
                :IDVendedor
            );
        """)

      
        for idx, it in enumerate(itens, start=1):
            cod_ponto = int(it["cod_ponto"])
            cod_face = (it["cod_face"] or "").strip()
            data_inicio = it["data_inicio"]
            data_fim = it["data_fim"]
            loop = (it["loop"] or "").strip()

       
            info_face = (
                db.session.query(
                    DimFacesPaineis.TipoPainel.label("TipoPainel"),
                    DimFacesPaineis.IDDimPaineisEuromidia.label("IDDimPaineisEuromidia"),
                )
                .filter(
                    DimFacesPaineis.CodPonto == cod_ponto,
                    DimFacesPaineis.CodFace == cod_face,
                )
                .first()
            )
            if not info_face:
                raise ValueError(f"Item #{idx}: CodFace inválido para o CodPonto selecionado.")

            tipo_painel_face = (info_face.TipoPainel or "").strip().upper()
            eh_digital = (tipo_painel_face == "PAINEL DIGITAL")

            if eh_digital:
                if loop not in LOOPS_PERMITIDOS:
                    raise ValueError(f"Item #{idx}: Loop inválido. Use: {', '.join(LOOPS_PERMITIDOS)}.")
            else:
                loop = "2MIN"

            db.session.execute(
                sql_insert_item,
                {
                    "NumeroContrato": numero_contrato,
                    "NumeroPrevia": numero_previa,
                    "DataLancamento": data_lancamento,
                    "DataInicioPrevisto": data_inicio,
                    "DataTerminoPrevisto": data_fim,
                    "CidadeExibicao": cidade_exibicao,
                    "Tipo": loop, 
                    "Origem": origem,
                    "TipoDocumento": tipo_documento,
                    "RazaoSocial": razao_social,
                    "CNPJ": cnpj,
                    "CPF": cpf,
                    "MarcaExibida": marca_exibida,
                    "Vendedor": nome_vendedor,
                    "Cota": int(it.get("cota") or 1),
                    "CodPonto": cod_ponto,
                    "CodFace": cod_face,
                    "QuantidadeParcelas": it.get("qtd_parcelas"),
                    "FaturamentoBrutoMensal": it.get("bruto_mensal"),
                    "FaturamentoLiquidoMensal": it.get("liquido_mensal"),
                    "Permuta": int(it.get("permuta_flag") or 0),
                    "ValorPermuta": it.get("valor_permuta"),
                    "IDFatoControleContratosEuromidia": int(id_fato_contrato),
                    "IDVendedor": vendedor_id,
                },
            )

        db.session.commit()
        flash(f"Contrato MANUAL cadastrado com sucesso. IDFatoControleContratosEuromidia={id_fato_contrato}", "success")
        return redirect(url_for("Paineis.cadastrar_ocupacao"))

    except Exception as e:
        db.session.rollback()
        flash(str(e), "danger")
        return redirect(url_for("Paineis.cadastrar_ocupacao"))




@paineis_bp.route("/api/pontos", methods=["GET"])
@limiter.limit("120 per minute", methods=["GET"])
@retry_get_view(db, attempts=2, base_delay=0.2, max_delay=0.8)
def api_pontos():
    q = (request.args.get("q") or "").strip()

    sql = text("""
        SELECT TOP (100)
            p.CodPonto,
            TipoPainel = MAX(NULLIF(LTRIM(RTRIM(COALESCE(p.Tipo, ''))), ''))
        FROM [Integracao].[Silver].[DimPaineisEuromidia] p
        WHERE
            p.CodPonto IS NOT NULL
            AND (
                :q = ''
                OR CAST(p.CodPonto AS varchar(30)) LIKE :q_like
            )
        GROUP BY p.CodPonto
        ORDER BY p.CodPonto ASC
    """)

    rows = db.session.execute(
        sql,
        {
            "q": q,
            "q_like": f"{q}%"
        }
    ).mappings().all()

    itens = []

    for row in rows:
        cod_ponto = row["CodPonto"]
        if cod_ponto is None:
            continue

        tipo_painel = _normalizar_texto_checking(row["TipoPainel"]) or "Sem tipo"

        try:
            cod_ponto_int = int(cod_ponto)
        except Exception:
            continue

        itens.append({
            "cod_ponto": cod_ponto_int,
            "tipo_painel": tipo_painel,
            "texto": f"{cod_ponto_int} | {tipo_painel}",
        })

    return jsonify({"items": itens})


@paineis_bp.route("/api/pontos/<int:codponto>/faces", methods=["GET"])
@limiter.limit("120 per minute", methods=["GET"])
@retry_get_view(db, attempts=2, base_delay=0.2, max_delay=0.8)
def api_faces_do_ponto(codponto: int):
    sql = text("""
        SELECT TOP (200)
            f.CodFace,
            TipoFace = MAX(
                NULLIF(
                    LTRIM(RTRIM(COALESCE(f.Tipo, p.Tipo, ''))),
                    ''
                )
            )
        FROM [Integracao].[Silver].[DimFacesPaineis] f
        LEFT JOIN [Integracao].[Silver].[DimPaineisEuromidia] p
            ON p.IDDimPaineisEuromidia = f.IDDimPaineisEuromidia
        WHERE
            f.CodPonto = :codponto
            AND NULLIF(LTRIM(RTRIM(COALESCE(f.CodFace, ''))), '') IS NOT NULL
        GROUP BY f.CodFace
        ORDER BY f.CodFace ASC
    """)

    rows = db.session.execute(
        sql,
        {"codponto": codponto}
    ).mappings().all()

    itens = []

    for row in rows:
        cod_face = _normalizar_texto_checking(row["CodFace"])
        if not cod_face:
            continue

        tipo_face = _normalizar_texto_checking(row["TipoFace"]) or "Sem tipo"

        itens.append({
            "cod_face": cod_face,
            "tipo_face": tipo_face,
            "texto": f"{cod_face} | {tipo_face}",
        })

    return jsonify({
        "codponto": codponto,
        "items": itens
    })




CAPACIDADE_DIGITAL_FIXA = 16


def _primeiro_ultimo_dia_mes(ano: int, mes: int):
    primeiro = date(ano, mes, 1)
    ultimo = date(ano, mes, calendar.monthrange(ano, mes)[1])
    return primeiro, ultimo


def _fim_efetivo_item(dt_fim_previsto, dt_cancelamento):
    df = dt_fim_previsto
    dc = dt_cancelamento

    if df is None and dc is None:
        return None
    if df is None:
        return dc
    if dc is None:
        return df
    return dc if dc < df else df


def _span_por_cota(cota) -> int:
    try:
        c = int(cota)
    except:
        c = None

    if c == 1:
        return 2
    if c == 2:
        return 1
    return 1


def _calcular_ocupacao_e_conflitos_por_grupo(
    ocupacoes,
    tipo_por_face,
    tipo_por_idcadastro,
    capacidade_digital_por_cp,
    faces_por_cp_tipo,
    dt_ini,
    dt_fim,
):

    def _daterange(d0: date, d1: date):
        cur = d0
        while cur <= d1:
            yield cur
            cur += timedelta(days=1)


    out = {}


    spans_por_cp_por_dia = {}


    intervalos_por_face = {}


    faces_ocupadas_nao_digital = set()


    for cp, cf, cota, di, df, idcad in ocupacoes:
        if cp is None or di is None or df is None:
            continue

        
        cf_norm = (str(cf) or "").strip() if cf is not None else ""


        ini = max(di, dt_ini)
        fim = min(df, dt_fim)
        if fim < ini:
            continue

        tipo = ""
        if idcad not in (None, ""):
            try:
                tipo = (tipo_por_idcadastro.get(int(idcad)) or "").strip().upper()
            except:
                tipo = ""

        if not tipo and cf_norm:
            tipo = (tipo_por_face.get((cp, cf_norm)) or "").strip().upper()

      
        if not tipo:
            if (cp, "PAINEL DIGITAL") in faces_por_cp_tipo or int(capacidade_digital_por_cp.get(cp) or 0) > 0:
                tipo = "PAINEL DIGITAL"

        if not tipo:
            continue

        if tipo == "PAINEL DIGITAL":
            span = _span_por_cota(cota)
            spans_por_cp_por_dia.setdefault(cp, {})
            for d in _daterange(ini, fim):
                spans_por_cp_por_dia[cp][d] = spans_por_cp_por_dia[cp].get(d, 0) + span
        else:
        
            if not cf_norm:
                continue
            faces_ocupadas_nao_digital.add((cp, tipo, cf_norm))
            intervalos_por_face.setdefault((cp, tipo, cf_norm), []).append((ini, fim))


    for cp, por_dia in spans_por_cp_por_dia.items():
    
        capacidade_cadastro = int(capacidade_digital_por_cp.get(cp) or 0)

        pico_mes = 0
        tem_conflito = False

        for d, soma_span in por_dia.items():
            if soma_span > pico_mes:
                pico_mes = soma_span
            if capacidade_cadastro > 0 and soma_span > capacidade_cadastro:
                tem_conflito = True


        denominador = capacidade_cadastro
        if denominador <= 0:
            denominador = int(pico_mes) if pico_mes > 0 else 0
            tem_conflito = False

        key = (cp, "PAINEL DIGITAL")
        out[key] = {
            "ocupadas": int(pico_mes),
            "denominador": int(denominador),
            "conflitos": 1 if tem_conflito else 0,
        }

    ocupadas_por_grupo = {}
    for (cp, tipo, cf) in faces_ocupadas_nao_digital:
        ocupadas_por_grupo[(cp, tipo)] = ocupadas_por_grupo.get((cp, tipo), 0) + 1


    conflitos_por_grupo = {}
    for (cp, tipo, cf), intervalos in intervalos_por_face.items():
        if len(intervalos) <= 1:
            continue

        intervalos = sorted(intervalos, key=lambda x: (x[0], x[1]))
        fim_atual = intervalos[0][1]
        conflito = False

        for i in range(1, len(intervalos)):
            ini, fim = intervalos[i]
            if ini <= fim_atual:
                conflito = True
                break
            fim_atual = max(fim_atual, fim)

        if conflito:
            conflitos_por_grupo[(cp, tipo)] = conflitos_por_grupo.get((cp, tipo), 0) + 1


    for (cp, tipo), faces_set in faces_por_cp_tipo.items():
        if tipo == "PAINEL DIGITAL":
            continue
        out.setdefault((cp, tipo), {})
        out[(cp, tipo)]["ocupadas"] = int(ocupadas_por_grupo.get((cp, tipo), 0))
        out[(cp, tipo)]["denominador"] = int(len(faces_set))
        out[(cp, tipo)]["conflitos"] = int(conflitos_por_grupo.get((cp, tipo), 0))

    return out


def _intersecao_mes(di, df, dt_ini, dt_fim):
    """Retorna (dia_inicio, dia_fim) dentro do mês (1..last_day) ou (None,None) se não intersecta."""
    if di is None or df is None:
        return None, None
    if di > dt_fim or df < dt_ini:
        return None, None

    ini = di if di >= dt_ini else dt_ini
    fim = df if df <= dt_fim else dt_fim
    return ini.day, fim.day


def _marcar_conflitos_por_face(ocupacoes_por_face):
    """
    ocupacoes_por_face: dict { "118AD": [ {DataInicio, DataFim, ...}, ... ] }
    Retorna set faces_em_conflito e também marca cada item com 'BitConflito' (True/False) se overlap existir.
    """
    faces_em_conflito = set()

    for face, itens in ocupacoes_por_face.items():
        if not itens or len(itens) <= 1:
            for it in itens:
                it["BitConflito"] = False
            continue

        itens_sorted = sorted(itens, key=lambda x: (x["DataInicio"], x["DataFim"]))
        conflito = False
        fim_atual = itens_sorted[0]["DataFim"]

        for i in range(1, len(itens_sorted)):
            ini = itens_sorted[i]["DataInicio"]
            fim = itens_sorted[i]["DataFim"]
            if ini <= fim_atual:
                conflito = True
                break
            if fim > fim_atual:
                fim_atual = fim

        if conflito:
            faces_em_conflito.add(face)

        for it in itens:
            it["BitConflito"] = conflito

    return faces_em_conflito




@paineis_bp.route("/", methods=["GET"])
@login_required
@limiter.limit("80 per minute", methods=["GET"])
@retry_get_view(db, attempts=6, base_delay=0.2, max_delay=1.5)
def lista_paineis():

    def _parse_iso_date(s: str):
        try:
            if not s:
                return None
            return datetime.strptime(s.strip(), "%Y-%m-%d").date()
        except:
            return None

    def _somente_digitos(s: str) -> str:
        try:
            if not s:
                return ""
            return "".join([c for c in str(s) if c.isdigit()])
        except:
            return ""

    def _parse_float_ptbr(s: str):
        try:
            if s is None:
                return None
            s = str(s).strip()
            if not s:
                return None

            s = s.replace(".", "").replace(",", ".") if ("," in s and "." in s) else s.replace(",", ".")
            return float(s)
        except:
            return None

    def _calcular_periodo_rapido(periodo_rapido: str, hoje_ref: date):
        try:
            p = (periodo_rapido or "").strip().lower()
            if not p:
                return None, None

            if p in ("dia", "hoje"):
                return hoje_ref, hoje_ref

            if p == "semana":
                inicio = hoje_ref - timedelta(days=hoje_ref.weekday())
                fim = inicio + timedelta(days=6)
                return inicio, fim

            if p == "mes":
                inicio = date(hoje_ref.year, hoje_ref.month, 1)
                fim = hoje_ref
                return inicio, fim

            if p == "ano":
                inicio = date(hoje_ref.year, 1, 1)
                fim = hoje_ref
                return inicio, fim

            return None, None
        except:
            return None, None

    def _parse_lista_tokens_busca(raw: str):
        try:
            s = (raw or "").strip()
            if not s:
                return {"tipo": None, "codpontos": [], "cnpjs": [], "codfaces": [], "erro": ""}

            parts = []
            for chunk in s.replace(",", " ").split():
                c = (chunk or "").strip()
                if c:
                    parts.append(c)

            if not parts:
                return {"tipo": None, "codpontos": [], "cnpjs": [], "codfaces": [], "erro": ""}

            tipos_encontrados = set()
            codpontos = []
            cnpjs = []
            codfaces = []

            for p in parts:
                p_str = str(p).strip()
                p_digits = _somente_digitos(p_str)

                if p_digits and p_digits == p_str:
                    if len(p_digits) == 14:
                        tipos_encontrados.add("cnpj")
                        cnpjs.append(p_digits)
                    else:
                        tipos_encontrados.add("codponto")
                        try:
                            codpontos.append(int(p_digits))
                        except:
                            return {
                                "tipo": None,
                                "codpontos": [],
                                "cnpjs": [],
                                "codfaces": [],
                                "erro": f"Token inválido para CodPonto: '{p_str}'",
                            }
                else:
                    tipos_encontrados.add("codface")
                    codfaces.append(p_str)

            if len(tipos_encontrados) > 1:
                return {
                    "tipo": None,
                    "codpontos": [],
                    "cnpjs": [],
                    "codfaces": [],
                    "erro": "Busca inválida: você misturou tipos (CodPonto / CodFace / CNPJ). Use apenas um tipo por vez.",
                }

            t = next(iter(tipos_encontrados)) if tipos_encontrados else None

            def _dedup(seq):
                seen = set()
                out = []
                for x in seq:
                    if x in seen:
                        continue
                    seen.add(x)
                    out.append(x)
                return out

            codpontos = _dedup(codpontos)
            cnpjs = _dedup(cnpjs)
            codfaces = _dedup(codfaces)

            return {"tipo": t, "codpontos": codpontos, "cnpjs": cnpjs, "codfaces": codfaces, "erro": ""}
        except:
            return {"tipo": None, "codpontos": [], "cnpjs": [], "codfaces": [], "erro": "Erro ao processar a busca."}

    def _getlist_limpo(chave: str):
        try:
            vals = request.args.getlist(chave)
            out = []
            seen = set()
            for v in vals:
                vv = (v or "").strip()
                if not vv:
                    continue
                if vv in seen:
                    continue
                seen.add(vv)
                out.append(vv)
            return out
        except:
            return []


    def _getlist_args_expandido(chave: str):
        try:
            valores = []
            for bruto in request.args.getlist(chave):
                if bruto is None:
                    continue

                s_bruto = str(bruto).strip()
                if not s_bruto:
                    continue

                partes = [s_bruto]
                if "|" in s_bruto:
                    partes = s_bruto.split("|")
                elif ";" in s_bruto:
                    partes = s_bruto.split(";")
                elif "," in s_bruto:
                    partes = s_bruto.split(",")

                for parte in partes:
                    vv = (parte or "").strip()
                    if vv:
                        valores.append(vv)

            out = []
            seen = set()
            for v in valores:
                chave_seen = v.upper()
                if chave_seen in seen:
                    continue
                seen.add(chave_seen)
                out.append(v)

            return out
        except:
            return []

    def _get_faces_selecionadas():
        try:
            codfaces_raw = _getlist_args_expandido("selected_codface")
            codpontos_raw = _getlist_args_expandido("selected_codponto")

            pares = []
            codfaces = []
            codpontos = []

            seen_pares = set()
            seen_faces = set()
            seen_pontos = set()

            if codfaces_raw and codpontos_raw and len(codpontos_raw) == len(codfaces_raw):
                for cp_raw, cf_raw in zip(codpontos_raw, codfaces_raw):
                    try:
                        cp_int = int(str(cp_raw).strip())
                    except:
                        continue

                    cf_txt = (str(cf_raw) or "").strip()
                    if not cf_txt:
                        continue

                    chave_par = (cp_int, cf_txt.upper())
                    if chave_par in seen_pares:
                        continue
                    seen_pares.add(chave_par)

                    pares.append((cp_int, cf_txt))

                    if cp_int not in seen_pontos:
                        seen_pontos.add(cp_int)
                        codpontos.append(cp_int)

                    chave_face = cf_txt.upper()
                    if chave_face not in seen_faces:
                        seen_faces.add(chave_face)
                        codfaces.append(cf_txt)
            else:
                for cf_raw in codfaces_raw:
                    cf_txt = (str(cf_raw) or "").strip()
                    if not cf_txt:
                        continue

                    chave_face = cf_txt.upper()
                    if chave_face in seen_faces:
                        continue
                    seen_faces.add(chave_face)
                    codfaces.append(cf_txt)

            return {
                "pares": pares,
                "codfaces": codfaces,
                "codfaces_norm": [x.upper() for x in codfaces],
                "codpontos": codpontos,
            }
        except:
            return {
                "pares": [],
                "codfaces": [],
                "codfaces_norm": [],
                "codpontos": [],
            }

    def _aplicar_filtro_faces_dim(query, pares_faces, codfaces_norm):
        try:
            if not pares_faces and not codfaces_norm:
                return query

            codface_dim_norm = func.upper(func.ltrim(func.rtrim(func.coalesce(DimFacesPaineis.CodFace, ""))))

            if pares_faces:
                filtros = []
                for cp_sel, cf_sel in pares_faces:
                    filtros.append(
                        and_(
                            DimFacesPaineis.CodPonto == cp_sel,
                            codface_dim_norm == str(cf_sel).strip().upper(),
                        )
                    )

                if filtros:
                    return query.filter(or_(*filtros))
                return query

            return query.filter(codface_dim_norm.in_(codfaces_norm))
        except:
            return query

    def _aplicar_filtro_faces_itens(query, pares_faces, codfaces_norm):
        try:
            if not pares_faces and not codfaces_norm:
                return query

            codface_item_norm = func.upper(
                func.ltrim(func.rtrim(func.coalesce(FatoControleContratosItensEuromidia.CodFace, "")))
            )

            if pares_faces:
                filtros = []
                for cp_sel, cf_sel in pares_faces:
                    filtros.append(
                        and_(
                            FatoControleContratosItensEuromidia.CodPonto == cp_sel,
                            codface_item_norm == str(cf_sel).strip().upper(),
                        )
                    )

                if filtros:
                    return query.filter(or_(*filtros))
                return query

            return query.filter(codface_item_norm.in_(codfaces_norm))
        except:
            return query

    def _resolver_periodo_por_dim_calendario(
        ano: str,
        anomes: str,
        semana_iso: str,
        quinzena: str,
        bisemana: str,
        fim_de_semana: str,
        dia_semana_iso: str,
        trimestre: str,
        mes: str,
        dia: str,
    ):
        try:
            def _to_int(v: str):
                vv = (v or "").strip()
                if not vv:
                    return None
                if vv.isdigit():
                    return int(vv)
                return None

            ano_i = _to_int(ano)
            anomes_i = _to_int(anomes)
            semana_i = _to_int(semana_iso)
            quinzena_i = _to_int(quinzena)
            bisemana_i = _to_int(bisemana)
            dia_semana_i = _to_int(dia_semana_iso)
            trimestre_i = _to_int(trimestre)
            mes_i = _to_int(mes)
            dia_i = _to_int(dia)

            fim_semana_flag = (fim_de_semana or "").strip()
            if fim_semana_flag not in ("", "0", "1"):
                fim_semana_flag = ""

            if not any(
                [
                    ano_i is not None,
                    anomes_i is not None,
                    semana_i is not None,
                    quinzena_i is not None,
                    bisemana_i is not None,
                    dia_semana_i is not None,
                    trimestre_i is not None,
                    mes_i is not None,
                    dia_i is not None,
                    fim_semana_flag in ("0", "1"),
                ]
            ):
                return None, None

            qcal = db.session.query(
                func.min(DimCalendario.data).label("dt_ini"),
                func.max(DimCalendario.data).label("dt_fim"),
            )

            if ano_i is not None:
                qcal = qcal.filter(DimCalendario.ano == ano_i)

            if anomes_i is not None:
                qcal = qcal.filter(DimCalendario.ano_mes == anomes_i)

            if mes_i is not None:
                qcal = qcal.filter(DimCalendario.mes == mes_i)

            if dia_i is not None:
                qcal = qcal.filter(DimCalendario.dia == dia_i)

            if trimestre_i is not None:
                qcal = qcal.filter(DimCalendario.trimestre == trimestre_i)

            if semana_i is not None:
                qcal = qcal.filter(DimCalendario.semana_ano_iso == semana_i)

            if dia_semana_i is not None:
                qcal = qcal.filter(DimCalendario.dia_semana_iso == dia_semana_i)

            if quinzena_i is not None:
                qcal = qcal.filter(DimCalendario.quinzena == quinzena_i)

            if bisemana_i is not None:
                qcal = qcal.filter(DimCalendario.bi_semana_numero == bisemana_i)

            if fim_semana_flag in ("0", "1"):
                qcal = qcal.filter(DimCalendario.eh_fim_de_semana == (fim_semana_flag == "1"))

            row = qcal.first()
            if not row:
                return None, None

            dt_ini = getattr(row, "dt_ini", None)
            dt_fim = getattr(row, "dt_fim", None)

            if dt_ini and dt_fim and dt_ini > dt_fim:
                dt_ini, dt_fim = dt_fim, dt_ini

            return dt_ini, dt_fim
        except:
            return None, None

    def _coerce_to_date(v):
        try:
            if v is None:
                return None
            if isinstance(v, datetime):
                return v.date()
            if isinstance(v, date):
                return v
            s = str(v).strip()
            if not s:
                return None

            try:
                return datetime.strptime(s, "%Y-%m-%d").date()
            except:
                pass

            try:
                return datetime.strptime(s, "%d/%m/%Y").date()
            except:
                pass

            try:
                return datetime.fromisoformat(s).date()
            except:
                return None
        except:
            return None

    def _fim_efetivo_item(df_prev, dc):
        try:
            dc_dt = _coerce_to_date(dc)
            if dc_dt:
                return dc_dt
            return _coerce_to_date(df_prev)
        except:
            return None

 
    def _slots_por_cota(cota):
        try:
            if cota is None:
                return 0.0

            if isinstance(cota, (int, float)):
                c = float(cota)
            else:
                s = str(cota).strip()
                if not s:
                    return 0.0
                d = _somente_digitos(s)
                if d:
                    c = float(d)
                else:
                    try:
                        c = float(s.replace(",", "."))
                    except:
                        return 0.0

            if c <= 0:
                return 0.0

            return float(1080.0 / float(c))
        except:
            return 0.0


    def _uso_no_periodo_por_intervalos(intervalos, dt_ini, dt_fim, denom_cap):
        try:
            if not intervalos or dt_ini is None or dt_fim is None:
                return 0.0, 0.0

            if dt_fim < dt_ini:
                dt_ini, dt_fim = dt_fim, dt_ini

            denom_cap = int(denom_cap or 0)
            denom_cap = max(1, denom_cap)

            eventos = []
            for di, df, slots in intervalos:
                if di is None or df is None:
                    continue

                di2 = di if di >= dt_ini else dt_ini
                df2 = df if df <= dt_fim else dt_fim

                if df2 < dt_ini or di2 > dt_fim:
                    continue
                if df2 < di2:
                    continue

                s = float(slots or 0.0)
                if s <= 0:
                    continue

                eventos.append((di2, +s))
                eventos.append((df2 + timedelta(days=1), -s))

            if not eventos:
                return 0.0, 0.0

            eventos.sort(key=lambda x: (x[0], 0 if x[1] < 0 else 1))

            atual = 0.0
            maximo = 0.0
            uso_slots_dias = 0.0

            cursor = dt_ini
            i = 0
            n = len(eventos)

            while i < n and eventos[i][0] < dt_ini:
                atual += float(eventos[i][1])
                if atual > maximo:
                    maximo = atual
                i += 1

            while cursor <= dt_fim:
                while i < n and eventos[i][0] == cursor:
                    atual += float(eventos[i][1])
                    if atual > maximo:
                        maximo = atual
                    i += 1

                proxima_data = dt_fim + timedelta(days=1)
                if i < n:
                    proxima_data = eventos[i][0]

                if proxima_data <= cursor:
                    cursor = cursor + timedelta(days=1)
                    continue

                fim_trecho_exclusivo = proxima_data
                if fim_trecho_exclusivo > (dt_fim + timedelta(days=1)):
                    fim_trecho_exclusivo = dt_fim + timedelta(days=1)

                dias_trecho = (fim_trecho_exclusivo - cursor).days
                if dias_trecho > 0:
                    ocupado = atual
                    if ocupado < 0:
                        ocupado = 0.0
                    if ocupado > float(denom_cap):
                        ocupado = float(denom_cap)

                    uso_slots_dias += float(ocupado) * float(dias_trecho)

                cursor = fim_trecho_exclusivo

            return float(uso_slots_dias), float(maximo)
        except:
            return 0.0, 0.0

    def _calcular_ocupacao_e_conflitos_por_face(
        ocupacoes,
        tipo_por_face,
        tipo_por_idcadastro,
        capacidade_digital_por_cp,
        dt_ini,
        dt_fim,
    ):
        try:
            por_face_tipo = {}

            for cp_int, cf_norm, cota, di, df_ef, idcad in (ocupacoes or []):
                if cp_int is None:
                    continue
                if not cf_norm:
                    continue

                di_dt = _coerce_to_date(di)
                df_dt = _coerce_to_date(df_ef)
                if di_dt is None or df_dt is None:
                    continue

                if dt_ini is not None and df_dt < dt_ini:
                    continue
                if dt_fim is not None and di_dt > dt_fim:
                    continue
                if df_dt < di_dt:
                    continue

                tp_up = tipo_por_face.get((cp_int, cf_norm))
                if (not tp_up) and (idcad not in (None, "")):
                    try:
                        tp_up = tipo_por_idcadastro.get(int(idcad))
                    except:
                        tp_up = None
                if not tp_up:
                    continue

                key = (cp_int, cf_norm, tp_up)

                slots = 1.0
                if tp_up == "PAINEL DIGITAL":
                    slots = _slots_por_cota(cota)

                por_face_tipo.setdefault(key, []).append((di_dt, df_dt, slots))

            out = {}

            if dt_ini is None or dt_fim is None:
                return out

            if dt_fim < dt_ini:
                dt_ini, dt_fim = dt_fim, dt_ini

            total_dias = int((dt_fim - dt_ini).days) + 1
            if total_dias <= 0:
                total_dias = 1

            for (cp_int, cf_norm, tp_up), intervalos in por_face_tipo.items():
                denom = 1
                if tp_up == "PAINEL DIGITAL":
                    denom = int(capacidade_digital_por_cp.get(cp_int) or 0) or int(CAPACIDADE_DIGITAL_FIXA)
                    denom = max(1, int(denom))

                uso_slots_dias, max_simult_sem_teto = _uso_no_periodo_por_intervalos(
                    intervalos=intervalos,
                    dt_ini=dt_ini,
                    dt_fim=dt_fim,
                    denom_cap=denom,
                )

                conflitos = 1 if (float(max_simult_sem_teto) > float(denom)) else 0

                uso_medio = float(uso_slots_dias) / float(total_dias)

                cap_total = float(denom) * float(total_dias)
                pct_uso = int(round((float(uso_slots_dias) / cap_total) * 100.0, 0)) if cap_total > 0 else 0
                if pct_uso < 0:
                    pct_uso = 0
                if pct_uso > 100:
                    pct_uso = 100

                ocupadas_vis = int(round(uso_medio, 0))
                if ocupadas_vis < 0:
                    ocupadas_vis = 0
                if ocupadas_vis > denom:
                    ocupadas_vis = denom

                out[(cp_int, cf_norm, tp_up)] = {
                    "ocupadas": int(ocupadas_vis),
                    "denominador": int(denom),
                    "conflitos": int(conflitos),
                    "pct_uso": int(pct_uso),
                    "uso_medio": float(uso_medio),
                    "uso_slots_dias": float(uso_slots_dias),
                    "total_dias": int(total_dias),
                    "max_simult_sem_teto": float(max_simult_sem_teto),
                }

            return out
        except:
            return {}

    def _is_admin() -> bool:
        try:
            if getattr(current_user, "has_permission", None):
                if current_user.has_permission("ADMIN_TUDO"):
                    return True
                if current_user.has_permission("PAINEIS_VER_EXIBIDORA"):
                    return True
                if current_user.has_permission("ADMIN"):
                    return True

            perfil = getattr(current_user, "perfil", None)
            nome_perfil = (getattr(perfil, "NomePerfil", "") or "").strip().upper()

            if nome_perfil in ("ADMIN", "ADMINISTRADOR", "ADM", "ADMIN_TUDO"):
                return True

            return False
        except Exception:
            return False

    hoje = date.today()
    primeiro_dia_mes_atual = date(hoje.year, hoje.month, 1)

    q = (request.args.get("q") or "").strip()

    faces_selecionadas_info = _get_faces_selecionadas()
    selected_face_pairs = faces_selecionadas_info["pares"]
    selected_codfaces = faces_selecionadas_info["codfaces"]
    selected_codfaces_norm = faces_selecionadas_info["codfaces_norm"]
    selected_codpontos = faces_selecionadas_info["codpontos"]
    tem_faces_selecionadas = bool(selected_face_pairs or selected_codfaces_norm)

    tipo_list = _getlist_limpo("tipo")
    exibidora_list = _getlist_limpo("exibidora")
    cidade_exibicao_list = _getlist_limpo("cidade_exibicao")
    tipo_documento_list = _getlist_limpo("tipo_documento")
    vendedor_list = _getlist_limpo("vendedor")
    marca_exibida_list = _getlist_limpo("marca_exibida")
    status_list = _getlist_limpo("status")


    formato_list = _getlist_limpo("formato")

    tipo_single = (request.args.get("tipo") or "").strip()
    exibidora_single = (request.args.get("exibidora") or "").strip()
    cidade_exibicao_single = (request.args.get("cidade_exibicao") or "").strip()
    tipo_documento_single = (request.args.get("tipo_documento") or "").strip()
    vendedor_single = (request.args.get("vendedor") or "").strip()
    marca_exibida_single = (request.args.get("marca_exibida") or "").strip()
    status_single = (request.args.get("status") or "").strip()
    formato_single = (request.args.get("formato") or "").strip()

    pode_ver_exibidora = _is_admin()
    if not pode_ver_exibidora:
        exibidora_list = []
        exibidora_single = ""

    if tipo_single and (tipo_single not in tipo_list):
        tipo_list.append(tipo_single)
    if exibidora_single and (exibidora_single not in exibidora_list):
        exibidora_list.append(exibidora_single)
    if cidade_exibicao_single and (cidade_exibicao_single not in cidade_exibicao_list):
        cidade_exibicao_list.append(cidade_exibicao_single)
    if tipo_documento_single and (tipo_documento_single not in tipo_documento_list):
        tipo_documento_list.append(tipo_documento_single)
    if vendedor_single and (vendedor_single not in vendedor_list):
        vendedor_list.append(vendedor_single)
    if marca_exibida_single and (marca_exibida_single not in marca_exibida_list):
        marca_exibida_list.append(marca_exibida_single)
    if status_single and (status_single not in status_list):
        status_list.append(status_single)
    if formato_single and (formato_single not in formato_list):
        formato_list.append(formato_single)


    altxlarg_legacy = (request.args.get("altxlarg") or "").strip()
    if altxlarg_legacy and (altxlarg_legacy not in formato_list):
        formato_list.append(altxlarg_legacy)

  
    altxlarg = ""
    if len(formato_list) == 1:
        altxlarg = formato_list[0]
    elif altxlarg_legacy:
        altxlarg = altxlarg_legacy

    iluminado = (request.args.get("iluminado") or "").strip()

    area_min_str = (request.args.get("area_min") or "").strip()
    area_max_str = (request.args.get("area_max") or "").strip()
    area_min = _parse_float_ptbr(area_min_str)
    area_max = _parse_float_ptbr(area_max_str)

    ponto_ativo_raw = request.args.get("ponto_ativo", None)
    ponto_ativo = (ponto_ativo_raw or "").strip().lower()

    if ponto_ativo in ("todos", "all"):
        ponto_ativo = ""
    elif ponto_ativo not in ("", "0", "1"):
        ponto_ativo = ""

    if ponto_ativo_raw is None:
        ponto_ativo = "1"

    dt_ini_str = (request.args.get("dt_ini") or "").strip()
    dt_fim_str = (request.args.get("dt_fim") or "").strip()

    mes_atual_flag = (request.args.get("mes_atual") or "").strip()
    tudo_flag = (request.args.get("tudo") or "").strip()

    mes_atual_user = mes_atual_flag == "1"
    tudo = tudo_flag == "1"

    periodo_rapido = (request.args.get("periodo") or "").strip().lower()

    ano_filtro = (request.args.get("ano") or "").strip()
    anomes_filtro = (request.args.get("anomes") or "").strip()
    mes_filtro = (request.args.get("mes") or "").strip()
    dia_filtro = (request.args.get("dia") or "").strip()
    trimestre_filtro = (request.args.get("trimestre") or "").strip()
    semana_iso_filtro = (request.args.get("semana_iso") or "").strip()
    dia_semana_iso_filtro = (request.args.get("dia_semana_iso") or "").strip()
    quinzena_filtro = (request.args.get("quinzena") or "").strip()

    # ✅ NOVO: aceita o novo nome bi_semana e mantém compatibilidade com bisemana
    bi_semana_filtro = (request.args.get("bi_semana") or request.args.get("bisemana") or "").strip()

    fim_de_semana_filtro = (request.args.get("fim_de_semana") or "").strip()

    try:
        chaves_args_reais = [
            k for k in request.args.keys()
            if k not in ("selected_codface", "selected_codponto")
        ]
        sem_args_reais = (len(chaves_args_reais) == 0)
    except:
        sem_args_reais = False

    if sem_args_reais:
        mes_atual_user = True

    try:
        page = int(request.args.get("page") or "1")
    except:
        page = 1

    try:
        per_page = int(request.args.get("per_page") or "20")
    except:
        per_page = 20

    per_page = max(5, min(per_page, 100))
    page = max(1, page)


    if tudo:
        page = 1
        ponto_ativo = ""
        dt_ini_filtro = None
        dt_fim_filtro = None
        dt_ini_str = ""
        dt_fim_str = ""
        mes_atual_user = False
        periodo_rapido = ""

        filtro_ativo_cancelamento = (FatoControleContratosItensEuromidia.AtivoCancelamento == "A")

        fim_efetivo_global_sql = func.coalesce(
            FatoControleContratosItensEuromidia.DataCancelamento,
            FatoControleContratosItensEuromidia.DataTerminoPrevisto,
        )

        agg_periodo_q = (
            db.session.query(
                func.min(FatoControleContratosItensEuromidia.DataInicioPrevisto).label("dt_min"),
                func.max(fim_efetivo_global_sql).label("dt_max"),
            )
            .filter(
                FatoControleContratosItensEuromidia.DataInicioPrevisto != None,
                filtro_ativo_cancelamento,
            )
        )

        agg_periodo_q = _aplicar_filtro_faces_itens(
            agg_periodo_q,
            pares_faces=selected_face_pairs,
            codfaces_norm=selected_codfaces_norm,
        )

        agg_periodo = agg_periodo_q.first()

        dt_min_base = getattr(agg_periodo, "dt_min", None) if agg_periodo else None
        dt_max_base = getattr(agg_periodo, "dt_max", None) if agg_periodo else None

        dt_ini_calc = dt_min_base or primeiro_dia_mes_atual
        dt_fim_calc = dt_max_base or hoje

        if dt_ini_calc and dt_fim_calc and dt_ini_calc > dt_fim_calc:
            dt_ini_calc, dt_fim_calc = dt_fim_calc, dt_ini_calc

        dt_ini_ocup = dt_ini_calc
        dt_fim_ocup = dt_fim_calc

    else:
        dt_ini_dim, dt_fim_dim = _resolver_periodo_por_dim_calendario(
            ano=ano_filtro,
            anomes=anomes_filtro,
            semana_iso=semana_iso_filtro,
            quinzena=quinzena_filtro,
            bisemana=bi_semana_filtro,
            fim_de_semana=fim_de_semana_filtro,
            dia_semana_iso=dia_semana_iso_filtro,
            trimestre=trimestre_filtro,
            mes=mes_filtro,
            dia=dia_filtro,
        )

        if dt_ini_dim and dt_fim_dim:
            dt_ini_user = dt_ini_dim
            dt_fim_user = dt_fim_dim

            mes_atual_user = False
            periodo_rapido = ""

            dt_ini_str = dt_ini_user.strftime("%Y-%m-%d")
            dt_fim_str = dt_fim_user.strftime("%Y-%m-%d")
        else:
            dt_ini_rap, dt_fim_rap = _calcular_periodo_rapido(periodo_rapido, hoje)

            if dt_ini_rap and dt_fim_rap:
                dt_ini_user = dt_ini_rap
                dt_fim_user = dt_fim_rap

                mes_atual_user = False

                dt_ini_str = dt_ini_user.strftime("%Y-%m-%d")
                dt_fim_str = dt_fim_user.strftime("%Y-%m-%d")
            else:
                dt_ini_user = _parse_iso_date(dt_ini_str)
                dt_fim_user = _parse_iso_date(dt_fim_str)

        dt_ini_ocup = primeiro_dia_mes_atual
        dt_fim_ocup = hoje

        user_definiu_periodo = False

        if mes_atual_user:
            dt_ini_filtro = primeiro_dia_mes_atual
            dt_fim_filtro = hoje
            user_definiu_periodo = True

            dt_ini_str = dt_ini_filtro.strftime("%Y-%m-%d")
            dt_fim_str = dt_fim_filtro.strftime("%Y-%m-%d")

            periodo_rapido = ""

        else:
            if dt_ini_user or dt_fim_user:
                dt_ini_calc = dt_ini_user or dt_fim_user or primeiro_dia_mes_atual
                dt_fim_calc = dt_fim_user or dt_ini_user or hoje

                if dt_ini_calc and dt_fim_calc and dt_ini_calc > dt_fim_calc:
                    dt_ini_calc, dt_fim_calc = dt_fim_calc, dt_ini_calc

                dt_ini_filtro = dt_ini_calc
                dt_fim_filtro = dt_fim_calc
                user_definiu_periodo = True

                dt_ini_str = dt_ini_filtro.strftime("%Y-%m-%d")
                dt_fim_str = dt_fim_filtro.strftime("%Y-%m-%d")

            else:
                dt_ini_filtro = None
                dt_fim_filtro = None
                dt_ini_str = ""
                dt_fim_str = ""
                mes_atual_user = False
                periodo_rapido = ""

        if user_definiu_periodo and dt_ini_filtro and dt_fim_filtro:
            dt_ini_ocup = dt_ini_filtro
            dt_fim_ocup = dt_fim_filtro

        dt_ini_calc = dt_ini_ocup
        dt_fim_calc = dt_fim_ocup


    bisemanas_select = []
    try:
        dt_ini_base_bi = dt_ini_calc or primeiro_dia_mes_atual
        dt_fim_base_bi = dt_fim_calc or hoje

        if dt_ini_base_bi and dt_fim_base_bi and dt_ini_base_bi > dt_fim_base_bi:
            dt_ini_base_bi, dt_fim_base_bi = dt_fim_base_bi, dt_ini_base_bi

        bis_raw = (
            db.session.query(
                DimCalendario.bi_semana_numero,
                DimCalendario.inicio_bi_semana,
                DimCalendario.fim_bi_semana,
            )
            .filter(
                DimCalendario.bi_semana_numero != None,
                DimCalendario.inicio_bi_semana != None,
                DimCalendario.fim_bi_semana != None,
                DimCalendario.inicio_bi_semana <= dt_fim_base_bi,
                DimCalendario.fim_bi_semana >= dt_ini_base_bi,
            )
            .group_by(
                DimCalendario.bi_semana_numero,
                DimCalendario.inicio_bi_semana,
                DimCalendario.fim_bi_semana,
            )
            .order_by(DimCalendario.inicio_bi_semana.asc())
            .all()
        )

        vistos_bi = set()
        for n, di_bi, df_bi in (bis_raw or []):
            if n is None or di_bi is None or df_bi is None:
                continue

            try:
                value = str(int(n))
            except:
                value = str(n).strip()

            chave_bi = (value, di_bi, df_bi)
            if chave_bi in vistos_bi:
                continue
            vistos_bi.add(chave_bi)

            bisemanas_select.append({
                "value": value,
                "label": f"{value} — {di_bi.strftime('%d/%m')} a {df_bi.strftime('%d/%m')}",
                "inicio": di_bi.strftime("%Y-%m-%d"),
                "fim": df_bi.strftime("%Y-%m-%d"),
            })
    except:
        bisemanas_select = []


    tipo_face_norm = func.upper(func.ltrim(func.rtrim(func.coalesce(DimFacesPaineis.Tipo, ""))))
    tipo_painel_norm = func.upper(func.ltrim(func.rtrim(func.coalesce(DimPaineisEuromidia.Tipo, ""))))

    rn_painel_tipo = func.row_number().over(
        partition_by=(DimPaineisEuromidia.CodPonto, tipo_painel_norm),
        order_by=(DimPaineisEuromidia.DataAtualizacao.desc(), DimPaineisEuromidia.IDDimPaineisEuromidia.desc()),
    ).label("rn")

    sub_ultimo_painel_tipo = (
        db.session.query(
            DimPaineisEuromidia.CodPonto.label("CodPonto"),
            tipo_painel_norm.label("TipoNorm"),
            DimPaineisEuromidia.Tipo.label("Tipo"),
            DimPaineisEuromidia.Exibidora.label("Exibidora"),
            DimPaineisEuromidia.Logradouro.label("Logradouro"),
            DimPaineisEuromidia.UF.label("UF"),
            DimPaineisEuromidia.Cidade.label("Cidade"),

           
            DimPaineisEuromidia.Bairro.label("Bairro"),
            DimPaineisEuromidia.Referencia.label("Referencia"),

            DimPaineisEuromidia.FormatoLxA.label("FormatoLxA"),
            DimPaineisEuromidia.AreaTotalm.label("AreaTotalm"),
            DimPaineisEuromidia.BitEnergia.label("BitEnergia"),
            DimPaineisEuromidia.BitInternet.label("BitInternet"),
            DimPaineisEuromidia.BitAtivo.label("BitAtivo"),
            DimPaineisEuromidia.BitAluguel.label("BitAluguel"),
            DimPaineisEuromidia.BitIluminado.label("BitIluminado"),
            DimPaineisEuromidia.QuantidadeFaces.label("QuantidadeFaces"),
            rn_painel_tipo,
        )
        .filter(DimPaineisEuromidia.CodPonto != None)
        .subquery()
    )

    ultimo_painel = (
        db.session.query(sub_ultimo_painel_tipo)
        .filter(sub_ultimo_painel_tipo.c.rn == 1)
        .subquery()
    )


    tipos = (
        db.session.query(func.upper(func.trim(DimFacesPaineis.Tipo)).label("Tipo"))
        .filter(DimFacesPaineis.Tipo != None, func.trim(DimFacesPaineis.Tipo) != "")
        .distinct()
        .order_by(func.upper(func.trim(DimFacesPaineis.Tipo)).asc())
        .all()
    )
    tipos = [r[0] for r in tipos if r and r[0]]

    if pode_ver_exibidora:
        exibidoras = (
            db.session.query(func.trim(ultimo_painel.c.Exibidora).label("Exibidora"))
            .filter(ultimo_painel.c.Exibidora != None, func.trim(ultimo_painel.c.Exibidora) != "")
            .distinct()
            .order_by(func.trim(ultimo_painel.c.Exibidora).asc())
            .all()
        )
        exibidoras = [r[0] for r in exibidoras if r and r[0]]
    else:
        exibidoras = []

    altxlarg_opcoes = (
        db.session.query(func.trim(ultimo_painel.c.FormatoLxA).label("FormatoLxA"))
        .filter(ultimo_painel.c.FormatoLxA != None, func.trim(ultimo_painel.c.FormatoLxA) != "")
        .distinct()
        .order_by(func.trim(ultimo_painel.c.FormatoLxA).asc())
        .all()
    )
    altxlarg_opcoes = [r[0] for r in altxlarg_opcoes if r and r[0]]

    iluminado_opcoes_raw = (
        db.session.query(ultimo_painel.c.BitIluminado)
        .filter(ultimo_painel.c.BitIluminado != None)
        .distinct()
        .order_by(ultimo_painel.c.BitIluminado.asc())
        .all()
    )
    iluminado_opcoes = []
    for r in iluminado_opcoes_raw:
        if not r:
            continue
        val = r[0]
        if val is None:
            continue
        iluminado_opcoes.append("1" if bool(val) else "0")
    iluminado_opcoes = sorted(list(set(iluminado_opcoes)))

    areas_total = (
        db.session.query(ultimo_painel.c.AreaTotalm)
        .filter(ultimo_painel.c.AreaTotalm != None)
        .distinct()
        .order_by(ultimo_painel.c.AreaTotalm.asc())
        .all()
    )
    areas_total = [float(r[0]) for r in areas_total if r and r[0] is not None]

    area_min_max = (
        db.session.query(
            func.min(ultimo_painel.c.AreaTotalm).label("min_area"),
            func.max(ultimo_painel.c.AreaTotalm).label("max_area"),
        )
        .filter(ultimo_painel.c.AreaTotalm != None)
        .first()
    )
    area_total_min_global = getattr(area_min_max, "min_area", None) if area_min_max else None
    area_total_max_global = getattr(area_min_max, "max_area", None) if area_min_max else None

    if (area_min is not None) and (area_max is not None) and (area_min > area_max):
        area_min, area_max = area_max, area_min
        area_min_str, area_max_str = area_max_str, area_min_str

    filtro_ativo_cancelamento = (FatoControleContratosItensEuromidia.AtivoCancelamento == "A")

    fim_efetivo_sql = func.coalesce(
        FatoControleContratosItensEuromidia.DataCancelamento,
        FatoControleContratosItensEuromidia.DataTerminoPrevisto,
        date(9999, 12, 31),
    )

    if tudo:
        filtro_periodo_itens_ocup = and_(
            FatoControleContratosItensEuromidia.DataInicioPrevisto != None,
            filtro_ativo_cancelamento,
        )
    else:
        filtro_periodo_itens_ocup = and_(
            FatoControleContratosItensEuromidia.DataInicioPrevisto != None,
            FatoControleContratosItensEuromidia.DataInicioPrevisto <= dt_fim_ocup,
            fim_efetivo_sql >= dt_ini_ocup,
            filtro_ativo_cancelamento,
        )

    if tudo:
        filtro_periodo_itens_facet = and_(
            FatoControleContratosItensEuromidia.DataInicioPrevisto != None,
            filtro_ativo_cancelamento,
        )
    else:
        if (dt_ini_filtro is not None) and (dt_fim_filtro is not None):
            filtro_periodo_itens_facet = and_(
                FatoControleContratosItensEuromidia.DataInicioPrevisto != None,
                FatoControleContratosItensEuromidia.DataInicioPrevisto <= dt_fim_filtro,
                fim_efetivo_sql >= dt_ini_filtro,
                filtro_ativo_cancelamento,
            )
        else:
            filtro_periodo_itens_facet = and_(
                FatoControleContratosItensEuromidia.DataInicioPrevisto != None,
                filtro_ativo_cancelamento,
            )

    if tudo:
        cidades_exibicao = []
        tipos_documento = []
        vendedores_contrato = []
        marcas_exibidas = []
    else:
        base_filtro_itens_periodo_q = (
            db.session.query(FatoControleContratosItensEuromidia.IDFatoControleContratosItensEuromidia.label("ID"))
            .filter(filtro_periodo_itens_facet)
        )

        base_filtro_itens_periodo_q = _aplicar_filtro_faces_itens(
            base_filtro_itens_periodo_q,
            pares_faces=selected_face_pairs,
            codfaces_norm=selected_codfaces_norm,
        )

        base_filtro_itens_periodo_sq = base_filtro_itens_periodo_q.subquery()
        ids_itens_periodo_select = select(base_filtro_itens_periodo_sq.c.ID)

        cidades_exibicao = (
            db.session.query(FatoControleContratosItensEuromidia.CidadeExibicao)
            .filter(
                FatoControleContratosItensEuromidia.CidadeExibicao != None,
                FatoControleContratosItensEuromidia.CidadeExibicao != "",
                FatoControleContratosItensEuromidia.IDFatoControleContratosItensEuromidia.in_(ids_itens_periodo_select),
            )
            .distinct()
            .order_by(FatoControleContratosItensEuromidia.CidadeExibicao.asc())
            .all()
        )
        cidades_exibicao = [r[0] for r in cidades_exibicao if r and r[0]]

        tipos_documento = (
            db.session.query(FatoControleContratosItensEuromidia.TipoDocumento)
            .filter(
                FatoControleContratosItensEuromidia.TipoDocumento != None,
                FatoControleContratosItensEuromidia.TipoDocumento != "",
                FatoControleContratosItensEuromidia.IDFatoControleContratosItensEuromidia.in_(ids_itens_periodo_select),
            )
            .distinct()
            .order_by(FatoControleContratosItensEuromidia.TipoDocumento.asc())
            .all()
        )
        tipos_documento = [r[0] for r in tipos_documento if r and r[0]]

        vendedores_contrato = (
            db.session.query(FatoControleContratosItensEuromidia.Vendedor)
            .filter(
                FatoControleContratosItensEuromidia.Vendedor != None,
                FatoControleContratosItensEuromidia.Vendedor != "",
                FatoControleContratosItensEuromidia.IDFatoControleContratosItensEuromidia.in_(ids_itens_periodo_select),
            )
            .distinct()
            .order_by(FatoControleContratosItensEuromidia.Vendedor.asc())
            .all()
        )
        vendedores_contrato = [r[0] for r in vendedores_contrato if r and r[0]]

        marca_escolhida_sql = func.nullif(FatoControleContratosItensEuromidia.MarcaExibida, "")
        marca_label = marca_escolhida_sql.label("Marca")

        marcas_exibidas = (
            db.session.query(marca_label)
            .filter(
                marca_label != None,
                marca_label != "",
                FatoControleContratosItensEuromidia.IDFatoControleContratosItensEuromidia.in_(ids_itens_periodo_select),
            )
            .distinct()
            .order_by(marca_label.asc())
            .all()
        )
        marcas_exibidas = [r[0] for r in marcas_exibidas if r and r[0]]

 
    base_q = (
        db.session.query(
            DimFacesPaineis.CodFace.label("CodFace"),
            DimFacesPaineis.CodPonto.label("CodPonto"),
            DimFacesPaineis.IDDimFacesPaineis.label("IDDimFacesPaineis"),
            func.upper(func.trim(DimFacesPaineis.Tipo)).label("TipoProd"),

            func.trim(ultimo_painel.c.Exibidora).label("EXIBIDORA"),
            func.trim(ultimo_painel.c.Logradouro).label("Logradouro"),
            ultimo_painel.c.QuantidadeFaces.label("NumFacesMax"),
            func.trim(ultimo_painel.c.UF).label("UF"),

            # ✅ NOVO
            func.trim(ultimo_painel.c.Bairro).label("Bairro"),
            func.trim(ultimo_painel.c.Referencia).label("Referencia"),

            cast(ultimo_painel.c.BitEnergia, db.Integer).label("BitEnergia"),
            cast(ultimo_painel.c.BitInternet, db.Integer).label("BitInternet"),
            cast(ultimo_painel.c.BitAtivo, db.Integer).label("BitAtivo"),
            cast(ultimo_painel.c.BitAluguel, db.Integer).label("BitAluguel"),

            func.trim(ultimo_painel.c.FormatoLxA).label("FormatoLxA"),
            ultimo_painel.c.AreaTotalm.label("AreaTotalm"),
            cast(ultimo_painel.c.BitIluminado, db.Integer).label("BitIluminado"),

            func.trim(ultimo_painel.c.Cidade).label("Municipio"),
            func.trim(ultimo_painel.c.UF).label("UF_COL"),
            func.trim(ultimo_painel.c.FormatoLxA).label("Formato"),
            ultimo_painel.c.QuantidadeFaces.label("NumFacesCadastro"),
        )
        .select_from(DimFacesPaineis)
        .join(
            ultimo_painel,
            and_(
                ultimo_painel.c.CodPonto == DimFacesPaineis.CodPonto,
                ultimo_painel.c.TipoNorm == tipo_face_norm,
            ),
        )
        .filter(DimFacesPaineis.CodFace != None, func.trim(DimFacesPaineis.CodFace) != "")
    )

    base_q = _aplicar_filtro_faces_dim(
        base_q,
        pares_faces=selected_face_pairs,
        codfaces_norm=selected_codfaces_norm,
    )

    if tipo_list:
        tipo_list_norm = [str(x).strip().upper() for x in tipo_list if str(x).strip()]
        base_q = base_q.filter(func.upper(func.trim(DimFacesPaineis.Tipo)).in_(tipo_list_norm))

    if pode_ver_exibidora and exibidora_list:
        base_q = base_q.filter(func.trim(ultimo_painel.c.Exibidora).in_(exibidora_list))

    # =========================================================
    # NOVO: filtro multi de formato
    # =========================================================
    if formato_list:
        formato_list_limpo = [str(x).strip() for x in formato_list if str(x).strip()]
        if formato_list_limpo:
            base_q = base_q.filter(func.trim(ultimo_painel.c.FormatoLxA).in_(formato_list_limpo))

    if iluminado in ("0", "1"):
        base_q = base_q.filter(ultimo_painel.c.BitIluminado == (iluminado == "1"))

    if area_min is not None:
        base_q = base_q.filter(ultimo_painel.c.AreaTotalm != None, ultimo_painel.c.AreaTotalm >= area_min)

    if area_max is not None:
        base_q = base_q.filter(ultimo_painel.c.AreaTotalm != None, ultimo_painel.c.AreaTotalm <= area_max)

    if ponto_ativo in ("0", "1"):
        base_q = base_q.filter(ultimo_painel.c.BitAtivo == (ponto_ativo == "1"))

    busca_info = _parse_lista_tokens_busca(q)
    busca_erro = (busca_info.get("erro") or "").strip()

    if q and (not busca_erro):
        t = busca_info.get("tipo")
        codpontos_busca = busca_info.get("codpontos") or []
        codfaces_busca = busca_info.get("codfaces") or []
        cnpjs_busca = busca_info.get("cnpjs") or []

        filtros_busca = []

        if t == "codponto":
            if codpontos_busca:
                filtros_busca.append(DimFacesPaineis.CodPonto.in_(codpontos_busca))

        elif t == "codface":
            if codfaces_busca:
                filtros_busca.append(DimFacesPaineis.CodFace.in_(codfaces_busca))

        elif t == "cnpj":
            if cnpjs_busca:
                cnpj_sql_normalizado = func.replace(
                    func.replace(func.replace(FatoControleContratosItensEuromidia.CNPJ, ".", ""), "/", ""),
                    "-",
                    "",
                )

                sub_faces_cnpj_q = (
                    db.session.query(FatoControleContratosItensEuromidia.CodFace)
                    .filter(
                        FatoControleContratosItensEuromidia.CodFace != None,
                        filtro_periodo_itens_facet,
                        cnpj_sql_normalizado.in_(cnpjs_busca),
                    )
                )

                sub_faces_cnpj_q = _aplicar_filtro_faces_itens(
                    sub_faces_cnpj_q,
                    pares_faces=selected_face_pairs,
                    codfaces_norm=selected_codfaces_norm,
                )

                sub_faces_cnpj = sub_faces_cnpj_q.distinct().subquery()
                filtros_busca.append(DimFacesPaineis.CodFace.in_(select(sub_faces_cnpj.c.CodFace)))

        if filtros_busca:
            base_q = base_q.filter(or_(*filtros_busca))

    if cidade_exibicao_list or tipo_documento_list or vendedor_list or marca_exibida_list:
        marca_escolhida_sql = func.nullif(FatoControleContratosItensEuromidia.MarcaExibida, "")

        sub_faces = (
            db.session.query(FatoControleContratosItensEuromidia.CodFace)
            .filter(FatoControleContratosItensEuromidia.CodFace != None, filtro_periodo_itens_facet)
        )

        sub_faces = _aplicar_filtro_faces_itens(
            sub_faces,
            pares_faces=selected_face_pairs,
            codfaces_norm=selected_codfaces_norm,
        )

        if cidade_exibicao_list:
            sub_faces = sub_faces.filter(FatoControleContratosItensEuromidia.CidadeExibicao.in_(cidade_exibicao_list))

        if tipo_documento_list:
            sub_faces = sub_faces.filter(FatoControleContratosItensEuromidia.TipoDocumento.in_(tipo_documento_list))

        if vendedor_list:
            sub_faces = sub_faces.filter(FatoControleContratosItensEuromidia.Vendedor.in_(vendedor_list))

        if marca_exibida_list:
            sub_faces = sub_faces.filter(marca_escolhida_sql.in_(marca_exibida_list))

        sub_faces = sub_faces.distinct().subquery()
        base_q = base_q.filter(DimFacesPaineis.CodFace.in_(select(sub_faces.c.CodFace)))

    total = base_q.count()

    rows = (
        base_q
        .order_by(
            func.trim(ultimo_painel.c.Cidade).asc(),
            DimFacesPaineis.CodFace.asc(),
            func.upper(func.trim(DimFacesPaineis.Tipo)).asc(),
        )
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    codpontos_pagina = sorted(list({int(r.CodPonto) for r in rows if r and r.CodPonto is not None}))
    codfaces_pagina = sorted(list({(str(r.CodFace) or "").strip() for r in rows if r and (r.CodFace is not None)}))
    ids_faces_pagina = sorted(list({int(r.IDDimFacesPaineis) for r in rows if r and getattr(r, "IDDimFacesPaineis", None) is not None}))

    def _montar_placeholders_sql(prefixo: str, valores: list, parametros: dict):
        placeholders = []
        for idx, valor in enumerate(valores or []):
            nome_param = f"{prefixo}_{idx}"
            placeholders.append(f":{nome_param}")
            parametros[nome_param] = valor
        return ", ".join(placeholders)

    mapa_url_imagem_por_id_face = {}
    mapa_url_imagem_por_codface = {}

    try:
        if ids_faces_pagina:
            parametros_ids = {}
            placeholders_ids = _montar_placeholders_sql("id_face", ids_faces_pagina, parametros_ids)

            sql_imagem_por_id_face = text(f"""
                WITH imagens_ordenadas AS (
                    SELECT
                        img.IDDimFacesPaineis,
                        img.UrlImagem,
                        ROW_NUMBER() OVER (
                            PARTITION BY img.IDDimFacesPaineis
                            ORDER BY img.DataAtualizacao DESC, img.IDDimImagemPainel DESC
                        ) AS rn
                    FROM [Integracao].[Silver].[DimImagemPainel] AS img
                    WHERE
                        img.BitAtivo = 1
                        AND img.NumeroImagem = 1
                        AND img.IDDimFacesPaineis IN ({placeholders_ids})
                )
                SELECT
                    IDDimFacesPaineis,
                    UrlImagem
                FROM imagens_ordenadas
                WHERE rn = 1
            """)

            rows_imagem_por_id = db.session.execute(sql_imagem_por_id_face, parametros_ids).mappings().all()
            for row_img in rows_imagem_por_id:
                try:
                    id_face_img = int(row_img.get("IDDimFacesPaineis") or 0)
                except:
                    id_face_img = 0

                url_img = (row_img.get("UrlImagem") or "").strip()
                if id_face_img > 0 and url_img:
                    mapa_url_imagem_por_id_face[id_face_img] = url_img

        codfaces_sem_imagem_por_id = []
        vistos_codfaces_sem_imagem = set()

        for r_img in rows:
            codface_tmp = (str(getattr(r_img, "CodFace", "") or "")).strip()
            if not codface_tmp:
                continue

            try:
                id_face_tmp = int(getattr(r_img, "IDDimFacesPaineis", None) or 0)
            except:
                id_face_tmp = 0

            if id_face_tmp > 0 and id_face_tmp in mapa_url_imagem_por_id_face:
                continue

            chave_cf_tmp = codface_tmp.upper()
            if chave_cf_tmp in vistos_codfaces_sem_imagem:
                continue

            vistos_codfaces_sem_imagem.add(chave_cf_tmp)
            codfaces_sem_imagem_por_id.append(chave_cf_tmp)

        if codfaces_sem_imagem_por_id:
            parametros_codfaces = {}
            placeholders_codfaces = _montar_placeholders_sql("codface_img", codfaces_sem_imagem_por_id, parametros_codfaces)

            sql_imagem_por_codface = text(f"""
                WITH imagens_ordenadas AS (
                    SELECT
                        UPPER(LTRIM(RTRIM(ISNULL(img.CodFace, '')))) AS CodFaceNormalizada,
                        img.UrlImagem,
                        ROW_NUMBER() OVER (
                            PARTITION BY UPPER(LTRIM(RTRIM(ISNULL(img.CodFace, ''))))
                            ORDER BY img.DataAtualizacao DESC, img.IDDimImagemPainel DESC
                        ) AS rn
                    FROM [Integracao].[Silver].[DimImagemPainel] AS img
                    WHERE
                        img.BitAtivo = 1
                        AND img.NumeroImagem = 1
                        AND UPPER(LTRIM(RTRIM(ISNULL(img.CodFace, '')))) IN ({placeholders_codfaces})
                )
                SELECT
                    CodFaceNormalizada,
                    UrlImagem
                FROM imagens_ordenadas
                WHERE rn = 1
            """)

            rows_imagem_por_codface = db.session.execute(sql_imagem_por_codface, parametros_codfaces).mappings().all()
            for row_img in rows_imagem_por_codface:
                codface_img = (row_img.get("CodFaceNormalizada") or "").strip().upper()
                url_img = (row_img.get("UrlImagem") or "").strip()
                if codface_img and url_img:
                    mapa_url_imagem_por_codface[codface_img] = url_img
    except:
        mapa_url_imagem_por_id_face = {}
        mapa_url_imagem_por_codface = {}

    tipo_por_face = {}
    tipo_por_idcadastro = {}
    faces_por_cp_tipo = {}
    capacidade_digital_por_cp = {}

    # =========================================================
    # ✅ capacidade digital por CodPonto (último por CodPonto+Tipo)
    # =========================================================
    if codpontos_pagina:
        try:
            tipo_painel_norm2 = func.upper(func.ltrim(func.rtrim(func.coalesce(DimPaineisEuromidia.Tipo, ""))))

            rn_dig = func.row_number().over(
                partition_by=(DimPaineisEuromidia.CodPonto, tipo_painel_norm2),
                order_by=(DimPaineisEuromidia.DataAtualizacao.desc(), DimPaineisEuromidia.IDDimPaineisEuromidia.desc()),
            ).label("rn")

            sub_digital = (
                db.session.query(
                    DimPaineisEuromidia.CodPonto.label("CodPonto"),
                    tipo_painel_norm2.label("TipoNorm"),
                    DimPaineisEuromidia.QuantidadeFaces.label("QuantidadeFaces"),
                    rn_dig,
                )
                .filter(DimPaineisEuromidia.CodPonto.in_(codpontos_pagina))
                .subquery()
            )

            ult_digital = (
                db.session.query(sub_digital.c.CodPonto, sub_digital.c.QuantidadeFaces)
                .filter(
                    sub_digital.c.rn == 1,
                    sub_digital.c.TipoNorm.like("%DIGITAL%")
                )
                .all()
            )

            for cp, qf in ult_digital:
                try:
                    cp_int = int(cp) if cp is not None else None
                    if cp_int is None:
                        continue
                    cap = int(qf) if qf is not None else int(CAPACIDADE_DIGITAL_FIXA)
                    cap = max(1, min(cap, 200))
                    capacidade_digital_por_cp[cp_int] = cap
                except:
                    continue
        except:
            pass


    if codpontos_pagina:
        inv = (
            db.session.query(
                DimFacesPaineis.IDDimFacesPaineis,
                DimFacesPaineis.CodPonto,
                DimFacesPaineis.CodFace,
                DimFacesPaineis.Tipo,
            )
            .filter(DimFacesPaineis.CodPonto.in_(codpontos_pagina))
            .all()
        )

        for iddim, cp, cf, tp in inv:
            cp_int = int(cp) if cp is not None else None
            if cp_int is None:
                continue

            cf_norm = (str(cf) or "").strip()
            tp_norm = (tp or "").strip()
            tp_up = tp_norm.upper()

            if cf_norm and tp_norm:
                tipo_por_face[(cp_int, cf_norm)] = tp_up
                faces_por_cp_tipo.setdefault((cp_int, tp_up), set()).add(cf_norm)

            if iddim not in (None, "") and tp_norm:
                try:
                    tipo_por_idcadastro[int(iddim)] = tp_up
                except:
                    pass

            if tp_up == "PAINEL DIGITAL" and (cp_int not in capacidade_digital_por_cp):
                capacidade_digital_por_cp[cp_int] = int(CAPACIDADE_DIGITAL_FIXA)

    ocupacoes = []
    if codpontos_pagina:
        marca_escolhida_sql = func.nullif(FatoControleContratosItensEuromidia.MarcaExibida, "")

        rows_itens_q = (
            db.session.query(
                FatoControleContratosItensEuromidia.CodPonto,
                FatoControleContratosItensEuromidia.CodFace,
                FatoControleContratosItensEuromidia.Cota,
                FatoControleContratosItensEuromidia.DataInicioPrevisto,
                FatoControleContratosItensEuromidia.DataTerminoPrevisto,
                FatoControleContratosItensEuromidia.DataCancelamento,
                DimFacesPaineis.IDDimFacesPaineis,
            )
            .outerjoin(
                DimFacesPaineis,
                and_(
                    DimFacesPaineis.CodPonto == FatoControleContratosItensEuromidia.CodPonto,
                    DimFacesPaineis.CodFace == FatoControleContratosItensEuromidia.CodFace,
                ),
            )
            .filter(
                FatoControleContratosItensEuromidia.CodPonto.in_(codpontos_pagina),
                filtro_periodo_itens_ocup,
            )
        )

        rows_itens_q = _aplicar_filtro_faces_itens(
            rows_itens_q,
            pares_faces=selected_face_pairs,
            codfaces_norm=selected_codfaces_norm,
        )

        if codfaces_pagina:
            rows_itens_q = rows_itens_q.filter(FatoControleContratosItensEuromidia.CodFace.in_(codfaces_pagina))

        if cidade_exibicao_list:
            rows_itens_q = rows_itens_q.filter(FatoControleContratosItensEuromidia.CidadeExibicao.in_(cidade_exibicao_list))

        if tipo_documento_list:
            rows_itens_q = rows_itens_q.filter(FatoControleContratosItensEuromidia.TipoDocumento.in_(tipo_documento_list))

        if vendedor_list:
            rows_itens_q = rows_itens_q.filter(FatoControleContratosItensEuromidia.Vendedor.in_(vendedor_list))

        if marca_exibida_list:
            rows_itens_q = rows_itens_q.filter(marca_escolhida_sql.in_(marca_exibida_list))

        rows_itens = rows_itens_q.all()

        for cp, cf, cota, di, df_prev, dc, idcad in rows_itens:
            if cp is None or di is None:
                continue

            cp_int = int(cp)
            cf_norm = (str(cf) or "").strip()
            if not cf_norm:
                continue

            df_ef = _fim_efetivo_item(df_prev, dc)
            di_dt = _coerce_to_date(di)

            if di_dt is None or df_ef is None:
                continue

            if df_ef < dt_ini_ocup:
                continue
            if di_dt > dt_fim_ocup:
                continue
            if df_ef < di_dt:
                continue

            item_tp_up = tipo_por_face.get((cp_int, cf_norm))
            if (not item_tp_up) and (idcad not in (None, "")):
                try:
                    item_tp_up = tipo_por_idcadastro.get(int(idcad))
                except:
                    item_tp_up = None

            if not item_tp_up:
                continue

            if item_tp_up != "PAINEL DIGITAL":
                faces_validas = faces_por_cp_tipo.get((cp_int, item_tp_up), set())
                if cf_norm not in faces_validas:
                    continue

            ocupacoes.append((cp_int, cf_norm, cota, di_dt, df_ef, idcad))

    mapa_face = _calcular_ocupacao_e_conflitos_por_face(
        ocupacoes=ocupacoes,
        tipo_por_face=tipo_por_face,
        tipo_por_idcadastro=tipo_por_idcadastro,
        capacidade_digital_por_cp=capacidade_digital_por_cp,
        dt_ini=dt_ini_ocup,
        dt_fim=dt_fim_ocup,
    )

    itens = []
    for r in rows:
        cf = (str(r.CodFace) or "").strip()
        cp = int(r.CodPonto) if getattr(r, "CodPonto", None) is not None else None

        tp = (r.TipoProd or "").strip()
        tp_up = tp.upper()

        denom_padrao = int(capacidade_digital_por_cp.get(cp) or CAPACIDADE_DIGITAL_FIXA) if tp_up == "PAINEL DIGITAL" else 1

        info = mapa_face.get(
            (cp, cf, tp_up),
            {"ocupadas": 0, "denominador": denom_padrao, "conflitos": 0, "pct_uso": 0, "uso_medio": 0.0, "uso_slots_dias": 0.0, "total_dias": 0, "max_simult_sem_teto": 0.0},
        )

        ocupadas = int(info.get("ocupadas") or 0)
        denominador = int(info.get("denominador") or 0)
        conflitos = int(info.get("conflitos") or 0)

        if denominador <= 0:
            denominador = denom_padrao

        pct = int(info.get("pct_uso") or 0)
        if pct < 0:
            pct = 0
        if pct > 100:
            pct = 100

        status_calc = "livre"
        if pct > 0:
            status_calc = "ocupado"
        if conflitos > 0:
            status_calc = "com_conflito"

        logr = (getattr(r, "Logradouro", None) or "").strip()
        uf = (r.UF or "").strip()

        endereco_linha = "—"
        if logr or uf:
            if logr and uf:
                endereco_linha = f"{logr} - {uf}"
            elif logr:
                endereco_linha = logr
            else:
                endereco_linha = uf

        municipio = (getattr(r, "Municipio", None) or "").strip()
        uf_col = (getattr(r, "UF_COL", None) or (r.UF or "")).strip()
        formato_col = (getattr(r, "Formato", None) or (getattr(r, "FormatoLxA", None) or "")).strip()

       
        bairro_col = (getattr(r, "Bairro", None) or "").strip()
        referencia_col = (getattr(r, "Referencia", None) or "").strip()

  
        num_faces_cadastro = int(getattr(r, "NumFacesCadastro", None) or 0) if getattr(r, "NumFacesCadastro", None) is not None else 0

        if tp_up == "PAINEL DIGITAL":
            num_faces_exibicao = int(denominador) 
        else:
          
            num_faces_exibicao = int(num_faces_cadastro or 0)
            if num_faces_exibicao <= 0:
                num_faces_exibicao = 1

        try:
            id_dim_faces_paineis = int(getattr(r, "IDDimFacesPaineis", None) or 0)
        except:
            id_dim_faces_paineis = 0

        url_imagem_painel = ""
        if id_dim_faces_paineis > 0:
            url_imagem_painel = (mapa_url_imagem_por_id_face.get(id_dim_faces_paineis) or "").strip()

        if (not url_imagem_painel) and cf:
            url_imagem_painel = (mapa_url_imagem_por_codface.get(cf.upper()) or "").strip()

        itens.append(
            {
                "CodFace": cf,
                "CodPonto": cp,
                "IDDimFacesPaineis": id_dim_faces_paineis,
                "UrlImagemPainel": url_imagem_painel,
                "TipoProd": tp,
                "Cidade": municipio,
                "Endereco": endereco_linha,
                "Logradouro": logr,
                "EXIBIDORA": (r.EXIBIDORA or "") if pode_ver_exibidora else "",
                "UF": (r.UF or ""),
                "BitEnergia": int(getattr(r, "BitEnergia", 0) or 0),
                "BitInternet": int(getattr(r, "BitInternet", 0) or 0),
                "BitAtivo": int(getattr(r, "BitAtivo", 0) or 0),
                "BitAluguel": int(getattr(r, "BitAluguel", 0) or 0),

          
                "Bairro": bairro_col,
                "Referencia": referencia_col,

              
                "NumFaces": int(num_faces_exibicao),

                
                "NumFacesOcupacao": int(denominador),

                "Ocupadas": ocupadas,
                "Pct": pct,
                "Conflitos": conflitos,
                "Status": status_calc,
                "AltxLarg": (getattr(r, "FormatoLxA", None) or ""),
                "AreaTotal": (getattr(r, "AreaTotalm", None)),
                "Iluminado": "1" if int(getattr(r, "BitIluminado", 0) or 0) == 1 else "0",
                "Municipio": municipio,
                "UF_COL": uf_col,
                "Formato": formato_col,

                "NumFacesCadastro": int(num_faces_cadastro),

                "UsoMedio": float(info.get("uso_medio") or 0.0),
                "UsoSlotsDias": float(info.get("uso_slots_dias") or 0.0),
                "TotalDiasPeriodo": int(info.get("total_dias") or 0),
                "MaxSimultaneoSemTeto": float(info.get("max_simult_sem_teto") or 0.0),
            }
        )

    status_list_norm = [(s or "").strip().lower() for s in (status_list or []) if (s or "").strip()]
    if status_list_norm and ("todos" not in status_list_norm):
        itens_filtrados = []
        for x in itens:
            if "com_conflito" in status_list_norm and x.get("Conflitos", 0) > 0:
                itens_filtrados.append(x)
                continue
            if "livre" in status_list_norm and int(x.get("Pct", 0) or 0) == 0:
                itens_filtrados.append(x)
                continue
            if "ocupado" in status_list_norm and (int(x.get("Pct", 0) or 0) > 0 and x.get("Conflitos", 0) == 0):
                itens_filtrados.append(x)
                continue
        itens = itens_filtrados

    total_pages = max(1, (total + per_page - 1) // per_page)

    if tudo:
        periodo_label = f"Tudo ({dt_ini_ocup.strftime('%d/%m/%Y')} até {dt_fim_ocup.strftime('%d/%m/%Y')})"
    else:
        if (dt_ini_filtro is not None) and (dt_fim_filtro is not None):
            periodo_label = f"{dt_ini_filtro.strftime('%d/%m/%Y')} até {dt_fim_filtro.strftime('%d/%m/%Y')}"
        else:
            periodo_label = f"Mês atual ({dt_ini_ocup.strftime('%d/%m/%Y')} até {dt_fim_ocup.strftime('%d/%m/%Y')})"

    ranges_rapidos = {
        "dia": (hoje, hoje),
        "hoje": (hoje, hoje),
        "semana": (hoje - timedelta(days=hoje.weekday()), (hoje - timedelta(days=hoje.weekday())) + timedelta(days=6)),
        "mes": (primeiro_dia_mes_atual, hoje),
        "ano": (date(hoje.year, 1, 1), hoje),
    }

    return render_template(
        "euromidia/paineis_lista.html",
        itens=itens,
        tipos=tipos,
        exibidoras=exibidoras if pode_ver_exibidora else [],
        cidades_exibicao=cidades_exibicao,
        tipos_documento=tipos_documento,
        vendedores_contrato=vendedores_contrato,
        marcas_exibidas=marcas_exibidas,
        altxlarg_opcoes=altxlarg_opcoes,
        iluminado_opcoes=iluminado_opcoes,
        areas_total=areas_total,
        area_total_min_global=area_total_min_global,
        area_total_max_global=area_total_max_global,
        ranges_rapidos=ranges_rapidos,
        pode_ver_exibidora=pode_ver_exibidora,
        bisemanas_select=bisemanas_select,

        filtros={
            "q": q,
            "tipo": tipo_list,
            "exibidora": exibidora_list if pode_ver_exibidora else [],
            "cidade_exibicao": cidade_exibicao_list,
            "tipo_documento": tipo_documento_list,
            "vendedor": vendedor_list,
            "marca_exibida": marca_exibida_list,
            "status": status_list_norm if status_list_norm else ["todos"],

      
            "formato": formato_list,

            "busca_erro": busca_erro,
            "altxlarg": altxlarg,
            "iluminado": iluminado,
            "area_min": area_min_str,
            "area_max": area_max_str,
            "ponto_ativo": ponto_ativo,
            "dt_ini": dt_ini_str,
            "dt_fim": dt_fim_str,
            "mes_atual": "1" if mes_atual_user else "0",
            "tudo": "1" if tudo else "0",
            "periodo": periodo_rapido,
            "ano": ano_filtro,
            "anomes": anomes_filtro,
            "mes": mes_filtro,
            "dia": dia_filtro,
            "trimestre": trimestre_filtro,
            "semana_iso": semana_iso_filtro,
            "dia_semana_iso": dia_semana_iso_filtro,
            "quinzena": quinzena_filtro,

           
            "bi_semana": bi_semana_filtro,
            "bisemana": bi_semana_filtro,

            "fim_de_semana": fim_de_semana_filtro,
            "periodo_label": periodo_label,
            "per_page": per_page,
            "pode_ver_exibidora": pode_ver_exibidora,
            "selected_codface": selected_codfaces,
            "selected_codponto": [str(x) for x in selected_codpontos],
            "tem_faces_selecionadas": tem_faces_selecionadas,
        },
        paginacao={
            "page": page,
            "per_page": per_page,
            "total": total,
            "total_pages": total_pages,
            "inicio": (page - 1) * per_page + 1 if total > 0 else 0,
            "fim": min(page * per_page, total),
        },
    )






@paineis_bp.route("/<int:codponto>/grade", methods=["GET", "POST"])
@login_required
@limiter.limit("80 per minute", methods=["GET"])
@retry_get_view(db, attempts=6, base_delay=0.2, max_delay=1.5)
def grade_painel(codponto: int):

    from sqlalchemy import text as sql_text, false as sa_false, bindparam

    mes_ref = (request.args.get("mes_ref") or "").strip()
    filtro_cliente = (request.args.get("cliente") or "").strip()
    filtro_tipo_prod = (request.args.get("tipo") or "").strip()

    mes_de = (request.args.get("mes_de") or "").strip()
    mes_ate = (request.args.get("mes_ate") or "").strip()

    modo = (request.args.get("modo") or "").strip().lower()
    data_ref_str = (request.args.get("data_ref") or "").strip()

    bi_semana_sel = (request.args.get("bi_semana") or "").strip()

    LOOPS_PERMITIDOS = [f"SPAN{n:02d}" for n in range(1, CAPACIDADE_DIGITAL_FIXA + 1)]
    hoje = date.today()

    def _parse_date_ymd(s: str):
        try:
            s = (s or "").strip()
            if len(s) >= 10 and s[4] == "-" and s[7] == "-":
                y = int(s[0:4])
                m = int(s[5:7])
                d = int(s[8:10])
                return date(y, m, d)
        except:
            return None
        return None

    def _eh_yyyy_mm(s: str) -> bool:
        s = (s or "").strip()
        return (len(s) == 7 and "-" in s)

    def _yyyy_mm_para_ano_mes(s: str):
        s = (s or "").strip()
        try:
            y = int(s.split("-")[0])
            m = int(s.split("-")[1])
            return y, m
        except:
            return None, None

    def _coerce_to_date(v):
        try:
            if v is None:
                return None
            if isinstance(v, datetime):
                return v.date()
            if isinstance(v, date):
                return v
            s = str(v).strip()
            if not s:
                return None

            try:
                return datetime.strptime(s, "%Y-%m-%d").date()
            except:
                pass

            try:
                return datetime.strptime(s, "%d/%m/%Y").date()
            except:
                pass

            try:
                return datetime.fromisoformat(s).date()
            except:
                return None
        except:
            return None

    def _fim_efetivo_item(df_prev, dc):
        try:
            dc_dt = _coerce_to_date(dc)
            if dc_dt:
                return dc_dt
            return _coerce_to_date(df_prev)
        except:
            return None

    def _like_param(s: str):
        s = (s or "").strip()
        if not s:
            return None
        return f"%{s}%"

    def _intersecta(di_a: date, df_a: date, di_b: date, df_b: date) -> bool:
        try:
            if di_a is None or df_a is None or di_b is None or df_b is None:
                return False
            return (di_a <= df_b) and (di_b <= df_a)
        except:
            return False

    def _normalizar_codface(s: str) -> str:
        try:
            return str(s or "").strip()
        except:
            return ""

    def _lista_args_limpa(nome_param: str):
        valores = []

        try:
            for v in request.args.getlist(nome_param):
                s = (str(v).strip() if v is not None else "")
                if s:
                    valores.append(s)
        except:
            pass

        if not valores:
            bruto = (request.args.get(nome_param) or "").strip()
            if bruto:
                if "|" in bruto:
                    partes = [x.strip() for x in bruto.split("|") if x.strip()]
                    valores.extend(partes)
                elif ";" in bruto:
                    partes = [x.strip() for x in bruto.split(";") if x.strip()]
                    valores.extend(partes)
                else:
                    valores.append(bruto)

        limpos = []
        vistos = set()
        for v in valores:
            vv = str(v or "").strip()
            if not vv:
                continue
            chave = vv.casefold()
            if chave in vistos:
                continue
            vistos.add(chave)
            limpos.append(vv)

        return limpos

    filtros_codface_brutos = _lista_args_limpa("codface")
    filtros_codface = []
    vistos_codface = set()

    for cf_tmp in (filtros_codface_brutos or []):
        cf_norm_tmp = _normalizar_codface(cf_tmp)
        if not cf_norm_tmp:
            continue

        chave_tmp = cf_norm_tmp.casefold()
        if chave_tmp in vistos_codface:
            continue

        vistos_codface.add(chave_tmp)
        filtros_codface.append(cf_norm_tmp)

    filtros_codface_ci = {cf.casefold() for cf in (filtros_codface or [])}
    tem_filtro_codface = len(filtros_codface) > 0

    filtro_codface = filtros_codface[0] if len(filtros_codface) == 1 else ""
    codface_selecionado = filtro_codface

    def _resolver_codponto_por_codface_global(codface_valor: str):
        """
        Resolve o CodPonto da CodFace em toda a DimFacesPaineis.

        Critério técnico:
        1) procura a CodFace em toda a dimensão;
        2) se houver registro no CodPonto atual, ele tem prioridade para evitar redirect desnecessário;
        3) caso não haja no painel atual, escolhe o registro mais recente pelo IDDimFacesPaineis DESC.

        Isso mantém comportamento determinístico sem remover nenhuma funcionalidade existente.
        """
        codface_valor = _normalizar_codface(codface_valor)
        if not codface_valor:
            return None

        codface_valor_up = codface_valor.upper()

        rows_face_global = (
            db.session.query(
                DimFacesPaineis.CodPonto,
                DimFacesPaineis.CodFace,
                DimFacesPaineis.IDDimFacesPaineis,
            )
            .filter(
                DimFacesPaineis.CodPonto != None,
                DimFacesPaineis.CodFace != None,
                DimFacesPaineis.CodFace != "",
                func.upper(func.ltrim(func.rtrim(DimFacesPaineis.CodFace))) == codface_valor_up,
            )
            .order_by(
                case((DimFacesPaineis.CodPonto == int(codponto), 0), else_=1).asc(),
                DimFacesPaineis.IDDimFacesPaineis.desc(),
            )
            .all()
        )

        if not rows_face_global:
            return None

        for cp_row, cf_row, _id_row in (rows_face_global or []):
            try:
                cp_int = int(cp_row) if cp_row is not None else None
            except:
                cp_int = None

            cf_norm = _normalizar_codface(cf_row)
            if cp_int is None or not cf_norm:
                continue

            if cp_int == int(codponto):
                return cp_int

        for cp_row, cf_row, _id_row in (rows_face_global or []):
            try:
                cp_int = int(cp_row) if cp_row is not None else None
            except:
                cp_int = None

            cf_norm = _normalizar_codface(cf_row)
            if cp_int is None or not cf_norm:
                continue

            return cp_int

        return None

    def _resolver_codponto_unico_por_codfaces_globais(codfaces_lista):
        """
        Quando há múltiplas faces na query, só existe redirect seguro se
        todas resolverem para o mesmo CodPonto. Caso contrário, a rota
        atual continua responsável apenas pelo CodPonto informado na URL.
        """
        try:
            codpontos_resolvidos = []

            for cf_val in (codfaces_lista or []):
                cp_res = _resolver_codponto_por_codface_global(cf_val)
                if cp_res is None:
                    return None
                codpontos_resolvidos.append(int(cp_res))

            if not codpontos_resolvidos:
                return None

            unicos = sorted(list(set(codpontos_resolvidos)))
            if len(unicos) == 1:
                return int(unicos[0])

            return None
        except:
            return None

    def _redirect_grade_preservando_query(codponto_destino: int, codfaces_override=None):
        """
        Redireciona para a URL canônica da grade do painel destino,
        preservando todos os parâmetros atuais da query string.
        """
        args_redirect = request.args.to_dict(flat=False)

        lista_codfaces = list(codfaces_override or [])
        if lista_codfaces:
            args_redirect["codface"] = lista_codfaces
        else:
            try:
                args_redirect.pop("codface", None)
            except:
                pass

        return redirect(
            url_for(
                "Paineis.grade_painel",
                codponto=int(codponto_destino),
                **args_redirect,
            )
        )

    filtros_vendedor_brutos = _lista_args_limpa("vendedor")
    filtro_vendedor = " | ".join(filtros_vendedor_brutos)

    vendedores_select = []
    nomes_vendedores_validos_ci = {}

    try:
        sql_vendedores = sql_text("""
            SELECT
                IDVendedor,
                NomeVendedor
            FROM [Integracao].[dbo].[Vendedores] WITH (NOLOCK)
            WHERE IDEmpresaProprietaria = 3
              AND NULLIF(LTRIM(RTRIM(NomeVendedor)), '') IS NOT NULL
            ORDER BY NomeVendedor
        """)

        rows_vendedores = db.session.execute(sql_vendedores).fetchall()

        vistos_vend = set()
        for rv in (rows_vendedores or []):
            try:
                id_vendedor = int(rv[0]) if rv[0] is not None else None
            except:
                id_vendedor = None

            try:
                nome_vendedor = str(rv[1] or "").strip()
            except:
                nome_vendedor = ""

            if not nome_vendedor:
                continue

            chave_ci = nome_vendedor.casefold()
            if chave_ci in vistos_vend:
                continue

            vistos_vend.add(chave_ci)

            vendedores_select.append({
                "IDVendedor": id_vendedor,
                "NomeVendedor": nome_vendedor,
            })
            nomes_vendedores_validos_ci[chave_ci] = nome_vendedor

    except:
        vendedores_select = []
        nomes_vendedores_validos_ci = {}

    vendedores_selecionados = []
    vendedores_termos_livres = []

    if filtros_vendedor_brutos:
        if vendedores_select:
            for nome_bruto in filtros_vendedor_brutos:
                nome_resolvido = nomes_vendedores_validos_ci.get(nome_bruto.casefold())
                if nome_resolvido:
                    if nome_resolvido not in vendedores_selecionados:
                        vendedores_selecionados.append(nome_resolvido)
                else:
                    vendedores_termos_livres.append(nome_bruto)
        else:
            vendedores_termos_livres = list(filtros_vendedor_brutos)

    filtros_segmento_brutos = _lista_args_limpa("segmento")
    filtro_segmento = " | ".join(filtros_segmento_brutos)

    segmentos_select = []
    nomes_segmentos_validos_ci = {}
    segmentos_selecionados = []
    segmentos_termos_livres = []

    dt_ini = None
    dt_fim = None


    if tem_filtro_codface:
        codponto_resolvido_pelas_faces = _resolver_codponto_unico_por_codfaces_globais(filtros_codface)

        if (
            codponto_resolvido_pelas_faces is not None
            and int(codponto_resolvido_pelas_faces) != int(codponto)
        ):
            return _redirect_grade_preservando_query(
                int(codponto_resolvido_pelas_faces),
                codfaces_override=filtros_codface,
            )

    if modo in ("dia", "semana", "mes", "ano"):
        dt_ref = _parse_date_ymd(data_ref_str) or hoje

        cal_ref = (
            db.session.query(DimCalendario)
            .filter(DimCalendario.data == dt_ref)
            .first()
        )

        if cal_ref:
            if modo == "dia":
                dt_ini = cal_ref.data
                dt_fim = cal_ref.data

            elif modo == "semana":
                dt_ini = cal_ref.inicio_semana
                dt_fim = cal_ref.fim_semana

            elif modo == "mes":
                ini_mes = (
                    db.session.query(func.min(DimCalendario.data))
                    .filter(DimCalendario.ano_mes == cal_ref.ano_mes)
                    .scalar()
                )
                fim_mes = (
                    db.session.query(func.max(DimCalendario.data))
                    .filter(DimCalendario.ano_mes == cal_ref.ano_mes)
                    .scalar()
                )
                dt_ini = ini_mes or date(int(cal_ref.ano), int(cal_ref.mes), 1)
                dt_fim = fim_mes or dt_ini

            elif modo == "ano":
                ini_ano = (
                    db.session.query(func.min(DimCalendario.data))
                    .filter(DimCalendario.ano == cal_ref.ano)
                    .scalar()
                )
                fim_ano = (
                    db.session.query(func.max(DimCalendario.data))
                    .filter(DimCalendario.ano == cal_ref.ano)
                    .scalar()
                )
                dt_ini = ini_ano or date(int(cal_ref.ano), 1, 1)
                dt_fim = fim_ano or date(int(cal_ref.ano), 12, 31)

        if dt_ini is not None and dt_fim is not None:
            if dt_fim < dt_ini:
                dt_ini, dt_fim = dt_fim, dt_ini

            mes_ref = f"{dt_ini.year:04d}-{dt_ini.month:02d}"
            mes_de = mes_ref
            mes_ate = f"{dt_fim.year:04d}-{dt_fim.month:02d}-{dt_fim.day:02d}"

    meses_select = []
    bisemanas_select = []

    try:
        meses_select = [
            (r[0] or "").strip()
            for r in (
                db.session.query(DimCalendario.ano_mes)
                .filter(DimCalendario.ano_mes != None, DimCalendario.ano_mes != "")
                .distinct()
                .order_by(DimCalendario.ano_mes.desc())
                .limit(48)
                .all()
            )
            if (r and r[0])
        ]
    except:
        meses_select = []

    if dt_ini is None or dt_fim is None:
        if mes_ref and _eh_yyyy_mm(mes_ref):
            ano, mes = _yyyy_mm_para_ano_mes(mes_ref)
            if not ano or not mes:
                ano = hoje.year
                mes = hoje.month
                mes_ref = f"{ano:04d}-{mes:02d}"
        else:
            ano = hoje.year
            mes = hoje.month
            mes_ref = f"{ano:04d}-{mes:02d}"

        if mes_de and _eh_yyyy_mm(mes_de):
            ano_de, mes_de_int = _yyyy_mm_para_ano_mes(mes_de)
            if ano_de and mes_de_int:
                dt_ini = date(ano_de, mes_de_int, 1)
            else:
                dt_ini, _ = _primeiro_ultimo_dia_mes(ano, mes)
                mes_de = mes_ref
        else:
            dt_ini, _ = _primeiro_ultimo_dia_mes(ano, mes)
            mes_de = mes_ref

        if mes_ate and len(mes_ate) >= 10 and "-" in mes_ate:
            try:
                y = int(mes_ate[0:4])
                m = int(mes_ate[5:7])
                d = int(mes_ate[8:10])
                dt_fim = date(y, m, d)
            except:
                _, dt_fim = _primeiro_ultimo_dia_mes(ano, mes)
                mes_ate = ""
        else:
            _, dt_fim = _primeiro_ultimo_dia_mes(ano, mes)

        if dt_fim < dt_ini:
            dt_ini, dt_fim = dt_fim, dt_ini

    try:
        bis_raw = (
            db.session.query(
                DimCalendario.bi_semana_numero,
                DimCalendario.inicio_bi_semana,
                DimCalendario.fim_bi_semana,
            )
            .filter(
                DimCalendario.bi_semana_numero != None,
                DimCalendario.inicio_bi_semana != None,
                DimCalendario.fim_bi_semana != None,
                DimCalendario.inicio_bi_semana <= dt_fim,
                DimCalendario.fim_bi_semana >= dt_ini,
            )
            .group_by(
                DimCalendario.bi_semana_numero,
                DimCalendario.inicio_bi_semana,
                DimCalendario.fim_bi_semana,
            )
            .order_by(DimCalendario.inicio_bi_semana.asc())
            .all()
        )
    except:
        bis_raw = []

    for n, di_bi, df_bi in (bis_raw or []):
        if di_bi is not None and df_bi is not None:
            if not _intersecta(dt_ini, dt_fim, di_bi, df_bi):
                continue

        try:
            n_int = int(n)
            value = str(n_int)
        except:
            value = str(n).strip()

        try:
            label = f"{value} — {di_bi.strftime('%d/%m')} a {df_bi.strftime('%d/%m')}"
        except:
            label = f"{value}"

        try:
            ini_str = di_bi.strftime("%Y-%m-%d") if di_bi else ""
        except:
            ini_str = ""

        try:
            fim_str = df_bi.strftime("%Y-%m-%d") if df_bi else ""
        except:
            fim_str = ""

        bisemanas_select.append({
            "value": value,
            "label": label,
            "inicio": ini_str,
            "fim": fim_str,
        })

    if bi_semana_sel:
        bi_num = None
        try:
            bi_num = int(str(bi_semana_sel).strip())
        except:
            bi_num = None

        if bi_num is not None:
            cal_bi = (
                db.session.query(
                    func.min(DimCalendario.inicio_bi_semana),
                    func.max(DimCalendario.fim_bi_semana),
                )
                .filter(
                    DimCalendario.bi_semana_numero == bi_num,
                    DimCalendario.inicio_bi_semana != None,
                    DimCalendario.fim_bi_semana != None,
                )
                .first()
            )

            if cal_bi and cal_bi[0] and cal_bi[1]:
                dt_ini = cal_bi[0]
                dt_fim = cal_bi[1]

                if dt_fim < dt_ini:
                    dt_ini, dt_fim = dt_fim, dt_ini

                mes_ref = f"{dt_ini.year:04d}-{dt_ini.month:02d}"
                mes_de = mes_ref
                mes_ate = f"{dt_fim.year:04d}-{dt_fim.month:02d}-{dt_fim.day:02d}"

                modo = ""
                data_ref_str = ""

    ano = dt_ini.year
    mes = dt_ini.month
    mes_ref = f"{ano:04d}-{mes:02d}"

    total_dias = (dt_fim - dt_ini).days + 1
    ultimo_dia = total_dias


    rows_codfaces_select = (
        db.session.query(DimFacesPaineis.CodFace)
        .filter(
            DimFacesPaineis.CodFace != None,
            DimFacesPaineis.CodFace != "",
        )
        .group_by(DimFacesPaineis.CodFace)
        .order_by(DimFacesPaineis.CodFace.asc())
        .all()
    )

    codfaces_select = []
    vistos_codfaces_select = set()

    for (cf_select,) in (rows_codfaces_select or []):
        cf_norm_tmp = _normalizar_codface(cf_select)
        if not cf_norm_tmp:
            continue

        chave_tmp = cf_norm_tmp.casefold()
        if chave_tmp in vistos_codfaces_select:
            continue

        vistos_codfaces_select.add(chave_tmp)
        codfaces_select.append(cf_norm_tmp)


    faces_info_raw = (
        db.session.query(
            DimFacesPaineis.IDDimPaineisEuromidia,
            DimFacesPaineis.CodFace,
            DimFacesPaineis.Tipo,
            DimPaineisEuromidia.Exibidora,
            DimPaineisEuromidia.QuantidadeFaces,
        )
        .outerjoin(
            DimPaineisEuromidia,
            DimPaineisEuromidia.IDDimPaineisEuromidia == DimFacesPaineis.IDDimPaineisEuromidia,
        )
        .filter(
            DimFacesPaineis.CodPonto == codponto,
            DimFacesPaineis.CodFace != None,
            DimFacesPaineis.CodFace != "",
        )
        .all()
    )

    tipo_por_codface = {}

    for idpainel, cf, tp, ex, qf in (faces_info_raw or []):
        cf_norm_tmp = _normalizar_codface(cf)
        if not cf_norm_tmp:
            continue

        chave_tmp = cf_norm_tmp.casefold()
        if chave_tmp not in tipo_por_codface:
            tipo_por_codface[chave_tmp] = (str(tp or "").strip())

    if len(filtros_codface) == 1:
        tipo_face_selecionada = tipo_por_codface.get(filtros_codface[0].casefold(), "").strip()
        if tipo_face_selecionada:
            filtro_tipo_prod = tipo_face_selecionada

    if tem_filtro_codface:
        faces_info_raw = [
            x for x in faces_info_raw
            if (_normalizar_codface(x[1]).casefold() in filtros_codface_ci)
        ]

    faces_info = []
    tipos_distintos = set()
    for idpainel, cf, tp, ex, qf in (faces_info_raw or []):
        tp_up = (tp or "").strip().upper()
        if tp_up:
            tipos_distintos.add(tp_up)

    tipo_filtro_up = (filtro_tipo_prod or "").strip().upper()
    if tipo_filtro_up:
        for idpainel, cf, tp, ex, qf in (faces_info_raw or []):
            tp_up = (tp or "").strip().upper()
            if tp_up == tipo_filtro_up:
                faces_info.append((idpainel, cf, tp, ex, qf))
    else:
        if len(tipos_distintos) > 1:
            if "PAINEL DIGITAL" in tipos_distintos:
                tipo_filtro_up = "PAINEL DIGITAL"
            else:
                tipo_filtro_up = sorted(list(tipos_distintos))[0]
            for idpainel, cf, tp, ex, qf in (faces_info_raw or []):
                tp_up = (tp or "").strip().upper()
                if tp_up == tipo_filtro_up:
                    faces_info.append((idpainel, cf, tp, ex, qf))
        else:
            faces_info = list(faces_info_raw)

    faces = []
    exibidora = ""
    qtd_faces_painel_max = 0

    for idpainel, cf, tp, ex, qf in (faces_info or []):
        cf_norm = _normalizar_codface(cf)
        if not cf_norm:
            continue

        faces.append(cf_norm)

        if not exibidora:
            exibidora = (ex or "")

        if qf not in (None, ""):
            try:
                qf_int = int(qf)
                if qf_int > qtd_faces_painel_max:
                    qtd_faces_painel_max = qf_int
            except:
                pass

    faces = sorted(list(dict.fromkeys(faces)), key=lambda x: x.casefold())

    if tem_filtro_codface:
        faces = [f for f in faces if f.casefold() in filtros_codface_ci]

    eh_digital = any(((tp or "").strip().upper() == "PAINEL DIGITAL") for (_, _, tp, _, _) in (faces_info or []))

    if eh_digital:
        num_faces = int(CAPACIDADE_DIGITAL_FIXA) if int(CAPACIDADE_DIGITAL_FIXA or 0) > 0 else 0
        tipo_prod = "PAINEL DIGITAL"
    else:
        if qtd_faces_painel_max and int(qtd_faces_painel_max) > 0:
            num_faces = int(qtd_faces_painel_max)
        else:
            num_faces = len(faces) if faces else 0

        tipo_prod = ""
        for (_, _, tp, _, _) in (faces_info or []):
            tp_txt = (tp or "").strip()
            if tp_txt:
                tipo_prod = tp_txt
                break


    def _valor_texto_painel(v):
        try:
            txt = str(v or "").strip()
            return txt if txt else ""
        except:
            return ""

    painel_info_rows = (
        db.session.query(
            DimPaineisEuromidia.Tipo,
            DimPaineisEuromidia.Cidade,
            DimPaineisEuromidia.UF,
            DimPaineisEuromidia.Bairro,
            DimPaineisEuromidia.Logradouro,
        )
        .filter(DimPaineisEuromidia.CodPonto == codponto)
        .order_by(
            DimPaineisEuromidia.DataAtualizacao.desc(),
            DimPaineisEuromidia.IDDimPaineisEuromidia.desc(),
        )
        .all()
    )

    def _primeiro_valor_painel_por_indice(indice_coluna: int, valor_padrao: str = "") -> str:
        for row in (painel_info_rows or []):
            try:
                valor = _valor_texto_painel(row[indice_coluna])
            except:
                valor = ""

            if valor:
                return valor

        return valor_padrao

    tipo_painel_info = _primeiro_valor_painel_por_indice(
        0,
        (str(tipo_prod).strip() if tipo_prod not in (None, "") else ""),
    )
    cidade_painel = _primeiro_valor_painel_por_indice(1, "")
    uf_painel = _primeiro_valor_painel_por_indice(2, "")
    bairro_painel = _primeiro_valor_painel_por_indice(3, "")
    logradouro_painel = _primeiro_valor_painel_por_indice(4, "")

    try:
        partes_sql_segmentos = [
            """
            SELECT DISTINCT
                LTRIM(RTRIM(cn.Classe)) AS Classe
            FROM [Integracao].[Silver].[FatoControleContratosItensEuromidia] i WITH (NOLOCK)
            INNER JOIN [Integracao].[Silver].[FatoControleContratosEuromidia] c WITH (NOLOCK)
                ON c.IDFatoControleContratosEuromidia = i.IDFatoControleContratoEuromidia
            LEFT JOIN [Integracao].[Silver].[DimEmpresas] e WITH (NOLOCK)
                ON e.IDEmpresa = c.IDEmpresa
            LEFT JOIN [Integracao].[Silver].[DimCnaes] cn WITH (NOLOCK)
                ON REPLACE(REPLACE(REPLACE(REPLACE(LTRIM(RTRIM(ISNULL(e.CNAE, ''))), '.', ''), '-', ''), '/', ''), ' ', '')
                 = REPLACE(REPLACE(REPLACE(REPLACE(LTRIM(RTRIM(ISNULL(cn.cnaepadrao, ''))), '.', ''), '-', ''), '/', ''), ' ', '')
            WHERE i.CodPonto = :codponto
              AND i.AtivoCancelamento = 'A'
              AND i.DataInicioPrevisto IS NOT NULL
              AND i.DataInicioPrevisto <= :dt_fim
              AND COALESCE(i.DataCancelamento, i.DataTerminoPrevisto, CONVERT(date,'9999-12-31')) >= :dt_ini
              AND NULLIF(LTRIM(RTRIM(cn.Classe)), '') IS NOT NULL
            """
        ]

        params_segmentos = {
            "codponto": int(codponto),
            "dt_ini": dt_ini,
            "dt_fim": dt_fim,
        }

        binds_segmentos = []

        if tem_filtro_codface:
            partes_sql_segmentos.append(" AND LTRIM(RTRIM(ISNULL(i.CodFace, ''))) IN :codfaces ")
            params_segmentos["codfaces"] = list(filtros_codface)
            binds_segmentos.append(bindparam("codfaces", expanding=True))

        if filtro_cliente:
            partes_sql_segmentos.append(" AND i.MarcaExibida LIKE :cliente_like ")
            params_segmentos["cliente_like"] = f"%{filtro_cliente}%"

        filtros_sql_segmentos_vendedor = []

        if vendedores_selecionados:
            partes_sql_segmentos.append(" AND i.Vendedor IN :vendedores_exatos_segmentos ")
            params_segmentos["vendedores_exatos_segmentos"] = list(vendedores_selecionados)
            binds_segmentos.append(bindparam("vendedores_exatos_segmentos", expanding=True))

        for idx_v, termo_v in enumerate(vendedores_termos_livres or []):
            chave_v = f"vendedor_like_segmento_{idx_v}"
            filtros_sql_segmentos_vendedor.append(f" i.Vendedor LIKE :{chave_v} ")
            params_segmentos[chave_v] = f"%{termo_v}%"

        if filtros_sql_segmentos_vendedor:
            partes_sql_segmentos.append(" AND (" + " OR ".join(filtros_sql_segmentos_vendedor) + ") ")

        partes_sql_segmentos.append(" ORDER BY Classe ")

        stmt_segmentos = sql_text("".join(partes_sql_segmentos))
        if binds_segmentos:
            stmt_segmentos = stmt_segmentos.bindparams(*binds_segmentos)

        rows_segmentos = db.session.execute(stmt_segmentos, params_segmentos).fetchall()

        vistos_seg = set()
        for rs in (rows_segmentos or []):
            try:
                classe = str(rs[0] or "").strip()
            except:
                classe = ""

            if not classe:
                continue

            chave_seg = classe.casefold()
            if chave_seg in vistos_seg:
                continue

            vistos_seg.add(chave_seg)
            segmentos_select.append(classe)
            nomes_segmentos_validos_ci[chave_seg] = classe

    except:
        segmentos_select = []
        nomes_segmentos_validos_ci = {}

    if filtros_segmento_brutos:
        if segmentos_select:
            for seg_bruto in filtros_segmento_brutos:
                seg_resolvido = nomes_segmentos_validos_ci.get(seg_bruto.casefold())
                if seg_resolvido:
                    if seg_resolvido not in segmentos_selecionados:
                        segmentos_selecionados.append(seg_resolvido)
                else:
                    segmentos_termos_livres.append(seg_bruto)
        else:
            segmentos_termos_livres = list(filtros_segmento_brutos)

    contratos_segmento_permitidos = None

    if segmentos_selecionados or segmentos_termos_livres:
        try:
            partes_sql_ids_segmento = [
                """
                SELECT DISTINCT
                    i.IDFatoControleContratoEuromidia
                FROM [Integracao].[Silver].[FatoControleContratosItensEuromidia] i WITH (NOLOCK)
                INNER JOIN [Integracao].[Silver].[FatoControleContratosEuromidia] c WITH (NOLOCK)
                    ON c.IDFatoControleContratosEuromidia = i.IDFatoControleContratoEuromidia
                LEFT JOIN [Integracao].[Silver].[DimEmpresas] e WITH (NOLOCK)
                    ON e.IDEmpresa = c.IDEmpresa
                LEFT JOIN [Integracao].[Silver].[DimCnaes] cn WITH (NOLOCK)
                    ON REPLACE(REPLACE(REPLACE(REPLACE(LTRIM(RTRIM(ISNULL(e.CNAE, ''))), '.', ''), '-', ''), '/', ''), ' ', '')
                     = REPLACE(REPLACE(REPLACE(REPLACE(LTRIM(RTRIM(ISNULL(cn.cnaepadrao, ''))), '.', ''), '-', ''), '/', ''), ' ', '')
                WHERE i.CodPonto = :codponto
                  AND i.AtivoCancelamento = 'A'
                  AND i.DataInicioPrevisto IS NOT NULL
                  AND i.DataInicioPrevisto <= :dt_fim
                  AND COALESCE(i.DataCancelamento, i.DataTerminoPrevisto, CONVERT(date,'9999-12-31')) >= :dt_ini
                  AND NULLIF(LTRIM(RTRIM(cn.Classe)), '') IS NOT NULL
                """
            ]

            params_ids_segmento = {
                "codponto": int(codponto),
                "dt_ini": dt_ini,
                "dt_fim": dt_fim,
            }

            binds_ids_segmento = []

            if tem_filtro_codface:
                partes_sql_ids_segmento.append(" AND LTRIM(RTRIM(ISNULL(i.CodFace, ''))) IN :codfaces ")
                params_ids_segmento["codfaces"] = list(filtros_codface)
                binds_ids_segmento.append(bindparam("codfaces", expanding=True))

            if filtro_cliente:
                partes_sql_ids_segmento.append(" AND i.MarcaExibida LIKE :cliente_like ")
                params_ids_segmento["cliente_like"] = f"%{filtro_cliente}%"

            if vendedores_selecionados:
                partes_sql_ids_segmento.append(" AND i.Vendedor IN :vendedores_exatos_segmento_ids ")
                params_ids_segmento["vendedores_exatos_segmento_ids"] = list(vendedores_selecionados)
                binds_ids_segmento.append(bindparam("vendedores_exatos_segmento_ids", expanding=True))

            filtros_sql_ids_vendedor = []
            for idx_v, termo_v in enumerate(vendedores_termos_livres or []):
                chave_v = f"vendedor_like_segmento_ids_{idx_v}"
                filtros_sql_ids_vendedor.append(f" i.Vendedor LIKE :{chave_v} ")
                params_ids_segmento[chave_v] = f"%{termo_v}%"

            if filtros_sql_ids_vendedor:
                partes_sql_ids_segmento.append(" AND (" + " OR ".join(filtros_sql_ids_vendedor) + ") ")

            filtros_sql_ids_segmento = []

            if segmentos_selecionados:
                filtros_sql_ids_segmento.append(" LTRIM(RTRIM(cn.Classe)) IN :segmentos_exatos ")
                params_ids_segmento["segmentos_exatos"] = list(segmentos_selecionados)
                binds_ids_segmento.append(bindparam("segmentos_exatos", expanding=True))

            for idx_s, termo_s in enumerate(segmentos_termos_livres or []):
                chave_s = f"segmento_like_{idx_s}"
                filtros_sql_ids_segmento.append(f" LTRIM(RTRIM(cn.Classe)) LIKE :{chave_s} ")
                params_ids_segmento[chave_s] = f"%{termo_s}%"

            if filtros_sql_ids_segmento:
                partes_sql_ids_segmento.append(" AND (" + " OR ".join(filtros_sql_ids_segmento) + ") ")

            stmt_ids_segmento = sql_text("".join(partes_sql_ids_segmento))
            if binds_ids_segmento:
                stmt_ids_segmento = stmt_ids_segmento.bindparams(*binds_ids_segmento)

            rows_ids_segmento = db.session.execute(stmt_ids_segmento, params_ids_segmento).fetchall()

            contratos_segmento_permitidos = []
            for rr in (rows_ids_segmento or []):
                try:
                    if rr[0] is not None:
                        contratos_segmento_permitidos.append(int(rr[0]))
                except:
                    pass

            contratos_segmento_permitidos = sorted(list(dict.fromkeys(contratos_segmento_permitidos)))

        except:
            contratos_segmento_permitidos = []

    fim_efetivo_sql = func.coalesce(
        FatoControleContratosItensEuromidia.DataCancelamento,
        FatoControleContratosItensEuromidia.DataTerminoPrevisto,
        date(9999, 12, 31),
    )

    q_oc = (
        db.session.query(
            FatoControleContratosItensEuromidia.IDFatoControleContratosItensEuromidia,
            FatoControleContratosItensEuromidia.IDFatoControleContratoEuromidia,
            FatoControleContratosItensEuromidia.CodFace,
            FatoControleContratosItensEuromidia.MarcaExibida,
            FatoControleContratosItensEuromidia.Vendedor,
            FatoControleContratosItensEuromidia.DataInicioPrevisto,
            FatoControleContratosItensEuromidia.DataTerminoPrevisto,
            FatoControleContratosItensEuromidia.DataCancelamento,
            FatoControleContratosItensEuromidia.Cota,
            FatoControleContratosItensEuromidia.NumeroContrato,
            FatoControleContratosItensEuromidia.NumeroPrevia,
            FatoControleContratosItensEuromidia.FaturamentoLiquidoFinalMensal,
            FatoControleContratosItensEuromidia.TotalLiquidoContratoAGBRCTACORDO,
        )
        .filter(
            FatoControleContratosItensEuromidia.CodPonto == codponto,
            FatoControleContratosItensEuromidia.AtivoCancelamento == "A",
            FatoControleContratosItensEuromidia.DataInicioPrevisto != None,
            FatoControleContratosItensEuromidia.DataInicioPrevisto <= dt_fim,
            fim_efetivo_sql >= dt_ini,
        )
    )

    if faces:
        if tem_filtro_codface:
            q_oc = q_oc.filter(FatoControleContratosItensEuromidia.CodFace.in_(faces))
        else:
            q_oc = q_oc.filter(
                or_(
                    FatoControleContratosItensEuromidia.CodFace.in_(faces),
                    FatoControleContratosItensEuromidia.CodFace.is_(None),
                )
            )

    if filtro_cliente:
        q_oc = q_oc.filter(FatoControleContratosItensEuromidia.MarcaExibida.like(f"%{filtro_cliente}%"))

    if vendedores_selecionados or vendedores_termos_livres:
        filtros_sql_vendedor = []

        if vendedores_selecionados:
            filtros_sql_vendedor.append(
                FatoControleContratosItensEuromidia.Vendedor.in_(vendedores_selecionados)
            )

        for termo_v in (vendedores_termos_livres or []):
            termo_v = (termo_v or "").strip()
            if termo_v:
                filtros_sql_vendedor.append(
                    FatoControleContratosItensEuromidia.Vendedor.like(f"%{termo_v}%")
                )

        if filtros_sql_vendedor:
            if len(filtros_sql_vendedor) == 1:
                q_oc = q_oc.filter(filtros_sql_vendedor[0])
            else:
                q_oc = q_oc.filter(or_(*filtros_sql_vendedor))

    if contratos_segmento_permitidos is not None:
        if contratos_segmento_permitidos:
            q_oc = q_oc.filter(
                FatoControleContratosItensEuromidia.IDFatoControleContratoEuromidia.in_(contratos_segmento_permitidos)
            )
        else:
            q_oc = q_oc.filter(sa_false())

    rows = (
        q_oc.order_by(
            FatoControleContratosItensEuromidia.CodFace.asc(),
            FatoControleContratosItensEuromidia.DataInicioPrevisto.asc(),
        )
        .all()
    )

    reserva_id_original_por_iditem = {}
    spanqtd_por_iditem_reserva = {}

    rows_reservas_raw = (
        db.session.query(
            FatoOcupacaoPaineisEuromidia.IDFatoOcupacaoPaineisEuromidia,
            FatoOcupacaoPaineisEuromidia.IDFatoControleContratos,
            FatoOcupacaoPaineisEuromidia.CodFace,
            FatoOcupacaoPaineisEuromidia.MarcaExibida,
            FatoOcupacaoPaineisEuromidia.Vendedor,
            FatoOcupacaoPaineisEuromidia.DataInicio,
            FatoOcupacaoPaineisEuromidia.DataFim,
            FatoOcupacaoPaineisEuromidia.Cota,
            FatoOcupacaoPaineisEuromidia.NumeroContrato,
            FatoOcupacaoPaineisEuromidia.NumeroPrevia,
            FatoOcupacaoPaineisEuromidia.Status,
            FatoOcupacaoPaineisEuromidia.Origem,
            FatoOcupacaoPaineisEuromidia.LoopInicio,
            FatoOcupacaoPaineisEuromidia.LoopFim,
            FatoOcupacaoPaineisEuromidia.SpanQtd,
        )
        .filter(
            FatoOcupacaoPaineisEuromidia.CodPonto == codponto,
            FatoOcupacaoPaineisEuromidia.Origem == "RESERVA",
            FatoOcupacaoPaineisEuromidia.Status == "RESERVADO",
            FatoOcupacaoPaineisEuromidia.DataInicio != None,
            FatoOcupacaoPaineisEuromidia.DataInicio <= dt_fim,
            FatoOcupacaoPaineisEuromidia.DataFim >= dt_ini,
        )
        .all()
    )

    if faces:
        if tem_filtro_codface:
            rows_reservas_raw = [
                rr for rr in (rows_reservas_raw or [])
                if (_normalizar_codface(rr[2]) in faces)
            ]
        else:
            rows_reservas_raw = [
                rr for rr in (rows_reservas_raw or [])
                if (_normalizar_codface(rr[2]) in faces) or (rr[2] is None)
            ]

    if filtro_cliente:
        filtro_cliente_low = filtro_cliente.lower()
        rows_reservas_raw = [
            rr for rr in (rows_reservas_raw or [])
            if filtro_cliente_low in str(rr[3] or "").lower()
        ]

    if vendedores_selecionados or vendedores_termos_livres:
        vendedores_sel_ci = {str(v).strip().casefold() for v in (vendedores_selecionados or []) if str(v).strip()}
        termos_livres_ci = [str(v).strip().casefold() for v in (vendedores_termos_livres or []) if str(v).strip()]

        rows_reservas_raw_filtrado = []
        for rr in (rows_reservas_raw or []):
            vendedor_rr = str(rr[4] or "").strip()
            vendedor_rr_ci = vendedor_rr.casefold()

            ok_nome_exato = False
            ok_termo_livre = False

            if vendedores_sel_ci and vendedor_rr_ci in vendedores_sel_ci:
                ok_nome_exato = True

            if termos_livres_ci:
                for termo_ci in termos_livres_ci:
                    if termo_ci in vendedor_rr_ci:
                        ok_termo_livre = True
                        break

            if vendedores_sel_ci and termos_livres_ci:
                if ok_nome_exato or ok_termo_livre:
                    rows_reservas_raw_filtrado.append(rr)
            elif vendedores_sel_ci:
                if ok_nome_exato:
                    rows_reservas_raw_filtrado.append(rr)
            elif termos_livres_ci:
                if ok_termo_livre:
                    rows_reservas_raw_filtrado.append(rr)

        rows_reservas_raw = rows_reservas_raw_filtrado

    if contratos_segmento_permitidos is not None:
        if contratos_segmento_permitidos:
            contratos_segmento_permitidos_set = set(int(x) for x in contratos_segmento_permitidos)
            rows_reservas_raw = [
                rr for rr in (rows_reservas_raw or [])
                if rr[1] is not None and int(rr[1]) in contratos_segmento_permitidos_set
            ]
        else:
            rows_reservas_raw = []

    rows_reservas = []
    for rr in (rows_reservas_raw or []):
        _id_res = rr[0]
        _id_ctr = rr[1]
        _cf = rr[2]
        _marca = rr[3]
        _vend = rr[4]
        _di = rr[5]
        _df = rr[6]
        _cota = rr[7]
        _num_contrato = rr[8]
        _num_previa = rr[9]
        _span_qtd = rr[14]

        try:
            id_res_int = int(_id_res)
            id_item_reserva = -abs(id_res_int)
        except:
            id_item_reserva = -abs(hash(str(_id_res)))

        try:
            reserva_id_original_por_iditem[id_item_reserva] = int(_id_res)
        except:
            pass

        try:
            if _span_qtd is not None:
                spanqtd_por_iditem_reserva[id_item_reserva] = int(_span_qtd)
        except:
            pass

        rows_reservas.append(
            (
                id_item_reserva,
                _id_ctr,
                _cf,
                _marca,
                _vend,
                _di,
                _df,
                None,
                _cota,
                _num_contrato,
                _num_previa,
                None,
                None,
            )
        )

    if rows_reservas:
        rows = list(rows or [])
        rows.extend(rows_reservas)

    def _limpa_str(x):
        x = (x or "")
        try:
            x = str(x)
        except:
            x = ""
        x = x.strip()
        return x if x else None

    opcoes_clientes = sorted({_limpa_str(r[3]) for r in (rows or []) if _limpa_str(r[3])})
    opcoes_vendedores_grade = sorted({_limpa_str(r[4]) for r in (rows or []) if _limpa_str(r[4])})

    opcoes_vendedores = [
        (str(v["NomeVendedor"]).strip())
        for v in (vendedores_select or [])
        if str(v.get("NomeVendedor") or "").strip()
    ]

    if not opcoes_vendedores:
        opcoes_vendedores = list(opcoes_vendedores_grade)

    opcoes_segmentos = list(segmentos_select or [])
    opcoes_contratos = sorted({_limpa_str(r[9]) for r in (rows or []) if _limpa_str(r[9])})
    opcoes_previas = sorted({_limpa_str(r[10]) for r in (rows or []) if _limpa_str(r[10])})

    ocupacoes_por_slot = {}
    slots_conflito = set()
    ocupacoes_por_face = {}
    faces_conflito = set()

    def _texto_barra(cota, marca, vendedor) -> str:
        cota_txt = f"Cota {cota}" if cota not in (None, "") else "Cota"
        marca_txt = (marca or "").strip()
        vend_txt = (vendedor or "").strip()

        if marca_txt and vend_txt:
            return f"{cota_txt} {marca_txt} - {vend_txt}".strip()
        if marca_txt:
            return f"{cota_txt} {marca_txt}".strip()
        if vend_txt:
            return f"{cota_txt} - {vend_txt}".strip()
        return cota_txt

    def _intersecao_periodo(di_item: date, df_item: date, dt_ini_periodo: date, dt_fim_periodo: date):
        if di_item is None or df_item is None:
            return (None, None)

        ini_eff = di_item if di_item >= dt_ini_periodo else dt_ini_periodo
        fim_eff = df_item if df_item <= dt_fim_periodo else dt_fim_periodo

        if fim_eff < ini_eff:
            return (None, None)

        dia_ini_offset = (ini_eff - dt_ini_periodo).days + 1
        dia_fim_offset = (fim_eff - dt_ini_periodo).days + 1
        return (dia_ini_offset, dia_fim_offset)

    def _meses_no_periodo(dt_ini_local: date, dt_fim_local: date):
        meses = []
        y = dt_ini_local.year
        m = dt_ini_local.month
        while True:
            meses.append((y, m))
            if y == dt_fim_local.year and m == dt_fim_local.month:
                break
            m += 1
            if m == 13:
                m = 1
                y += 1
        return meses

    def _dias_no_mes(ano_local: int, mes_local: int) -> int:
        if mes_local in (1, 3, 5, 7, 8, 10, 12):
            return 31
        if mes_local in (4, 6, 9, 11):
            return 30
        eh_bissexto = (ano_local % 4 == 0 and (ano_local % 100 != 0 or ano_local % 400 == 0))
        return 29 if eh_bissexto else 28

    def _intersecao_mes(dt_ini_local: date, dt_fim_local: date, ano_local: int, mes_local: int):
        di_mes = date(ano_local, mes_local, 1)
        df_mes = date(ano_local, mes_local, _dias_no_mes(ano_local, mes_local))

        ini_eff = dt_ini_local if dt_ini_local >= di_mes else di_mes
        fim_eff = dt_fim_local if dt_fim_local <= df_mes else df_mes

        if fim_eff < ini_eff:
            return None, None, 0
        dias = (fim_eff - ini_eff).days + 1
        return ini_eff, fim_eff, dias

    def _buscar_custo_painel_ano_referencia(codponto_local: int, ano_ref: int, origem=None):
        try:
            q = (
                db.session.query(
                    DimCustoPainel.Ano,
                    DimCustoPainel.Valor,
                    DimCustoPainel.Origem,
                    DimCustoPainel.IDDimCustoPainel,
                    DimCustoPainel.DataCarga,
                )
                .filter(DimCustoPainel.CodPonto == int(codponto_local))
            )

            if origem:
                q = q.filter(DimCustoPainel.Origem == str(origem).strip())

            row = (
                q.order_by(
                    case((DimCustoPainel.Ano <= int(ano_ref), 0), else_=1).asc(),
                    func.abs(DimCustoPainel.Ano - int(ano_ref)).asc(),
                    DimCustoPainel.Ano.desc(),
                    DimCustoPainel.DataCarga.desc(),
                    DimCustoPainel.IDDimCustoPainel.desc(),
                )
                .first()
            )

            if not row:
                return None, None, None

            ano_db = row[0]
            valor_db = row[1]
            origem_db = row[2]

            try:
                ano_int = int(ano_db) if ano_db is not None else None
            except:
                ano_int = None

            try:
                valor_float = float(valor_db) if valor_db is not None else None
            except:
                valor_float = None

            origem_str = (str(origem_db).strip() if origem_db is not None else None) or None
            return ano_int, valor_float, origem_str

        except:
            return None, None, None

    def _span_por_cota(cota_val):
        try:
            c = int(str(cota_val).strip())
        except:
            return 1

        if c == 1080:
            return 2

        if c == 540:
            return 1

        return 1

    slots_total = None
    slots_ocupados = 0
    ocupacao_pct = None

    custo_total = None
    custos_mes = []

    margem_pct = None
    rentabilidade_valor = None

    meses_periodo = _meses_no_periodo(dt_ini, dt_fim)

    if eh_digital:
        for f in faces:
            for lp in LOOPS_PERMITIDOS:
                ocupacoes_por_slot[(f, lp)] = []

        itens_por_face = {f: [] for f in faces}

        for r in (rows or []):
            _id_item = r[0]
            _id_contrato = r[1]
            cf = _normalizar_codface(r[2]) if r[2] is not None else ""
            marca = r[3] or ""
            vend = r[4] or ""
            di = r[5]
            df_prev = r[6]
            dc = r[7]
            cota = r[8]
            num_contrato = r[9] or ""
            num_previa = r[10] or ""

            if di is None:
                continue

            df = _fim_efetivo_item(df_prev, dc)
            if df is None:
                continue

            if not cf:
                if tem_filtro_codface:
                    continue
                if faces:
                    cf = faces[0]
                else:
                    continue

            if cf not in itens_por_face:
                continue

            spans_grade = int(_span_por_cota(cota) or 1)
            if spans_grade <= 0:
                spans_grade = 1
            if spans_grade > len(LOOPS_PERMITIDOS):
                spans_grade = len(LOOPS_PERMITIDOS)

            eh_reserva = False
            try:
                eh_reserva = int(_id_item) < 0
            except:
                eh_reserva = False

            spans_kpi = spans_grade
            if eh_reserva:
                try:
                    spans_kpi_tmp = spanqtd_por_iditem_reserva.get(int(_id_item))
                    if spans_kpi_tmp is not None:
                        spans_kpi = int(spans_kpi_tmp)
                except:
                    spans_kpi = spans_grade

            dia_ini, dia_fim = _intersecao_periodo(di, df, dt_ini, dt_fim)
            if dia_ini is None:
                continue

            itens_por_face[cf].append(
                {
                    "ID": _id_item,
                    "IDFatoControleContratos": _id_contrato,
                    "CodFace": cf,
                    "MarcaExibida": marca,
                    "Loop": f"COTA {cota}",
                    "Vendedor": vend,
                    "DataInicio": di,
                    "DataFim": df,
                    "DiaInicio": dia_ini,
                    "DiaFim": dia_fim,
                    "TextoOriginal": f"CONTRATO:{num_contrato} | PRÉVIA:{num_previa}",
                    "Cota": cota,
                    "Spans": spans_grade,
                    "SpansKpi": spans_kpi,
                    "BarraTexto": _texto_barra(cota, marca, vend),
                    "EhReserva": bool(eh_reserva),
                    "OrigemItem": ("RESERVA" if eh_reserva else "CONTRATO"),
                    "CorBarra": ("#92400e" if eh_reserva else None),
                    "StatusDb": ("RESERVADO" if eh_reserva else None),
                    "OrigemDb": ("RESERVA" if eh_reserva else None),
                    "ReservaIDOriginal": (reserva_id_original_por_iditem.get(int(_id_item)) if eh_reserva else None),
                }
            )

        for cf, itens_face in (itens_por_face or {}).items():
            if not itens_face:
                continue

            itens_face = sorted(itens_face, key=lambda x: (x["DataInicio"], x["DataFim"], x["ID"]))
            fim_por_slot = {lp: None for lp in LOOPS_PERMITIDOS}

            for it in itens_face:
                ini = it["DataInicio"]
                fim = it["DataFim"]

                spans = int(it.get("Spans") or 1)
                if spans <= 0:
                    spans = 1
                if spans > len(LOOPS_PERMITIDOS):
                    spans = len(LOOPS_PERMITIDOS)

                idx_escolhido = None

                for idx in range(0, len(LOOPS_PERMITIDOS) - spans + 1):
                    ok = True
                    for off in range(spans):
                        lp = LOOPS_PERMITIDOS[idx + off]
                        fim_atual = fim_por_slot.get(lp)
                        if fim_atual is not None and ini <= fim_atual:
                            ok = False
                            break
                    if ok:
                        idx_escolhido = idx
                        break

                conflito_forcado = False

                if idx_escolhido is None:
                    conflito_forcado = True

                    melhor_idx = 0
                    melhor_val = None
                    for idx in range(0, len(LOOPS_PERMITIDOS) - spans + 1):
                        val = max(
                            [
                                (fim_por_slot.get(LOOPS_PERMITIDOS[idx + off]) or date(1900, 1, 1))
                                for off in range(spans)
                            ]
                        )
                        if melhor_val is None or val < melhor_val:
                            melhor_val = val
                            melhor_idx = idx

                    idx_escolhido = melhor_idx

                for off in range(spans):
                    lp = LOOPS_PERMITIDOS[idx_escolhido + off]

                    fim_atual = fim_por_slot.get(lp)
                    if fim_atual is None or fim > fim_atual:
                        fim_por_slot[lp] = fim

                    item_slot = dict(it)
                    item_slot["SpanAltura"] = spans
                    item_slot["SpanOffset"] = off
                    item_slot["SpanInicio"] = (off == 0)
                    item_slot["BitConflito"] = bool(conflito_forcado)

                    ocupacoes_por_slot[(cf, lp)].append(item_slot)

                    if conflito_forcado:
                        slots_conflito.add((cf, lp))

        for (cf, lp), itens in (ocupacoes_por_slot or {}).items():
            if not itens or len(itens) <= 1:
                continue

            itens_sorted = sorted(itens, key=lambda x: (x["DataInicio"], x["DataFim"], x["ID"], x.get("SpanOffset", 0)))
            conflito = False
            fim_atual = itens_sorted[0]["DataFim"]

            for i in range(1, len(itens_sorted)):
                ini = itens_sorted[i]["DataInicio"]
                fim = itens_sorted[i]["DataFim"]
                if ini <= fim_atual:
                    conflito = True
                    break
                if fim > fim_atual:
                    fim_atual = fim

            if conflito:
                slots_conflito.add((cf, lp))
                for it in itens:
                    it["BitConflito"] = True

        faces_conflito = sorted(list({cf for (cf, lp) in slots_conflito}))

        ocupacoes_por_face = {}
        for f in faces:
            agg = []
            for lp in LOOPS_PERMITIDOS:
                for it in ocupacoes_por_slot.get((f, lp), []):
                    if it.get("SpanInicio"):
                        agg.append(it)
            ocupacoes_por_face[f] = agg

    else:
        for f in faces:
            ocupacoes_por_face[f] = []

        for r in (rows or []):
            _id_item = r[0]
            _id_contrato = r[1]
            cf = _normalizar_codface(r[2]) if r[2] is not None else ""
            if not cf:
                continue

            if tem_filtro_codface and (cf.casefold() not in filtros_codface_ci):
                continue

            marca = r[3] or ""
            vend = r[4] or ""
            di = r[5]
            df_prev = r[6]
            dc = r[7]
            cota = r[8]
            num_contrato = r[9] or ""
            num_previa = r[10] or ""

            if di is None:
                continue

            df = _fim_efetivo_item(df_prev, dc)
            if df is None:
                continue

            dia_ini, dia_fim = _intersecao_periodo(di, df, dt_ini, dt_fim)
            if dia_ini is None:
                continue

            eh_reserva = False
            try:
                eh_reserva = int(_id_item) < 0
            except:
                eh_reserva = False

            ocupacoes_por_face.setdefault(cf, []).append(
                {
                    "ID": _id_item,
                    "IDFatoControleContratos": _id_contrato,
                    "CodFace": cf,
                    "MarcaExibida": marca,
                    "Loop": f"COTA {cota}",
                    "Vendedor": vend,
                    "DataInicio": di,
                    "DataFim": df,
                    "DiaInicio": dia_ini,
                    "DiaFim": dia_fim,
                    "TextoOriginal": f"CONTRATO:{num_contrato} | PRÉVIA:{num_previa}",
                    "BarraTexto": _texto_barra(cota, marca, vend),
                    "EhReserva": bool(eh_reserva),
                    "OrigemItem": ("RESERVA" if eh_reserva else "CONTRATO"),
                    "CorBarra": ("#92400e" if eh_reserva else None),
                    "StatusDb": ("RESERVADO" if eh_reserva else None),
                    "OrigemDb": ("RESERVA" if eh_reserva else None),
                    "ReservaIDOriginal": (reserva_id_original_por_iditem.get(int(_id_item)) if eh_reserva else None),
                }
            )

        faces_conflito = sorted(list(_marcar_conflitos_por_face(ocupacoes_por_face)))
        slots_conflito = set()
        ocupacoes_por_slot = {}

    from pathlib import Path
    from flask import current_app
    caminho_sql = Path(current_app.root_path) / "euromidia" / "querys" / "sql_periodo_grade.sql"
    SQL_KPI_PERIODO = caminho_sql.read_text(encoding="utf-8")

    k_digital = int(CAPACIDADE_DIGITAL_FIXA or 16)
    face_padrao = (faces[0] if faces else "").strip()

    cliente_like = _like_param(filtro_cliente)

    vendedor_like = None
    if len(vendedores_selecionados) == 1 and not vendedores_termos_livres:
        vendedor_like = _like_param(vendedores_selecionados[0])
    elif not vendedores_selecionados and len(vendedores_termos_livres) == 1:
        vendedor_like = _like_param(vendedores_termos_livres[0])

    faces_csv = ",".join([str(f).strip() for f in (faces or []) if str(f).strip()])

    kpi_params_tuple = (
        int(codponto),
        dt_ini,
        dt_fim,
        (filtro_codface or None),
        cliente_like,
        vendedor_like,
        int(k_digital),
        1 if eh_digital else 0,
        (faces_csv or None),
        (face_padrao or None),
    )

    conn = db.session.connection().connection
    cursor = conn.cursor()
    cursor.execute(SQL_KPI_PERIODO, kpi_params_tuple)

    row_kpi = cursor.fetchone()

    rows_mes = []
    try:
        cursor.nextset()
        rows_mes = cursor.fetchall() or []
    except Exception:
        rows_mes = []

    row_fin = None
    try:
        cursor.nextset()
        row_fin = cursor.fetchone()
    except Exception:
        row_fin = None

    row_cdi = None
    try:
        avancou = cursor.nextset()
        if avancou:
            row_cdi = cursor.fetchone()
        else:
            row_cdi = None
    except Exception:
        row_cdi = None

    rows_taxas_diarias = []
    try:
        avancou = cursor.nextset()
        if avancou:
            rows_taxas_diarias = cursor.fetchall() or []
        else:
            rows_taxas_diarias = []
    except Exception:
        rows_taxas_diarias = []

    slots_total = 0
    slots_ocupados = 0
    ocupacao_pct = None

    if row_kpi:
        slots_total = int(row_kpi[2] or 0)
        ocupacao_pct = float(row_kpi[5]) if row_kpi[5] is not None else None
        slots_ocupados = int(row_kpi[6] or 0)

    receita_periodo_sql = None
    custo_periodo_sql = None
    rentabilidade_valor_sql = None
    margem_pct_sql = None

    if row_fin:
        try:
            receita_periodo_sql = float(row_fin[0]) if row_fin[0] is not None else None
        except:
            receita_periodo_sql = None

        try:
            custo_periodo_sql = float(row_fin[1]) if row_fin[1] is not None else None
        except:
            custo_periodo_sql = None

        try:
            rentabilidade_valor_sql = float(row_fin[2]) if row_fin[2] is not None else None
        except:
            rentabilidade_valor_sql = None

        try:
            margem_pct_sql = float(row_fin[3]) if row_fin[3] is not None else None
        except:
            margem_pct_sql = None

    cdi_info = {
        "QtdDiasCdi": None,
        "CdiSomaPercentDia": None,
        "CdiFatorPeriodo": None,
        "CdiPercentPeriodo": None,
    }

    if row_cdi:
        try:
            cdi_info["QtdDiasCdi"] = int(row_cdi[0]) if row_cdi[0] is not None else None
        except:
            cdi_info["QtdDiasCdi"] = None

        try:
            cdi_info["CdiSomaPercentDia"] = float(row_cdi[1]) if row_cdi[1] is not None else None
        except:
            cdi_info["CdiSomaPercentDia"] = None

        try:
            cdi_info["CdiFatorPeriodo"] = float(row_cdi[2]) if row_cdi[2] is not None else None
        except:
            cdi_info["CdiFatorPeriodo"] = None

        try:
            cdi_info["CdiPercentPeriodo"] = float(row_cdi[3]) if row_cdi[3] is not None else None
        except:
            cdi_info["CdiPercentPeriodo"] = None

    taxas_diarias = []
    for rr in (rows_taxas_diarias or []):
        try:
            taxas_diarias.append(
                {
                    "DataReferencia": rr[0],
                    "CdiPercentDiaRaw": rr[1],
                    "CdiPercentDia": rr[2],
                    "CdiPercentAno": rr[3],
                    "SelicPercentDiaRaw": rr[4],
                    "SelicPercentDia": rr[5],
                    "SelicPercentAno": rr[6],
                    "DataAtualizacao": rr[7],
                }
            )
        except:
            pass

    try:
        while True:
            avancou = cursor.nextset()
            if not avancou:
                break
            if cursor.description:
                try:
                    cursor.fetchall()
                except:
                    pass
    except:
        pass

    try:
        cursor.close()
    except:
        pass

    def _buscar_retorno_indice_por_periodo(nome_tabela_qualificado: str, dt_ini_: date, dt_fim_: date):
        try:
            sql = f"""
            SET NOCOUNT ON;

            SELECT
                DataInicio = (
                    SELECT TOP 1 [Data]
                    FROM {nome_tabela_qualificado} WITH (NOLOCK)
                    WHERE [Data] >= ? AND [Data] <= ?
                    ORDER BY [Data] ASC
                ),
                DataFim = (
                    SELECT TOP 1 [Data]
                    FROM {nome_tabela_qualificado} WITH (NOLOCK)
                    WHERE [Data] >= ? AND [Data] <= ?
                    ORDER BY [Data] DESC
                ),
                ValorInicio = (
                    SELECT TOP 1 TRY_CONVERT(decimal(18,6), [PeriodoAnterior])
                    FROM {nome_tabela_qualificado} WITH (NOLOCK)
                    WHERE [Data] >= ? AND [Data] <= ?
                    ORDER BY [Data] ASC
                ),
                ValorFim = (
                    SELECT TOP 1 TRY_CONVERT(decimal(18,6), [PeriodoAtual])
                    FROM {nome_tabela_qualificado} WITH (NOLOCK)
                    WHERE [Data] >= ? AND [Data] <= ?
                    ORDER BY [Data] DESC
                ),
                QtdDias = (
                    SELECT COUNT(1)
                    FROM {nome_tabela_qualificado} WITH (NOLOCK)
                    WHERE [Data] >= ? AND [Data] <= ?
                );
            """

            cur_idx = None
            try:
                cur_idx = conn.cursor()
                params = (
                    dt_ini_, dt_fim_,
                    dt_ini_, dt_fim_,
                    dt_ini_, dt_fim_,
                    dt_ini_, dt_fim_,
                    dt_ini_, dt_fim_,
                )
                cur_idx.execute(sql, params)
                r = cur_idx.fetchone()
            finally:
                try:
                    if cur_idx is not None:
                        cur_idx.close()
                except:
                    pass

            if not r:
                return {
                    "QtdDias": None,
                    "FatorPeriodo": None,
                    "RetornoPercentPeriodo": None,
                    "ValorInicio": None,
                    "ValorFim": None,
                    "DataInicio": None,
                    "DataFim": None,
                }

            d_ini = r[0]
            d_fim = r[1]

            try:
                v_ini = float(r[2]) if r[2] is not None else None
            except:
                v_ini = None

            try:
                v_fim = float(r[3]) if r[3] is not None else None
            except:
                v_fim = None

            try:
                qtd = int(r[4]) if r[4] is not None else None
            except:
                qtd = None

            fator = None
            ret_pct = None
            try:
                if (v_ini is not None) and (v_fim is not None) and (float(v_ini) != 0.0):
                    fator = float(v_fim) / float(v_ini)
                    ret_pct = (fator - 1.0) * 100.0
            except:
                fator = None
                ret_pct = None

            return {
                "QtdDias": qtd,
                "FatorPeriodo": fator,
                "RetornoPercentPeriodo": ret_pct,
                "ValorInicio": v_ini,
                "ValorFim": v_fim,
                "DataInicio": d_ini,
                "DataFim": d_fim,
            }

        except:
            return {
                "QtdDias": None,
                "FatorPeriodo": None,
                "RetornoPercentPeriodo": None,
                "ValorInicio": None,
                "ValorFim": None,
                "DataInicio": None,
                "DataFim": None,
            }

    ooh_info = _buscar_retorno_indice_por_periodo(
        "[Integracao].[Silver].[FatoIndiceOOHDiario]",
        dt_ini,
        dt_fim
    )

    ooh_global_info = _buscar_retorno_indice_por_periodo(
        "[Integracao].[Silver].[FatoIndiceOOHGlobal]",
        dt_ini,
        dt_fim
    )

    custo_total = None
    custos_mes = []

    margem_pct = None
    rentabilidade_valor = None
    receita_liquida_periodo = None
    qtd_contratos = 0
    ticket_medio = None

    dt_fim_exclusivo = dt_fim + timedelta(days=1)

    ultimo_dia_com_contrato = None
    try:
        for r in (rows or []):
            try:
                _id_item = int(r[0]) if r[0] is not None else None
            except:
                _id_item = None
            if _id_item is not None and _id_item < 0:
                continue

            di = r[5]
            df_prev = r[6]

            if di is None or df_prev is None:
                continue

            di_dt = _coerce_to_date(di)
            df_dt = _coerce_to_date(df_prev)

            if di_dt is None or df_dt is None:
                continue

            if df_dt < dt_ini or di_dt > dt_fim:
                continue

            df_limite = df_dt if df_dt <= dt_fim else dt_fim

            if (ultimo_dia_com_contrato is None) or (df_limite > ultimo_dia_com_contrato):
                ultimo_dia_com_contrato = df_limite
    except:
        ultimo_dia_com_contrato = None

    if (total_dias <= 0) or (ultimo_dia_com_contrato is None):
        custo_total = None
        receita_liquida_periodo = None
        rentabilidade_valor = None
        margem_pct = None
        custos_mes = []
        qtd_contratos = 0
        ticket_medio = None
    else:
        dt_fim_exclusivo_ate_contrato = ultimo_dia_com_contrato + timedelta(days=1)

        ano_ref_custo = int(dt_ini.year)
        ano_custo, valor_mensal, origem_custo = _buscar_custo_painel_ano_referencia(
            int(codponto),
            ano_ref_custo,
            None
        )

        if (ano_custo is not None) and (valor_mensal is not None):
            try:
                custos_mes = [{
                    "Ano": int(ano_custo),
                    "Origem": (origem_custo or ""),
                    "ValorMensal": float(valor_mensal),
                }]
            except:
                custos_mes = []

        try:
            caminho_sql_fin = Path(current_app.root_path) / "euromidia" / "querys" / "retorna_rentabilidade_margem_grade_painel.sql"
            SQL_FIN_GRADE = caminho_sql_fin.read_text(encoding="utf-8")

            conn = db.session.connection().connection
            cursor_fin = conn.cursor()

            params_fin = (
                dt_ini,
                dt_fim_exclusivo_ate_contrato,
                int(codponto),
            )

            cursor_fin.execute(SQL_FIN_GRADE, params_fin)
            row_fin_grade = cursor_fin.fetchone()

            if row_fin_grade:

                try:
                    qtd_contratos = int(row_fin_grade[0]) if row_fin_grade[0] is not None else 0
                except:
                    qtd_contratos = 0

                if qtd_contratos <= 0:
                    custo_total = None
                    receita_liquida_periodo = None
                    rentabilidade_valor = None
                    margem_pct = None
                    ticket_medio = None
                else:
                    try:
                        receita_liquida_periodo = float(row_fin_grade[1]) if row_fin_grade[1] is not None else None
                    except:
                        receita_liquida_periodo = None

                    try:
                        custo_total = float(row_fin_grade[2]) if row_fin_grade[2] is not None else None
                    except:
                        custo_total = None

                    try:
                        rentabilidade_valor = float(row_fin_grade[3]) if row_fin_grade[3] is not None else None
                    except:
                        rentabilidade_valor = None

                    try:
                        margem_pct = float(row_fin_grade[4]) if row_fin_grade[4] is not None else None
                    except:
                        margem_pct = None

                    try:
                        if (receita_liquida_periodo is not None) and (qtd_contratos > 0):
                            ticket_medio = float(receita_liquida_periodo) / float(qtd_contratos)
                        else:
                            ticket_medio = None
                    except:
                        ticket_medio = None
            else:
                custo_total = None
                receita_liquida_periodo = None
                rentabilidade_valor = None
                margem_pct = None
                qtd_contratos = 0
                ticket_medio = None

        except:
            custo_total = None
            receita_liquida_periodo = None
            rentabilidade_valor = None
            margem_pct = None
            qtd_contratos = 0
            ticket_medio = None

    ano_prev, mes_prev = ano, mes - 1
    if mes_prev == 0:
        mes_prev = 12
        ano_prev -= 1
    mes_ref_prev = f"{ano_prev:04d}-{mes_prev:02d}"

    ano_next, mes_next = ano, mes + 1
    if mes_next == 13:
        mes_next = 1
        ano_next += 1
    mes_ref_next = f"{ano_next:04d}-{mes_next:02d}"

    dias = []
    dt_cursor = dt_ini
    idx = 1
    while dt_cursor <= dt_fim:
        is_fds = dt_cursor.weekday() >= 5
        dias.append(
            {
                "dia": idx,
                "label": dt_cursor.strftime("%d/%m"),
                "is_fds": is_fds,
                "data": dt_cursor,
            }
        )
        dt_cursor = dt_cursor + timedelta(days=1)
        idx += 1

    return render_template(
        "euromidia/painel_grade.html",
        codponto=codponto,
        tipo_prod=tipo_prod,
        exibidora=exibidora,
        num_faces=num_faces,
        mes_ref=mes_ref,
        mes_ref_prev=mes_ref_prev,
        mes_ref_next=mes_ref_next,
        mes_de=mes_de,
        mes_ate=mes_ate,
        dias=dias,
        ultimo_dia=ultimo_dia,
        total_dias=total_dias,
        faces=faces,
        codfaces_select=codfaces_select,
        codface_selecionado=codface_selecionado,
        codfaces_selecionadas=filtros_codface,
        ocupacoes_por_slot=ocupacoes_por_slot,
        slots_conflito=sorted(list(slots_conflito)),
        loops_permitidos=LOOPS_PERMITIDOS,
        eh_digital=eh_digital,
        ocupacoes_por_face=ocupacoes_por_face,
        faces_conflito=faces_conflito,
        filtros={
            "cliente": filtro_cliente,
            "vendedor": filtro_vendedor,
            "vendedores": vendedores_selecionados,
            "vendedores_texto_livre": vendedores_termos_livres,
            "segmento": filtro_segmento,
            "segmentos": segmentos_selecionados,
            "segmentos_texto_livre": segmentos_termos_livres,
            "tipo": (filtro_tipo_prod or ""),
            "codface": (filtro_codface or ""),
            "codfaces": list(filtros_codface),
        },
        ocupacao_pct=ocupacao_pct,
        slots_ocupados=slots_ocupados,
        slots_total=slots_total,
        custo_total=custo_total,
        margem_pct=margem_pct,
        rentabilidade_valor=rentabilidade_valor,
        receita_total=receita_liquida_periodo,
        qtd_contratos=qtd_contratos,
        ticket_medio=ticket_medio,
        opcoes_clientes=opcoes_clientes,
        opcoes_vendedores=opcoes_vendedores,
        opcoes_vendedores_grade=opcoes_vendedores_grade,
        opcoes_segmentos=opcoes_segmentos,
        opcoes_contratos=opcoes_contratos,
        opcoes_previas=opcoes_previas,
        vendedores_select=vendedores_select,
        vendedores_selecionados=vendedores_selecionados,
        vendedores_termos_livres=vendedores_termos_livres,
        segmentos_select=segmentos_select,
        segmentos_selecionados=segmentos_selecionados,
        segmentos_termos_livres=segmentos_termos_livres,
        cdi_info=cdi_info,
        taxas_diarias=taxas_diarias,
        meses_select=meses_select,
        bisemanas_select=bisemanas_select,
        ooh_info=ooh_info,
        indice_ooh_info=ooh_info,
        ooh_global_info=ooh_global_info,
        indice_ooh_global_info=ooh_global_info,
        tipo_painel_info=tipo_painel_info,
        cidade_painel=cidade_painel,
        uf_painel=uf_painel,
        bairro_painel=bairro_painel,
        logradouro_painel=logradouro_painel,
    )


















@paineis_bp.route("/grade/multi", methods=["GET"])
@login_required
@limiter.limit("80 per minute", methods=["GET"])
@retry_get_view(db, attempts=6, base_delay=0.2, max_delay=1.5)
def grade_painel_multi():

    from urllib.parse import urlencode
    from pathlib import Path
    from flask import current_app
    from datetime import date, datetime, timedelta
    import calendar
    import math
    from sqlalchemy import func, or_

    try:
        CAPACIDADE_DIGITAL_PADRAO = int(globals().get("CAPACIDADE_DIGITAL_FIXA", 16) or 16)
    except:
        CAPACIDADE_DIGITAL_PADRAO = 16

    def _normalizar_texto(valor):
        try:
            return str(valor or "").strip()
        except:
            return ""

    def _normalizar_codface(valor):
        return _normalizar_texto(valor)

    def _coerce_to_date(valor):
        if valor is None:
            return None

        if isinstance(valor, datetime):
            return valor.date()

        if isinstance(valor, date):
            return valor

        texto = _normalizar_texto(valor)
        if not texto:
            return None

        formatos = ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f")
        for fmt in formatos:
            try:
                return datetime.strptime(texto, fmt).date()
            except:
                pass

        try:
            return datetime.fromisoformat(texto).date()
        except:
            return None

    def _extrair_mes_referencia(valor):
        texto = _normalizar_texto(valor)
        if not texto:
            return ""

        if len(texto) >= 7 and texto[4] == "-":
            ano_txt = texto[:4]
            mes_txt = texto[5:7]
            if ano_txt.isdigit() and mes_txt.isdigit():
                ano_int = int(ano_txt)
                mes_int = int(mes_txt)
                if 1 <= mes_int <= 12:
                    return f"{ano_int:04d}-{mes_int:02d}"

        return ""

    def _somar_meses(ano: int, mes: int, delta: int):
        indice = (ano * 12) + (mes - 1) + delta
        novo_ano = indice // 12
        novo_mes = (indice % 12) + 1
        return novo_ano, novo_mes

    def _resolver_mes_ref_base():
        hoje = date.today()

        for chave in ("mes_ref", "mes_de", "mes_ate", "data_ref"):
            mes_ref_tmp = _extrair_mes_referencia(request.args.get(chave))
            if mes_ref_tmp:
                return mes_ref_tmp

        return f"{hoje.year:04d}-{hoje.month:02d}"

    def _montar_info_periodo():
        mes_ref_base = _resolver_mes_ref_base()

        try:
            ano = int(mes_ref_base[:4])
            mes = int(mes_ref_base[5:7])
        except:
            hoje = date.today()
            ano = hoje.year
            mes = hoje.month
            mes_ref_base = f"{ano:04d}-{mes:02d}"

        dt_ini_mes = date(ano, mes, 1)
        ultimo_dia_mes = calendar.monthrange(ano, mes)[1]
        dt_fim_mes = date(ano, mes, ultimo_dia_mes)

        dt_ini = dt_ini_mes
        dt_fim = dt_fim_mes

        mes_de_arg = _coerce_to_date(request.args.get("mes_de"))
        mes_ate_arg = _coerce_to_date(request.args.get("mes_ate"))
        data_ref_arg = _coerce_to_date(request.args.get("data_ref"))
        modo = _normalizar_texto(request.args.get("modo")).lower()

        if mes_de_arg and mes_ate_arg:
            dt_ini = mes_de_arg
            dt_fim = mes_ate_arg
            if dt_fim < dt_ini:
                dt_ini, dt_fim = dt_fim, dt_ini
        elif mes_de_arg:
            dt_ini = mes_de_arg
            dt_fim = dt_fim_mes
            if dt_fim < dt_ini:
                dt_fim = dt_ini
        elif mes_ate_arg:
            dt_ini = dt_ini_mes
            dt_fim = mes_ate_arg
            if dt_fim < dt_ini:
                dt_ini = dt_fim
        elif modo == "dia" and data_ref_arg:
            dt_ini = data_ref_arg
            dt_fim = data_ref_arg
        elif modo == "semana" and data_ref_arg:
            inicio_semana = data_ref_arg - timedelta(days=data_ref_arg.weekday())
            fim_semana = inicio_semana + timedelta(days=6)
            dt_ini = inicio_semana
            dt_fim = fim_semana
        else:
            dt_ini = dt_ini_mes
            dt_fim = dt_fim_mes

        dias = []
        dt_cursor = dt_ini
        idx = 1
        while dt_cursor <= dt_fim:
            dias.append({
                "dia": idx,
                "label": dt_cursor.strftime("%d/%m"),
                "is_fds": dt_cursor.weekday() >= 5,
                "data": dt_cursor,
            })
            dt_cursor = dt_cursor + timedelta(days=1)
            idx += 1

        ano_prev, mes_prev = _somar_meses(ano, mes, -1)
        ano_next, mes_next = _somar_meses(ano, mes, 1)

        return {
            "mes_ref": f"{ano:04d}-{mes:02d}",
            "mes_ref_prev": f"{ano_prev:04d}-{mes_prev:02d}",
            "mes_ref_next": f"{ano_next:04d}-{mes_next:02d}",
            "ano": ano,
            "mes": mes,
            "dt_ini": dt_ini,
            "dt_fim": dt_fim,
            "dias": dias,
            "ultimo_dia": len(dias),
            "total_dias": len(dias),
        }

    def _lista_args_limpa(nome_param):
        valores = []

        try:
            for v in request.args.getlist(nome_param):
                s = _normalizar_texto(v)
                if s:
                    valores.append(s)
        except:
            pass

        if not valores:
            bruto = _normalizar_texto(request.args.get(nome_param))
            if bruto:
                if "|" in bruto:
                    partes = [x.strip() for x in bruto.split("|") if x.strip()]
                    valores.extend(partes)
                elif ";" in bruto:
                    partes = [x.strip() for x in bruto.split(";") if x.strip()]
                    valores.extend(partes)
                else:
                    valores.append(bruto)

        limpos = []
        vistos = set()

        for v in valores:
            vv = _normalizar_texto(v)
            if not vv:
                continue

            chave = vv.casefold()
            if chave in vistos:
                continue

            vistos.add(chave)
            limpos.append(vv)

        return limpos

    def _limpar_valores_reais(valores):
        saida = []
        vistos = set()

        for v in (valores or []):
            vv = _normalizar_texto(v)
            if not vv:
                continue

            vv_low = vv.casefold()
            if vv_low.startswith("todos"):
                continue
            if vv_low.startswith("todas"):
                continue

            if vv_low in vistos:
                continue

            vistos.add(vv_low)
            saida.append(vv)

        return saida

    def _lista_codpontos_validos():
        codpontos = []
        vistos = set()

        for bruto in (_lista_args_limpa("codponto") or []):
            try:
                cp_int = int(str(bruto).strip())
            except:
                continue

            if cp_int <= 0:
                continue

            if cp_int in vistos:
                continue

            vistos.add(cp_int)
            codpontos.append(cp_int)

        return codpontos

    def _lista_pares_codponto_codface():
        pares = []

        codfaces_raw = _lista_args_limpa("codface")
        codpontos_raw = _lista_args_limpa("codponto")

        if not codfaces_raw or not codpontos_raw:
            return pares

        if len(codfaces_raw) != len(codpontos_raw):
            return pares

        for cf_bruto, cp_bruto in zip(codfaces_raw, codpontos_raw):
            cf_norm = _normalizar_codface(cf_bruto)
            if not cf_norm:
                continue

            try:
                cp_int = int(str(cp_bruto).strip())
            except:
                continue

            if cp_int <= 0:
                continue

            pares.append((cp_int, cf_norm))

        return pares

    def _montar_slots_digitais(capacidade: int):
        try:
            capacidade_int = int(capacidade or 0)
        except:
            capacidade_int = 0

        if capacidade_int <= 0:
            capacidade_int = CAPACIDADE_DIGITAL_PADRAO

        return [f"SPAN{i:02d}" for i in range(1, capacidade_int + 1)]

    def _resolver_codponto_por_codface_global(codface_valor: str, codponto_hint: int = None, codpontos_preferidos=None):
        codface_norm = _normalizar_codface(codface_valor)
        if not codface_norm:
            return None

        codface_up = codface_norm.upper()
        codpontos_preferidos_set = set()

        for cp in (codpontos_preferidos or []):
            try:
                cp_int = int(cp)
            except:
                continue

            if cp_int > 0:
                codpontos_preferidos_set.add(cp_int)

        rows_face_global = (
            db.session.query(
                DimFacesPaineis.CodPonto,
                DimFacesPaineis.CodFace,
                DimFacesPaineis.IDDimFacesPaineis,
            )
            .filter(
                DimFacesPaineis.CodPonto != None,
                DimFacesPaineis.CodFace != None,
                DimFacesPaineis.CodFace != "",
                func.upper(func.ltrim(func.rtrim(DimFacesPaineis.CodFace))) == codface_up,
            )
            .order_by(
                DimFacesPaineis.IDDimFacesPaineis.desc(),
            )
            .all()
        )

        if not rows_face_global:
            return None

        if codponto_hint is not None:
            try:
                codponto_hint = int(codponto_hint)
            except:
                codponto_hint = None

        if codponto_hint is not None:
            for cp_row, cf_row, _id_row in (rows_face_global or []):
                try:
                    cp_int = int(cp_row) if cp_row is not None else None
                except:
                    cp_int = None

                cf_norm = _normalizar_codface(cf_row)
                if cp_int is None or not cf_norm:
                    continue

                if cp_int == codponto_hint:
                    return cp_int

        if codpontos_preferidos_set:
            for cp_row, cf_row, _id_row in (rows_face_global or []):
                try:
                    cp_int = int(cp_row) if cp_row is not None else None
                except:
                    cp_int = None

                cf_norm = _normalizar_codface(cf_row)
                if cp_int is None or not cf_norm:
                    continue

                if cp_int in codpontos_preferidos_set:
                    return cp_int

        for cp_row, cf_row, _id_row in (rows_face_global or []):
            try:
                cp_int = int(cp_row) if cp_row is not None else None
            except:
                cp_int = None

            cf_norm = _normalizar_codface(cf_row)
            if cp_int is None or not cf_norm:
                continue

            return cp_int

        return None

    def _montar_url_grade_unica(codponto_destino: int, codfaces_painel=None):
        parametros = []

        mes_ref = _normalizar_texto(request.args.get("mes_ref"))
        cliente = _normalizar_texto(request.args.get("cliente"))
        tipo = _normalizar_texto(request.args.get("tipo"))
        mes_de = _normalizar_texto(request.args.get("mes_de"))
        mes_ate = _normalizar_texto(request.args.get("mes_ate"))
        modo = _normalizar_texto(request.args.get("modo"))
        data_ref = _normalizar_texto(request.args.get("data_ref"))
        bi_semana = _normalizar_texto(request.args.get("bi_semana"))

        if mes_ref:
            parametros.append(("mes_ref", mes_ref))
        if cliente:
            parametros.append(("cliente", cliente))
        if tipo:
            parametros.append(("tipo", tipo))
        if mes_de:
            parametros.append(("mes_de", mes_de))
        if mes_ate:
            parametros.append(("mes_ate", mes_ate))
        if modo:
            parametros.append(("modo", modo))
        if data_ref:
            parametros.append(("data_ref", data_ref))
        if bi_semana:
            parametros.append(("bi_semana", bi_semana))

        vendedores = _lista_args_limpa("vendedor")
        segmentos = _lista_args_limpa("segmento")

        for vendedor in (vendedores or []):
            parametros.append(("vendedor", vendedor))

        for segmento in (segmentos or []):
            parametros.append(("segmento", segmento))

        lista_codfaces = []
        vistos_codfaces = set()

        for cf in (codfaces_painel or []):
            cf_norm = _normalizar_codface(cf)
            if not cf_norm:
                continue

            chave = cf_norm.casefold()
            if chave in vistos_codfaces:
                continue

            vistos_codfaces.add(chave)
            lista_codfaces.append(cf_norm)

        for cf in lista_codfaces:
            parametros.append(("codface", cf))

        if lista_codfaces:
            parametros.append(("face_principal", lista_codfaces[0]))

        url_base = url_for("Paineis.grade_painel", codponto=int(codponto_destino))

        if not parametros:
            return url_base

        return f"{url_base}?{urlencode(parametros, doseq=True)}"

    def _agrupar_faces_por_painel(codfaces_lista, codpontos_explicitos):
        faces_por_painel = {}
        faces_nao_resolvidas = []

        for cp in (codpontos_explicitos or []):
            try:
                cp_int = int(cp)
            except:
                continue

            if cp_int > 0 and cp_int not in faces_por_painel:
                faces_por_painel[cp_int] = []

        pares_explicitos = _lista_pares_codponto_codface()
        pares_processados_ci = set()

        for cp_int, cf_norm in (pares_explicitos or []):
            if cp_int <= 0 or not cf_norm:
                continue

            if cp_int not in faces_por_painel:
                faces_por_painel[cp_int] = []

            chave_par = (cp_int, cf_norm.casefold())
            if chave_par in pares_processados_ci:
                continue

            pares_processados_ci.add(chave_par)

            ja_tem = {x.casefold() for x in faces_por_painel[cp_int]}
            if cf_norm.casefold() not in ja_tem:
                faces_por_painel[cp_int].append(cf_norm)

        for idx, cf in enumerate(codfaces_lista or []):
            cf_norm = _normalizar_codface(cf)
            if not cf_norm:
                continue

            ja_coberta = False
            for cp_existente, lista_faces_existente in (faces_por_painel or {}).items():
                if cf_norm.casefold() in {x.casefold() for x in (lista_faces_existente or [])}:
                    ja_coberta = True
                    break

            if ja_coberta:
                continue

            codponto_hint = None
            if idx < len(codpontos_explicitos or []):
                try:
                    codponto_hint = int((codpontos_explicitos or [])[idx])
                except:
                    codponto_hint = None

            cp_resolvido = _resolver_codponto_por_codface_global(
                codface_valor=cf_norm,
                codponto_hint=codponto_hint,
                codpontos_preferidos=codpontos_explicitos,
            )

            if cp_resolvido is None:
                faces_nao_resolvidas.append(cf_norm)
                continue

            cp_int = int(cp_resolvido)

            if cp_int not in faces_por_painel:
                faces_por_painel[cp_int] = []

            ja_tem = {x.casefold() for x in faces_por_painel[cp_int]}
            if cf_norm.casefold() not in ja_tem:
                faces_por_painel[cp_int].append(cf_norm)

        return faces_por_painel, faces_nao_resolvidas

    def _buscar_resumo_painel(codponto_local: int):
        tipo_painel_norm = func.upper(func.ltrim(func.rtrim(func.coalesce(DimPaineisEuromidia.Tipo, ""))))

        rn = func.row_number().over(
            partition_by=(DimPaineisEuromidia.CodPonto, tipo_painel_norm),
            order_by=(
                DimPaineisEuromidia.DataAtualizacao.desc(),
                DimPaineisEuromidia.IDDimPaineisEuromidia.desc(),
            ),
        ).label("rn")

        sub = (
            db.session.query(
                DimPaineisEuromidia.CodPonto.label("CodPonto"),
                DimPaineisEuromidia.Tipo.label("Tipo"),
                DimPaineisEuromidia.Exibidora.label("Exibidora"),
                DimPaineisEuromidia.Cidade.label("Cidade"),
                DimPaineisEuromidia.UF.label("UF"),
                DimPaineisEuromidia.Bairro.label("Bairro"),
                DimPaineisEuromidia.Logradouro.label("Logradouro"),
                rn,
            )
            .filter(DimPaineisEuromidia.CodPonto == int(codponto_local))
            .subquery()
        )

        row = (
            db.session.query(sub)
            .filter(sub.c.rn == 1)
            .first()
        )

        if not row:
            return {
                "CodPonto": int(codponto_local),
                "Tipo": "",
                "Exibidora": "",
                "Cidade": "",
                "UF": "",
                "Bairro": "",
                "Logradouro": "",
            }

        return {
            "CodPonto": int(codponto_local),
            "Tipo": (_normalizar_texto(getattr(row, "Tipo", ""))),
            "Exibidora": (_normalizar_texto(getattr(row, "Exibidora", ""))),
            "Cidade": (_normalizar_texto(getattr(row, "Cidade", ""))),
            "UF": (_normalizar_texto(getattr(row, "UF", ""))),
            "Bairro": (_normalizar_texto(getattr(row, "Bairro", ""))),
            "Logradouro": (_normalizar_texto(getattr(row, "Logradouro", ""))),
        }

    def _montar_localizacao(resumo_painel: dict):
        partes = []

        bairro = _normalizar_texto(resumo_painel.get("Bairro"))
        cidade = _normalizar_texto(resumo_painel.get("Cidade"))
        uf = _normalizar_texto(resumo_painel.get("UF"))
        logradouro = _normalizar_texto(resumo_painel.get("Logradouro"))

        if bairro:
            partes.append(bairro)

        if cidade and uf:
            partes.append(f"{cidade}/{uf}")
        elif cidade:
            partes.append(cidade)
        elif uf:
            partes.append(uf)

        if logradouro:
            partes.append(logradouro)

        return " • ".join([p for p in partes if p])

    def _fim_efetivo_item(data_fim_prevista, data_cancelamento):
        dc = _coerce_to_date(data_cancelamento)
        if dc is not None:
            return dc

        df = _coerce_to_date(data_fim_prevista)
        if df is not None:
            return df

        return None

    def _intersecao_periodo(data_inicio, data_fim, dt_ini, dt_fim):
        di = _coerce_to_date(data_inicio)
        df = _coerce_to_date(data_fim)

        if di is None or df is None:
            return None, None

        if df < di:
            return None, None

        ini_clip = max(di, dt_ini)
        fim_clip = min(df, dt_fim)

        if fim_clip < ini_clip:
            return None, None

        dia_ini = int((ini_clip - dt_ini).days) + 1
        dia_fim = int((fim_clip - dt_ini).days) + 1
        return dia_ini, dia_fim

    def _texto_barra(cota, marca, vend):
        partes = []

        try:
            if cota not in (None, "", 0, "0"):
                partes.append(f"COTA {int(cota)}")
        except:
            if cota not in (None, ""):
                partes.append(f"COTA {cota}")

        marca_txt = _normalizar_texto(marca)
        vend_txt = _normalizar_texto(vend)

        if marca_txt:
            partes.append(marca_txt)
        if vend_txt:
            partes.append(vend_txt)

        return " • ".join([p for p in partes if p])

    def _marcar_conflitos_por_face(ocupacoes_por_face):
        faces_conflitadas = set()

        for cf, itens in (ocupacoes_por_face or {}).items():
            if not itens or len(itens) <= 1:
                continue

            itens_ordenados = sorted(
                itens,
                key=lambda x: (
                    _coerce_to_date(x.get("DataInicio")) or date(1900, 1, 1),
                    _coerce_to_date(x.get("DataFim")) or date(1900, 1, 1),
                    x.get("ID") or 0,
                ),
            )

            for i in range(len(itens_ordenados) - 1):
                atual = itens_ordenados[i]
                proximo = itens_ordenados[i + 1]

                fim_atual = _coerce_to_date(atual.get("DataFim"))
                ini_prox = _coerce_to_date(proximo.get("DataInicio"))

                if fim_atual is None or ini_prox is None:
                    continue

                if ini_prox <= fim_atual:
                    faces_conflitadas.add(cf)
                    atual["BitConflito"] = True
                    proximo["BitConflito"] = True

        return faces_conflitadas

    def _span_por_cota(cota_val):
        try:
            c = int(str(cota_val).strip())
        except:
            return 1

        if c == 1080:
            return 2

        if c == 540:
            return 1

        return 1

    def _slots_por_cota(cota_val):
        try:
            c = int(str(cota_val).strip())
        except:
            return 1.0

        if c == 1080:
            return 2.0

        if c == 540:
            return 1.0

        return 1.0

    def _uso_no_periodo_por_intervalos(intervalos, dt_ini, dt_fim, denom_cap):
        try:
            if not intervalos:
                return 0.0, 0.0

            eventos = []

            for di, df, slots in (intervalos or []):
                di_dt = _coerce_to_date(di)
                df_dt = _coerce_to_date(df)

                if di_dt is None or df_dt is None:
                    continue

                if df_dt < di_dt:
                    continue

                ini_clip = max(di_dt, dt_ini)
                fim_clip = min(df_dt, dt_fim)

                if fim_clip < ini_clip:
                    continue

                eventos.append((ini_clip, float(slots)))
                eventos.append((fim_clip + timedelta(days=1), -float(slots)))

            if not eventos:
                return 0.0, 0.0

            eventos.sort(key=lambda x: (x[0], 0 if x[1] > 0 else 1))

            cursor = dt_ini
            atual = 0.0
            maximo = 0.0
            uso_slots_dias = 0.0

            for data_evento, delta in eventos:
                if data_evento > (dt_fim + timedelta(days=1)):
                    data_evento = dt_fim + timedelta(days=1)

                if data_evento > cursor:
                    dias_trecho = (data_evento - cursor).days
                    if dias_trecho > 0:
                        ocupado = atual
                        if ocupado < 0:
                            ocupado = 0.0
                        if ocupado > float(denom_cap):
                            ocupado = float(denom_cap)

                        uso_slots_dias += float(ocupado) * float(dias_trecho)

                    cursor = data_evento

                atual += float(delta)
                if atual > maximo:
                    maximo = atual

            if cursor < (dt_fim + timedelta(days=1)):
                dias_trecho = ((dt_fim + timedelta(days=1)) - cursor).days
                if dias_trecho > 0:
                    ocupado = atual
                    if ocupado < 0:
                        ocupado = 0.0
                    if ocupado > float(denom_cap):
                        ocupado = float(denom_cap)

                    uso_slots_dias += float(ocupado) * float(dias_trecho)

            return float(uso_slots_dias), float(maximo)
        except:
            return 0.0, 0.0

    def _buscar_financeiro_periodo(codponto_local, dt_ini, dt_fim, possui_itens):
        qtd_contratos = 0
        receita_liquida_periodo = None
        custo_total = None
        rentabilidade_valor = None
        margem_pct = None
        ticket_medio = None

        if not possui_itens:
            return {
                "qtd_contratos": 0,
                "receita_total": None,
                "custo_total": None,
                "rentabilidade_valor": None,
                "margem_pct": None,
                "ticket_medio": None,
            }

        try:
            caminho_sql_fin = Path(current_app.root_path) / "euromidia" / "querys" / "retorna_rentabilidade_margem_grade_painel.sql"
            sql_fin = caminho_sql_fin.read_text(encoding="utf-8")

            dt_fim_exclusivo = dt_fim + timedelta(days=1)

            raw_conn_fin = db.engine.raw_connection()
            cursor_fin = raw_conn_fin.cursor()

            try:
                cursor_fin.execute(
                    sql_fin,
                    (
                        dt_ini,
                        dt_fim_exclusivo,
                        int(codponto_local),
                    ),
                )
                row_fin_grade = cursor_fin.fetchone()

                try:
                    while cursor_fin.nextset():
                        pass
                except:
                    pass

            finally:
                try:
                    cursor_fin.close()
                except:
                    pass

                try:
                    raw_conn_fin.close()
                except:
                    pass

            if row_fin_grade:
                try:
                    qtd_contratos = int(row_fin_grade[0]) if row_fin_grade[0] is not None else 0
                except:
                    qtd_contratos = 0

                if qtd_contratos > 0:
                    try:
                        receita_liquida_periodo = float(row_fin_grade[1]) if row_fin_grade[1] is not None else None
                    except:
                        receita_liquida_periodo = None

                    try:
                        custo_total = float(row_fin_grade[2]) if row_fin_grade[2] is not None else None
                    except:
                        custo_total = None

                    try:
                        rentabilidade_valor = float(row_fin_grade[3]) if row_fin_grade[3] is not None else None
                    except:
                        rentabilidade_valor = None

                    try:
                        margem_pct = float(row_fin_grade[4]) if row_fin_grade[4] is not None else None
                    except:
                        margem_pct = None

                    try:
                        if (receita_liquida_periodo is not None) and (qtd_contratos > 0):
                            ticket_medio = float(receita_liquida_periodo) / float(qtd_contratos)
                        else:
                            ticket_medio = None
                    except:
                        ticket_medio = None

            return {
                "qtd_contratos": qtd_contratos,
                "receita_total": receita_liquida_periodo,
                "custo_total": custo_total,
                "rentabilidade_valor": rentabilidade_valor,
                "margem_pct": margem_pct,
                "ticket_medio": ticket_medio,
            }

        except Exception:
            current_app.logger.exception(
                "Erro ao buscar financeiro do período. codponto=%s dt_ini=%s dt_fim=%s",
                codponto_local,
                dt_ini,
                dt_fim,
            )
            return {
                "qtd_contratos": 0,
                "receita_total": None,
                "custo_total": None,
                "rentabilidade_valor": None,
                "margem_pct": None,
                "ticket_medio": None,
            }

    def _buscar_indicadores_periodo(dt_ini, dt_fim):
        cdi_info = {
            "DataInicio": None,
            "DataFim": None,
            "QtdDias": None,
            "CdiAcumulado": None,
            "CdiFatorPeriodo": None,
            "CdiPercentPeriodo": None,
        }
        taxas_diarias = []

        ooh_info = {
            "QtdDias": None,
            "FatorPeriodo": None,
            "RetornoPercentPeriodo": None,
            "ValorInicio": None,
            "ValorFim": None,
            "DataInicio": None,
            "DataFim": None,
        }

        ooh_global_info = {
            "QtdDias": None,
            "FatorPeriodo": None,
            "RetornoPercentPeriodo": None,
            "ValorInicio": None,
            "ValorFim": None,
            "DataInicio": None,
            "DataFim": None,
        }

        def _to_int(valor):
            try:
                return int(valor) if valor is not None else None
            except:
                return None

        def _to_float(valor):
            try:
                return float(valor) if valor is not None else None
            except:
                return None

        def _normalizar_dict_indicador(ind):
            base = {
                "QtdDias": None,
                "FatorPeriodo": None,
                "RetornoPercentPeriodo": None,
                "ValorInicio": None,
                "ValorFim": None,
                "DataInicio": None,
                "DataFim": None,
            }
            if not ind:
                return dict(base)

            saida = dict(base)
            try:
                for k, v in dict(ind).items():
                    saida[k] = v
            except:
                pass
            return saida

        def _drenar_resultsets(cursor_local):
            if cursor_local is None:
                return
            try:
                while True:
                    possui_mais = cursor_local.nextset()
                    if not possui_mais:
                        break
            except:
                pass

        def _buscar_retorno_indice_por_periodo(conn_local, nome_tabela_qualificado: str, dt_ini_: date, dt_fim_: date):
            cur_idx = None

            try:
                sql_idx = f"""
                SELECT
                    DataInicio = (
                        SELECT TOP 1 [Data]
                        FROM {nome_tabela_qualificado} WITH (NOLOCK)
                        WHERE [Data] >= ? AND [Data] <= ?
                        ORDER BY [Data] ASC
                    ),
                    DataFim = (
                        SELECT TOP 1 [Data]
                        FROM {nome_tabela_qualificado} WITH (NOLOCK)
                        WHERE [Data] >= ? AND [Data] <= ?
                        ORDER BY [Data] DESC
                    ),
                    ValorInicio = (
                        SELECT TOP 1 TRY_CONVERT(decimal(18,6), [PeriodoAnterior])
                        FROM {nome_tabela_qualificado} WITH (NOLOCK)
                        WHERE [Data] >= ? AND [Data] <= ?
                        ORDER BY [Data] ASC
                    ),
                    ValorFim = (
                        SELECT TOP 1 TRY_CONVERT(decimal(18,6), [PeriodoAtual])
                        FROM {nome_tabela_qualificado} WITH (NOLOCK)
                        WHERE [Data] >= ? AND [Data] <= ?
                        ORDER BY [Data] DESC
                    ),
                    QtdDias = (
                        SELECT COUNT(1)
                        FROM {nome_tabela_qualificado} WITH (NOLOCK)
                        WHERE [Data] >= ? AND [Data] <= ?
                    );
                """

                params_idx = (
                    dt_ini_, dt_fim_,
                    dt_ini_, dt_fim_,
                    dt_ini_, dt_fim_,
                    dt_ini_, dt_fim_,
                    dt_ini_, dt_fim_,
                )

                cur_idx = conn_local.cursor()
                cur_idx.execute(sql_idx, params_idx)
                row_idx = cur_idx.fetchone()
                _drenar_resultsets(cur_idx)

                if not row_idx:
                    return _normalizar_dict_indicador(None)

                data_inicio_idx = row_idx[0]
                data_fim_idx = row_idx[1]
                valor_inicio_idx = _to_float(row_idx[2])
                valor_fim_idx = _to_float(row_idx[3])
                qtd_dias_idx = _to_int(row_idx[4])

                fator_periodo_idx = None
                retorno_percent_periodo_idx = None

                if (
                    valor_inicio_idx is not None
                    and valor_fim_idx is not None
                    and float(valor_inicio_idx) != 0.0
                ):
                    try:
                        fator_periodo_idx = float(valor_fim_idx) / float(valor_inicio_idx)
                        retorno_percent_periodo_idx = (float(fator_periodo_idx) - 1.0) * 100.0
                    except:
                        fator_periodo_idx = None
                        retorno_percent_periodo_idx = None

                return _normalizar_dict_indicador({
                    "QtdDias": qtd_dias_idx,
                    "FatorPeriodo": fator_periodo_idx,
                    "RetornoPercentPeriodo": retorno_percent_periodo_idx,
                    "ValorInicio": valor_inicio_idx,
                    "ValorFim": valor_fim_idx,
                    "DataInicio": data_inicio_idx,
                    "DataFim": data_fim_idx,
                })

            except Exception:
                current_app.logger.exception(
                    "Erro ao buscar índice por período. tabela=%s dt_ini=%s dt_fim=%s",
                    nome_tabela_qualificado,
                    dt_ini_,
                    dt_fim_,
                )
                return _normalizar_dict_indicador(None)
            finally:
                try:
                    if cur_idx is not None:
                        cur_idx.close()
                except:
                    pass

        if dt_ini is None or dt_fim is None:
            return cdi_info, taxas_diarias, ooh_info, ooh_global_info

        raw_conn = None
        cur_cdi = None
        cur_taxas = None

        try:
            raw_conn = db.engine.raw_connection()

            sql_cdi_resumo = """
            SELECT
                MIN(DataReferencia) AS DataInicio,
                MAX(DataReferencia) AS DataFim,
                COUNT(1) AS QtdDias,
                CAST(EXP(SUM(LOG(1 + TRY_CONVERT(float, CdiPercentDia) / 100.0))) - 1 AS decimal(18,10)) AS CdiAcumulado,
                CAST(EXP(SUM(LOG(1 + TRY_CONVERT(float, CdiPercentDia) / 100.0))) AS decimal(18,10)) AS CdiFatorPeriodo,
                CAST((EXP(SUM(LOG(1 + TRY_CONVERT(float, CdiPercentDia) / 100.0))) - 1) * 100.0 AS decimal(18,6)) AS CdiPercentPeriodo
            FROM [Integracao].[Silver].[DimTaxaJurosDiaria] WITH (NOLOCK)
            WHERE DataReferencia >= ?
              AND DataReferencia <= ?;
            """

            cur_cdi = raw_conn.cursor()
            cur_cdi.execute(sql_cdi_resumo, (dt_ini, dt_fim))
            row_cdi = cur_cdi.fetchone()
            _drenar_resultsets(cur_cdi)

            if row_cdi:
                cdi_info["DataInicio"] = row_cdi[0]
                cdi_info["DataFim"] = row_cdi[1]
                cdi_info["QtdDias"] = _to_int(row_cdi[2])
                cdi_info["CdiAcumulado"] = _to_float(row_cdi[3])
                cdi_info["CdiFatorPeriodo"] = _to_float(row_cdi[4])
                cdi_info["CdiPercentPeriodo"] = _to_float(row_cdi[5])

            sql_taxas = """
            SELECT
                DataReferencia,
                CdiPercentDiaRaw,
                CdiPercentDia,
                CdiPercentAno,
                SelicPercentDiaRaw,
                SelicPercentDia,
                SelicPercentAno,
                DataAtualizacao
            FROM [Integracao].[Silver].[DimTaxaJurosDiaria] WITH (NOLOCK)
            WHERE DataReferencia >= ?
              AND DataReferencia <= ?
            ORDER BY DataReferencia ASC;
            """

            cur_taxas = raw_conn.cursor()
            cur_taxas.execute(sql_taxas, (dt_ini, dt_fim))
            rows_taxas_diarias = cur_taxas.fetchall() or []
            _drenar_resultsets(cur_taxas)

            for rr in (rows_taxas_diarias or []):
                try:
                    taxas_diarias.append(
                        {
                            "DataReferencia": rr[0],
                            "CdiPercentDiaRaw": rr[1],
                            "CdiPercentDia": rr[2],
                            "CdiPercentAno": rr[3],
                            "SelicPercentDiaRaw": rr[4],
                            "SelicPercentDia": rr[5],
                            "SelicPercentAno": rr[6],
                            "DataAtualizacao": rr[7],
                        }
                    )
                except:
                    pass

            ooh_info = _normalizar_dict_indicador(
                _buscar_retorno_indice_por_periodo(
                    conn_local=raw_conn,
                    nome_tabela_qualificado="[Integracao].[Silver].[FatoIndiceOOHDiario]",
                    dt_ini_=dt_ini,
                    dt_fim_=dt_fim,
                )
            )

            ooh_global_info = _normalizar_dict_indicador(
                _buscar_retorno_indice_por_periodo(
                    conn_local=raw_conn,
                    nome_tabela_qualificado="[Integracao].[Silver].[FatoIndiceOOHGlobal]",
                    dt_ini_=dt_ini,
                    dt_fim_=dt_fim,
                )
            )

            current_app.logger.info(
                "Indicadores carregados grade multi. dt_ini=%s dt_fim=%s cdi=%s ooh=%s ooh_global=%s",
                dt_ini,
                dt_fim,
                cdi_info.get("CdiPercentPeriodo"),
                ooh_info.get("RetornoPercentPeriodo"),
                ooh_global_info.get("RetornoPercentPeriodo"),
            )

        except Exception:
            current_app.logger.exception(
                "Erro ao buscar indicadores do período da grade multi. dt_ini=%s dt_fim=%s",
                dt_ini,
                dt_fim,
            )
        finally:
            try:
                if cur_cdi is not None:
                    cur_cdi.close()
                pass
            except:
                pass

            try:
                if cur_taxas is not None:
                    cur_taxas.close()
                pass
            except:
                pass

            try:
                if raw_conn is not None:
                    raw_conn.close()
                pass
            except:
                pass

        return cdi_info, taxas_diarias, ooh_info, ooh_global_info

    def _copiar_dict_indicador(valor):
        if not valor:
            return {}
        try:
            return dict(valor)
        except:
            return {}

    def _copiar_lista_taxas(lista):
        saida = []
        for item in (lista or []):
            if isinstance(item, dict):
                saida.append(dict(item))
            else:
                saida.append(item)
        return saida

    def _deduplicar_rows_grade(rows):
        saida = []
        vistos = set()

        for r in (rows or []):
            try:
                chave = (
                    r[0],
                    _normalizar_codface(r[2]),
                    _coerce_to_date(r[5]),
                    _coerce_to_date(r[6]),
                    _coerce_to_date(r[7]),
                    _normalizar_texto(r[3]),
                    _normalizar_texto(r[4]),
                    _normalizar_texto(r[8]),
                    _normalizar_texto(r[9]),
                    _normalizar_texto(r[10]),
                    _normalizar_texto(r[11]) if len(r) > 11 else "",
                )
            except:
                chave = tuple(r)

            if chave in vistos:
                continue

            vistos.add(chave)
            saida.append(r)

        return saida

    def _montar_payload_grade_painel_multi(
        codponto_local,
        codfaces_local,
        resumo_painel,
        info_periodo,
        cdi_info_painel=None,
        taxas_diarias_painel=None,
        ooh_info_painel=None,
        ooh_global_info_painel=None,
    ):
        dt_ini = info_periodo["dt_ini"]
        dt_fim = info_periodo["dt_fim"]
        dias = list(info_periodo["dias"])

        filtro_cliente = _normalizar_texto(request.args.get("cliente"))
        vendedores_selecionados = _limpar_valores_reais(_lista_args_limpa("vendedor"))

        faces_db_raw = (
            db.session.query(
                DimFacesPaineis.CodFace,
                DimFacesPaineis.Tipo,
                DimFacesPaineis.IDDimPaineisEuromidia,
            )
            .filter(
                DimFacesPaineis.CodPonto == int(codponto_local),
                DimFacesPaineis.CodFace != None,
                DimFacesPaineis.CodFace != "",
            )
            .all()
        )

        tipo_por_face = {}
        tipo_por_idcadastro = {}
        faces_disponiveis_db = []

        for cf_db, tipo_db, idcad_db in (faces_db_raw or []):
            cf_norm = _normalizar_codface(cf_db)
            if not cf_norm:
                continue

            faces_disponiveis_db.append(cf_norm)
            tipo_up = _normalizar_texto(tipo_db).upper()

            tipo_por_face[(int(codponto_local), cf_norm)] = tipo_up

            if idcad_db not in (None, ""):
                try:
                    tipo_por_idcadastro[int(idcad_db)] = tipo_up
                except:
                    pass

        faces_disponiveis_db = sorted(list(dict.fromkeys(faces_disponiveis_db)), key=lambda x: x.casefold())

        faces = []
        vistos_faces = set()

        for cf_sel in (codfaces_local or []):
            cf_norm = _normalizar_codface(cf_sel)
            if not cf_norm:
                continue

            chave = cf_norm.casefold()
            if chave in vistos_faces:
                continue

            vistos_faces.add(chave)
            faces.append(cf_norm)

        if not faces:
            for cf_db in (faces_disponiveis_db or []):
                chave = cf_db.casefold()
                if chave in vistos_faces:
                    continue
                vistos_faces.add(chave)
                faces.append(cf_db)

        tipo_painel_info = _normalizar_texto(resumo_painel.get("Tipo"))
        if not tipo_painel_info:
            tipos_distintos = []
            vistos_tipos = set()
            for _k, tipo_up in (tipo_por_face or {}).items():
                if not tipo_up:
                    continue
                if tipo_up in vistos_tipos:
                    continue
                vistos_tipos.add(tipo_up)
                tipos_distintos.append(tipo_up)
            if tipos_distintos:
                tipo_painel_info = tipos_distintos[0]

        eh_digital = "PAINEL DIGITAL" == _normalizar_texto(tipo_painel_info).upper()

        if eh_digital:
            capacidade_digital = int(CAPACIDADE_DIGITAL_PADRAO or 16)
            loops_permitidos_multi = _montar_slots_digitais(capacidade_digital)
        else:
            capacidade_digital = 0
            loops_permitidos_multi = []

        ocupacoes_por_face = {f: [] for f in (faces or [])}
        ocupacoes_por_slot = {}
        faces_conflito = []
        slots_conflito = []

        if eh_digital:
            for f in (faces or []):
                for lp in loops_permitidos_multi:
                    ocupacoes_por_slot[(f, lp)] = []

        fim_efetivo_sql = func.coalesce(
            FatoControleContratosItensEuromidia.DataCancelamento,
            FatoControleContratosItensEuromidia.DataTerminoPrevisto,
            date(9999, 12, 31),
        )

        q_itens = (
            db.session.query(
                FatoControleContratosItensEuromidia.IDFatoControleContratosItensEuromidia,
                FatoControleContratosItensEuromidia.IDFatoControleContratoEuromidia,
                FatoControleContratosItensEuromidia.CodFace,
                FatoControleContratosItensEuromidia.MarcaExibida,
                FatoControleContratosItensEuromidia.Vendedor,
                FatoControleContratosItensEuromidia.DataInicioPrevisto,
                FatoControleContratosItensEuromidia.DataTerminoPrevisto,
                FatoControleContratosItensEuromidia.DataCancelamento,
                FatoControleContratosItensEuromidia.Cota,
                FatoControleContratosItensEuromidia.NumeroContrato,
                FatoControleContratosItensEuromidia.NumeroPrevia,
                FatoControleContratosItensEuromidia.CodPonto,
            )
            .filter(
                FatoControleContratosItensEuromidia.CodPonto == int(codponto_local),
                FatoControleContratosItensEuromidia.DataInicioPrevisto != None,
                FatoControleContratosItensEuromidia.AtivoCancelamento == "A",
                FatoControleContratosItensEuromidia.DataInicioPrevisto <= dt_fim,
                fim_efetivo_sql >= dt_ini,
            )
        )

        if faces:
            q_itens = q_itens.filter(FatoControleContratosItensEuromidia.CodFace.in_(faces))

        if filtro_cliente:
            padrao_cliente = f"%{filtro_cliente.upper()}%"
            q_itens = q_itens.filter(
                func.upper(func.coalesce(FatoControleContratosItensEuromidia.MarcaExibida, "")).like(padrao_cliente)
            )

        if vendedores_selecionados:
            filtros_vend_or = []
            for nome_vend in (vendedores_selecionados or []):
                nome_up = _normalizar_texto(nome_vend).upper()
                if not nome_up:
                    continue
                filtros_vend_or.append(
                    func.upper(func.coalesce(FatoControleContratosItensEuromidia.Vendedor, "")).like(f"%{nome_up}%")
                )

            if filtros_vend_or:
                q_itens = q_itens.filter(or_(*filtros_vend_or))

        rows = (
            q_itens
            .order_by(
                FatoControleContratosItensEuromidia.DataInicioPrevisto.asc(),
                FatoControleContratosItensEuromidia.IDFatoControleContratosItensEuromidia.asc(),
            )
            .all()
        )

        rows = _deduplicar_rows_grade(rows)

        itens_agregados_kpi = []
        possui_itens = False

        if eh_digital:
            itens_por_face = {f: [] for f in (faces or [])}

            for r in (rows or []):
                _id_item = r[0]
                _id_contrato = r[1]
                cf = _normalizar_codface(r[2]) if r[2] is not None else ""
                marca = r[3] or ""
                vend = r[4] or ""
                di = _coerce_to_date(r[5])
                df_prev = r[6]
                dc = r[7]
                cota = r[8]
                num_contrato = r[9] or ""
                num_previa = r[10] or ""

                if di is None:
                    continue

                df = _fim_efetivo_item(df_prev, dc)
                if df is None:
                    continue

                if not cf:
                    if faces:
                        cf = faces[0]
                    else:
                        continue

                if cf not in itens_por_face:
                    continue

                dia_ini, dia_fim = _intersecao_periodo(di, df, dt_ini, dt_fim)
                if dia_ini is None:
                    continue

                possui_itens = True

                spans_grade = int(_span_por_cota(cota) or 1)
                if spans_grade <= 0:
                    spans_grade = 1
                if spans_grade > len(loops_permitidos_multi):
                    spans_grade = len(loops_permitidos_multi)

                item_base = {
                    "ID": _id_item,
                    "IDFatoControleContratos": _id_contrato,
                    "CodFace": cf,
                    "MarcaExibida": marca,
                    "Loop": f"COTA {cota}",
                    "Vendedor": vend,
                    "DataInicio": di,
                    "DataFim": df,
                    "DiaInicio": dia_ini,
                    "DiaFim": dia_fim,
                    "TextoOriginal": f"CONTRATO:{num_contrato} | PRÉVIA:{num_previa}",
                    "BarraTexto": _texto_barra(cota, marca, vend),
                    "EhReserva": False,
                    "OrigemItem": "CONTRATO",
                    "CorBarra": None,
                    "StatusDb": None,
                    "OrigemDb": None,
                    "ReservaIDOriginal": None,
                    "NumeroContrato": num_contrato,
                    "NumeroPrevia": num_previa,
                    "SpanAltura": spans_grade,
                }

                itens_por_face.setdefault(cf, []).append(item_base)
                itens_agregados_kpi.append((di, df, float(_slots_por_cota(cota) or 1.0)))

            slots_conflito_set = set()

            for cf, itens_face in (itens_por_face or {}).items():
                if cf not in ocupacoes_por_face:
                    ocupacoes_por_face[cf] = []

                if cf and loops_permitidos_multi:
                    for lp in loops_permitidos_multi:
                        ocupacoes_por_slot.setdefault((cf, lp), [])

                itens_ordenados = sorted(
                    itens_face,
                    key=lambda x: (
                        _coerce_to_date(x.get("DataInicio")) or date(1900, 1, 1),
                        _coerce_to_date(x.get("DataFim")) or date(1900, 1, 1),
                        x.get("ID") or 0,
                    ),
                )

                fim_por_slot = {lp: date(1900, 1, 1) for lp in loops_permitidos_multi}

                for it in (itens_ordenados or []):
                    spans = int(it.get("SpanAltura") or 1)
                    spans = max(1, min(spans, len(loops_permitidos_multi)))

                    idx_escolhido = None
                    conflito_forcado = False

                    for idx_loop in range(0, len(loops_permitidos_multi) - spans + 1):
                        bloco = loops_permitidos_multi[idx_loop: idx_loop + spans]
                        cabe = True

                        for lp in bloco:
                            fim_atual = fim_por_slot.get(lp) or date(1900, 1, 1)
                            if _coerce_to_date(it.get("DataInicio")) <= fim_atual:
                                cabe = False
                                break

                        if cabe:
                            idx_escolhido = idx_loop
                            break

                    if idx_escolhido is None:
                        conflito_forcado = True
                        melhor_idx = 0
                        melhor_val = None

                        for idx_loop in range(0, len(loops_permitidos_multi) - spans + 1):
                            bloco = loops_permitidos_multi[idx_loop: idx_loop + spans]
                            val = max([(fim_por_slot.get(lp) or date(1900, 1, 1)) for lp in bloco])

                            if melhor_val is None or val < melhor_val:
                                melhor_val = val
                                melhor_idx = idx_loop

                        idx_escolhido = melhor_idx

                    for off in range(spans):
                        lp = loops_permitidos_multi[idx_escolhido + off]

                        fim_item = _coerce_to_date(it.get("DataFim")) or date(1900, 1, 1)
                        fim_atual = fim_por_slot.get(lp)
                        if fim_atual is None or fim_item > fim_atual:
                            fim_por_slot[lp] = fim_item

                        item_slot = dict(it)
                        item_slot["SpanAltura"] = spans
                        item_slot["SpanOffset"] = off
                        item_slot["SpanInicio"] = (off == 0)
                        item_slot["BitConflito"] = bool(conflito_forcado)

                        ocupacoes_por_slot.setdefault((cf, lp), []).append(item_slot)

                        if conflito_forcado:
                            slots_conflito_set.add((cf, lp))

            for (cf, lp), itens in (ocupacoes_por_slot or {}).items():
                if not itens or len(itens) <= 1:
                    continue

                itens_sorted = sorted(
                    itens,
                    key=lambda x: (
                        _coerce_to_date(x.get("DataInicio")) or date(1900, 1, 1),
                        _coerce_to_date(x.get("DataFim")) or date(1900, 1, 1),
                        x.get("ID") or 0,
                        x.get("SpanOffset", 0),
                    ),
                )

                conflito = False
                fim_atual = _coerce_to_date(itens_sorted[0].get("DataFim")) or date(1900, 1, 1)

                for i in range(1, len(itens_sorted)):
                    ini = _coerce_to_date(itens_sorted[i].get("DataInicio"))
                    fim = _coerce_to_date(itens_sorted[i].get("DataFim"))

                    if ini is None or fim is None:
                        continue

                    if ini <= fim_atual:
                        conflito = True
                        break

                    if fim > fim_atual:
                        fim_atual = fim

                if conflito:
                    slots_conflito_set.add((cf, lp))
                    for it in itens:
                        it["BitConflito"] = True

            faces_conflito = sorted(list({cf for (cf, _lp) in slots_conflito_set}))
            slots_conflito = sorted(list(slots_conflito_set), key=lambda x: (x[0], x[1]))

            for f in (faces or []):
                agg = []
                for lp in loops_permitidos_multi:
                    for it in ocupacoes_por_slot.get((f, lp), []):
                        if it.get("SpanInicio"):
                            agg.append(it)
                ocupacoes_por_face[f] = agg

        else:
            for f in (faces or []):
                ocupacoes_por_face[f] = []

            for r in (rows or []):
                _id_item = r[0]
                _id_contrato = r[1]
                cf = _normalizar_codface(r[2]) if r[2] is not None else ""
                marca = r[3] or ""
                vend = r[4] or ""
                di = _coerce_to_date(r[5])
                df_prev = r[6]
                dc = r[7]
                cota = r[8]
                num_contrato = r[9] or ""
                num_previa = r[10] or ""

                if di is None:
                    continue

                df = _fim_efetivo_item(df_prev, dc)
                if df is None:
                    continue

                if not cf:
                    continue

                if faces and (cf not in faces):
                    continue

                dia_ini, dia_fim = _intersecao_periodo(di, df, dt_ini, dt_fim)
                if dia_ini is None:
                    continue

                possui_itens = True
                itens_agregados_kpi.append((di, df, 1.0))

                eh_reserva = False
                try:
                    eh_reserva = int(_id_item) < 0
                except:
                    eh_reserva = False

                ocupacoes_por_face.setdefault(cf, []).append(
                    {
                        "ID": _id_item,
                        "IDFatoControleContratos": _id_contrato,
                        "CodFace": cf,
                        "MarcaExibida": marca,
                        "Loop": f"COTA {cota}",
                        "Vendedor": vend,
                        "DataInicio": di,
                        "DataFim": df,
                        "DiaInicio": dia_ini,
                        "DiaFim": dia_fim,
                        "TextoOriginal": f"CONTRATO:{num_contrato} | PRÉVIA:{num_previa}",
                        "BarraTexto": _texto_barra(cota, marca, vend),
                        "EhReserva": bool(eh_reserva),
                        "OrigemItem": ("RESERVA" if eh_reserva else "CONTRATO"),
                        "CorBarra": ("#92400e" if eh_reserva else None),
                        "StatusDb": ("RESERVADO" if eh_reserva else None),
                        "OrigemDb": ("RESERVA" if eh_reserva else None),
                        "ReservaIDOriginal": None,
                        "NumeroContrato": num_contrato,
                        "NumeroPrevia": num_previa,
                    }
                )

            faces_conflito = sorted(list(_marcar_conflitos_por_face(ocupacoes_por_face)))
            slots_conflito = []
            ocupacoes_por_slot = {}

        total_dias = int(info_periodo["total_dias"] or 0)
        if total_dias <= 0:
            total_dias = 1

        if eh_digital:
            slots_total = int(capacidade_digital) * max(1, len(faces))
        else:
            slots_total = max(1, len(faces))

        uso_slots_dias, max_simultaneo = _uso_no_periodo_por_intervalos(
            intervalos=itens_agregados_kpi,
            dt_ini=dt_ini,
            dt_fim=dt_fim,
            denom_cap=max(1, slots_total),
        )

        ocupacao_pct = None
        if slots_total > 0 and total_dias > 0:
            capacidade_total_periodo = float(slots_total) * float(total_dias)
            if capacidade_total_periodo > 0:
                ocupacao_pct = float(uso_slots_dias) / float(capacidade_total_periodo) * 100.0

        try:
            slots_ocupados = int(min(float(slots_total), math.ceil(float(max_simultaneo)))) if max_simultaneo > 0 else 0
        except:
            slots_ocupados = 0

        financeiros = _buscar_financeiro_periodo(
            codponto_local=int(codponto_local),
            dt_ini=dt_ini,
            dt_fim=dt_fim,
            possui_itens=possui_itens,
        )

        tipo = _normalizar_texto(resumo_painel.get("Tipo"))
        exibidora = _normalizar_texto(resumo_painel.get("Exibidora"))
        cidade = _normalizar_texto(resumo_painel.get("Cidade"))
        uf = _normalizar_texto(resumo_painel.get("UF"))
        bairro = _normalizar_texto(resumo_painel.get("Bairro"))
        logradouro = _normalizar_texto(resumo_painel.get("Logradouro"))
        localizacao = _montar_localizacao(resumo_painel)

        titulo_partes = [f"Painel {int(codponto_local)}"]
        if tipo:
            titulo_partes.append(tipo)
        titulo = " • ".join(titulo_partes)

        subtitulo_partes = []
        if exibidora:
            subtitulo_partes.append(exibidora)
        if localizacao:
            subtitulo_partes.append(localizacao)
        subtitulo = " • ".join(subtitulo_partes)

        url_grade = _montar_url_grade_unica(
            codponto_destino=int(codponto_local),
            codfaces_painel=list(faces or []),
        )

        cdi_info_payload = _copiar_dict_indicador(cdi_info_painel)
        taxas_diarias_payload = _copiar_lista_taxas(taxas_diarias_painel)
        ooh_info_payload = _copiar_dict_indicador(ooh_info_painel)
        ooh_global_info_payload = _copiar_dict_indicador(ooh_global_info_painel)

        cdi_percent_periodo_payload = None
        ooh_percent_periodo_payload = None
        ooh_global_percent_periodo_payload = None

        try:
            cdi_percent_periodo_payload = cdi_info_payload.get("CdiPercentPeriodo")
        except:
            cdi_percent_periodo_payload = None

        try:
            ooh_percent_periodo_payload = ooh_info_payload.get("RetornoPercentPeriodo")
        except:
            ooh_percent_periodo_payload = None

        try:
            ooh_global_percent_periodo_payload = ooh_global_info_payload.get("RetornoPercentPeriodo")
        except:
            ooh_global_percent_periodo_payload = None

        return {
            "codponto": int(codponto_local),
            "painel_id": int(codponto_local),
            "id": int(codponto_local),

            "codfaces": list(faces or []),
            "faces": list(faces or []),
            "total_codfaces": len(list(faces or [])),
            "qtd_faces": len(list(faces or [])),
            "num_faces": len(list(faces or [])),
            "codfaces_label": ", ".join(list(faces or [])) if faces else "",

            "url_grade": url_grade,
            "href": url_grade,
            "url": url_grade,
            "iframe_url": url_grade,
            "embed_url": url_grade,

            "tipo": tipo,
            "tipo_prod": tipo,
            "tipo_painel_info": tipo_painel_info or tipo,
            "exibidora": exibidora,

            "cidade": cidade,
            "uf": uf,
            "bairro": bairro,
            "logradouro": logradouro,
            "localizacao": localizacao,

            "cidade_painel": cidade,
            "uf_painel": uf,
            "bairro_painel": bairro,
            "logradouro_painel": logradouro,

            "titulo": titulo,
            "subtitulo": subtitulo,

            "tem_faces": len(list(faces or [])) > 0,
            "eh_digital": eh_digital,

            "dias": dias,
            "total_dias": int(info_periodo["total_dias"]),
            "ultimo_dia": int(info_periodo["ultimo_dia"]),

            "ocupacoes_por_face": ocupacoes_por_face,
            "faces_conflito": faces_conflito,

            "loops_permitidos": (list(loops_permitidos_multi) if eh_digital else []),
            "capacidade_digital": int(capacidade_digital or 0),
            "ocupacoes_por_slot": ocupacoes_por_slot,
            "slots_conflito": slots_conflito,

            "ocupacao_pct": ocupacao_pct,
            "slots_ocupados": slots_ocupados,
            "slots_total": slots_total,

            "margem_pct": financeiros.get("margem_pct"),
            "rentabilidade_valor": financeiros.get("rentabilidade_valor"),
            "custo_total": financeiros.get("custo_total"),
            "receita_total": financeiros.get("receita_total"),
            "qtd_contratos": financeiros.get("qtd_contratos"),
            "ticket_medio": financeiros.get("ticket_medio"),

            "cdi_info": cdi_info_payload,
            "taxas_diarias": taxas_diarias_payload,
            "ooh_info": ooh_info_payload,
            "ooh_global_info": ooh_global_info_payload,
            "indice_ooh_info": dict(ooh_info_payload),
            "indice_ooh_global_info": dict(ooh_global_info_payload),

            "cdi_percent_periodo": cdi_percent_periodo_payload,
            "ooh_percent_periodo": ooh_percent_periodo_payload,
            "ooh_global_percent_periodo": ooh_global_percent_periodo_payload,
        }

    filtros_codface_brutos = _lista_args_limpa("codface")
    filtros_codface = []
    vistos_codface = set()

    for cf in (filtros_codface_brutos or []):
        cf_norm = _normalizar_codface(cf)
        if not cf_norm:
            continue

        chave = cf_norm.casefold()
        if chave in vistos_codface:
            continue

        vistos_codface.add(chave)
        filtros_codface.append(cf_norm)

    codpontos_explicitos = _lista_codpontos_validos()

    faces_por_painel, faces_nao_resolvidas = _agrupar_faces_por_painel(
        codfaces_lista=filtros_codface,
        codpontos_explicitos=codpontos_explicitos,
    )

    codpontos_alvo = []
    vistos_codpontos_alvo = set()

    for cp in (list(codpontos_explicitos or []) + list((faces_por_painel or {}).keys())):
        try:
            cp_int = int(cp)
        except:
            continue

        if cp_int <= 0:
            continue

        if cp_int in vistos_codpontos_alvo:
            continue

        vistos_codpontos_alvo.add(cp_int)
        codpontos_alvo.append(cp_int)

    if not codpontos_alvo:
        flash("Nenhuma face válida foi encontrada para abrir a multi-grade.", "warning")
        return redirect(url_for("Paineis.lista_paineis"))

    if len(codpontos_alvo) == 1:
        codponto_unico = int(codpontos_alvo[0])
        return redirect(
            _montar_url_grade_unica(
                codponto_destino=codponto_unico,
                codfaces_painel=faces_por_painel.get(codponto_unico, []),
            )
        )

    info_periodo = _montar_info_periodo()

    cdi_info, taxas_diarias, ooh_info, ooh_global_info = _buscar_indicadores_periodo(
        dt_ini=info_periodo["dt_ini"],
        dt_fim=info_periodo["dt_fim"],
    )

    cdi_percent_periodo_global = None
    ooh_percent_periodo_global = None
    ooh_global_percent_periodo_global = None

    try:
        cdi_percent_periodo_global = cdi_info.get("CdiPercentPeriodo")
    except:
        cdi_percent_periodo_global = None

    try:
        ooh_percent_periodo_global = ooh_info.get("RetornoPercentPeriodo")
    except:
        ooh_percent_periodo_global = None

    try:
        ooh_global_percent_periodo_global = ooh_global_info.get("RetornoPercentPeriodo")
    except:
        ooh_global_percent_periodo_global = None

    current_app.logger.info(
        "Grade multi antes de montar cards. mes_ref=%s cdi=%s ooh=%s ooh_global=%s",
        info_periodo["mes_ref"],
        cdi_percent_periodo_global,
        ooh_percent_periodo_global,
        ooh_global_percent_periodo_global,
    )

    grades = []

    for ordem, codponto_destino in enumerate(codpontos_alvo, start=1):
        resumo_painel = _buscar_resumo_painel(int(codponto_destino))
        codfaces_painel = list(faces_por_painel.get(int(codponto_destino), []) or [])

        grade_item = _montar_payload_grade_painel_multi(
            codponto_local=int(codponto_destino),
            codfaces_local=codfaces_painel,
            resumo_painel=resumo_painel,
            info_periodo=info_periodo,
            cdi_info_painel=cdi_info,
            taxas_diarias_painel=taxas_diarias,
            ooh_info_painel=ooh_info,
            ooh_global_info_painel=ooh_global_info,
        )
        grade_item["ordem"] = int(ordem)

        current_app.logger.info(
            "Card multi montado. codponto=%s cdi=%s ooh=%s ooh_global=%s",
            grade_item.get("codponto"),
            grade_item.get("cdi_percent_periodo"),
            grade_item.get("ooh_percent_periodo"),
            grade_item.get("ooh_global_percent_periodo"),
        )

        grades.append(grade_item)

    if filtros_codface:
        total_faces = len(filtros_codface)
    else:
        total_faces = 0
        for g in (grades or []):
            try:
                total_faces += int(g.get("total_codfaces") or 0)
            except:
                pass

    total_grades = len(grades)

    return render_template(
        "euromidia/painel_grade_multi.html",

        paineis_grade=grades,
        grades=grades,
        painel_cards=grades,
        grades_renderizadas=grades,

        total_grades=total_grades,
        total_paineis=total_grades,
        total_faces=total_faces,

        mes_ref=info_periodo["mes_ref"],
        mes_ref_prev=info_periodo["mes_ref_prev"],
        mes_ref_next=info_periodo["mes_ref_next"],

        codfaces_selecionadas=list(filtros_codface or []),
        codpontos_alvo=codpontos_alvo,
        faces_por_painel=faces_por_painel,
        faces_nao_resolvidas=faces_nao_resolvidas,

        cdi_info=cdi_info,
        taxas_diarias=taxas_diarias,
        ooh_info=ooh_info,
        ooh_global_info=ooh_global_info,
        indice_ooh_info=ooh_info,
        indice_ooh_global_info=ooh_global_info,

        cdi_percent_periodo=cdi_percent_periodo_global,
        ooh_percent_periodo=ooh_percent_periodo_global,
        ooh_global_percent_periodo=ooh_global_percent_periodo_global,

        filtros={
            "mes_ref": info_periodo["mes_ref"],
            "cliente": _normalizar_texto(request.args.get("cliente")),
            "tipo": _normalizar_texto(request.args.get("tipo")),
            "mes_de": _normalizar_texto(request.args.get("mes_de")),
            "mes_ate": _normalizar_texto(request.args.get("mes_ate")),
            "modo": _normalizar_texto(request.args.get("modo")),
            "data_ref": _normalizar_texto(request.args.get("data_ref")),
            "bi_semana": _normalizar_texto(request.args.get("bi_semana")),
            "vendedores": _lista_args_limpa("vendedor"),
            "segmentos": _lista_args_limpa("segmento"),
            "codfaces": list(filtros_codface or []),

            "codpontos": list(codpontos_alvo or []),
            "codpontos_explicitos": list(codpontos_explicitos or []),
        },
    )















@paineis_bp.get("/contratos/<int:id_fato_controle_contratos>", strict_slashes=False)
@retry_get_view(db, attempts=6, base_delay=0.2, max_delay=1.5)
def contrato_detalhe(id_fato_controle_contratos: int):

    sql_contrato = text("""
        SELECT
            c.IDFatoControleContratosEuromidia,
            c.DataAtualizacao,
            c.Referencia,
            c.NumeroContrato,
            c.NumeroPrevia,
            c.CNPJ,
            c.DataAssinaturaRenovacao,
            c.IDTrimestre,
            c.DataLancamento,
            c.RazaoSocial,
            c.CPF,
            c.MarcaExibida,
            c.Vendedor,
            c.TipoDocumento,
            c.Origem,
            c.SDR,
            c.Agencia,
            c.CnpjAgencia,
            c.Bureau,
            c.CnpjBureau,
            c.Intermediario,
            c.CnpjIntermediario,
            c.QuantidadePontos,
            c.QuantidadeFaces,

            c.TotalBrutoContrato,
            c.TotalLiquidoContratoAGBRCTACORDO,
            c.TotalLiquidoContratoAGBRVENDGERCOOR,

            c.TotalValorMensalAgencia,
            c.TotalValorBureauMensal,
            c.TotalValorCartaAcordoMensal,
            c.TotalValorOutrasComissoes,
            c.TotalFaturamentoLiquidoMensal,

            c.TotalPercentualAgencia,
            c.TotalPercentualBureau,
            c.TotalPercentualCartaAcordo,
            c.TotalPercentualComissaoVendedor,
            c.TotalPercentualComissaoCoordenacao,

            c.TotalValorVendedor,
            c.ValorVendedorTotal,

            c.IDEmpresa,
            c.IDCategoriaMarca

        FROM [Integracao].[Silver].[FatoControleContratosEuromidia] AS c
        WHERE c.IDFatoControleContratosEuromidia = :id
    """)

    contrato = db.session.execute(
        sql_contrato,
        {"id": id_fato_controle_contratos}
    ).mappings().first()

    if not contrato:
        abort(404)

    sql_itens = text(r"""
    ;WITH ItensContrato AS (
        SELECT
            i.*,

            i.DataInicioPrevisto  AS DataInicioPrevisto_dt,
            i.DataTerminoPrevisto AS DataTerminoPrevisto_dt,
            i.DataCancelamento    AS DataCancelamento_dt

        FROM [Integracao].[Silver].[FatoControleContratosItensEuromidia] i
        WHERE i.IDFatoControleContratoEuromidia = :id
    ),
    MargemFace AS (
        SELECT
            m.IDFatoControleContratosEuromidia AS IDFatoControleContratos,
            m.CodFace AS CodFace,

            SUM(TRY_CONVERT(decimal(19,4), m.FaturamentoLiquidoMensalFinal)) AS FatLiquidoFinal_Soma,
            SUM(TRY_CONVERT(decimal(19,4), m.CustoMensalAlocado))            AS CustoMensalAlocado_Soma,

            CAST(
                SUM(TRY_CONVERT(decimal(19,4), m.FaturamentoLiquidoMensalFinal)) -
                SUM(TRY_CONVERT(decimal(19,4), m.CustoMensalAlocado))
                AS decimal(19,4)
            ) AS MargemR_Face_Soma,

            CAST(
                CASE
                    WHEN NULLIF(SUM(TRY_CONVERT(decimal(19,4), m.FaturamentoLiquidoMensalFinal)), 0) IS NULL THEN NULL
                    ELSE
                        100.0 *
                        (
                            SUM(TRY_CONVERT(decimal(19,4), m.FaturamentoLiquidoMensalFinal)) -
                            SUM(TRY_CONVERT(decimal(19,4), m.CustoMensalAlocado))
                        )
                        / NULLIF(SUM(TRY_CONVERT(decimal(19,4), m.FaturamentoLiquidoMensalFinal)), 0)
                END
                AS decimal(10,2)
            ) AS MargemPct_Face

        FROM [Integracao].[Silver].[DimMargemPaineisEuromidia] m
        WHERE m.IDFatoControleContratosEuromidia = :id
        GROUP BY
            m.IDFatoControleContratosEuromidia,
            m.CodFace
    )
    SELECT
        i.IDFatoControleContratosItensEuromidia,
        i.IDFatoControleContratoEuromidia,
        i.DataAtualizacao,
        i.Referencia,
        i.NumeroContrato,
        i.NumeroPrevia,
        i.CNPJ,

        i.DataLancamento,
        i.DataAssinaturaRenovacao,
        i.IDTrimestre,

        i.DataInicioPrevisto_dt  AS DataInicioPrevisto,
        i.DataTerminoPrevisto_dt AS DataTerminoPrevisto,
        i.DataCancelamento_dt    AS DataCancelamento,

        i.CidadeExibicao,
        i.Tipo,
        i.Origem,
        i.TipoDocumento,
        i.RazaoSocial,
        i.CPF,
        i.MarcaExibida,
        i.Vendedor,
        i.SDR,

        i.Cota,
        i.CodPonto,
        i.CodFace,

        i.TexmpoExposicao AS TempoExposicaoDias,
        i.NumeroParcelas  AS QuantidadeParcelas,

        i.FaturamentoBrutoMensal,
        i.ValorPermuta,
        i.FaturamentoLiquidoPermuta,

        i.TotalBrutoContrato,
        i.TotalLiquidoContratoAGBRCTACORDO,
        i.TotalLiquidoContratoAGBRVENDGERCOOR,

        i.ValorMensalAgencia,
        i.ValorBureauMensal,
        i.ValorCartaAcordoMensal AS ValorAcordoMensal,
        i.ValorOutrasComissoes   AS OutrasComissoes,

        i.FaturamentoLiquidoMensal,
        i.FaturamentoLiquidoFinalMensal AS FaturamentoLiquidoMensalFinal,

        i.ValorVendedor,
        i.ValorVendedorTotal,
        i.ValorCoordenador,
        i.ValorCoordenadorTotal,
        i.ValorGerencia,
        i.ValorGerenciaTotal,

        i.PercentualAgencia,
        i.PercentualBureau,
        i.PercentualCartaAcordo,
        i.PercentualComissaoVendedor,
        i.PercentualComissaoCoordenacao,
        i.PercentualComissaoGerencia,

        i.Status,

        mf.FatLiquidoFinal_Soma,
        mf.CustoMensalAlocado_Soma,
        mf.MargemR_Face_Soma,
        mf.MargemPct_Face

    FROM ItensContrato i
    LEFT JOIN MargemFace mf
        ON mf.IDFatoControleContratos = i.IDFatoControleContratoEuromidia
       AND mf.CodFace = i.CodFace

    ORDER BY
        i.CodFace ASC,
        i.DataInicioPrevisto_dt ASC,
        i.IDFatoControleContratosItensEuromidia ASC
""")


    itens = db.session.execute(
        sql_itens,
        {"id": id_fato_controle_contratos}
    ).mappings().all()

    return render_template(
        "euromidia/contrato_detalhe.html",
        contrato=contrato,
        itens=itens,
    )







def _sql_bool(v):
  
    if v is None:
        return None
    if v is True or v is False:
        return v
    try:
        iv = int(v)
    except (TypeError, ValueError):
        return None
    if iv == 1:
        return True
    if iv == 0:
        return False
    return None


def _somente_digitos(valor) -> str:
   
    if valor is None:
        return ""
    return re.sub(r"\D+", "", str(valor))


def _normalizar_cnpj(cnpj: str) -> str:
    return _somente_digitos(cnpj)

def _s(v) -> str:

    if v is None:
        return ""
    return str(v).strip()


def _to_bool(v) -> bool:
    return bool(v) is True

def _to_decimal_19_2(v):
 
    if v is None:
        return None
    txt = str(v).strip()
    if txt == "":
        return None

   
    txt = txt.replace(".", "").replace(",", ".") if (txt.count(",") == 1 and txt.count(".") >= 1) else txt.replace(",", ".")
    try:
        d = Decimal(txt)
    except (InvalidOperation, ValueError):
        return None

    
    return d.quantize(Decimal("0.01"))

def _bool_from_situacao_receita(descricao_situacao: str):
    
    ds = _s(descricao_situacao).upper()
    if ds == "ATIVA":
        return True
    if ds in ("BAIXADA", "SUSPENSA", "INAPTA", "NULA"):
        return False
    return None




@paineis_bp.get("/api/receita/<cnpj>")
@retry_get_view(db, attempts=6, base_delay=0.2, max_delay=1.5)
def paineis_api_buscar_receita(cnpj: str):
    cnpj_limpo = _somente_digitos(cnpj)

    if len(cnpj_limpo) != 14:
        return jsonify({"ok": False, "error": "CNPJ inválido (precisa ter 14 dígitos)."}), 400

    url = f"https://minhareceita.org/{cnpj_limpo}"

    try:
        resp = requests.get(url, timeout=12)
    except requests.RequestException:
        return jsonify({"ok": False, "error": "Falha ao consultar a Receita."}), 502

    if resp.status_code == 404:
        return jsonify({"ok": False, "error": "CNPJ não encontrado."}), 404

    if resp.status_code == 429:
        return jsonify({"ok": False, "error": "Limite de consultas atingido. Aguarde e tente novamente."}), 429

    if not (200 <= resp.status_code < 300):
        return jsonify({"ok": False, "error": f"Erro na Receita (HTTP {resp.status_code})."}), 502

    try:
        data = resp.json()
    except ValueError:
        return jsonify({"ok": False, "error": "Resposta inválida da Receita."}), 502

   
    descricao_situacao = _s(data.get("descricao_situacao_cadastral"))

    payload = {
        "id_cliente": None,

        "razao_social": _s(data.get("razao_social")),
        "nome_fantasia": _s(data.get("nome_fantasia")),
        "porte": _s(data.get("porte")),

        "email": _s(data.get("email")),
        "telefone": _somente_digitos(data.get("ddd_telefone_1") or ""),

        "cnpj": _somente_digitos(data.get("cnpj", "")),
        "cnae": _somente_digitos(data.get("cnae_fiscal") or data.get("cnae") or ""),
        "descricao_cnae": _s(data.get("cnae_fiscal_descricao") or data.get("descricao_cnae") or ""),

        "logradouro": _s(((data.get("descricao_tipo_de_logradouro") or "") + " " + (data.get("logradouro") or ""))),
        "numero": _s(data.get("numero")),
        "complemento": _s(data.get("complemento")),

        "cidade": _s(data.get("municipio")),
        "estado": "",  
        "uf": _s(data.get("uf")),

        "capital_social": data.get("capital_social") or "",

        "bit_situacao_cadastral": _bool_from_situacao_receita(descricao_situacao),
        "descricao_matriz_filial": _s(data.get("descricao_matriz_filial") or data.get("matriz_filial") or ""),
        "identificador_matriz_filial": _s(data.get("identificador_matriz_filial") or ""),

        "bit_ativo": _bool_from_situacao_receita(descricao_situacao),

        "id_empresa_proprietaria": None
    }

    return jsonify({"ok": True, "data": payload}), 200





@paineis_bp.get("/api/clientes/buscar/<cnpj>")
@limiter.limit("60 per minute", methods=["GET"])
def paineis_api_clientes_buscar(cnpj: str):
    cnpj_limpo = _normalizar_cnpj(cnpj)

    if len(cnpj_limpo) != 14:
        return jsonify({
            "ok": False,
            "status": "invalid",
            "message": "CNPJ inválido (precisa ter 14 dígitos).",
            "data": None,
        }), 400

    sql_cliente = text("""
        SELECT TOP (1)
            c.IDCliente,
            c.RazaoSocial,
            c.NomeFantasia,
            c.Porte,
            c.Email,
            c.Telefone,
            c.CNPJ,
            c.CNAE,
            c.Logradouro,
            c.Numero,
            c.Complemento,
            c.Cidade,
            c.Estado,
            c.UF,
            c.CapitalSocial,
            c.BitSituacaoCadastral,
            c.DescricaoMatrizFilial,
            c.DescricaoCnae,
            c.BitAtivo,
            c.IDEmpresaProprietaria,
            c.IdentificadorMatrizFilial
        FROM [Integracao].[Silver].[DimClientes] c
        WHERE REPLACE(REPLACE(REPLACE(c.CNPJ, '.', ''), '/', ''), '-', '') = :cnpj
    """)

    row = db.session.execute(sql_cliente, {"cnpj": cnpj_limpo}).mappings().first()

    if row:
        data = {
            "id_cliente": row.get("IDCliente"),
            "razao_social": row.get("RazaoSocial"),
            "nome_fantasia": row.get("NomeFantasia"),
            "porte": row.get("Porte"),

            "email": row.get("Email"),
            "telefone": row.get("Telefone"),

            "cnpj": _somente_digitos(row.get("CNPJ") or ""),
            "cnae": _somente_digitos(row.get("CNAE") or ""),
            "descricao_cnae": row.get("DescricaoCnae"),

            "logradouro": row.get("Logradouro"),
            "numero": row.get("Numero"),
            "complemento": row.get("Complemento"),

            "cidade": row.get("Cidade"),
            "estado": row.get("Estado"),
            "uf": row.get("UF"),

            "capital_social": str(row.get("CapitalSocial")) if row.get("CapitalSocial") is not None else "",

            "bit_situacao_cadastral": _sql_bool(row.get("BitSituacaoCadastral")),
            "descricao_matriz_filial": row.get("DescricaoMatrizFilial"),
            "identificador_matriz_filial": row.get("IdentificadorMatrizFilial"),

            "bit_ativo": _sql_bool(row.get("BitAtivo")),

            "id_empresa_proprietaria": row.get("IDEmpresaProprietaria"),
        }

        return jsonify({
            "ok": True,
            "status": "exists",
            "message": "Cliente já existente. Você pode editar e salvar.",
            "data": data,
        }), 200

    url = f"https://minhareceita.org/{cnpj_limpo}"

    try:
        resp = requests.get(url, timeout=12)
    except requests.RequestException:
        return jsonify({
            "ok": False,
            "status": "api_error",
            "message": "Falha ao consultar a Receita.",
            "data": None,
        }), 502

    if resp.status_code == 404:
        return jsonify({
            "ok": True,
            "status": "not_found",
            "message": "Não encontrado no banco nem na Receita. Preencha manualmente e clique em Cadastrar.",
            "data": {"cnpj": cnpj_limpo},
        }), 200

    if resp.status_code == 429:
        return jsonify({
            "ok": False,
            "status": "rate_limited",
            "message": "Limite de consultas atingido. Aguarde e tente novamente.",
            "data": None,
        }), 429

    if not (200 <= resp.status_code < 300):
        return jsonify({
            "ok": False,
            "status": "api_error",
            "message": f"Erro na Receita (HTTP {resp.status_code}).",
            "data": None,
        }), 502

    try:
        data = resp.json()
    except ValueError:
        return jsonify({
            "ok": False,
            "status": "api_error",
            "message": "Resposta inválida da Receita.",
            "data": None,
        }), 502

    descricao_situacao = _s(data.get("descricao_situacao_cadastral"))

    payload = {
        "id_cliente": None,

        "razao_social": _s(data.get("razao_social")),
        "nome_fantasia": _s(data.get("nome_fantasia")),
        "porte": _s(data.get("porte")),

        "email": _s(data.get("email")),
        "telefone": _somente_digitos(data.get("ddd_telefone_1") or ""),

        "cnpj": _somente_digitos(data.get("cnpj", "")),
        "cnae": _somente_digitos(data.get("cnae_fiscal") or data.get("cnae") or ""),
        "descricao_cnae": _s(data.get("cnae_fiscal_descricao") or data.get("descricao_cnae") or ""),

        "logradouro": _s(((data.get("descricao_tipo_de_logradouro") or "") + " " + (data.get("logradouro") or ""))),
        "numero": _s(data.get("numero")),
        "complemento": _s(data.get("complemento")),

        "cidade": _s(data.get("municipio")),
        "estado": "",
        "uf": _s(data.get("uf")),

        "capital_social": data.get("capital_social") or "",

        "bit_situacao_cadastral": _bool_from_situacao_receita(descricao_situacao),
        "descricao_matriz_filial": _s(data.get("descricao_matriz_filial") or data.get("matriz_filial") or ""),
        "identificador_matriz_filial": _s(data.get("identificador_matriz_filial") or ""),

        "bit_ativo": _bool_from_situacao_receita(descricao_situacao),

        "id_empresa_proprietaria": None
    }

    return jsonify({
        "ok": True,
        "status": "from_api",
        "message": "Cliente não encontrado no banco. Você pode editar e cadastrar.",
        "data": payload,
    }), 200










def _s(v) -> str:
    if v is None:
        return ""
    return str(v).strip()

def _somente_digitos(v) -> str:
    return "".join(ch for ch in _s(v) if ch.isdigit())

def _normalizar_cnpj(v) -> str:
    return _somente_digitos(v)

def _to_decimal_19_2(v):
    """
    Aceita: "10000.00", "10000,00", "10.000,00", 10000, Decimal
    Retorna Decimal com 2 casas ou None.
    """
    raw = _s(v)
    if raw == "":
        return None



    if "," in raw and "." in raw:
        raw = raw.replace(".", "").replace(",", ".")
    else:
        raw = raw.replace(",", ".")

    try:
        d = Decimal(raw)
    except (InvalidOperation, ValueError):
        raise ValueError("Capital Social inválido (use um número).")

    return d.quantize(Decimal("0.01"))

def _to_bool_ou_none(v):
    if v is True:
        return True
    if v is False:
        return False
    return None






@paineis_bp.post("/api/clientes/salvar")
@limiter.limit("20 per minute", methods=["POST"])
@retry_get_view(db, attempts=6, base_delay=0.2, max_delay=1.5)
def paineis_api_clientes_salvar():
    body = request.get_json(silent=True) or {}

    cnpj = _normalizar_cnpj(body.get("cnpj"))
    if len(cnpj) != 14:
        return jsonify({"ok": False, "message": "CNPJ inválido (precisa ter 14 dígitos)."}), 400

    razao_social = _s(body.get("razao_social"))
    if not razao_social:
        return jsonify({"ok": False, "message": "Razão Social é obrigatória."}), 400

 
    try:
        capital_social = _to_decimal_19_2(body.get("capital_social"))
    except ValueError as e:
        return jsonify({"ok": False, "message": str(e)}), 400

 
    bit_situacao = _to_bool_ou_none(body.get("bit_situacao_cadastral"))
    bit_ativo = _to_bool_ou_none(body.get("bit_ativo"))

  
    id_empresa_raw = _s(body.get("id_empresa_proprietaria"))
    id_empresa_int = None
    if id_empresa_raw != "":
        if not id_empresa_raw.isdigit():
            return jsonify({"ok": False, "message": "ID Empresa Proprietária inválido (precisa ser número)."}), 400
        id_empresa_int = int(id_empresa_raw)

    uf = _s(body.get("uf")).upper()
    if uf and len(uf) != 2:
        return jsonify({"ok": False, "message": "UF inválida (precisa ter 2 letras)."}), 400

    payload = {
        "CNPJ": cnpj,
        "RazaoSocial": razao_social,
        "NomeFantasia": _s(body.get("nome_fantasia")),
        "Porte": _s(body.get("porte")),
        "Email": _s(body.get("email")),
        "Telefone": _somente_digitos(body.get("telefone")),

        "CNAE": _somente_digitos(body.get("cnae")),
        "Logradouro": _s(body.get("logradouro")),
        "Numero": _s(body.get("numero")),
        "Complemento": _s(body.get("complemento")),

        "Cidade": _s(body.get("cidade")),
        "Estado": _s(body.get("estado")),
        "UF": uf,

        "CapitalSocial": capital_social,

        "BitSituacaoCadastral": bit_situacao,
        "DescricaoMatrizFilial": _s(body.get("descricao_matriz_filial")),
        "DescricaoCnae": _s(body.get("descricao_cnae")),

        "BitAtivo": bit_ativo,
        "IDEmpresaProprietaria": id_empresa_int,
        "IdentificadorMatrizFilial": _s(body.get("identificador_matriz_filial")),
    }


    sql_upsert = text("""
        SET NOCOUNT ON;

        DECLARE @id INT;

        -- trava a linha do CNPJ, se existir
        SELECT TOP (1) @id = IDCliente
        FROM [Integracao].[Silver].[DimClientes] WITH (UPDLOCK, HOLDLOCK)
        WHERE REPLACE(REPLACE(REPLACE(CNPJ, '.', ''), '/', ''), '-', '') = :CNPJ;

        IF @id IS NULL
        BEGIN
            -- trava tabela para garantir MAX+1 seguro
            SELECT @id = ISNULL(MAX(IDCliente), 0) + 1
            FROM [Integracao].[Silver].[DimClientes] WITH (TABLOCKX, HOLDLOCK);

            INSERT INTO [Integracao].[Silver].[DimClientes]
            (
                IDCliente,
                RazaoSocial,
                NomeFantasia,
                Porte,
                Email,
                Telefone,
                CNPJ,
                CNAE,
                Logradouro,
                Numero,
                Complemento,
                Cidade,
                Estado,
                UF,
                CapitalSocial,
                BitSituacaoCadastral,
                DescricaoMatrizFilial,
                DescricaoCnae,
                BitAtivo,
                IDEmpresaProprietaria,
                IdentificadorMatrizFilial
            )
            VALUES
            (
                @id,
                :RazaoSocial,
                :NomeFantasia,
                :Porte,
                :Email,
                :Telefone,
                :CNPJ,
                :CNAE,
                :Logradouro,
                :Numero,
                :Complemento,
                :Cidade,
                :Estado,
                :UF,
                :CapitalSocial,
                :BitSituacaoCadastral,
                :DescricaoMatrizFilial,
                :DescricaoCnae,
                :BitAtivo,
                :IDEmpresaProprietaria,
                :IdentificadorMatrizFilial
            );

            SELECT 'insert' AS Confirmacao, @id AS IDCliente;
        END
        ELSE
        BEGIN
            UPDATE [Integracao].[Silver].[DimClientes]
            SET
                RazaoSocial = :RazaoSocial,
                NomeFantasia = :NomeFantasia,
                Porte = :Porte,
                Email = :Email,
                Telefone = :Telefone,
                CNPJ = :CNPJ,
                CNAE = :CNAE,
                Logradouro = :Logradouro,
                Numero = :Numero,
                Complemento = :Complemento,
                Cidade = :Cidade,
                Estado = :Estado,
                UF = :UF,
                CapitalSocial = :CapitalSocial,
                BitSituacaoCadastral = :BitSituacaoCadastral,
                DescricaoMatrizFilial = :DescricaoMatrizFilial,
                DescricaoCnae = :DescricaoCnae,
                BitAtivo = :BitAtivo,
                IDEmpresaProprietaria = :IDEmpresaProprietaria,
                IdentificadorMatrizFilial = :IdentificadorMatrizFilial
            WHERE IDCliente = @id;

            SELECT 'update' AS Confirmacao, @id AS IDCliente;
        END
    """)

    try:
        row = db.session.execute(sql_upsert, payload).mappings().first()
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "message": f"Erro ao salvar no banco: {str(e)}"}), 500

    confirmacao = (row.get("Confirmacao") if row else None) or "salvo"
    id_cliente = (row.get("IDCliente") if row else None)

    msg = "Cadastro cadastrado com sucesso." if confirmacao == "insert" else "Cadastro atualizado com sucesso."

    return jsonify({
        "ok": True,
        "confirmacao": confirmacao,
        "message": msg,
        "data": {"id_cliente": id_cliente, "cnpj": cnpj}
    }), 200



@paineis_bp.get("/clientes/novo")
def paineis_clientes_novo():
    form = FormCadastroCliente()
    return render_template("euromidia/novo.html", form=form)





from sqlalchemy.orm import load_only, noload


@paineis_bp.get("/<int:codponto>/detalhe")
@login_required
@requer_permissao("PAINEIS_DETALHE_VER")
@retry_get_view(db, attempts=6, base_delay=0.2, max_delay=1.5)
def detalhe_painel(codponto: int):

    painel = (
        db.session.query(DimPaineisEuromidia)
        .options(
            noload("*"),
            load_only(
                DimPaineisEuromidia.IDDimPaineisEuromidia,
                DimPaineisEuromidia.DataAtualizacao,
                DimPaineisEuromidia.CodPonto,
                DimPaineisEuromidia.QuantidadeFaces,
                DimPaineisEuromidia.Tipo,
                DimPaineisEuromidia.Cidade,
                DimPaineisEuromidia.UF,
                DimPaineisEuromidia.Logradouro,
                DimPaineisEuromidia.Sentido,
                DimPaineisEuromidia.Bairro,
                DimPaineisEuromidia.Referencia,
                DimPaineisEuromidia.Numero,
                DimPaineisEuromidia.CEP,
                DimPaineisEuromidia.Latitude,
                DimPaineisEuromidia.Longitude,
                DimPaineisEuromidia.FormatoLxA,
                DimPaineisEuromidia.FormatoLonaAcabadaLxAm,
                DimPaineisEuromidia.AreaTotalm,
                DimPaineisEuromidia.BitIluminado,
                DimPaineisEuromidia.Restricoes,
                DimPaineisEuromidia.TipoSolo,
                DimPaineisEuromidia.DataInstalacao,
                DimPaineisEuromidia.DataRetirada,
                DimPaineisEuromidia.Exibidora,
                DimPaineisEuromidia.BitProprio,
                DimPaineisEuromidia.BitAtivo,
                DimPaineisEuromidia.BitAluguel,
                DimPaineisEuromidia.BitEnergia,
                DimPaineisEuromidia.BitInternet,
                DimPaineisEuromidia.IDProduto,
            ),
        )
        .filter(DimPaineisEuromidia.CodPonto == codponto)
        .first()
    )

    if not painel:
        abort(404, description=f"Painel CodPonto {codponto} não encontrado.")


    def _sim_nao(valor):
        return "Sim" if bool(valor) else "Não"

    def _texto(v, padrao="—"):
        if v is None:
            return padrao
        s = str(v).strip()
        return s if s else padrao

    def _to_date(v):
        if v is None:
            return None
        if isinstance(v, date) and not isinstance(v, datetime):
            return v
        if isinstance(v, datetime):
            return v.date()
        try:
            return v.date()
        except:
            return None

    def _parse_ddmmyyyy(s: str):
        """
        Espera "DD-MM-YYYY" ou "DD/MM/YYYY".
        Retorna date ou None.
        """
        if not s:
            return None
        s = (str(s) or "").strip()
        if not s:
            return None
        s = s.replace("/", "-")
        try:
            parts = s.split("-")
            if len(parts) != 3:
                return None
            dd = int(parts[0])
            mm = int(parts[1])
            yy = int(parts[2])
            return date(yy, mm, dd)
        except:
            return None

    def _primeiro_dia_ano(dt: date):
        return date(dt.year, 1, 1)

    def _meses_entre(dt_ini: date, dt_fim: date):
        if dt_ini is None or dt_fim is None:
            return []
        if dt_fim < dt_ini:
            return []

        y = dt_ini.year
        m = dt_ini.month
        y2 = dt_fim.year
        m2 = dt_fim.month

        out = []
        while (y < y2) or (y == y2 and m <= m2):
            out.append((y, m))
            m += 1
            if m > 12:
                m = 1
                y += 1
        return out

    def _add_months(dt: date, delta_meses: int):
        if dt is None:
            return None

        y = dt.year
        m = dt.month + int(delta_meses)

        while m > 12:
            m -= 12
            y += 1
        while m < 1:
            m += 12
            y -= 1

        if m == 12:
            prox = date(y + 1, 1, 1)
        else:
            prox = date(y, m + 1, 1)
        ultimo_dia = (prox - timedelta(days=1)).day

        dd = dt.day
        if dd > ultimo_dia:
            dd = ultimo_dia

        return date(y, m, dd)


    endereco_linha_1 = " ".join([
        _texto(painel.Logradouro, "").strip()
    ]).strip()

    if _texto(painel.Numero, "").strip():
        endereco_linha_1 = f"{endereco_linha_1} nº {_texto(painel.Numero, '').strip()}".strip()

    endereco_linha_2 = " • ".join([
        _texto(painel.Bairro),
        f"CEP {_texto(painel.CEP)}",
        _texto(painel.UF),
    ])


    hoje = date.today()

    dt_de_txt = (request.args.get("dt_de") or "").strip()
    dt_ate_txt = (request.args.get("dt_ate") or "").strip()

    dt_de = _parse_ddmmyyyy(dt_de_txt)
    dt_ate = _parse_ddmmyyyy(dt_ate_txt)


    if dt_de is None and dt_ate is None:
        dt_ate = hoje
        dt_ate_mes = date(dt_ate.year, dt_ate.month, 1)
        dt_de_mes = _add_months(dt_ate_mes, -11)
        dt_de = dt_de_mes
    else:
        if dt_ate is None:
            dt_ate = hoje

        if dt_de is None:
            dt_de = _primeiro_dia_ano(dt_ate)

    if dt_de > dt_ate:
        dt_de, dt_ate = dt_ate, dt_de


    sql_matriz = text("""
        SELECT
            CodPonto,
            Ano,
            Jan, Fev, Mar, Abr, Mai, Jun,
            Jul, Ago, [Set], [Out], Nov, Dez,
            NoAno,
            Acumulado
        FROM [Integracao].[Silver].[DimPontosHistoricoVendas]
        WHERE CodPonto = :codponto
        ORDER BY Ano DESC;
    """)

    rows_matriz = db.session.execute(sql_matriz, {"codponto": codponto}).fetchall()

    matriz_faturamento = []
    for r in rows_matriz:
        try:
            yy = int(r.Ano)
        except:
            continue

        def _v(x):
            try:
                return float(x or 0.0)
            except:
                return 0.0

        linha = {
            "Ano": yy,
            "Meses": {
                1: _v(r.Jan),
                2: _v(r.Fev),
                3: _v(r.Mar),
                4: _v(r.Abr),
                5: _v(r.Mai),
                6: _v(r.Jun),
                7: _v(r.Jul),
                8: _v(r.Ago),
                9: _v(getattr(r, "Set")),
                10: _v(getattr(r, "Out")),
                11: _v(r.Nov),
                12: _v(r.Dez),
            },
            "NoAno": _v(r.NoAno),
            "Acumulado": _v(r.Acumulado),
        }
        matriz_faturamento.append(linha)


    dt_de_mes = date(dt_de.year, dt_de.month, 1)
    dt_ate_mes = date(dt_ate.year, dt_ate.month, 1)

    sql_serie = text("""
        ;WITH Base AS (
            SELECT
                d.CodPonto,
                d.Ano,
                v.Mes,
                v.Valor,
                DATEFROMPARTS(d.Ano, v.Mes, 1) AS MesRef
            FROM [Integracao].[Silver].[DimPontosHistoricoVendas] d
            CROSS APPLY (VALUES
                (1, d.Jan),
                (2, d.Fev),
                (3, d.Mar),
                (4, d.Abr),
                (5, d.Mai),
                (6, d.Jun),
                (7, d.Jul),
                (8, d.Ago),
                (9, d.[Set]),
                (10, d.[Out]),
                (11, d.Nov),
                (12, d.Dez)
            ) v(Mes, Valor)
            WHERE d.CodPonto = :codponto
        )
        SELECT
            Ano,
            Mes,
            CAST(Valor AS decimal(19,2)) AS Valor
        FROM Base
        WHERE MesRef >= :dt_de_mes
          AND MesRef <= :dt_ate_mes
        ORDER BY Ano, Mes;
    """)

    rows_serie = db.session.execute(sql_serie, {
        "codponto": codponto,
        "dt_de_mes": dt_de_mes,
        "dt_ate_mes": dt_ate_mes,
    }).fetchall()

    serie_temporal = []
    for r in rows_serie:
        try:
            yy = int(r.Ano)
            mm = int(r.Mes)
            vv = float(r.Valor or 0.0)
        except:
            continue

        serie_temporal.append({
            "x": f"{yy:04d}-{mm:02d}",
            "y": vv,
        })


    return render_template(
        "euromidia/painel_detalhe.html",
        painel=painel,
        endereco_linha_1=_texto(endereco_linha_1),
        endereco_linha_2=_texto(endereco_linha_2),
        bit_ativo=bool(painel.BitAtivo),
        aluguel_txt=_sim_nao(painel.BitAluguel),
        energia_txt=_sim_nao(painel.BitEnergia),
        internet_txt=_sim_nao(painel.BitInternet),
        proprio_txt="Próprio" if bool(painel.BitProprio) else "Terceiro",

        faturamento_matriz=matriz_faturamento,
        faturamento_serie=serie_temporal,

        filtros_faturamento={
            "dt_de": dt_de.strftime("%d-%m-%Y") if dt_de else "",
            "dt_ate": dt_ate.strftime("%d-%m-%Y") if dt_ate else "",
        },
    )








def _limpar_email(email: str) -> str:
    return (email or "").strip().lower()

def _parse_data_yyyy_mm_dd(txt: str):
    s = (txt or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d")
    except Exception:
        return None

def _senha_hash(senha_plana: str) -> str:
  
   
    return generate_password_hash(senha_plana, method="pbkdf2:sha256")




@paineis_bp.get("/admin/usuarios")
@login_required
@requer_permissao("USUARIOS_GERENCIAR")
def admin_usuarios_lista():
    q = (request.args.get("q") or "").strip()

    query = (
        db.session.query(DimUsuarios)
        .join(DimPerfilUsuario, DimUsuarios.IDDimPerfilUsuario == DimPerfilUsuario.IDDimPerfilUsuario)
    )

    if q:
        q_like = f"%{q}%"
        query = query.filter(
            func.lower(DimUsuarios.Email).like(func.lower(q_like))
            | func.lower(DimUsuarios.NomeUsuario).like(func.lower(q_like))
            | func.lower(DimPerfilUsuario.NomePerfil).like(func.lower(q_like))
        )

    usuarios = query.order_by(DimUsuarios.IDDimUsuarios.desc()).all()
    return render_template("euromidia/usuarios_lista.html", usuarios=usuarios, q=q)











@paineis_bp.get("/admin/usuarios/novo")
@login_required
@requer_permissao("ADMIN_TUDO")
def admin_usuarios_novo_form():
    perfis = db.session.query(DimPerfilUsuario).order_by(DimPerfilUsuario.NomePerfil.asc()).all()

    form_usuario = FormUsuarioNovo()
    form_usuario.id_perfil.choices = [
        (p.IDDimPerfilUsuario, p.NomePerfil or f"Perfil #{p.IDDimPerfilUsuario}")
        for p in perfis
    ]
    form_usuario.ativo.data = "1"

    return render_template(
        "euromidia/usuarios_form.html",
        modo="novo",
        usuario=None,
        perfis=perfis,
        permissoes_todas=[],
        extras=[],
        form_usuario=form_usuario,
        form_senha=None,
        form_perm_extra=None,
        remover_forms={},
    )


@paineis_bp.post("/admin/usuarios/novo")
@login_required
@limiter.limit("10 per minute", methods=["POST"])
@requer_permissao("ADMIN_TUDO")
def admin_usuarios_novo_post():
    perfis = db.session.query(DimPerfilUsuario).order_by(DimPerfilUsuario.NomePerfil.asc()).all()

    form_usuario = FormUsuarioNovo()
    form_usuario.id_perfil.choices = [
        (p.IDDimPerfilUsuario, p.NomePerfil or f"Perfil #{p.IDDimPerfilUsuario}")
        for p in perfis
    ]

    if not form_usuario.validate_on_submit():
        flash("Revise os campos do formulário.", "danger")
        return render_template(
            "euromidia/usuarios_form.html",
            modo="novo",
            usuario=None,
            perfis=perfis,
            permissoes_todas=[],
            extras=[],
            form_usuario=form_usuario,
            form_senha=None,
            form_perm_extra=None,
            remover_forms={},
        )

    nome = (form_usuario.nome.data or "").strip()
    email = _limpar_email(form_usuario.email.data)
    senha = form_usuario.senha.data or ""
    id_perfil_int = int(form_usuario.id_perfil.data)
    ativo = (form_usuario.ativo.data or "1").strip()

    existe = db.session.query(DimUsuarios).filter(func.lower(DimUsuarios.Email) == email).first()
    if existe:
        flash("Já existe usuário com esse email.", "danger")
        return render_template(
            "euromidia/usuarios_form.html",
            modo="novo",
            usuario=None,
            perfis=perfis,
            permissoes_todas=[],
            extras=[],
            form_usuario=form_usuario,
            form_senha=None,
            form_perm_extra=None,
            remover_forms={},
        )

    novo = DimUsuarios(
        IDDimPerfilUsuario=id_perfil_int,
        NomeUsuario=nome or None,
        Email=email,
        HashSenha=_senha_hash(senha),
        BitAtivo=True if ativo == "1" else False,
        CreatedAt=datetime.now(),
        UpdateAt=datetime.now(),
        UltimoLogin=None,
    )

    db.session.add(novo)
    db.session.commit()

    flash("Usuário criado com sucesso.", "success")
    return redirect(url_for("Paineis.admin_usuarios_editar_form", id_usuario=novo.IDDimUsuarios))







@paineis_bp.get("/admin/usuarios/<int:id_usuario>/editar")
@login_required
@requer_permissao("USUARIOS_GERENCIAR")
def admin_usuarios_editar_form(id_usuario: int):
    usuario = (
        db.session.query(DimUsuarios)
        .filter(DimUsuarios.IDDimUsuarios == id_usuario)
        .first()
    )
    if not usuario:
        abort(404)

    perfis = db.session.query(DimPerfilUsuario).order_by(DimPerfilUsuario.NomePerfil.asc()).all()

    permissoes_todas = (
        db.session.query(DimPermissoes)
        .order_by(DimPermissoes.CodigoPermissao.asc())
        .all()
    )

    extras = (
        db.session.query(PermissoesUsuario)
        .filter(PermissoesUsuario.IDDimUsuarios == id_usuario)
        .all()
    )

 
    form_usuario = FormUsuarioEditar(obj=usuario)
    form_usuario.id_perfil.choices = [(p.IDDimPerfilUsuario, p.NomePerfil or f"Perfil #{p.IDDimPerfilUsuario}") for p in perfis]
    form_usuario.id_perfil.data = int(usuario.IDDimPerfilUsuario)
    form_usuario.ativo.data = "1" if bool(usuario.BitAtivo) else "0"

    form_senha = FormTrocarSenha(prefix="senha")

    form_perm_extra = FormPermissaoExtraUpsert(prefix="pex")
    form_perm_extra.id_permissao.choices = [
        (perm.IDDimPermissoes, f"{(perm.CodigoPermissao or '').strip()} — {(perm.Descricao or '').strip()}")
        for perm in permissoes_todas
    ]


    remover_forms = {}
    for row in extras:
        f = FormPermissaoExtraRemover(prefix=f"rm_{row.IDDimPermissoes}")
        f.id_permissao.data = str(row.IDDimPermissoes)
        remover_forms[row.IDDimPermissoes] = f

    return render_template(
        "euromidia/usuarios_form.html",
        modo="editar",
        usuario=usuario,
        perfis=perfis,
        permissoes_todas=permissoes_todas,
        extras=extras,
        form_usuario=form_usuario,
        form_senha=form_senha,
        form_perm_extra=form_perm_extra,
        remover_forms=remover_forms,
    )












@paineis_bp.post("/admin/usuarios/<int:id_usuario>/editar")
@login_required
@limiter.limit("20 per minute", methods=["POST"])
@requer_permissao("ADMIN_TUDO")
def admin_usuarios_editar_post(id_usuario: int):
    usuario = db.session.query(DimUsuarios).filter(DimUsuarios.IDDimUsuarios == id_usuario).first()
    if not usuario:
        abort(404)

    perfis = db.session.query(DimPerfilUsuario).order_by(DimPerfilUsuario.NomePerfil.asc()).all()

    form_usuario = FormUsuarioEditar()
    form_usuario.id_perfil.choices = [
        (p.IDDimPerfilUsuario, p.NomePerfil or f"Perfil #{p.IDDimPerfilUsuario}")
        for p in perfis
    ]

    if not form_usuario.validate_on_submit():
        flash("Revise os campos do formulário.", "danger")

        # Recarrega as outras coisas do GET (pra não quebrar o template)
        permissoes_todas = db.session.query(DimPermissoes).order_by(DimPermissoes.CodigoPermissao.asc()).all()
        extras = db.session.query(PermissoesUsuario).filter(PermissoesUsuario.IDDimUsuarios == id_usuario).all()

        form_senha = FormTrocarSenha(prefix="senha")
        form_perm_extra = FormPermissaoExtraUpsert(prefix="pex")
        form_perm_extra.id_permissao.choices = [
            (perm.IDDimPermissoes, f"{(perm.CodigoPermissao or '').strip()} — {(perm.Descricao or '').strip()}")
            for perm in permissoes_todas
        ]

        remover_forms = {}
        for row in extras:
            f = FormPermissaoExtraRemover(prefix=f"rm_{row.IDDimPermissoes}")
            f.id_permissao.data = str(row.IDDimPermissoes)
            remover_forms[row.IDDimPermissoes] = f

        return render_template(
            "euromidia/usuarios_form.html",
            modo="editar",
            usuario=usuario,
            perfis=perfis,
            permissoes_todas=permissoes_todas,
            extras=extras,
            form_usuario=form_usuario,
            form_senha=form_senha,
            form_perm_extra=form_perm_extra,
            remover_forms=remover_forms,
        )

    nome = (form_usuario.nome.data or "").strip()
    email = _limpar_email(form_usuario.email.data)
    id_perfil_int = int(form_usuario.id_perfil.data)
    ativo = (form_usuario.ativo.data or "1").strip()

    existe = (
        db.session.query(DimUsuarios)
        .filter(func.lower(DimUsuarios.Email) == email)
        .filter(DimUsuarios.IDDimUsuarios != id_usuario)
        .first()
    )
    if existe:
        flash("Já existe outro usuário com esse email.", "danger")
        return redirect(url_for("Paineis.admin_usuarios_editar_form", id_usuario=id_usuario))

    usuario.NomeUsuario = nome or None
    usuario.Email = email
    usuario.IDDimPerfilUsuario = id_perfil_int
    usuario.BitAtivo = True if ativo == "1" else False
    usuario.UpdateAt = datetime.now()

    db.session.commit()

    flash("Usuário atualizado com sucesso.", "success")
    return redirect(url_for("Paineis.admin_usuarios_editar_form", id_usuario=id_usuario))








@paineis_bp.post("/admin/usuarios/<int:id_usuario>/trocar-senha")
@login_required
@limiter.limit("10 per minute", methods=["POST"])
@requer_permissao("USUARIOS_GERENCIAR")
def admin_usuarios_trocar_senha(id_usuario: int):
    usuario = db.session.query(DimUsuarios).filter(DimUsuarios.IDDimUsuarios == id_usuario).first()
    if not usuario:
        abort(404)

    form_senha = FormTrocarSenha(prefix="senha")

    if not form_senha.validate_on_submit():
        flash("Senha inválida. Mínimo 6 caracteres.", "danger")
        return redirect(url_for("Paineis.admin_usuarios_editar_form", id_usuario=id_usuario))

    senha = form_senha.senha.data or ""
    usuario.HashSenha = _senha_hash(senha)
    usuario.UpdateAt = datetime.now()

    db.session.commit()

    flash("Senha atualizada com sucesso.", "success")
    return redirect(url_for("Paineis.admin_usuarios_editar_form", id_usuario=id_usuario))







@paineis_bp.post("/admin/usuarios/<int:id_usuario>/permissoes-extras/aplicar")
@login_required
@limiter.limit("30 per minute", methods=["POST"])
@requer_permissao("USUARIOS_GERENCIAR")
def admin_usuarios_add_perm_extra(id_usuario: int):
    usuario = db.session.query(DimUsuarios).filter(DimUsuarios.IDDimUsuarios == id_usuario).first()
    if not usuario:
        abort(404)

    permissoes_todas = db.session.query(DimPermissoes).order_by(DimPermissoes.CodigoPermissao.asc()).all()

    form_perm_extra = FormPermissaoExtraUpsert(prefix="pex")
    form_perm_extra.id_permissao.choices = [
        (perm.IDDimPermissoes, f"{(perm.CodigoPermissao or '').strip()} — {(perm.Descricao or '').strip()}")
        for perm in permissoes_todas
    ]

    if not form_perm_extra.validate_on_submit():
        flash("Dados de permissão extra inválidos.", "danger")
        return redirect(url_for("Paineis.admin_usuarios_editar_form", id_usuario=id_usuario))

    id_perm = int(form_perm_extra.id_permissao.data)
    tipo = (form_perm_extra.tipo.data or "").strip().upper()
    data_exp = form_perm_extra.data_expiracao.data
    obs = (form_perm_extra.observacao.data or "").strip()

    perm = db.session.query(DimPermissoes).filter(DimPermissoes.IDDimPermissoes == id_perm).first()
    if not perm:
        flash("Permissão não encontrada.", "danger")
        return redirect(url_for("Paineis.admin_usuarios_editar_form", id_usuario=id_usuario))

    row = (
        db.session.query(PermissoesUsuario)
        .filter(PermissoesUsuario.IDDimUsuarios == id_usuario)
        .filter(PermissoesUsuario.IDDimPermissoes == id_perm)
        .first()
    )

    if row:
        row.TipoAtribuicao = tipo
        row.DataExpiracao = data_exp
        row.Observacao = obs or None
        row.DataAtualizacao = datetime.now()
    else:
        row = PermissoesUsuario(
            IDDimUsuarios=id_usuario,
            IDDimPermissoes=id_perm,
            TipoAtribuicao=tipo,
            DataExpiracao=data_exp,
            Observacao=obs or None,
            DataAtualizacao=datetime.now(),
            CriadoPorIDDimUsuarios=None,
        )
        db.session.add(row)

    db.session.commit()

    flash("Permissão extra aplicada com sucesso.", "success")
    return redirect(url_for("Paineis.admin_usuarios_editar_form", id_usuario=id_usuario))






@paineis_bp.post("/admin/usuarios/<int:id_usuario>/permissoes-extras/remover")
@login_required
@limiter.limit("30 per minute", methods=["POST"])
@requer_permissao("USUARIOS_GERENCIAR")
def admin_usuarios_remover_perm_extra(id_usuario: int):


    id_perm_raw = (request.form.get("id_permissao") or "").strip()
    if not id_perm_raw:
        flash("Permissão inválida.", "danger")
        return redirect(url_for("Paineis.admin_usuarios_editar_form", id_usuario=id_usuario))

    prefix = f"rm_{id_perm_raw}"
    form_remover = FormPermissaoExtraRemover(prefix=prefix)

    if not form_remover.validate_on_submit():
        flash("Falha de CSRF ou formulário inválido.", "danger")
        return redirect(url_for("Paineis.admin_usuarios_editar_form", id_usuario=id_usuario))

    try:
        id_perm = int(form_remover.id_permissao.data)
    except Exception:
        flash("Permissão inválida.", "danger")
        return redirect(url_for("Paineis.admin_usuarios_editar_form", id_usuario=id_usuario))

    row = (
        db.session.query(PermissoesUsuario)
        .filter(PermissoesUsuario.IDDimUsuarios == id_usuario)
        .filter(PermissoesUsuario.IDDimPermissoes == id_perm)
        .first()
    )
    if not row:
        flash("Permissão extra não encontrada.", "warning")
        return redirect(url_for("Paineis.admin_usuarios_editar_form", id_usuario=id_usuario))

    db.session.delete(row)
    db.session.commit()

    flash("Permissão extra removida.", "success")
    return redirect(url_for("Paineis.admin_usuarios_editar_form", id_usuario=id_usuario))










def _get_checkout_codpontos() -> list[int]:
    """'Lê a lista do checkout na session e devolve lista de ints únicos.'"""
    bruto = session.get("checkout_codpontos", [])
    if not isinstance(bruto, list):
        bruto = []

    unicos = []
    vistos = set()

    for v in bruto:
        try:
            cod = int(v)
        except Exception:
            continue

        if cod <= 0:
            continue

        if cod in vistos:
            continue

        vistos.add(cod)
        unicos.append(cod)

    return unicos


def _set_checkout_codpontos(cods: list[int]) -> None:
    """'Escreve lista única no session e marca como modificada.'"""
    unicos = []
    vistos = set()

    for v in cods:
        try:
            cod = int(v)
        except Exception:
            continue

        if cod <= 0:
            continue

        if cod in vistos:
            continue

        vistos.add(cod)
        unicos.append(cod)

    session["checkout_codpontos"] = unicos
    session.modified = True


@paineis_bp.app_context_processor
def inject_checkout_qtd():
    """'Injeta checkout_qtd no Jinja para o badge.'"""
    cods = _get_checkout_codpontos()
    return {"checkout_qtd": len(cods)}



@paineis_bp.app_errorhandler(CSRFError)
def handle_csrf_error(e):
    """'Garante que chamadas do checkout recebam JSON em erro de CSRF.'"""
    path = (request.path or "").lower()

    wants_json = (
        path.startswith("/paineis/checkout")
        or request.headers.get("X-Requested-With") == "XMLHttpRequest"
        or "application/json" in (request.headers.get("Accept") or "")
        or request.is_json
    )

    if wants_json:
        return jsonify({"ok": False, "erro": "CSRF inválido/ausente. Recarregue a página e tente novamente."}), 400


    return e, 400


@paineis_bp.get("/checkout")
@login_required
def checkout_view():

    cods = _get_checkout_codpontos()

    itens = []
    faces_map = {}  

    if cods:
        paineis = (
            db.session.query(DimPaineisEuromidia)
            .filter(DimPaineisEuromidia.CodPonto.in_(cods))
            .all()
        )

        mapa = {int(p.CodPonto): p for p in paineis if p and p.CodPonto is not None}

        faces = (
            db.session.query(DimFacesPaineis)
            .filter(DimFacesPaineis.CodPonto.in_(cods))
            .order_by(DimFacesPaineis.CodPonto.asc(), DimFacesPaineis.CodFace.asc())
            .all()
        )
        for f in faces:
            try:
                cp = int(f.CodPonto or 0)
            except:
                continue
            if cp <= 0:
                continue
            faces_map.setdefault(cp, []).append({
                "CodFace": (f.CodFace or "").strip(),
                "TipoPainel": (f.TipoPainel or "").strip(),
            })

        for cod in cods:
            p = mapa.get(int(cod))
            if not p:
                continue

            itens.append({
                "CodPonto": int(p.CodPonto),
                "UF": (p.UF or "").strip(),
                "Bairro": (p.Bairro or "").strip(),
                "Exibidora": (p.Exibidora or "").strip(),
                "Endereco": (p.Logradouro or "").strip(),
                "CEP": (p.CEP or "").strip(),
                "Faces": (p.Faces if p.Faces is not None else None),
            })


    loops_permitidos = ["1MIN", "2MIN"]

    return render_template(
        "euromidia/checkout.html",
        itens=itens,
        faces_map=faces_map,
        loops_permitidos=loops_permitidos
    )


@paineis_bp.get("/checkout/status")
@login_required
def checkout_status():

    cods = _get_checkout_codpontos()
    return jsonify({"ok": True, "qtd": len(cods), "codpontos": cods})


@paineis_bp.post("/checkout/adicionar")
@login_required
def checkout_adicionar():

    payload = request.get_json(silent=True) or {}
    cod_raw = payload.get("codponto", None)

    try:
        cod = int(cod_raw)
    except Exception:
        return jsonify({"ok": False, "erro": "codponto inválido"}), 400

    if cod <= 0:
        return jsonify({"ok": False, "erro": "codponto inválido"}), 400

    existe = (
        db.session.query(DimPaineisEuromidia.CodPonto)
        .filter(DimPaineisEuromidia.CodPonto == cod)
        .first()
    )
    if not existe:
        return jsonify({"ok": False, "erro": f"CodPonto {cod} não encontrado"}), 404

    cods = _get_checkout_codpontos()
    ja_tinha = (cod in cods)

    if not ja_tinha:
        cods.append(cod)
        _set_checkout_codpontos(cods)

    cods_final = _get_checkout_codpontos()
    return jsonify({
        "ok": True,
        "adicionado": (not ja_tinha),
        "qtd": len(cods_final),
        "codpontos": cods_final
    })






@paineis_bp.post("/checkout/remover")
@login_required
def checkout_remover():
    """'Remove CodPonto do checkout.'"""
    payload = request.get_json(silent=True) or {}
    codponto = payload.get("codponto")

    try:
        codponto = int(codponto)
    except:
        return jsonify(ok=False, erro="codponto inválido"), 400

    ch = session.get("checkout_codpontos") or []
    ch = [int(x) for x in ch if str(x).isdigit()]

    antes = len(ch)
    ch = [x for x in ch if x != codponto]
    depois = len(ch)

    session["checkout_codpontos"] = ch
    session.modified = True

    removido = (depois < antes)

    return jsonify(
        ok=True,
        removido=removido,
        codponto=codponto,
        qtd=len(ch),
        codpontos=ch
    )


@paineis_bp.post("/checkout/limpar")
@login_required
def checkout_limpar():
    """'Limpa checkout.'"""
    _set_checkout_codpontos([])
    return jsonify({"ok": True, "qtd": 0, "codpontos": []})







@paineis_bp.get("/checkout/search/contratos")
@login_required
def checkout_search_contratos():
    q = (request.args.get("q") or "").strip()
    if len(q) < 2:
        return jsonify({"ok": True, "items": []})

    like = f"%{q}%"

    rows = (
        db.session.query(
            FatoControleContratos.IDFatoControleContratos,
            FatoControleContratos.NumeroContrato,
            FatoControleContratos.RazaoSocial,
            FatoControleContratos.CNPJ,
            FatoControleContratos.TotalFaturamentoLiquidoMensalFinal,
            FatoControleContratos.TotalFaturamentoLiquidoMensal,
            FatoControleContratos.TotalBrutoContrato,
        )
        .filter(
            (FatoControleContratos.NumeroContrato.ilike(like)) |
            (FatoControleContratos.RazaoSocial.ilike(like)) |
            (FatoControleContratos.CNPJ.ilike(like))
        )
        .order_by(FatoControleContratos.IDFatoControleContratos.desc())
        .limit(20)
        .all()
    )

    items = []
    for r in rows:
        items.append({
            "id": int(r[0]),
            "numero": (r[1] or "").strip(),
            "razao": (r[2] or "").strip(),
            "cnpj": (r[3] or "").strip(),
            "total_mensal_final": (str(r[4]) if r[4] is not None else ""),
            "total_mensal": (str(r[5]) if r[5] is not None else ""),
            "total_bruto": (str(r[6]) if r[6] is not None else ""),
        })

    return jsonify({"ok": True, "items": items})






@paineis_bp.get("/checkout/contrato/<int:id_contrato>")
@login_required
def checkout_get_contrato(id_contrato: int):
    r = (
        db.session.query(FatoControleContratos)
        .filter(FatoControleContratos.IDFatoControleContratos == id_contrato)
        .first()
    )
    if not r:
        return jsonify({"ok": False, "erro": "Contrato não encontrado"}), 404

    return jsonify({
        "ok": True,
        "contrato": {
            "id": int(r.IDFatoControleContratos),
            "numero": (r.NumeroContrato or "").strip(),
            "razao": (r.RazaoSocial or "").strip(),
            "cnpj": (r.CNPJ or "").strip(),
            "total_bruto": (str(r.TotalBrutoContrato) if r.TotalBrutoContrato is not None else ""),
            "total_mensal": (str(r.TotalFaturamentoLiquidoMensal) if r.TotalFaturamentoLiquidoMensal is not None else ""),
            "total_mensal_final": (str(r.TotalFaturamentoLiquidoMensalFinal) if r.TotalFaturamentoLiquidoMensalFinal is not None else ""),
        }
    })


@paineis_bp.get("/checkout/search/clientes")
@login_required
def checkout_search_clientes():
    q = (request.args.get("q") or "").strip()
    if len(q) < 2:
        return jsonify({"ok": True, "items": []})

    like = f"%{q}%"

    rows = (
        db.session.query(
            DimClientes.IDCliente,
            DimClientes.RazaoSocial,
            DimClientes.NomeFantasia,
            DimClientes.CNPJ,
        )
        .filter(
            (DimClientes.RazaoSocial.ilike(like)) |
            (DimClientes.NomeFantasia.ilike(like)) |
            (DimClientes.CNPJ.ilike(like))
        )
        .order_by(DimClientes.RazaoSocial.asc())
        .limit(20)
        .all()
    )

    items = []
    for r in rows:
        items.append({
            "id": int(r[0]),
            "razao": (r[1] or "").strip(),
            "fantasia": (r[2] or "").strip(),
            "cnpj": (r[3] or "").strip(),
        })

    return jsonify({"ok": True, "items": items})


@paineis_bp.get("/checkout/cliente/<int:id_cliente>")
@login_required
def checkout_get_cliente(id_cliente: int):
    r = (
        db.session.query(DimClientes)
        .filter(DimClientes.IDCliente == id_cliente)
        .first()
    )
    if not r:
        return jsonify({"ok": False, "erro": "Cliente não encontrado"}), 404

    return jsonify({
        "ok": True,
        "cliente": {
            "id": int(r.IDCliente),
            "razao": (r.RazaoSocial or "").strip(),
            "fantasia": (r.NomeFantasia or "").strip(),
            "cnpj": (r.CNPJ or "").strip(),
        }
    })





def _month_iter(start_y: int, start_m: int, count: int):
    """'Gera YYYY-MM sequencial por count meses.'"""
    y = start_y
    m = start_m
    for _ in range(count):
        yield (y, m)
        m += 1
        if m > 12:
            m = 1
            y += 1


def _to_month_key(d: date) -> str:
    """'YYYY-MM'"""
    return f"{d.year:04d}-{d.month:02d}"


@paineis_bp.get("/checkout/disponibilidade")
@login_required
def checkout_disponibilidade():
    """
    'Retorna meses disponíveis (YYYY-MM) para DataInicio/DataFim,
     removendo meses que encostam em intervalos ocupados.'
    """
    codponto = request.args.get("codponto")
    codface = (request.args.get("codface") or "").strip()

    try:
        codponto = int(codponto)
    except:
        return jsonify({"ok": False, "erro": "codponto inválido"}), 400

    if codponto <= 0 or not codface:
        return jsonify({"ok": False, "erro": "codponto/codface inválidos"}), 400

 
    ocup = (
        db.session.query(FatoOcupacaoPaineisEuromidia.DataInicio, FatoOcupacaoPaineisEuromidia.DataFim)
        .filter(FatoOcupacaoPaineisEuromidia.CodPonto == codponto)
        .filter(FatoOcupacaoPaineisEuromidia.CodFace == codface)
        .all()
    )


    bloqueados = set()

    for di, df in ocup:
        if not di or not df:
            continue

  
        y = di.year
        m = di.month

        while True:
            key = f"{y:04d}-{m:02d}"
            bloqueados.add(key)

            if y == df.year and m == df.month:
                break

            m += 1
            if m > 12:
                m = 1
                y += 1

   
    hoje = date.today()
    candidatos = []
    for (y, m) in _month_iter(hoje.year, hoje.month, 24):
        key = f"{y:04d}-{m:02d}"
        if key not in bloqueados:
            candidatos.append(key)

    return jsonify({"ok": True, "meses": candidatos, "bloqueados": sorted(list(bloqueados))})



import os
import time
from types import SimpleNamespace

from flask import request, render_template, current_app, url_for
from flask_login import login_required
from sqlalchemy import or_, desc, func, text


@paineis_bp.get("/contratos")
@login_required
def contratos_lista():
    tempo_inicio = time.perf_counter()

    q = (request.args.get("q") or "").strip()

    try:
        page = int(request.args.get("page") or 1)
    except ValueError:
        page = 1
    if page < 1:
        page = 1

    try:
        per_page = int(request.args.get("per_page") or 25)
    except ValueError:
        per_page = 25

    if per_page < 10:
        per_page = 10
    if per_page > 100:
        per_page = 100

    calcular_total = (request.args.get("calcular_total") == "1")

    consulta_base_ids = db.session.query(
        FatoControleContratosEuromidia.IDFatoControleContratosEuromidia
    )

    if q:
        like_prefixo = f"{q}%"
        like_contendo = f"%{q}%"

        texto_sem_mascara = (
            q.replace(".", "")
             .replace("/", "")
             .replace("-", "")
             .replace(" ", "")
        )

        busca_documental = texto_sem_mascara.isdigit()

        if busca_documental:
            consulta_base_ids = consulta_base_ids.filter(
                or_(
                    FatoControleContratosEuromidia.NumeroContrato.like(like_prefixo),
                    FatoControleContratosEuromidia.NumeroPrevia.like(like_prefixo),
                    FatoControleContratosEuromidia.CNPJ.like(like_prefixo),
                    FatoControleContratosEuromidia.CPF.like(like_prefixo),
                    func.cast(
                        FatoControleContratosEuromidia.IDFatoControleContratosEuromidia,
                        db.String
                    ).like(like_prefixo),
                )
            )
        else:
            consulta_base_ids = consulta_base_ids.filter(
                or_(
                    FatoControleContratosEuromidia.NumeroContrato.like(like_prefixo),
                    FatoControleContratosEuromidia.NumeroPrevia.like(like_prefixo),
                    FatoControleContratosEuromidia.CNPJ.like(like_prefixo),
                    FatoControleContratosEuromidia.CPF.like(like_prefixo),
                    FatoControleContratosEuromidia.RazaoSocial.like(like_contendo),
                    FatoControleContratosEuromidia.MarcaExibida.like(like_contendo),
                    FatoControleContratosEuromidia.Vendedor.like(like_contendo),
                    FatoControleContratosEuromidia.Agencia.like(like_contendo),
                    FatoControleContratosEuromidia.Bureau.like(like_contendo),
                    FatoControleContratosEuromidia.Intermediario.like(like_contendo),
                    FatoControleContratosEuromidia.TipoDocumento.like(like_contendo),
                    FatoControleContratosEuromidia.Origem.like(like_contendo),
                    FatoControleContratosEuromidia.SDR.like(like_contendo),
                )
            )

    total = None
    total_pages = None

    tempo_count = 0.0
    if calcular_total:
        t_count0 = time.perf_counter()

        total = consulta_base_ids.order_by(None).count()

        total_pages = max(1, (total + per_page - 1) // per_page)

        if page > total_pages:
            page = total_pages

        tempo_count = time.perf_counter() - t_count0

    offset = (page - 1) * per_page

    t_ids0 = time.perf_counter()

    ids_plus = [
        linha[0]
        for linha in (
            consulta_base_ids
            .order_by(
                desc(FatoControleContratosEuromidia.DataLancamento),
                desc(FatoControleContratosEuromidia.IDFatoControleContratosEuromidia),
            )
            .offset(offset)
            .limit(per_page + 1)
            .all()
        )
    ]

    tempo_ids = time.perf_counter() - t_ids0

    has_next = len(ids_plus) > per_page
    ids_pagina = ids_plus[:per_page]
    has_prev = (page > 1)

    contratos = []

    tempo_logo = 0.0
    tempo_detalhes = 0.0

    logo_empresa_url = ""
    logo_empresa_raw = ""

    if ids_pagina:
        t_logo0 = time.perf_counter()

        row_logo = db.session.execute(
            text("""
                SELECT TOP 1
                    Logo
                FROM [Integracao].[dbo].[EmpresaProprietaria]
                WHERE IDEmpresaProprietaria = 3
            """)
        ).mappings().first()

        logo_empresa_raw = (row_logo.get("Logo") if row_logo else "") or ""
        logo_empresa_raw = logo_empresa_raw.strip()

        if logo_empresa_raw:
            nome_arquivo_logo = os.path.basename(logo_empresa_raw.replace("\\", "/"))
            logo_empresa_url = url_for(
                "static",
                filename=f"LogoEmpresaProprietaria/{nome_arquivo_logo}",
            )

        tempo_logo = time.perf_counter() - t_logo0

        t_detalhes0 = time.perf_counter()

        subquery_cidades = (
            db.session.query(
                FatoControleContratosItensEuromidia.IDFatoControleContratoEuromidia.label("id_contrato"),
                func.min(FatoControleContratosItensEuromidia.CidadeExibicao).label("CidadeExibicao"),
            )
            .filter(
                FatoControleContratosItensEuromidia.IDFatoControleContratoEuromidia.in_(ids_pagina)
            )
            .group_by(
                FatoControleContratosItensEuromidia.IDFatoControleContratoEuromidia
            )
            .subquery()
        )

        linhas_contratos = (
            db.session.query(
                FatoControleContratosEuromidia.IDFatoControleContratosEuromidia.label("IDContrato"),
                FatoControleContratosEuromidia.IDFatoControleContratosEuromidia.label("IDFatoControleContratosEuromidia"),
                FatoControleContratosEuromidia.DataLancamento.label("DataLancamento"),
                FatoControleContratosEuromidia.RazaoSocial.label("RazaoSocial"),
                FatoControleContratosEuromidia.MarcaExibida.label("MarcaExibida"),
                subquery_cidades.c.CidadeExibicao.label("CidadeExibicao"),
            )
            .outerjoin(
                subquery_cidades,
                subquery_cidades.c.id_contrato == FatoControleContratosEuromidia.IDFatoControleContratosEuromidia
            )
            .filter(
                FatoControleContratosEuromidia.IDFatoControleContratosEuromidia.in_(ids_pagina)
            )
            .all()
        )

        tempo_detalhes = time.perf_counter() - t_detalhes0

        mapa_contratos = {}

        for linha in linhas_contratos:
            dados = dict(linha._mapping)
            dados["LogoEmpresaProprietaria"] = logo_empresa_raw
            dados["LogoEmpresaProprietariaUrl"] = logo_empresa_url
            mapa_contratos[dados["IDFatoControleContratosEuromidia"]] = SimpleNamespace(**dados)

        contratos = [
            mapa_contratos[id_contrato]
            for id_contrato in ids_pagina
            if id_contrato in mapa_contratos
        ]

    if total is None:
        inicio = 0 if len(contratos) == 0 else (offset + 1)
        fim = offset + len(contratos)
    else:
        inicio = 0 if total == 0 else (offset + 1)
        fim = min(offset + per_page, total)

    if total_pages is None:
        total_pages = page + 1 if has_next else page

    paginacao = {
        "page": page,
        "per_page": per_page,
        "total": total,
        "total_pages": total_pages,
        "inicio": inicio,
        "fim": fim,
        "has_next": has_next,
        "has_prev": has_prev,
        "calcular_total": calcular_total,
    }

    filtros = {
        "q": q,
        "per_page": per_page,
    }

    tempo_total = time.perf_counter() - tempo_inicio

    current_app.logger.warning(
        "contratos_lista | q=%r | page=%s | per_page=%s | total=%s | "
        "count=%.3fs | ids=%.3fs | logo=%.3fs | detalhes=%.3fs | total_req=%.3fs",
        q,
        page,
        per_page,
        total,
        tempo_count,
        tempo_ids,
        tempo_logo,
        tempo_detalhes,
        tempo_total,
    )

    return render_template(
        "euromidia/contratos_lista.html",
        contratos=contratos,
        paginacao=paginacao,
        filtros=filtros,
    )



@paineis_bp.get("/contratos/<int:id_contrato>")
@login_required
def contratos_detalhe(id_contrato: int):
    contrato = (
        db.session.query(FatoControleContratosEuromidia)
        .filter(FatoControleContratosEuromidia.IDFatoControleContratosEuromidia == id_contrato)
        .first()
    )

    if not contrato:
        abort(404, description="Contrato não encontrado.")

    
    itens = list(contrato.Itens or [])

    return render_template(
        "euromidia/contratos_detalhe.html",
        contrato=contrato,
        itens=itens,
    )













import hashlib
import json
import os
import re

from sqlalchemy import String, cast, func
from sqlalchemy.orm import aliased

from ..extensions import cache


TEMPO_CACHE_TOTAL_CLIENTES_SEGUNDOS = 120
TEMPO_CACHE_ITENS_CLIENTES_SEGUNDOS = 120
TEMPO_CACHE_FILTROS_CLIENTES_SEGUNDOS = 300
TEMPO_CACHE_OPCOES_CLIENTE_SEGUNDOS = 300


def _so_digitos(v: str) -> str:
    if not v:
        return ""
    return (
        str(v)
        .replace(".", "")
        .replace("-", "")
        .replace("/", "")
        .replace(" ", "")
        .strip()
    )


def _paginacao_basica(page: int, per_page: int, total: int):
    """Cria objeto simples de paginação igual o padrão do seu template."""
    if per_page <= 0:
        per_page = 20

    total_pages = max((total + per_page - 1) // per_page, 1)
    page = max(min(page, total_pages), 1)

    inicio = (page - 1) * per_page + 1 if total > 0 else 0
    fim = min(page * per_page, total) if total > 0 else 0

    return {
        "page": page,
        "per_page": per_page,
        "total": total,
        "total_pages": total_pages,
        "inicio": inicio,
        "fim": fim,
        "has_prev": page > 1,
        "has_next": page < total_pages,
        "prev_page": page - 1 if page > 1 else 1,
        "next_page": page + 1 if page < total_pages else total_pages,
    }


def _filtro_anti_lixo_clientes():
    return (
        (DimEmpresas.CNPJ.isnot(None))
        & (DimEmpresas.CNPJ != "")
        & (DimEmpresas.CNPJ != "00000000000000")
        & (DimEmpresas.CNPJ != "99999999999999")
        & (DimEmpresas.RazaoSocial.isnot(None))
        & (DimEmpresas.RazaoSocial != "")
        & (DimEmpresas.RazaoSocial != "None")
        & (DimEmpresas.UF.isnot(None))
        & (DimEmpresas.UF != "")
        & (DimEmpresas.Municipio.isnot(None))
        & (DimEmpresas.Municipio != "")
    )


def _aplicar_from_clientes(query):
    cls = DimClassificacacaoClientes
    rec = DimRecorrencia
    pub = DimPublicoAlvo

    return (
        query
        .select_from(DimEmpresas)
        .outerjoin(DimCnaes, DimCnaes.cnaepadrao == DimEmpresas.CNAE)
        .outerjoin(
            DimEmpresaProprietaria,
            DimEmpresaProprietaria.IDEmpresaProprietaria == DimEmpresas.IDEmpresaProprietaria,
        )
        .outerjoin(cls, cls.IDEmpresa == DimEmpresas.IDEmpresa)
        .outerjoin(
            rec,
            (rec.IDEmpresa == DimEmpresas.IDEmpresa)
            & (
                (rec.IDEmpresaProprietaria.is_(None))
                | (rec.IDEmpresaProprietaria == DimEmpresas.IDEmpresaProprietaria)
            ),
        )
        .outerjoin(pub, pub.IDDimPublicoAlvo == cls.IDDimPublicoAlvo)
        .filter(_filtro_anti_lixo_clientes())
    )


def _query_clientes_base_ids():
    return _aplicar_from_clientes(
        db.session.query(DimEmpresas.IDEmpresa.label("IDDimClientesEuromidia"))
    )


def _query_clientes_lista():
    cls = DimClassificacacaoClientes
    rec = DimRecorrencia
    pub = DimPublicoAlvo

    return _aplicar_from_clientes(
        db.session.query(
            DimEmpresas.IDEmpresa.label("IDDimClientesEuromidia"),
            DimEmpresas.CNPJ.label("CNPJ"),
            DimEmpresas.RazaoSocial.label("RazaoSocial"),
            DimEmpresas.Porte.label("Porte"),
            DimEmpresas.UF.label("UF"),
            DimEmpresas.Municipio.label("Municipio"),
            DimEmpresas.DescricaoSituacaoCadastral.label("DescricaoSituacaoCadastral"),
            DimEmpresas.DescricaoIdentificadorMatrizFilial.label("DescricaoMatrizFilial"),
            DimEmpresas.BitCliente.label("BitCliente"),
            DimCnaes.Classe.label("Classe"),
            DimCnaes.Setor.label("Setor"),
            DimCnaes.ScoreSetor.label("ScoreSetor"),
            DimCnaes.ClassificacaoMacro.label("ClassificacaoMacro"),
            cls.ClusterGrupoCliente.label("ClusterGrupoCliente"),
            cls.ScorePerfilEmpresa.label("ScorePerfilEmpresa"),
            cls.ClassificacaoPerfilEmpresa.label("ClassificacaoPerfilEmpresa"),
            DimCnaes.SubClasse.label("SubClasse"),
            DimEmpresas.IDEmpresaProprietaria.label("IDEmpresaProprietaria"),
            DimEmpresaProprietaria.Logo.label("LogoEmpresaProprietaria"),
            DimEmpresaProprietaria.RazaoSocial.label("RazaoSocialEmpresaProprietaria"),
            cls.ClasseValor.label("ClasseValor"),
            cls.TipoEscalaOperacional.label("TipoEscalaOperacional"),
            cls.ClasseEstrutural.label("ClasseEstrutural"),
            cls.ClasseGeo.label("ClasseGeo"),
            rec.Frequencia12M.label("Frequencia12M"),
            rec.ClasseFrequencia.label("ClasseFrequencia"),
            rec.DataUltimaAquisicao.label("DataUltimaAquisicao"),
            rec.DiasDesdeUltimaAquisicao.label("DiasDesdeUltimaAquisicao"),
            rec.ClasseRecencia.label("ClasseRecencia"),
            pub.IDDimPublicoAlvo.label("IDDimPublicoAlvo"),
            pub.NomePerfil.label("NomePerfilPublico"),
            pub.TipoUsoTerritorio.label("TipoUsoTerritorioPublico"),
            pub.FaixaEconomica.label("FaixaEconomicaPublico"),
            pub.TipoDemanda.label("TipoDemandaPublico"),
        )
    )


DEFINICOES_FILTROS_CLIENTES = {
    "municipio": {
        "coluna": DimEmpresas.Municipio,
        "multiplo": False,
    },
    "porte": {
        "coluna": DimEmpresas.Porte,
        "multiplo": False,
    },
    "classe": {
        "coluna": DimCnaes.Classe,
        "multiplo": True,
    },
    "setor": {
        "coluna": DimCnaes.Setor,
        "multiplo": True,
    },
    "subclasse": {
        "coluna": DimCnaes.SubClasse,
        "multiplo": True,
    },
    "empresa_proprietaria": {
        "coluna": DimEmpresaProprietaria.RazaoSocial,
        "multiplo": True,
    },
    "classe_valor": {
        "coluna": DimClassificacacaoClientes.ClasseValor,
        "multiplo": True,
    },
    "tipo_escala_operacional": {
        "coluna": DimClassificacacaoClientes.TipoEscalaOperacional,
        "multiplo": True,
    },
    "classe_estrutural": {
        "coluna": DimClassificacacaoClientes.ClasseEstrutural,
        "multiplo": True,
    },
    "classe_geo": {
        "coluna": DimClassificacacaoClientes.ClasseGeo,
        "multiplo": True,
    },
    "classe_frequencia": {
        "coluna": DimRecorrencia.ClasseFrequencia,
        "multiplo": True,
    },
    "classe_recencia": {
        "coluna": DimRecorrencia.ClasseRecencia,
        "multiplo": True,
    },
    "nome_perfil_publico": {
        "coluna": DimPublicoAlvo.NomePerfil,
        "multiplo": True,
    },
    "tipo_uso_territorio": {
        "coluna": DimPublicoAlvo.TipoUsoTerritorio,
        "multiplo": True,
    },
    "classificacao_macro": {
        "coluna": DimCnaes.ClassificacaoMacro,
        "multiplo": True,
    },
}


def _deduplicar_textos_preservando_ordem(valores):
    saida = []
    vistos = set()

    for valor in (valores or []):
        texto = str(valor or "").strip()
        if not texto:
            continue

        chave = texto.casefold()
        if chave in vistos:
            continue

        vistos.add(chave)
        saida.append(texto)

    return saida


def _ordenar_classificacao_macro(valores):
    ordem = {
        "Muito Favorável": 1,
        "Favorável": 2,
        "Neutro": 3,
        "Desfavorável": 4,
        "Muito Desfavorável": 5,
    }

    return sorted(
        _deduplicar_textos_preservando_ordem(valores),
        key=lambda valor: (ordem.get(valor, 999), valor.casefold()),
    )


def _normalizar_filtros_clientes_para_backend(filtros):
    saida = {}

    for nome, meta in DEFINICOES_FILTROS_CLIENTES.items():
        valor = filtros.get(nome)

        if meta["multiplo"]:
            saida[nome] = [
                str(x).strip()
                for x in (valor or [])
                if str(x or "").strip()
            ]
        else:
            saida[nome] = str(valor or "").strip()

    saida["q"] = str(filtros.get("q") or "").strip()
    saida["cliente"] = str(filtros.get("cliente") or "todos").strip().lower()

    if saida["cliente"] not in {"todos", "1", "0"}:
        saida["cliente"] = "todos"

    return saida


def _serializar_para_chave_cache(valor) -> str:
    return json.dumps(valor, sort_keys=True, ensure_ascii=False, default=str, separators=(",", ":"))


def _gerar_chave_cache_clientes(prefixo: str, payload: dict) -> str:
    texto = _serializar_para_chave_cache(payload)
    hash_texto = hashlib.sha256(texto.encode("utf-8")).hexdigest()
    return f"clientes:{prefixo}:{hash_texto}"


def _aplicar_filtros_clientes(query, filtros, excluir=None):
    excluir = set(excluir or [])
    filtros = _normalizar_filtros_clientes_para_backend(filtros)

    cls = DimClassificacacaoClientes
    rec = DimRecorrencia
    pub = DimPublicoAlvo

    if "q" not in excluir and filtros["q"]:
        like = f"%{filtros['q']}%"
        query = query.filter(
            (DimEmpresas.RazaoSocial.like(like))
            | (DimEmpresas.NomeFantasia.like(like))
            | (DimEmpresas.CNPJ.like(like))
            | (DimEmpresas.Porte.like(like))
            | (DimEmpresas.Municipio.like(like))
            | (DimEmpresas.UF.like(like))
            | (DimEmpresas.DescricaoSituacaoCadastral.like(like))
            | (DimEmpresas.DescricaoIdentificadorMatrizFilial.like(like))
            | (DimCnaes.Setor.like(like))
            | (DimCnaes.Classe.like(like))
            | (DimCnaes.SubClasse.like(like))
            | (DimCnaes.ClassificacaoMacro.like(like))
            | (cls.ClusterGrupoCliente.like(like))
            | (cast(cls.ScorePerfilEmpresa, String).like(like))
            | (cls.ClassificacaoPerfilEmpresa.like(like))
            | (DimEmpresaProprietaria.RazaoSocial.like(like))
            | (cls.ClasseValor.like(like))
            | (cls.TipoEscalaOperacional.like(like))
            | (cls.ClasseEstrutural.like(like))
            | (cls.ClasseGeo.like(like))
            | (rec.ClasseFrequencia.like(like))
            | (rec.ClasseRecencia.like(like))
            | (pub.NomePerfil.like(like))
            | (pub.TipoUsoTerritorio.like(like))
        )

    for nome, meta in DEFINICOES_FILTROS_CLIENTES.items():
        if nome in excluir:
            continue

        coluna = meta["coluna"]
        valor = filtros.get(nome)

        if meta["multiplo"]:
            if valor:
                query = query.filter(coluna.in_(valor))
        else:
            if valor:
                query = query.filter(coluna == valor)

    if "cliente" not in excluir:
        if filtros["cliente"] == "1":
            query = query.filter(func.coalesce(DimEmpresas.BitCliente, 0) == 1)
        elif filtros["cliente"] == "0":
            query = query.filter(func.coalesce(DimEmpresas.BitCliente, 0) == 0)

    return query


def _obter_valores_distintos_filtro_clientes(nome_filtro, filtros):
    filtros_normalizados = _normalizar_filtros_clientes_para_backend(filtros)

    payload_cache = {
        "nome_filtro": nome_filtro,
        "filtros": filtros_normalizados,
        "excluir": [nome_filtro],
    }
    chave_cache = _gerar_chave_cache_clientes("filtro_distinto", payload_cache)

    valores_cache = cache.get(chave_cache)
    if valores_cache is not None:
        return valores_cache

    meta = DEFINICOES_FILTROS_CLIENTES[nome_filtro]
    coluna = meta["coluna"]

    query = _query_clientes_base_ids()
    query = _aplicar_filtros_clientes(query, filtros_normalizados, excluir={nome_filtro})

    valores_banco = [
        row[0]
        for row in (
            query.with_entities(coluna)
            .filter(coluna.isnot(None))
            .filter(coluna != "")
            .distinct()
            .order_by(coluna.asc())
            .all()
        )
    ]

    valores_banco = _deduplicar_textos_preservando_ordem(valores_banco)

    valor_selecionado = filtros_normalizados.get(nome_filtro)
    if meta["multiplo"]:
        selecionados = [
            str(x).strip()
            for x in (valor_selecionado or [])
            if str(x or "").strip()
        ]
    else:
        selecionados = [str(valor_selecionado).strip()] if str(valor_selecionado or "").strip() else []

    valores = _deduplicar_textos_preservando_ordem(valores_banco + selecionados)

    if nome_filtro == "classificacao_macro":
        valores = _ordenar_classificacao_macro(valores)
    else:
        valores = sorted(valores, key=lambda valor: valor.casefold())

    cache.set(
        chave_cache,
        valores,
        timeout=TEMPO_CACHE_FILTROS_CLIENTES_SEGUNDOS,
    )
    return valores


def _obter_opcoes_cliente_dinamicas(filtros):
    filtros_normalizados = _normalizar_filtros_clientes_para_backend(filtros)

    payload_cache = {
        "tipo": "opcoes_cliente",
        "filtros": filtros_normalizados,
        "excluir": ["cliente"],
    }
    chave_cache = _gerar_chave_cache_clientes("opcoes_cliente", payload_cache)

    opcoes_cache = cache.get(chave_cache)
    if opcoes_cache is not None:
        return opcoes_cache

    query = _query_clientes_base_ids()
    query = _aplicar_filtros_clientes(query, filtros_normalizados, excluir={"cliente"})

    tem_cliente = (
        query.filter(func.coalesce(DimEmpresas.BitCliente, 0) == 1)
        .with_entities(DimEmpresas.IDEmpresa)
        .first()
        is not None
    )

    tem_nao_cliente = (
        query.filter(func.coalesce(DimEmpresas.BitCliente, 0) == 0)
        .with_entities(DimEmpresas.IDEmpresa)
        .first()
        is not None
    )

    opcoes = [{"valor": "todos", "rotulo": "Todos"}]

    if tem_cliente:
        opcoes.append({"valor": "1", "rotulo": "Cliente"})

    if tem_nao_cliente:
        opcoes.append({"valor": "0", "rotulo": "Não cliente"})

    cache.set(
        chave_cache,
        opcoes,
        timeout=TEMPO_CACHE_OPCOES_CLIENTE_SEGUNDOS,
    )
    return opcoes


def _obter_listas_filtros_clientes_dinamicas(filtros):
    filtros = _normalizar_filtros_clientes_para_backend(filtros)

    return {
        "municipios": _obter_valores_distintos_filtro_clientes("municipio", filtros),
        "portes": _obter_valores_distintos_filtro_clientes("porte", filtros),
        "classes": _obter_valores_distintos_filtro_clientes("classe", filtros),
        "setores": _obter_valores_distintos_filtro_clientes("setor", filtros),
        "subclasses": _obter_valores_distintos_filtro_clientes("subclasse", filtros),
        "empresas_proprietarias": _obter_valores_distintos_filtro_clientes("empresa_proprietaria", filtros),
        "classes_valor": _obter_valores_distintos_filtro_clientes("classe_valor", filtros),
        "tipos_escala_operacional": _obter_valores_distintos_filtro_clientes("tipo_escala_operacional", filtros),
        "classes_estruturais": _obter_valores_distintos_filtro_clientes("classe_estrutural", filtros),
        "classes_geo": _obter_valores_distintos_filtro_clientes("classe_geo", filtros),
        "classes_frequencia": _obter_valores_distintos_filtro_clientes("classe_frequencia", filtros),
        "classes_recencia": _obter_valores_distintos_filtro_clientes("classe_recencia", filtros),
        "nomes_perfil_publico": _obter_valores_distintos_filtro_clientes("nome_perfil_publico", filtros),
        "tipos_uso_territorio": _obter_valores_distintos_filtro_clientes("tipo_uso_territorio", filtros),
        "classificacoes_macro": _obter_valores_distintos_filtro_clientes("classificacao_macro", filtros),
        "opcoes_cliente": _obter_opcoes_cliente_dinamicas(filtros),
    }


def _obter_total_clientes_cacheado(filtros):
    filtros_normalizados = _normalizar_filtros_clientes_para_backend(filtros)
    chave_cache = _gerar_chave_cache_clientes(
        "total",
        {
            "filtros": filtros_normalizados,
        },
    )

    valor_cache = cache.get(chave_cache)
    if valor_cache is not None:
        return int(valor_cache)

    base = _query_clientes_lista()
    base = _aplicar_filtros_clientes(base, filtros_normalizados)

    total = (
        base.order_by(None)
        .with_entities(func.count(func.distinct(DimEmpresas.IDEmpresa)))
        .scalar()
    ) or 0

    cache.set(
        chave_cache,
        int(total),
        timeout=TEMPO_CACHE_TOTAL_CLIENTES_SEGUNDOS,
    )
    return int(total)


def _obter_itens_clientes_cacheados(filtros, page: int, per_page: int):
    filtros_normalizados = _normalizar_filtros_clientes_para_backend(filtros)

    payload_cache = {
        "filtros": filtros_normalizados,
        "page": int(page),
        "per_page": int(per_page),
    }
    chave_cache = _gerar_chave_cache_clientes("itens", payload_cache)

    valor_cache = cache.get(chave_cache)
    if valor_cache is not None:
        return valor_cache

    base = _query_clientes_lista()
    base = _aplicar_filtros_clientes(base, filtros_normalizados)

    rows = (
        base.order_by(DimEmpresas.RazaoSocial.asc(), DimEmpresas.IDEmpresa.asc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    itens = []
    for r in rows:
        d = dict(r._asdict())

        logo_path = (d.get("LogoEmpresaProprietaria") or "").strip()
        if logo_path:
            nome_arquivo = os.path.basename(logo_path.replace("\\", "/"))
            d["LogoEmpresaProprietariaUrl"] = url_for(
                "static",
                filename=f"LogoEmpresaProprietaria/{nome_arquivo}",
            )
        else:
            d["LogoEmpresaProprietariaUrl"] = ""

        itens.append(d)

    cache.set(
        chave_cache,
        itens,
        timeout=TEMPO_CACHE_ITENS_CLIENTES_SEGUNDOS,
    )
    return itens


@paineis_bp.get("/clientes")
@login_required
def clientes_lista():
    try:
        page = int((request.args.get("page") or "1").strip())
    except:
        page = 1
    page = max(page, 1)

    try:
        per_page = int((request.args.get("per_page") or "20").strip())
    except:
        per_page = 20
    per_page = max(5, min(per_page, 200))

    q = (request.args.get("q") or "").strip()
    municipio = (request.args.get("municipio") or "").strip()
    porte = (request.args.get("porte") or "").strip()

    classe = [x.strip() for x in (request.args.getlist("classe") or []) if (x or "").strip()]
    setor = [x.strip() for x in (request.args.getlist("setor") or []) if (x or "").strip()]
    subclasse = [x.strip() for x in (request.args.getlist("subclasse") or []) if (x or "").strip()]

    empresa_proprietaria = [
        x.strip()
        for x in (request.args.getlist("empresa_proprietaria") or [])
        if (x or "").strip()
    ]

    classe_valor = [
        x.strip()
        for x in (request.args.getlist("classe_valor") or [])
        if (x or "").strip()
    ]

    tipo_escala_operacional = [
        x.strip()
        for x in (request.args.getlist("tipo_escala_operacional") or [])
        if (x or "").strip()
    ]

    classe_estrutural = [
        x.strip()
        for x in (request.args.getlist("classe_estrutural") or [])
        if (x or "").strip()
    ]

    classe_geo = [
        x.strip()
        for x in (request.args.getlist("classe_geo") or [])
        if (x or "").strip()
    ]

    classe_frequencia = [
        x.strip()
        for x in (request.args.getlist("classe_frequencia") or [])
        if (x or "").strip()
    ]

    classe_recencia = [
        x.strip()
        for x in (request.args.getlist("classe_recencia") or [])
        if (x or "").strip()
    ]

    nome_perfil_publico = [
        x.strip()
        for x in (request.args.getlist("nome_perfil_publico") or [])
        if (x or "").strip()
    ]

    tipo_uso_territorio = [
        x.strip()
        for x in (request.args.getlist("tipo_uso_territorio") or [])
        if (x or "").strip()
    ]

    classificacao_macro = [
        x.strip()
        for x in (request.args.getlist("classificacao_macro") or [])
        if (x or "").strip()
    ]

    cliente = (request.args.get("cliente") or "todos").strip().lower()
    if cliente not in {"todos", "1", "0"}:
        cliente = "todos"

    filtros = {
        "q": q,
        "municipio": municipio,
        "porte": porte,
        "per_page": per_page,
        "classe": classe,
        "setor": setor,
        "subclasse": subclasse,
        "empresa_proprietaria": empresa_proprietaria,
        "classe_valor": classe_valor,
        "tipo_escala_operacional": tipo_escala_operacional,
        "classe_estrutural": classe_estrutural,
        "classe_geo": classe_geo,
        "classe_frequencia": classe_frequencia,
        "classe_recencia": classe_recencia,
        "nome_perfil_publico": nome_perfil_publico,
        "tipo_uso_territorio": tipo_uso_territorio,
        "classificacao_macro": classificacao_macro,
        "cliente": cliente,
    }

    total = _obter_total_clientes_cacheado(filtros)

    pag = _paginacao_basica(page, per_page, total)

    itens = _obter_itens_clientes_cacheados(
        filtros=filtros,
        page=pag["page"],
        per_page=pag["per_page"],
    )

    listas = _obter_listas_filtros_clientes_dinamicas(filtros)

    municipios = listas["municipios"]
    portes = listas["portes"]
    classes = listas["classes"]
    setores = listas["setores"]
    subclasses = listas["subclasses"]
    empresas_proprietarias = listas["empresas_proprietarias"]
    classes_valor = listas["classes_valor"]
    tipos_escala_operacional = listas["tipos_escala_operacional"]
    classes_estruturais = listas["classes_estruturais"]
    classes_geo = listas["classes_geo"]
    classes_frequencia = listas["classes_frequencia"]
    classes_recencia = listas["classes_recencia"]
    nomes_perfil_publico = listas["nomes_perfil_publico"]
    tipos_uso_territorio = listas["tipos_uso_territorio"]
    classificacoes_macro = listas["classificacoes_macro"]
    opcoes_cliente = listas["opcoes_cliente"]

    filtros["per_page"] = pag["per_page"]

    return render_template(
        "euromidia/clientes_lista.html",
        itens=itens,
        paginacao=pag,
        filtros=filtros,
        municipios=municipios,
        portes=portes,
        classes=classes,
        setores=setores,
        subclasses=subclasses,
        empresas_proprietarias=empresas_proprietarias,
        classes_valor=classes_valor,
        tipos_escala_operacional=tipos_escala_operacional,
        classes_estruturais=classes_estruturais,
        classes_geo=classes_geo,
        classes_frequencia=classes_frequencia,
        classes_recencia=classes_recencia,
        nomes_perfil_publico=nomes_perfil_publico,
        tipos_uso_territorio=tipos_uso_territorio,
        classificacoes_macro=classificacoes_macro,
        opcoes_cliente=opcoes_cliente,
    )


def _cnpj_somente_digitos(valor: str) -> str:
    if not valor:
        return ""
    return re.sub(r"\D+", "", str(valor))


def _cnpj_14(cnpj: str) -> str:
    d = _cnpj_somente_digitos(cnpj)
    if len(d) != 14:
        return d
    return f"{d[0:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:14]}"


def _matriz_filial_label(valor) -> str:
    """
    IdentificadorMatrizFilial (Receita):
      1 = Matriz
      2 = Filial
    """
    try:
        v = int(valor)
    except:
        return ""

    if v == 1:
        return "Matriz"

    if v == 2:
        return "Filial"

    return ""




@paineis_bp.get("/clientes/<int:id_empresa>")
def cliente_detalhe(id_empresa: int):

    row = (
        db.session.query(
            DimEmpresas.IDEmpresa.label("IDDimClientesEuromidia"),
            DimEmpresas.CNPJ.label("CNPJ"),
            DimEmpresas.RazaoSocial.label("RazaoSocial"),
            DimEmpresas.NomeFantasia.label("NomeFantasia"),

            DimEmpresas.Municipio.label("Municipio"),
            DimEmpresas.UF.label("UF"),
            DimEmpresas.CEP.label("CEP"),
            DimEmpresas.Bairro.label("Bairro"),
            DimEmpresas.Numero.label("Numero"),
            DimEmpresas.Logradouro.label("Logradouro"),
            DimEmpresas.Complemento.label("Complemento"),
            DimEmpresas.DescricaoTipoLogradouro.label("TipoLogradouro"),

            DimEmpresas.Porte.label("Porte"),
            DimEmpresas.IdentificadorMatrizFilial.label("MatrizFilial"),

            DimEmpresas.CNAE.label("CNAE"),
            DimEmpresas.DescricaoCnae.label("DescricaoCnae"),
            DimEmpresas.CapitalSocial.label("CapitalSocial"),

            DimEmpresas.DataInicioAtividades.label("DataInicioAtividade"),
            DimEmpresas.DataSituacaoCadastral.label("DataSituacaoCadastral"),

            DimEmpresas.IDEmpresaProprietaria.label("IDEmpresaProprietaria"),

            DimCnaes.Setor.label("Setor"),
            DimCnaes.MacroSetor.label("MacroSetor"),

            DimCnaes.ScoreSetor.label("ScoreSetor"),
            DimCnaes.ClassificacaoMacro.label("ClassificacaoSetor"),
            DimCnaes.ClassificacaoMacro.label("ClassificacaoMacro"),

            DimEmpresaProprietaria.RazaoSocial.label("RazaoSocialEmpresaProprietaria"),
            DimEmpresaProprietaria.CNPJ.label("CNPJEmpresaProprietaria"),
            DimEmpresaProprietaria.CNAE.label("CNAEEmpresaProprietaria"),
            DimEmpresaProprietaria.Logo.label("LogoEmpresaProprietaria"),
            DimEmpresaProprietaria.DescricaoCnae.label("DescricaoCnaeEmpresaProprietaria"),
            DimEmpresaProprietaria.BitAtivo.label("BitAtivoEmpresaProprietaria"),
        )
        .select_from(DimEmpresas)
        .outerjoin(DimCnaes, DimCnaes.cnaepadrao == DimEmpresas.CNAE)
        .outerjoin(
            DimEmpresaProprietaria,
            DimEmpresaProprietaria.IDEmpresaProprietaria == DimEmpresas.IDEmpresaProprietaria,
        )
        .filter(DimEmpresas.IDEmpresa == id_empresa)
        .first()
    )

    if not row:
        abort(404)

    empresa = dict(row._asdict())

    empresa["MatrizFilial"] = _matriz_filial_label(empresa.get("MatrizFilial"))

    cnpj_bd = _cnpj_somente_digitos(empresa.get("CNPJ") or "")
    cnpj14 = (cnpj_bd[-14:]).rjust(14, "0") if cnpj_bd else ""
    cnpj14_formatado = _cnpj_14(cnpj14)

    logo_path = (empresa.get("LogoEmpresaProprietaria") or "").strip()
    if logo_path:
        nome_arquivo = os.path.basename(logo_path.replace("\\", "/"))
        empresa["LogoEmpresaProprietariaUrl"] = url_for(
            "static",
            filename=f"LogoEmpresaProprietaria/{nome_arquivo}",
        )
    else:
        empresa["LogoEmpresaProprietariaUrl"] = ""

    classificacao = {}
    try:
        sql_classificacao_por_id = text("""
            SELECT TOP (1)
                c.*,
                pa.TipoUsoTerritorio AS TipoUsoTerritorio,
                pa.NomePerfil       AS NomePerfilPublicoAlvo,
                pa.FaixaEconomica   AS FaixaEconomicaPublicoAlvo,
                pa.TipoDemanda      AS TipoDemandaPublicoAlvo,
                pa.Descricao        AS DescricaoPublicoAlvo
            FROM Integracao.Silver.DimClassificacacaoClientes c
            LEFT JOIN Integracao.Silver.DimPublicoAlvo pa
                ON pa.IDDimPublicoAlvo = c.IDDimPublicoAlvo
            WHERE c.IDEmpresa = :id_empresa
            ORDER BY c.ReceitaTotal DESC
        """)

        classificacao_row = (
            db.session.execute(sql_classificacao_por_id, {"id_empresa": id_empresa})
            .mappings()
            .first()
        )

        if not classificacao_row and cnpj14:
            sql_classificacao_por_cnpj = text("""
                SELECT TOP (1)
                    c.*,
                    pa.TipoUsoTerritorio AS TipoUsoTerritorio,
                    pa.NomePerfil       AS NomePerfilPublicoAlvo,
                    pa.FaixaEconomica   AS FaixaEconomicaPublicoAlvo,
                    pa.TipoDemanda      AS TipoDemandaPublicoAlvo,
                    pa.Descricao        AS DescricaoPublicoAlvo
                FROM Integracao.Silver.DimClassificacacaoClientes c
                LEFT JOIN Integracao.Silver.DimPublicoAlvo pa
                    ON pa.IDDimPublicoAlvo = c.IDDimPublicoAlvo
                WHERE RIGHT(
                        '00000000000000' + REPLACE(REPLACE(REPLACE(LTRIM(RTRIM(c.CNPJ)),'/',''),'.',''),'-',''),
                        14
                      ) = :cnpj14
                ORDER BY c.ReceitaTotal DESC
            """)
            classificacao_row = (
                db.session.execute(sql_classificacao_por_cnpj, {"cnpj14": cnpj14})
                .mappings()
                .first()
            )

        classificacao = dict(classificacao_row) if classificacao_row else {}

        if classificacao:
            if "BitCompraOutrasEmpresas" not in classificacao and "BitComprasOutrasEmpresas" in classificacao:
                classificacao["BitCompraOutrasEmpresas"] = classificacao.get("BitComprasOutrasEmpresas")

            if "BitComprasOutrasEmpresas" not in classificacao and "BitCompraOutrasEmpresas" in classificacao:
                classificacao["BitComprasOutrasEmpresas"] = classificacao.get("BitCompraOutrasEmpresas")

            tipo_uso = (classificacao.get("TipoUsoTerritorio") or "").strip()
            perfil_atual = (classificacao.get("PerfilPublicoAlvo") or "").strip()

            if (not perfil_atual) and tipo_uso:
                classificacao["PerfilPublicoAlvo"] = tipo_uso

            if not (classificacao.get("PerfilPublicoAlvo") or "").strip():
                nome_perfil = (classificacao.get("NomePerfilPublicoAlvo") or "").strip()
                if nome_perfil:
                    classificacao["PerfilPublicoAlvo"] = nome_perfil

            for k in (
                "ReceitaTotal",
                "PercReceitaAcumulada",
                "ValorUltimaCompra",
                "Receita12M",
                "ScoreImportancia",
                "ScoreRetornoCluster",
                "ScoreRetornoTecnico",
                "ScorePerfilEmpresa",
            ):
                if k in classificacao and classificacao[k] is not None:
                    try:
                        classificacao[k] = float(classificacao[k])
                    except Exception:
                        pass

            for k in ("DiasDesdeUltimaCompra", "TotalItens", "ItensMesmaCidade"):
                if k in classificacao and classificacao[k] is not None:
                    try:
                        classificacao[k] = int(classificacao[k])
                    except Exception:
                        pass

    except Exception:
        classificacao = {}

    def _to_float_ou_none(valor):
        try:
            if valor is None:
                return None
            if isinstance(valor, str) and not valor.strip():
                return None
            return float(valor)
        except Exception:
            return None

    score_setor = _to_float_ou_none(empresa.get("ScoreSetor"))
    classificacao_setor = (
        (empresa.get("ClassificacaoSetor") or empresa.get("ClassificacaoMacro") or "").strip()
    )
    grupo_empresa = (classificacao.get("ClusterGrupoCliente") or "").strip()
    score_perfil = _to_float_ou_none(classificacao.get("ScorePerfilEmpresa"))
    classificacao_perfil = (classificacao.get("ClassificacaoPerfilEmpresa") or "").strip()

    segmentacao = {
        "ScoreSetor": score_setor,
        "ClassificacaoSetor": classificacao_setor,
        "ClassificacaoMacro": classificacao_setor,
        "GrupoEmpresa": grupo_empresa,
        "ClusterGrupoCliente": grupo_empresa,
        "ScorePerfil": score_perfil,
        "ScorePerfilEmpresa": score_perfil,
        "ClassificacaoPerfil": classificacao_perfil,
        "ClassificacaoPerfilEmpresa": classificacao_perfil,
    }

    classificacao["ScoreSetor"] = segmentacao["ScoreSetor"]
    classificacao["ClassificacaoSetor"] = segmentacao["ClassificacaoSetor"]
    classificacao["ClassificacaoMacro"] = segmentacao["ClassificacaoMacro"]
    classificacao["GrupoEmpresa"] = segmentacao["GrupoEmpresa"]
    classificacao["ClusterGrupoCliente"] = segmentacao["ClusterGrupoCliente"]
    classificacao["ScorePerfil"] = segmentacao["ScorePerfil"]
    classificacao["ScorePerfilEmpresa"] = segmentacao["ScorePerfilEmpresa"]
    classificacao["ClassificacaoPerfil"] = segmentacao["ClassificacaoPerfil"]
    classificacao["ClassificacaoPerfilEmpresa"] = segmentacao["ClassificacaoPerfilEmpresa"]

    valor_total_contratos_cliente = 0
    try:
        sql_total_contratos_cliente = text("""
            SELECT
                COALESCE(SUM(TRY_CONVERT(DECIMAL(18,2), TotalLiquidoContratoAGBRCTACORDO)), 0) AS ValorTotal
            FROM Integracao.Silver.FatoControleContratosItensEuromidia
            WHERE RIGHT('00000000000000' + REPLACE(REPLACE(REPLACE(LTRIM(RTRIM(CNPJ)),'/',''),'.',''),'-',''), 14) = :cnpj14
        """)
        tot_row = (
            db.session.execute(sql_total_contratos_cliente, {"cnpj14": cnpj14})
            .mappings()
            .first()
        )
        valor_total_contratos_cliente = float((tot_row or {}).get("ValorTotal") or 0)
    except Exception:
        valor_total_contratos_cliente = 0

    try:
        contratos_page = int((request.args.get("contratos_page") or "1").strip())
    except Exception:
        contratos_page = 1

    if contratos_page < 1:
        contratos_page = 1

    contratos_per_page = 6
    contratos_offset = (contratos_page - 1) * contratos_per_page

    contratos_cab = []
    contratos_cab_total = 0
    contratos_cab_total_pages = 1

    try:
        sql_contratos_cab_total = text("""
            ;WITH base AS (
                SELECT
                    IDFatoControleContratosEuromidia,
                    DataLancamento,
                    NumeroContrato,
                    NumeroPrevia,
                    RazaoSocial,
                    CNPJ,
                    CPF,
                    MarcaExibida,
                    Vendedor,
                    TipoDocumento,
                    Origem,
                    SDR,
                    Agencia,
                    CnpjAgencia,
                    Bureau,
                    CnpjBureau,
                    Intermediario,
                    CnpjIntermediario,
                    QuantidadePontos,
                    QuantidadeFaces,
                    TotalFaturamentoBrutoMensal,
                    TotalPercentualPermuta,
                    TotalCotaOportunidade,
                    TotalValorPermuta,
                    TotalFaturamentoLiquidoPermuta,
                    TotalBrutoContrato,
                    TotalLiquidoContratoAGBRCTACORDO,
                    TotalLiquidoContratoAGBRVENDGERCOOR,
                    TotalPercentualAgencia,
                    TotalValorMensalAgencia,
                    TotalPercentualBureau,
                    TotalValorBureauMensal,
                    TotalPercentualCartaAcordo,
                    TotalValorCartaAcordoMensal,
                    TotalValorOutrasComissoes,
                    TotalFaturamentoLiquidoMensal,
                    TotalPercentualComissaoVendedor,
                    TotalValorVendedor,
                    ValorVendedorTotal,
                    TotalPercentualComissaoCoordenacao,
                    IDEmpresa,
                    IDCategoriaMarca,
                    Referencia,
                    ROW_NUMBER() OVER (
                        PARTITION BY IDFatoControleContratosEuromidia
                        ORDER BY DataLancamento DESC
                    ) AS rn
                FROM Integracao.Silver.FatoControleContratosEuromidia
                WHERE RIGHT('00000000000000' + REPLACE(REPLACE(REPLACE(LTRIM(RTRIM(CNPJ)),'/',''),'.',''),'-',''), 14) = :cnpj14
            )
            SELECT COUNT(1) AS Total
            FROM base
            WHERE rn = 1
        """)
        total_row = (
            db.session.execute(sql_contratos_cab_total, {"cnpj14": cnpj14})
            .mappings()
            .first()
        )
        contratos_cab_total = int((total_row or {}).get("Total") or 0)

        if contratos_cab_total <= 0:
            contratos_cab_total_pages = 1
        else:
            contratos_cab_total_pages = (contratos_cab_total + contratos_per_page - 1) // contratos_per_page

        if contratos_page > contratos_cab_total_pages:
            contratos_page = contratos_cab_total_pages
            if contratos_page < 1:
                contratos_page = 1
            contratos_offset = (contratos_page - 1) * contratos_per_page

        sql_contratos_cab_page = text("""
            ;WITH base AS (
                SELECT
                    IDFatoControleContratosEuromidia,
                    DataLancamento,
                    NumeroContrato,
                    NumeroPrevia,
                    RazaoSocial,
                    CNPJ,
                    CPF,
                    MarcaExibida,
                    Vendedor,
                    TipoDocumento,
                    Origem,
                    SDR,
                    Agencia,
                    CnpjAgencia,
                    Bureau,
                    CnpjBureau,
                    Intermediario,
                    CnpjIntermediario,
                    QuantidadePontos,
                    QuantidadeFaces,
                    TotalFaturamentoBrutoMensal,
                    TotalPercentualPermuta,
                    TotalCotaOportunidade,
                    TotalValorPermuta,
                    TotalFaturamentoLiquidoPermuta,
                    TotalBrutoContrato,
                    TotalLiquidoContratoAGBRCTACORDO,
                    TotalLiquidoContratoAGBRVENDGERCOOR,
                    TotalPercentualAgencia,
                    TotalValorMensalAgencia,
                    TotalPercentualBureau,
                    TotalValorBureauMensal,
                    TotalPercentualCartaAcordo,
                    TotalValorCartaAcordoMensal,
                    TotalValorOutrasComissoes,
                    TotalFaturamentoLiquidoMensal,
                    TotalPercentualComissaoVendedor,
                    TotalValorVendedor,
                    ValorVendedorTotal,
                    TotalPercentualComissaoCoordenacao,
                    IDEmpresa,
                    IDCategoriaMarca,
                    Referencia,
                    ROW_NUMBER() OVER (
                        PARTITION BY IDFatoControleContratosEuromidia
                        ORDER BY DataLancamento DESC
                    ) AS rn
                FROM Integracao.Silver.FatoControleContratosEuromidia
                WHERE RIGHT('00000000000000' + REPLACE(REPLACE(REPLACE(LTRIM(RTRIM(CNPJ)),'/',''),'.',''),'-',''), 14) = :cnpj14
            )
            SELECT *
            FROM base
            WHERE rn = 1
            ORDER BY DataLancamento DESC
            OFFSET :off ROWS FETCH NEXT :take ROWS ONLY
        """)
        contratos_cab = list(
            db.session.execute(
                sql_contratos_cab_page,
                {"cnpj14": cnpj14, "off": contratos_offset, "take": contratos_per_page},
            )
            .mappings()
            .all()
        )
        contratos_cab = [dict(r) for r in contratos_cab]

    except Exception:
        contratos_cab = []
        contratos_cab_total = 0
        contratos_cab_total_pages = 1

    contratos_itens = []
    contratos_itens_por_contrato = {}
    try:
        from sqlalchemy import bindparam

        ids_cab = [
            r.get("IDFatoControleContratosEuromidia")
            for r in contratos_cab
            if r.get("IDFatoControleContratosEuromidia") is not None
        ]
        ids_cab = list(dict.fromkeys(ids_cab))

        if ids_cab:
            sql_contratos_itens_page = text("""
                SELECT
                    IDFatoControleContratosItensEuromidia,
                    IDFatoControleContratoEuromidia,
                    DataAtualizacao,
                    Referencia,
                    NumeroContrato,
                    NumeroPrevia,
                    CNPJ,
                    CodPonto,
                    CodFace,
                    DataLancamento,
                    Cota,
                    CidadeExibicao,
                    Tipo,
                    Origem,
                    EmpresaEuro,
                    CnpjExibibora,
                    TipoDocumento,
                    RazaoSocial,
                    CPF,
                    MarcaExibida,
                    Vendedor,
                    SDR,
                    Agencia,
                    CnpjAgencia,
                    Bureau,
                    CnpjBureau,
                    Intermediario,
                    CnpjIntermediario,
                    DataAssinaturaRenovacao,
                    IDTrimestre,
                    TexmpoExposicao,
                    DataInicioPrevisto,
                    DataTerminoPrevisto,
                    InicioRenovacao,
                    FaturamentoBrutoMensal,
                    PercentualPermuta,
                    CotaOportunidade,
                    ValorPermuta,
                    FaturamentoLiquidoPermuta,
                    NumeroParcelas,
                    DataInicioVencimento,
                    TotalBrutoContrato,
                    TotalLiquidoContratoAGBRCTACORDO,
                    TotalLiquidoContratoAGBRVENDGERCOOR,
                    PercentualAgencia,
                    ValorMensalAgencia,
                    PercentualBureau,
                    ValorBureauMensal,
                    PercentualCartaAcordo,
                    ValorCartaAcordoMensal,
                    ValorOutrasComissoes,
                    FaturamentoLiquidoMensal,
                    PercentualComissaoVendedor,
                    ValorVendedor,
                    ValorVendedorTotal,
                    PercentualComissaoCoordenacao,
                    ValorCoordenador,
                    ValorCoordenadorTotal,
                    PercentualComissaoGerencia,
                    ValorGerencia,
                    ValorGerenciaTotal,
                    AtivoCancelamento,
                    FaturamentoLiquidoFinalMensal,
                    ComissaoGerenciaNordeste,
                    Faturamento,
                    DataCancelamento,
                    OBS,
                    IDVendedor,
                    IDPainelEuromidia,
                    IDDimFacesPaineis,
                    DataFimEfetiva,
                    Status,
                    CASE
                        WHEN UPPER(LTRIM(RTRIM(COALESCE(AtivoCancelamento,'')))) <> 'A'
                            THEN 'CANCELADO'
                        WHEN TRY_CONVERT(date, DataTerminoPrevisto) IS NOT NULL
                             AND TRY_CONVERT(date, DataTerminoPrevisto) < CAST(GETDATE() AS date)
                            THEN 'CONCLUIDO'
                        ELSE 'ATIVO'
                    END AS StatusCalculado
                FROM Integracao.Silver.FatoControleContratosItensEuromidia
                WHERE IDFatoControleContratoEuromidia IN :ids
                ORDER BY IDFatoControleContratoEuromidia ASC, DataLancamento DESC, DataInicioPrevisto DESC
            """).bindparams(bindparam("ids", expanding=True))

            contratos_itens = list(
                db.session.execute(sql_contratos_itens_page, {"ids": ids_cab})
                .mappings()
                .all()
            )
            contratos_itens = [dict(r) for r in contratos_itens]

            contratos_itens_por_contrato = {}
            for it in contratos_itens:
                k = it.get("IDFatoControleContratoEuromidia")
                contratos_itens_por_contrato.setdefault(k, []).append(it)

        else:
            contratos_itens = []
            contratos_itens_por_contrato = {}

    except Exception:
        contratos_itens = []
        contratos_itens_por_contrato = {}

    contratos_paginacao = {
        "page": contratos_page,
        "per_page": contratos_per_page,
        "total": contratos_cab_total,
        "total_pages": contratos_cab_total_pages,
        "inicio": (contratos_offset + 1) if contratos_cab_total > 0 else 0,
        "fim": min(contratos_offset + contratos_per_page, contratos_cab_total) if contratos_cab_total > 0 else 0,
    }

    return render_template(
        "euromidia/empresa.html",
        empresa=empresa,
        cnpj14=cnpj14_formatado,
        classificacao=classificacao,
        segmentacao=segmentacao,
        valor_total_contratos_cliente=valor_total_contratos_cliente,
        contratos_cab=contratos_cab,
        contratos_itens=contratos_itens,
        contratos_itens_por_contrato=contratos_itens_por_contrato,
        contratos_paginacao=contratos_paginacao,
    )






@paineis_bp.route("/exportar_excel", methods=["GET"])
@limiter.limit("3 per minute", methods=["GET"])
@login_required
@retry_get_view(db, attempts=6, base_delay=0.2, max_delay=1.5)
def exportar_excel_ano():
    try:
        ano = int((request.args.get("ano") or "").strip())
    except Exception:
        ano = date.today().year

    if ano < 2000 or ano > 2100:
        ano = date.today().year

    dt_ini_ano = date(ano, 1, 1)
    dt_fim_ano = date(ano, 12, 31)

    codpontos = (
        db.session.query(DimFacesPaineis.CodPonto)
        .filter(DimFacesPaineis.CodPonto != None)
        .distinct()
        .order_by(func.cast(DimFacesPaineis.CodPonto, db.Integer).asc())
        .all()
    )
    codpontos = [r[0] for r in codpontos if r and r[0] is not None]

    wb = Workbook()
    wb.remove(wb.active)

    estilos = _excel_estilos_basicos()

    for codponto in codpontos:
        ws = wb.create_sheet(title=_excel_nome_aba(codponto))

       
        faces_info_raw = (
            db.session.query(
                DimFacesPaineis.IDDimPaineisEuromidia,
                DimFacesPaineis.CodFace,
                DimFacesPaineis.TipoPainel,
                DimFacesPaineis.Exibidora,
                DimFacesPaineis.QuantidadeFaces,
            )
            .filter(
                DimFacesPaineis.CodPonto == codponto,
                DimFacesPaineis.CodFace != None,
                DimFacesPaineis.CodFace != "",
            )
            .all()
        )

        tipos_distintos = set()
        for idpainel, cf, tp, ex, qf in faces_info_raw:
            tp_up = (tp or "").strip().upper()
            if tp_up:
                tipos_distintos.add(tp_up)

        tipo_filtro_up = ""
        if len(tipos_distintos) > 1:
            if "PAINEL DIGITAL" in tipos_distintos:
                tipo_filtro_up = "PAINEL DIGITAL"
            else:
                tipo_filtro_up = sorted(list(tipos_distintos))[0]
        elif len(tipos_distintos) == 1:
            tipo_filtro_up = list(tipos_distintos)[0]
        else:
            tipo_filtro_up = ""

        faces_info = []
        if tipo_filtro_up:
            for idpainel, cf, tp, ex, qf in faces_info_raw:
                tp_up = (tp or "").strip().upper()
                if tp_up == tipo_filtro_up:
                    faces_info.append((idpainel, cf, tp, ex, qf))
        else:
            faces_info = list(faces_info_raw)

        faces = []
        exibidora = ""
        idpainel_set = set()

        capacidades_digitais = []
        capacidades_nao_digitais = []
        tipo_prod = ""

        for idpainel, cf, tp, ex, qf in faces_info:
            cf_norm = (str(cf) or "").strip()
            if not cf_norm:
                continue
            faces.append(cf_norm)

            if not exibidora:
                exibidora = (ex or "")

            if idpainel not in (None, ""):
                try:
                    idpainel_set.add(int(idpainel))
                except:
                    pass

            tp_txt = (tp or "").strip()
            tp_up = tp_txt.upper()
            if not tipo_prod and tp_txt:
                tipo_prod = tp_txt

            qf_int = int(qf) if qf not in (None, "") else 0
            if tp_up == "PAINEL DIGITAL":
                if qf_int > 0:
                    capacidades_digitais.append(qf_int)
            else:
                if qf_int > 0:
                    capacidades_nao_digitais.append(qf_int)

        faces = [f for f in faces if f]
        faces = sorted(list(dict.fromkeys(faces)))

        eh_digital = (tipo_filtro_up.strip().upper() == "PAINEL DIGITAL") if tipo_filtro_up else False

        if eh_digital:
            num_faces = sum(capacidades_digitais) if len(capacidades_digitais) > 0 else 0
            tipo_prod = "PAINEL DIGITAL"
            if num_faces <= 0:
                num_faces = 0
        else:
            num_faces = max(capacidades_nao_digitais) if len(capacidades_nao_digitais) > 0 else 0
           
            if not tipo_prod and tipo_filtro_up:
                tipo_prod = tipo_filtro_up

       
        q_oc = (
            db.session.query(
                FatoOcupacaoPaineisEuromidia.CodFace,
                FatoOcupacaoPaineisEuromidia.MarcaExibida,
                FatoOcupacaoPaineisEuromidia.Loop,
                FatoOcupacaoPaineisEuromidia.Vendedor,
                FatoOcupacaoPaineisEuromidia.DataInicio,
                FatoOcupacaoPaineisEuromidia.DataFim,
                FatoOcupacaoPaineisEuromidia.TextoOriginal,
                FatoOcupacaoPaineisEuromidia.IDDimPaineisEuromidia,
            )
            .filter(
                FatoOcupacaoPaineisEuromidia.CodPonto == codponto,
                FatoOcupacaoPaineisEuromidia.DataInicio <= dt_fim_ano,
                FatoOcupacaoPaineisEuromidia.DataFim >= dt_ini_ano,
            )
            .order_by(
                FatoOcupacaoPaineisEuromidia.CodFace.asc(),
                FatoOcupacaoPaineisEuromidia.DataInicio.asc(),
            )
        )

        if idpainel_set:
            q_oc = q_oc.filter(
                or_(
                    FatoOcupacaoPaineisEuromidia.IDDimPaineisEuromidia.in_(list(idpainel_set)),
                    and_(
                        FatoOcupacaoPaineisEuromidia.IDDimPaineisEuromidia.is_(None),
                        FatoOcupacaoPaineisEuromidia.CodFace.in_(faces) if faces else True,
                    ),
                )
            )
        else:
            if faces:
                q_oc = q_oc.filter(FatoOcupacaoPaineisEuromidia.CodFace.in_(faces))

        rows_oc = q_oc.all()

        ocupacoes_por_face = {f: [] for f in faces}
        for r in rows_oc:
            cf = (r.CodFace or "").strip()
            if not cf:
                continue
            if cf not in ocupacoes_por_face:
                ocupacoes_por_face[cf] = []

            ocupacoes_por_face[cf].append(
                {
                    "CodFace": cf,
                    "MarcaExibida": r.MarcaExibida or "",
                    "Loop": r.Loop or "",
                    "Vendedor": r.Vendedor or "",
                    "DataInicio": r.DataInicio,
                    "DataFim": r.DataFim,
                    "TextoOriginal": r.TextoOriginal or "",
                }
            )

        _marcar_conflitos_por_face(ocupacoes_por_face)

        _excel_montar_aba_grade_ano(
            ws=ws,
            estilos=estilos,
            ano=ano,
            codponto=codponto,
            tipo_prod=tipo_prod,
            exibidora=exibidora,
            num_faces=num_faces,
            faces=faces,
            ocupacoes_por_face=ocupacoes_por_face,
        )

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    nome = f"grade_paineis_{ano}.xlsx"
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=nome,
    )



def _excel_nome_aba(codponto):
    """
    Excel tem limite de 31 chars e não pode: : \ / ? * [ ]
    """
    s = str(codponto)
    for ch in [":", "\\", "/", "?", "*", "[", "]"]:
        s = s.replace(ch, "-")
    s = s.strip()
    if not s:
        s = "Painel"
    if len(s) > 31:
        s = s[:31]
    return s


def _excel_estilos_basicos():
    """
    Mantém os imports AQUI dentro, para o helper não depender do endpoint.
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "border": border,
        "font_titulo": Font(bold=True, size=13, color="002884"),
        "font_sub": Font(bold=True, size=10, color="64748B"),
        "font_face": Font(bold=True, size=10, color="002884"),
        "font_cab": Font(bold=True, size=10, color="0F172A"),
        "font_item": Font(bold=True, size=9, color="0F172A"),
        "al_center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "al_left": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "fill_mes": PatternFill("solid", fgColor="FFF3C4"),
        "fill_cab": PatternFill("solid", fgColor="F8FAFC"),
        "fill_fds": PatternFill("solid", fgColor="EEF2F7"),
        "fill_barra": PatternFill("solid", fgColor="FFB300"),
        "fill_barra2": PatternFill("solid", fgColor="FFD766"),
        "fill_conflito": PatternFill("solid", fgColor="FCA5A5"),
    }


def _intersecao_mes(di, df, dt_ini_mes, dt_fim_mes):
    """
    Recebe:
      di, df: datas da ocupação
      dt_ini_mes, dt_fim_mes: limite do mês
    Retorna:
      (dia_ini, dia_fim) dentro do mês (inteiros 1..ultimo_dia)
      ou (None, None) se não intersecta.
    """
    if di is None or df is None:
        return (None, None)

    ini = di if di >= dt_ini_mes else dt_ini_mes
    fim = df if df <= dt_fim_mes else dt_fim_mes

    if fim < dt_ini_mes or ini > dt_fim_mes:
        return (None, None)
    if fim < ini:
        return (None, None)

    return (ini.day, fim.day)




def _excel_montar_aba_grade_ano(
    ws,
    estilos,
    ano: int,
    codponto,
    tipo_prod: str,
    exibidora: str,
    num_faces: int,
    faces: list[str],
    ocupacoes_por_face: dict,
):
    """
    Layout:
    - Linha 1: Título
    - Linha 2: Subtítulo (tipo/exibidora/numfaces)
    - A partir da linha 4: blocos de meses na horizontal
      Cada bloco: colunas [Face] + dias 1..ultimo_dia
      Linhas: 1 por face (com "barras" pintadas)
    """
    import calendar
    from datetime import date
    from openpyxl.utils import get_column_letter


    ws["A1"] = f"Grade Anual - CodPonto {codponto} ({ano})"
    ws["A1"].font = estilos["font_titulo"]
    ws["A1"].alignment = estilos["al_left"]

    ws["A2"] = f"Tipo: {tipo_prod}   •   Exibidora: {exibidora}   •   Nº Faces: {num_faces}"
    ws["A2"].font = estilos["font_sub"]
    ws["A2"].alignment = estilos["al_left"]

    ws.freeze_panes = "A5"


    linha_topo = 4
    altura_face = 1
    col_inicio = 1
    largura_col_face = 12
    largura_col_dia = 3.2
    gap_entre_meses = 2


    if not faces:
        faces = ["(sem faces)"]
        ocupacoes_por_face = {"(sem faces)": []}


    col = col_inicio
    for mes in range(1, 13):
        _, ultimo_dia = calendar.monthrange(ano, mes)
        dt_ini_mes = date(ano, mes, 1)
        dt_fim_mes = date(ano, mes, ultimo_dia)

        col_face = col
        col_d_last = col + ultimo_dia


        ws.merge_cells(
            start_row=linha_topo,
            start_column=col_face,
            end_row=linha_topo,
            end_column=col_d_last,
        )
        cel = ws.cell(row=linha_topo, column=col_face)
        cel.value = f"{calendar.month_abbr[mes].upper()}/{str(ano)[2:]}"
        cel.font = estilos["font_cab"]
        cel.alignment = estilos["al_center"]
        cel.fill = estilos["fill_mes"]


        cface = ws.cell(row=linha_topo + 1, column=col_face)
        cface.value = "Face"
        cface.font = estilos["font_cab"]
        cface.alignment = estilos["al_center"]
        cface.fill = estilos["fill_cab"]


        for d in range(1, ultimo_dia + 1):
            c = ws.cell(row=linha_topo + 1, column=col + d)
            c.value = d
            c.font = estilos["font_cab"]
            c.alignment = estilos["al_center"]

            dt = date(ano, mes, d)
            c.fill = estilos["fill_fds"] if dt.weekday() >= 5 else estilos["fill_cab"]


        ws.column_dimensions[get_column_letter(col_face)].width = largura_col_face
        for d in range(1, ultimo_dia + 1):
            ws.column_dimensions[get_column_letter(col + d)].width = largura_col_dia


        for idx, face in enumerate(faces):
            linha = (linha_topo + 2) + idx * altura_face


            c = ws.cell(row=linha, column=col_face)
            c.value = face
            c.font = estilos["font_face"]
            c.alignment = estilos["al_left"]
            c.fill = estilos["fill_cab"]


            for d in range(1, ultimo_dia + 1):
                cday = ws.cell(row=linha, column=col + d)
                cday.alignment = estilos["al_center"]
                dt = date(ano, mes, d)
                if dt.weekday() >= 5:
                    cday.fill = estilos["fill_fds"]


            itens = ocupacoes_por_face.get(face, [])
            if itens:
                cor_toggle = False

                for it in itens:
                    di = it.get("DataInicio")
                    df = it.get("DataFim")

                    dia_ini, dia_fim = _intersecao_mes(di, df, dt_ini_mes, dt_fim_mes)
                    if dia_ini is None:
                        continue

                    fill_barra = estilos["fill_barra2"] if cor_toggle else estilos["fill_barra"]
                    cor_toggle = not cor_toggle


                    for dd in range(dia_ini, dia_fim + 1):
                        cc = ws.cell(row=linha, column=col + dd)
                        cc.fill = estilos["fill_conflito"] if it.get("BitConflito") else fill_barra


                    texto = f"{it.get('MarcaExibida','')}".strip()
                    loop = (it.get("Loop") or "").strip()
                    vend = (it.get("Vendedor") or "").strip()
                    if loop or vend:
                        texto = f"{texto} • {loop} • {vend}".strip(" •")

                    ctexto = ws.cell(row=linha, column=col + dia_ini)
                    ctexto.value = texto[:120]
                    ctexto.font = estilos["font_item"]
                    ctexto.alignment = estilos["al_left"]


        r1 = linha_topo
        r2 = (linha_topo + 1) + 1 + (len(faces))
        c1 = col_face
        c2 = col_d_last
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                ws.cell(row=rr, column=cc).border = estilos["border"]


        col = col_d_last + gap_entre_meses


    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[linha_topo].height = 18
    ws.row_dimensions[linha_topo + 1].height = 18
    for i in range(len(faces)):
        ws.row_dimensions[(linha_topo + 2) + i].height = 18





@paineis_bp.route("/painel-detalhes/<int:codponto>", methods=["GET"])
def painel_detalhes(codponto: int):
    def _texto_limpo(valor):
        try:
            return str(valor or "").strip()
        except Exception:
            return ""

    def _to_bool_sql(valor, padrao=None):
        """Eu converto valores do SQL Server para bool de forma robusta."""
        if valor is None:
            return padrao

        if isinstance(valor, bool):
            return valor

        if isinstance(valor, (int, float)):
            return bool(int(valor))

        txt = _texto_limpo(valor).lower()
        if txt in {"1", "true", "t", "sim", "s", "yes", "y"}:
            return True
        if txt in {"0", "false", "f", "nao", "não", "n", "no"}:
            return False

        return padrao

    def _to_int_seguro(valor, padrao):
        try:
            return int(str(valor).strip())
        except Exception:
            return padrao

    def _normalizar_url_imagem(url):
        """Eu normalizo a URL da imagem para o navegador."""
        txt = _texto_limpo(url)
        if not txt:
            return None

        txt_lower = txt.lower()

        if txt_lower.startswith("http://") or txt_lower.startswith("https://") or txt.startswith("//"):
            return txt

        if txt.startswith("/"):
            return txt

        if txt_lower.startswith("static/"):
            return f"/{txt}"

        return f"/{txt.lstrip('/')}"

    def _primeiro_dia_mes(iso: str, fallback: str) -> str:
        try:
            d = datetime.strptime(iso, "%Y-%m-%d")
            return f"{d.year:04d}-{d.month:02d}-01"
        except Exception:
            return fallback

    def _to_float_br(v):
        if v is None:
            return None

        if isinstance(v, (int, float)):
            return float(v)

        s = str(v).strip()
        if not s:
            return None

        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s and "." not in s:
            s = s.replace(",", ".")

        try:
            return float(s)
        except Exception:
            return None

    def _carregar_imagens_painel(cod_ponto: int) -> list[dict]:
        """
        Eu busco as imagens do painel e devolvo já ordenadas para a galeria.
        Regras:
        - uso somente imagens ativas
        - ordeno por NumeroImagem
        - se não houver imagem válida, devolvo fallback
        """
        sql_imagens = text("""
            SELECT
                TRY_CONVERT(int, IDDimImagemPainel) AS IDDimImagemPainel,
                TRY_CONVERT(int, IDDimFacesPaineis) AS IDDimFacesPaineis,
                CAST(ISNULL(UrlImagem, '') AS varchar(600)) AS UrlImagem,
                TRY_CONVERT(int, NumeroImagem) AS NumeroImagem,
                CAST(ISNULL(CodFace, '') AS varchar(60)) AS CodFace,
                CAST(ISNULL(CodPonto, '') AS varchar(60)) AS CodPonto,
                CAST(ISNULL(BitAtivo, 1) AS bit) AS BitAtivo
            FROM [Integracao].[Silver].[DimImagemPainel]
            WHERE TRY_CONVERT(int, CodPonto) = :cod_ponto
              AND ISNULL(BitAtivo, 1) = 1
              AND NULLIF(LTRIM(RTRIM(ISNULL(UrlImagem, ''))), '') IS NOT NULL
            ORDER BY
                CASE WHEN TRY_CONVERT(int, NumeroImagem) IS NULL THEN 1 ELSE 0 END,
                TRY_CONVERT(int, NumeroImagem) ASC,
                TRY_CONVERT(int, IDDimImagemPainel) ASC
        """)

        linhas_imagens = db.session.execute(
            sql_imagens,
            {"cod_ponto": cod_ponto}
        ).mappings().all()

        imagens = []
        urls_vistas = set()

        for linha in linhas_imagens:
            url_normalizada = _normalizar_url_imagem(linha.get("UrlImagem"))
            if not url_normalizada:
                continue

            url_chave = url_normalizada.lower()
            if url_chave in urls_vistas:
                continue
            urls_vistas.add(url_chave)

            numero_imagem = linha.get("NumeroImagem")
            try:
                numero_imagem = int(numero_imagem) if numero_imagem is not None else None
            except Exception:
                numero_imagem = None

            imagens.append(
                {
                    "id_imagem": linha.get("IDDimImagemPainel"),
                    "id_face_painel": linha.get("IDDimFacesPaineis"),
                    "url": url_normalizada,
                    "numero": numero_imagem,
                    "cod_face": _texto_limpo(linha.get("CodFace")),
                    "cod_ponto": _texto_limpo(linha.get("CodPonto")),
                    "eh_fallback": False,
                }
            )

        if imagens:
            return imagens

        url_fallback = url_for("static", filename="imagens/painel-publicitario.png")
        return [
            {
                "id_imagem": None,
                "id_face_painel": None,
                "url": url_fallback,
                "numero": 1,
                "cod_face": "",
                "cod_ponto": str(cod_ponto),
                "eh_fallback": True,
            }
        ]

    sql_painel_dim = text("""
        SELECT TOP 1
            TRY_CONVERT(int, CodPonto) AS CodPonto,
            TRY_CONVERT(int, QuantidadeFaces) AS QuantidadeFaces,
            CAST(ISNULL(Tipo,'') AS varchar(120)) AS Tipo,
            CAST(ISNULL(Cidade,'') AS varchar(150)) AS Municipio,
            CAST(ISNULL(UF,'') AS varchar(10)) AS UF,
            CAST(ISNULL(Logradouro,'') AS varchar(250)) AS Logradouro,
            CAST(ISNULL(Numero,'') AS varchar(50)) AS Numero,
            CAST(ISNULL(Bairro,'') AS varchar(150)) AS Bairro,
            CAST(ISNULL(CEP,'') AS varchar(20)) AS CEP,
            CAST(ISNULL(Sentido,'') AS varchar(200)) AS Sentido,
            TRY_CONVERT(float, REPLACE(CAST(Latitude AS varchar(64)), ',', '.')) AS lat,
            TRY_CONVERT(float, REPLACE(CAST(Longitude AS varchar(64)), ',', '.')) AS lng,
            CAST(ISNULL(FormatoLxA,'') AS varchar(80)) AS FormatoLxA,
            CAST(ISNULL(Exibidora,'') AS varchar(150)) AS Exibidora,
            CAST(ISNULL(BitIluminado, 0) AS bit) AS BitIluminado,
            CAST(ISNULL(BitAtivo, 1) AS bit) AS BitAtivo,
            CAST(ISNULL(BitAluguel, 0) AS bit) AS BitAluguel
        FROM [Integracao].[Silver].[DimPaineisEuromidia]
        WHERE TRY_CONVERT(int, CodPonto) = :cod_ponto
    """)
    row_painel = db.session.execute(sql_painel_dim, {"cod_ponto": codponto}).mappings().first()

    if (not row_painel) or (row_painel.get("lat") is None) or (row_painel.get("lng") is None):
        sql_painel_fallback = text("""
            SELECT TOP 1
                TRY_CONVERT(int, CodPonto) AS CodPonto,
                CAST(ISNULL(Logr_Pr,'') AS varchar(200)) AS Logradouro,
                CAST(ISNULL(Tipo_Logr_Pr,'') AS varchar(80)) AS Tipo_Logradouro,
                CAST(ISNULL(BAIRRO,'') AS varchar(150)) AS Bairro,
                CAST(ISNULL([MUNICÍPIO],'') AS varchar(150)) AS Municipio,
                CAST(ISNULL(UF,'') AS varchar(10)) AS UF,
                CAST(ISNULL(CEP,'') AS varchar(20)) AS CEP,
                CAST(ISNULL(Sentido,'') AS varchar(200)) AS Sentido,
                CAST(ISNULL(TipoProd,'') AS varchar(120)) AS Tipo,
                CAST(ISNULL(CodFace,'') AS varchar(50)) AS CodFace,
                CAST(ISNULL(Iluminado,'') AS varchar(50)) AS Iluminado,
                TRY_CONVERT(float, REPLACE(CAST(LatD AS varchar(64)), ',', '.')) AS lat,
                TRY_CONVERT(float, REPLACE(CAST(LonD AS varchar(64)), ',', '.')) AS lng
            FROM [DataMining].[dbo].[CadastroPaineisEuromidia]
            WHERE TRY_CONVERT(int, CodPonto) = :cod_ponto
        """)
        row_fb = db.session.execute(sql_painel_fallback, {"cod_ponto": codponto}).mappings().first()

        if not row_fb or row_fb.get("lat") is None or row_fb.get("lng") is None:
            abort(404, description=f"Painel {codponto} não encontrado ou sem Latitude/Longitude válidos.")

        painel = {
            "id_painel": int(row_fb["CodPonto"]),
            "nome": f"Painel {int(row_fb['CodPonto'])}",
            "endereco": (
                f"{row_fb.get('Tipo_Logradouro','')} {row_fb.get('Logradouro','')}, "
                f"{row_fb.get('Bairro','')}, {row_fb.get('Municipio','')}-{row_fb.get('UF','')}"
            ).strip().strip(","),
            "formato": row_fb.get("Tipo") or "",
            "status": "Disponível",
            "lat": float(row_fb["lat"]),
            "lng": float(row_fb["lng"]),
            "url_ficha": f"/admin/paineis/{int(row_fb['CodPonto'])}",
            "uf": _texto_limpo(row_fb.get("UF")),
            "municipio": _texto_limpo(row_fb.get("Municipio")),
            "bairro": _texto_limpo(row_fb.get("Bairro")),
            "cep": _texto_limpo(row_fb.get("CEP")),
            "sentido": _texto_limpo(row_fb.get("Sentido")),
            "tipo": _texto_limpo(row_fb.get("Tipo")),
            "quantidade_faces": None,
            "exibidora": None,
            "bit_iluminado": _to_bool_sql(row_fb.get("Iluminado"), None),
            "bit_ativo": None,
            "bit_aluguel": None,
            "formato_lxa": None,
        }
    else:
        endereco_painel = (
            f"{row_painel.get('Logradouro','')}"
            + (f", {row_painel.get('Numero','')}" if _texto_limpo(row_painel.get("Numero")) else "")
            + (f" • {row_painel.get('Bairro','')}" if _texto_limpo(row_painel.get("Bairro")) else "")
            + (f" • {row_painel.get('Municipio','')}-{row_painel.get('UF','')}" if _texto_limpo(row_painel.get("Municipio")) else "")
            + (f" • CEP {row_painel.get('CEP','')}" if _texto_limpo(row_painel.get("CEP")) else "")
        ).strip()

        bit_ativo = _to_bool_sql(row_painel.get("BitAtivo"), True)
        bit_iluminado = _to_bool_sql(row_painel.get("BitIluminado"), False)
        bit_aluguel = _to_bool_sql(row_painel.get("BitAluguel"), False)

        painel = {
            "id_painel": int(row_painel["CodPonto"]),
            "nome": f"Painel {int(row_painel['CodPonto'])}",
            "endereco": endereco_painel,
            "formato": row_painel.get("Tipo") or "",
            "status": "Disponível" if bit_ativo else "Inativo",
            "lat": float(row_painel["lat"]),
            "lng": float(row_painel["lng"]),
            "url_ficha": f"/admin/paineis/{int(row_painel['CodPonto'])}",
            "uf": _texto_limpo(row_painel.get("UF")),
            "municipio": _texto_limpo(row_painel.get("Municipio")),
            "bairro": _texto_limpo(row_painel.get("Bairro")),
            "cep": _texto_limpo(row_painel.get("CEP")),
            "sentido": _texto_limpo(row_painel.get("Sentido")),
            "tipo": _texto_limpo(row_painel.get("Tipo")),
            "quantidade_faces": int(row_painel["QuantidadeFaces"]) if row_painel.get("QuantidadeFaces") is not None else None,
            "exibidora": _texto_limpo(row_painel.get("Exibidora")),
            "bit_iluminado": bit_iluminado,
            "bit_ativo": bit_ativo,
            "bit_aluguel": bit_aluguel,
            "formato_lxa": _texto_limpo(row_painel.get("FormatoLxA")),
        }

    imagens_painel = _carregar_imagens_painel(int(painel["id_painel"]))

    painel["imagens"] = imagens_painel
    painel["imagem_principal"] = imagens_painel[0]["url"] if imagens_painel else url_for(
        "static",
        filename="imagens/painel-publicitario.png"
    )
    painel["quantidade_imagens"] = len(imagens_painel)
    painel["tem_imagem_real"] = any(not img.get("eh_fallback") for img in imagens_painel)

    painel["produto"] = painel.get("tipo") or ""
    painel["faces"] = painel.get("quantidade_faces")

    try:
        dt_ini_str = request.args.get("dt_ini") or request.args.get("dtIni") or request.args.get("data_ini") or ""
        dt_fim_str = request.args.get("dt_fim") or request.args.get("dtFim") or request.args.get("data_fim") or ""

        if not dt_ini_str:
            dt_ini_str = "2024-01-01"
        if not dt_fim_str:
            dt_fim_str = "2026-12-01"

        dt_ini_mes = _primeiro_dia_mes(dt_ini_str, "2024-01-01")
        dt_fim_mes = _primeiro_dia_mes(dt_fim_str, "2026-12-01")

        sql_financeiro = text("""
            ;WITH BaseItens AS (
                SELECT
                    TRY_CONVERT(int, i.CodPonto) AS CodPonto,
                    DATEFROMPARTS(
                        YEAR(COALESCE(TRY_CONVERT(date, i.Referencia), TRY_CONVERT(date, i.DataLancamento))),
                        MONTH(COALESCE(TRY_CONVERT(date, i.Referencia), TRY_CONVERT(date, i.DataLancamento))),
                        1
                    ) AS DataRef,
                    TRY_CONVERT(float, i.FaturamentoLiquidoFinalMensal) AS ReceitaLiquidaMensal
                FROM [Integracao].[Silver].[FatoControleContratosItensEuromidia] i
                WHERE TRY_CONVERT(int, i.CodPonto) = :cod_ponto
                  AND COALESCE(TRY_CONVERT(date, i.Referencia), TRY_CONVERT(date, i.DataLancamento)) IS NOT NULL
            ),
            ReceitaMes_Full AS (
                SELECT
                    CodPonto,
                    DataRef,
                    SUM(COALESCE(ReceitaLiquidaMensal, 0.0)) AS ReceitaMes
                FROM BaseItens
                GROUP BY CodPonto, DataRef
            ),
            ReceitaMes_Periodo AS (
                SELECT
                    CodPonto,
                    DataRef,
                    SUM(COALESCE(ReceitaLiquidaMensal, 0.0)) AS ReceitaMes
                FROM BaseItens
                WHERE DataRef >= :dt_ini
                  AND DataRef <= :dt_fim
                GROUP BY CodPonto, DataRef
            )
            SELECT
                r.CodPonto,
                r.DataRef,
                YEAR(r.DataRef) AS Ano,
                MONTH(r.DataRef) AS Mes,
                r.ReceitaMes AS ReceitaMes_Full,
                rp.ReceitaMes AS ReceitaMes_Periodo,
                ca.ValorMensal AS CustoMensal,
                ca.Ano AS AnoCusto,
                ca.Mes AS MesCusto
            FROM ReceitaMes_Full r
            LEFT JOIN ReceitaMes_Periodo rp
                ON rp.CodPonto = r.CodPonto
               AND rp.DataRef  = r.DataRef
            OUTER APPLY (
                SELECT TOP 1
                    TRY_CONVERT(int, c.Ano) AS Ano,
                    TRY_CONVERT(int, c.Mes) AS Mes,
                    TRY_CONVERT(float, REPLACE(CAST(c.ValorMensal AS varchar(64)), ',', '.')) AS ValorMensal
                FROM [Integracao].[Silver].[DimCustoMensalPainel] c
                WHERE TRY_CONVERT(int, c.CodPonto) = r.CodPonto
                  AND (
                        (TRY_CONVERT(int, c.Ano) * 100 + TRY_CONVERT(int, c.Mes))
                        <=
                        (YEAR(r.DataRef) * 100 + MONTH(r.DataRef))
                  )
                ORDER BY (TRY_CONVERT(int, c.Ano) * 100 + TRY_CONVERT(int, c.Mes)) DESC
            ) ca
            ORDER BY r.DataRef ASC;
        """)

        linhas_fin = db.session.execute(
            sql_financeiro,
            {
                "cod_ponto": int(painel["id_painel"]),
                "dt_ini": dt_ini_mes,
                "dt_fim": dt_fim_mes,
            }
        ).mappings().all()

        serie_fin = []
        serie_hist = []

        for r in linhas_fin:
            data_ref_iso = (r.get("DataRef").isoformat() if r.get("DataRef") else None)
            ano_i = int(r.get("Ano") or 0)
            mes_i = int(r.get("Mes") or 0)

            receita_full = float(r.get("ReceitaMes_Full") or 0.0)

            custo_raw = r.get("CustoMensal")
            custo = float(custo_raw) if custo_raw is not None else None

            margem_full = None
            if receita_full > 0 and custo is not None:
                margem_full = ((receita_full - custo) / receita_full) * 100.0

            serie_hist.append({
                "data_ref": data_ref_iso,
                "ano": ano_i,
                "mes": mes_i,
                "receita": receita_full,
                "custo": custo,
                "margem_pct": float(margem_full) if margem_full is not None else None,
            })

            receita_periodo = r.get("ReceitaMes_Periodo")
            if receita_periodo is not None:
                receita_p = float(receita_periodo or 0.0)

                margem_p = None
                if receita_p > 0 and custo is not None:
                    margem_p = ((receita_p - custo) / receita_p) * 100.0

                serie_fin.append({
                    "data_ref": data_ref_iso,
                    "ano": ano_i,
                    "mes": mes_i,
                    "receita": receita_p,
                    "custo": custo,
                    "margem_pct": float(margem_p) if margem_p is not None else None,
                })

        sql_custo_categoria = text("""
            ;WITH BaseCat AS (
                SELECT
                    TRY_CONVERT(int, CodPonto) AS CodPonto,
                    TRY_CONVERT(int, Ano) AS Ano,
                    TRY_CONVERT(int, Mes) AS Mes,
                    DATEFROMPARTS(TRY_CONVERT(int, Ano), TRY_CONVERT(int, Mes), 1) AS DataRef,
                    CAST(ISNULL(Categoria,'') AS varchar(160)) AS Categoria,
                    TRY_CONVERT(float, ValorMensal) AS ValorMensal
                FROM [Integracao].[Silver].[DimCustoCategoriaMensalPainel]
                WHERE TRY_CONVERT(int, CodPonto) = :cod_ponto
                  AND TRY_CONVERT(int, Ano) IS NOT NULL
                  AND TRY_CONVERT(int, Mes) IS NOT NULL
            ),
            DentroPeriodo AS (
                SELECT *
                FROM BaseCat
                WHERE DataRef >= :dt_ini_mes
                  AND DataRef <= :dt_fim_mes
            ),
            UltimoMes AS (
                SELECT TOP 1 Ano, Mes
                FROM DentroPeriodo
                ORDER BY (Ano * 100 + Mes) DESC
            )
            SELECT
                d.Categoria,
                MAX(COALESCE(d.ValorMensal, 0.0)) AS Valor
            FROM DentroPeriodo d
            INNER JOIN UltimoMes u
                ON d.Ano = u.Ano AND d.Mes = u.Mes
            GROUP BY d.Categoria
            HAVING MAX(COALESCE(d.ValorMensal, 0.0)) > 0
            ORDER BY MAX(COALESCE(d.ValorMensal, 0.0)) DESC;
        """)

        linhas_cat = db.session.execute(
            sql_custo_categoria,
            {
                "cod_ponto": int(painel["id_painel"]),
                "dt_ini_mes": dt_ini_mes,
                "dt_fim_mes": dt_fim_mes,
            }
        ).mappings().all()

        custos_painel = []
        for r in linhas_cat:
            nome = _texto_limpo(r.get("Categoria")) or "Sem Categoria"
            valor = _to_float_br(r.get("Valor")) or 0.0
            custos_painel.append({"nome": nome, "valor": float(valor)})

        anos_presentes = sorted({x["ano"] for x in serie_hist if x.get("ano")}, reverse=True)

        matriz = []
        acumulado_geral = 0.0

        for ano in anos_presentes:
            linha = {
                "ano": ano,
                "painel": painel.get("nome") or f"Painel {painel.get('id_painel')}",
                "meses": {m: {"receita": None, "margem_pct": None} for m in range(1, 13)},
                "total_ano": 0.0,
                "margem_media_ano": None,
                "acumulado": None,
            }

            itens_ano = [x for x in serie_hist if x.get("ano") == ano and x.get("mes")]
            soma_receita_ano = 0.0
            soma_lucro_ano = 0.0
            soma_receita_margem = 0.0

            for it in itens_ano:
                mes = int(it["mes"])
                receita_it = float(it["receita"] or 0.0)

                custo_it = it.get("custo")
                custo_it_num = float(custo_it) if custo_it is not None else None

                linha["meses"][mes]["receita"] = receita_it
                linha["meses"][mes]["margem_pct"] = it.get("margem_pct")

                soma_receita_ano += receita_it

                if custo_it_num is not None:
                    soma_lucro_ano += (receita_it - custo_it_num)
                    soma_receita_margem += receita_it

            linha["total_ano"] = soma_receita_ano

            if soma_receita_margem > 0:
                linha["margem_media_ano"] = (soma_lucro_ano / soma_receita_margem) * 100.0

            acumulado_geral += soma_receita_ano
            linha["acumulado"] = acumulado_geral

            matriz.append(linha)

        fin_json = {
            "dt_ini": dt_ini_str,
            "dt_fim": dt_fim_str,
            "serie": serie_fin,
            "matriz": matriz,
            "custos": custos_painel
        }

    except Exception:
        fin_json = {
            "dt_ini": None,
            "dt_fim": None,
            "serie": [],
            "matriz": [],
            "custos": [],
        }

    prospects_mock = [
        {
            "id_empresa": 102,
            "nome": "Empresa Teste 2",
            "segmento": "Supermercado",
            "lat": -22.8995573,
            "lng": -47.0285836,
            "score": 72,
            "ultimo_contato": None,
            "url_ficha": "/admin/empresas/102",
            "contratou_este_painel": False,
            "ja_e_cliente_euromidia": False,
        },
        {
            "id_empresa": 103,
            "nome": "Empresa Teste 3",
            "segmento": "Automotivo",
            "lat": -23.55240,
            "lng": -46.63460,
            "score": 66,
            "ultimo_contato": "2025-12-18",
            "url_ficha": "/admin/empresas/103",
            "contratou_este_painel": False,
            "ja_e_cliente_euromidia": False,
        },
    ]

    sql_clientes_euromidia = text("""
        ;WITH ContratouEstePainel AS (
            SELECT DISTINCT
                c.IDEmpresa AS IDEmpresa
            FROM [Integracao].[Silver].[FatoControleContratosEuromidia] c
            INNER JOIN [Integracao].[Silver].[FatoControleContratosItensEuromidia] i
                ON i.IDFatoControleContratoEuromidia = c.IDFatoControleContratosEuromidia
            WHERE TRY_CONVERT(int, i.CodPonto) = :cod_ponto
        )
        SELECT
            e.IDEmpresa,
            CAST(COALESCE(e.NomeFantasia, e.RazaoSocial, '') AS nvarchar(200)) AS Nome,
            CAST(ISNULL(e.DescricaoCnae, '') AS nvarchar(250)) AS Segmento,
            e.CNPJ,
            CAST(ISNULL(e.UF,'') AS varchar(10)) AS UF,
            CAST(ISNULL(e.Municipio,'') AS nvarchar(150)) AS Municipio,
            CAST(ISNULL(e.Bairro,'') AS nvarchar(150)) AS Bairro,
            CAST(ISNULL(e.CEP,'') AS varchar(20)) AS CEP,
            CASE WHEN p.IDEmpresa IS NOT NULL THEN 1 ELSE 0 END AS ContratouEstePainel
        FROM [Integracao].[Silver].[DimEmpresas] e
        LEFT JOIN ContratouEstePainel p
            ON p.IDEmpresa = e.IDEmpresa
        WHERE e.IDEmpresaProprietaria = 3
    """)

    linhas_clientes = db.session.execute(
        sql_clientes_euromidia,
        {"cod_ponto": int(painel["id_painel"])}
    ).mappings().all()

    clientes_reais = []
    for r in linhas_clientes:
        clientes_reais.append({
            "id_empresa": int(r["IDEmpresa"]),
            "nome": r.get("Nome") or "",
            "segmento": r.get("Segmento") or "",
            "lat": None,
            "lng": None,
            "score": None,
            "ultimo_contato": None,
            "url_ficha": f"/admin/empresas/{int(r['IDEmpresa'])}",
            "contratou_este_painel": bool(r.get("ContratouEstePainel") == 1),
            "ja_e_cliente_euromidia": True,
            "cnpj": r.get("CNPJ"),
            "uf": r.get("UF"),
            "municipio": r.get("Municipio"),
            "bairro": r.get("Bairro"),
            "cep": r.get("CEP"),
        })

    empresas = clientes_reais + prospects_mock

    raio_m = max(0, _to_int_seguro(request.args.get("raio_m", "1000"), 1000))
    status = (request.args.get("status", "todos") or "todos").strip()
    segmento = (request.args.get("segmento", "todos") or "todos").strip()

    def _distancia_haversine_m(lat1, lng1, lat2, lng2):
        if lat1 is None or lng1 is None or lat2 is None or lng2 is None:
            return None
        R = 6371000.0
        lat1r = math.radians(float(lat1))
        lng1r = math.radians(float(lng1))
        lat2r = math.radians(float(lat2))
        lng2r = math.radians(float(lng2))
        dlat = lat2r - lat1r
        dlng = lng2r - lng1r
        a = (math.sin(dlat / 2) ** 2) + (math.cos(lat1r) * math.cos(lat2r) * (math.sin(dlng / 2) ** 2))
        c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
        return int(round(R * c))

    def _classificar_relacao(emp: dict) -> str:
        if emp.get("contratou_este_painel"):
            return "cliente_no_painel"
        if emp.get("ja_e_cliente_euromidia"):
            return "cliente_euromidia"
        return "prospect"

    def _classificar_proximidade(distancia_m: int, raio_metros: int) -> str:
        if distancia_m is None:
            return "sem_localizacao"
        return "dentro_do_raio" if distancia_m <= raio_metros else "fora_do_raio"

    def _definir_camada(status_relacao: str, proximidade: str) -> str:
        if status_relacao == "cliente_no_painel":
            return "clientes_painel"
        if status_relacao == "cliente_euromidia":
            return "clientes_euromidia"
        if proximidade == "dentro_do_raio":
            return "prospects_raio"
        return "fora_escopo"

    empresas_enriquecidas = []
    for emp in empresas:
        dist_m = _distancia_haversine_m(painel["lat"], painel["lng"], emp.get("lat"), emp.get("lng"))
        status_relacao = _classificar_relacao(emp)
        proximidade = _classificar_proximidade(dist_m, raio_m)
        camada = _definir_camada(status_relacao, proximidade)

        emp2 = dict(emp)
        emp2["distancia_m"] = dist_m
        emp2["status_relacao"] = status_relacao
        emp2["proximidade"] = proximidade
        emp2["camada"] = camada
        emp2["status"] = "cliente" if status_relacao in ("cliente_no_painel", "cliente_euromidia") else "prospect"
        empresas_enriquecidas.append(emp2)

    def _passa_filtro(emp: dict) -> bool:
        if emp.get("camada") == "fora_escopo":
            return False
        if segmento != "todos" and (emp.get("segmento") or "") != segmento:
            return False
        if status == "todos":
            return True
        if status == "cliente":
            return emp.get("status") == "cliente"
        if status == "prospect":
            return emp.get("status") == "prospect"
        if status == "cliente_no_painel":
            return emp.get("status_relacao") == "cliente_no_painel"
        if status == "cliente_euromidia":
            return emp.get("status_relacao") == "cliente_euromidia"
        return True

    empresas_filtradas = [e for e in empresas_enriquecidas if _passa_filtro(e)]

    camadas = {
        "clientes_painel": [e for e in empresas_filtradas if e.get("camada") == "clientes_painel"],
        "clientes_euromidia": [e for e in empresas_filtradas if e.get("camada") == "clientes_euromidia"],
        "prospects_raio": [e for e in empresas_filtradas if e.get("camada") == "prospects_raio"],
        "sem_localizacao": [e for e in empresas_filtradas if e.get("proximidade") == "sem_localizacao"],
    }

    empresas_no_raio = [
        e for e in empresas_enriquecidas
        if e.get("proximidade") == "dentro_do_raio" and e.get("camada") != "fora_escopo"
    ]
    clientes_no_raio = [e for e in empresas_no_raio if e.get("status") == "cliente"]

    cont_segmentos = {}
    for e in empresas_no_raio:
        seg = e.get("segmento") or "Sem Segmento"
        cont_segmentos[seg] = cont_segmentos.get(seg, 0) + 1

    segmento_top = None
    if cont_segmentos:
        segmento_top = sorted(cont_segmentos.items(), key=lambda x: x[1], reverse=True)[0][0]

    kpis = {
        "empresas_no_raio": len(empresas_no_raio),
        "clientes_no_raio": len(clientes_no_raio),
        "segmento_top": segmento_top,
        "clientes_sem_localizacao": len(camadas["sem_localizacao"]),
        "painel_faces": painel.get("quantidade_faces"),
        "painel_iluminado": painel.get("bit_iluminado"),
        "painel_ativo": painel.get("bit_ativo"),
        "painel_aluguel": painel.get("bit_aluguel"),
    }

    sql_bairros = text("""
        SELECT
            CAST(ISNULL(Bairro,'') AS varchar(160)) AS bairro_final,
            TRY_CONVERT(float, LatitudeBairro) AS lat,
            TRY_CONVERT(float, LongitudeBairro) AS lng,

            TRY_CONVERT(float, TotalEnderecos) AS total_enderecos,
            TRY_CONVERT(float, QuantidadeResidencial) AS qtd_residencial,
            TRY_CONVERT(float, QuantidadeNaoResidencial) AS qtd_nao_residencial,
            TRY_CONVERT(float, QuantidadeIndefinido) AS qtd_indefinido,

            TRY_CONVERT(float, PercentualResidencial) AS pct_residencial,
            TRY_CONVERT(float, PercentualNaoResidencial) AS pct_nao_residencial,
            TRY_CONVERT(float, PercentualIndefinido) AS pct_indefinido,

            TRY_CONVERT(float, PesoHeat) AS peso_heat,
            CAST(ISNULL(PerfilDominante,'') AS varchar(60)) AS perfil_dominante,

            TRY_CONVERT(float, QuantidadePaineis) AS qtd_paineis,
            CAST(ISNULL(ListaCodPonto,'') AS varchar(max)) AS lista_cod_ponto,
            CAST(ISNULL(BitPainelEuromidia, 0) AS bit) AS tem_painel_euromidia
        FROM [Integracao].[Silver].[DimPerfilRegiaoFull]
        WHERE UF = :uf
          AND Municipio = :municipio
          AND LatitudeBairro IS NOT NULL
          AND LongitudeBairro IS NOT NULL
    """)

    linhas_bairros = db.session.execute(
        sql_bairros,
        {"uf": painel["uf"], "municipio": painel["municipio"]}
    ).mappings().all()

    cep_points = []
    for r in linhas_bairros:
        lat_i = r.get("lat")
        lng_i = r.get("lng")
        if lat_i is None or lng_i is None:
            continue

        tot_end = float(r.get("total_enderecos") or 0.0)
        tot_nres = float(r.get("qtd_nao_residencial") or 0.0)

        dens_100 = 0.0
        if tot_end > 0:
            dens_100 = (tot_nres / tot_end) * 100.0

        cep_points.append({
            "lat": float(lat_i),
            "lng": float(lng_i),
            "props": {
                "BAIRRO": r.get("bairro_final") or "",
                "TOT_END": int(round(float(r.get("total_enderecos") or 0.0))),
                "TOT_DP": int(round(float(r.get("qtd_residencial") or 0.0))),
                "TOT_ESTAB_NRESID": int(round(float(r.get("qtd_nao_residencial") or 0.0))),
                "PCT_RESIDENCIAL_%": float(r.get("pct_residencial") or 0.0),
                "PCT_ESTAB_NRESID_%": float(r.get("pct_nao_residencial") or 0.0),
                "DENS_ESTAB_NRESID_100_END": float(dens_100),
                "PESO_HEAT": float(r.get("peso_heat") or 0.0),
                "TIPO_AREA": r.get("perfil_dominante") or "",
                "QTD_PAINEIS": int(round(float(r.get("qtd_paineis") or 0.0))),
                "TEM_PAINEL_EUROMIDIA": _to_bool_sql(r.get("tem_painel_euromidia"), False),
                "LISTA_COD_PONTO": r.get("lista_cod_ponto") or "",
                "LISTA_CEP_PAINEL": r.get("lista_cep_painel") or "",
            }
        })

    return render_template(
        "euromidia/mapa_mercado.html",
        painel_json=painel,
        empresas_json=empresas_filtradas,
        cep_points_json=cep_points,
        filtro_inicial={"raio_m": raio_m, "status": status, "segmento": segmento},
        camadas=camadas,
        kpis=kpis,
        fin_json=fin_json,
    )











@paineis_bp.route("/ocupacao/reserva/nova", methods=["GET"])
@login_required
def reserva_nova():
    codponto = (request.args.get("codponto") or "").strip()
    mes_ref  = (request.args.get("mes_ref") or "").strip()
    tipo = (request.args.get("tipo") or "").strip()
    cliente  = (request.args.get("cliente") or "").strip()
    vendedor = (request.args.get("vendedor") or "").strip()

    codface_sel = (request.args.get("codface") or request.args.get("cod_face") or "").strip()

    dt_ini_str = (request.args.get("dt_ini") or "").strip()
    dt_fim_str = (request.args.get("dt_fim") or "").strip()

    try:
        codponto_int = int(codponto) if codponto else None
    except Exception:
        codponto_int = None

    def parse_ymd(s):
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except Exception:
            return None

    dt_ini = parse_ymd(dt_ini_str)
    dt_fim = parse_ymd(dt_fim_str)

    if dt_ini is None or dt_fim is None:

        dt_base = None
        try:
            if mes_ref and len(mes_ref) == 7:
                dt_base = datetime.strptime(mes_ref + "-01", "%Y-%m-%d").date()
        except Exception:
            dt_base = None

        if dt_base is None:
            hoje = date.today()
            dt_base = date(hoje.year, hoje.month, 1)

        dt_ini = date(dt_base.year, dt_base.month, 1)

        if dt_base.month == 12:
            dt_fim = date(dt_base.year + 1, 1, 1) - timedelta(days=1)
        else:
            dt_fim = date(dt_base.year, dt_base.month + 1, 1) - timedelta(days=1)

    faces = []
    if codponto_int is not None:
        sql_faces = text("""
            SELECT CodFace, CodPonto, Face
            FROM [Silver].[DimFacesPaineis]
            WHERE CodPonto = :codponto
            ORDER BY TRY_CONVERT(int, Face), Face, CodFace
        """)
        faces = [dict(r._mapping) for r in db.session.execute(sql_faces, {"codponto": codponto_int})]

    if codface_sel:
        if not any((str(f.get("CodFace") or "").strip() == codface_sel) for f in faces):
            codface_sel = ""
    else:
        if faces:
            codface_sel = str(faces[0].get("CodFace") or "").strip()

    form = ReservaOcupacaoForm()
    form.cod_ponto.data = str(codponto_int or "")

    return render_template(
        "euromidia/reserva_nova.html",
        form=form,
        codponto=codponto,
        mes_ref=mes_ref,
        tipo=tipo,
        cliente=cliente,
        vendedor=vendedor,
        faces=faces,
        codface_sel=codface_sel,
        dt_ini=dt_ini.strftime("%Y-%m-%d"),
        dt_fim=dt_fim.strftime("%Y-%m-%d"),
    )








@paineis_bp.route("/api/ocupacao/calendario", methods=["GET"])
def api_ocupacao_calendario():

    cod_face = (request.args.get("cod_face") or request.args.get("codface") or "").strip()
    mes_ref  = (request.args.get("mes_ref") or "").strip()
    meses = request.args.get("meses", 24)

    try:
        meses = int(meses)
    except Exception:
        meses = 24

    if not cod_face:
        return jsonify({"ok": False, "erro": "cod_face obrigatório"}), 400

    sql = text("""
        DECLARE @CodFace varchar(20) = :cod_face;

        DECLARE @Inicio date =
        CASE
            WHEN :mes_ref IS NOT NULL AND LEN(:mes_ref) = 7
            THEN TRY_CONVERT(date, CONCAT(:mes_ref, '-01'))
            ELSE DATEFROMPARTS(YEAR(GETDATE()), MONTH(GETDATE()), 1)
        END;

        IF @Inicio IS NULL
            SET @Inicio = DATEFROMPARTS(YEAR(GETDATE()), MONTH(GETDATE()), 1);

        DECLARE @Fim date = DATEADD(MONTH, :meses, @Inicio);

        DECLARE @CodPonto int =
        (
            SELECT TOP (1) fo.CodPonto
            FROM [Integracao].[Silver].[FatoOcupacaoPaineisEuromidia] fo
            WHERE fo.CodFace = @CodFace
              AND fo.CodPonto IS NOT NULL
            ORDER BY fo.DataAtualizacao DESC
        );

        ;WITH
        Painel AS (
            SELECT TOP (1)
                p.CodPonto,
                TipoPainel = UPPER(LTRIM(RTRIM(p.Tipo))),
                QuantidadeFaces = NULLIF(p.QuantidadeFaces, 0),
                BitAtivo  = COALESCE(p.BitAtivo, 1)
            FROM [Integracao].[Silver].[DimPaineisEuromidia] p
            WHERE p.CodPonto = @CodPonto
        ),
        Capacidade AS (
            SELECT
                CodPonto = (SELECT CodPonto FROM Painel),
                TipoPainel = COALESCE((SELECT TipoPainel FROM Painel), 'DESCONHECIDO'),
                BitAtivo = (SELECT BitAtivo FROM Painel),

                EhDigital =
                    CASE
                      WHEN COALESCE((SELECT TipoPainel FROM Painel),'') LIKE '%DIGITAL%' THEN 1
                      ELSE 0
                    END,

                CapacidadeSlots =
                    CASE
                      WHEN COALESCE((SELECT TipoPainel FROM Painel),'') LIKE '%DIGITAL%'
                      THEN COALESCE((SELECT QuantidadeFaces FROM Painel), 16)
                      ELSE 1
                    END
        ),
        OcupacoesBase AS (
            SELECT
                fo.CodFace,
                DataInicio = CAST(fo.DataInicio AS date),
                DataFim    = CAST(fo.DataFim    AS date),
                SpanQtd  = fo.SpanQtd,
                Cota = fo.Cota,
                NumeroContrato = fo.NumeroContrato,
                NumeroPrevia   = fo.NumeroPrevia,
                DataAtualizacao= fo.DataAtualizacao
            FROM [Integracao].[Silver].[FatoOcupacaoPaineisEuromidia] fo
            WHERE fo.CodFace = @CodFace
              AND fo.DataInicio IS NOT NULL
              AND fo.DataFim    IS NOT NULL
              AND fo.CanceladoEm IS NULL
              AND fo.Status IN ('ATIVO','RESERVADO')
        ),
        OcupacoesDedup AS (
            SELECT *
            FROM (
                SELECT
                    b.*,
                    rn = ROW_NUMBER() OVER (
                        PARTITION BY
                            b.CodFace, b.DataInicio, b.DataFim,
                            ISNULL(b.NumeroContrato,''), ISNULL(b.NumeroPrevia,'')
                        ORDER BY b.DataAtualizacao DESC
                    )
                FROM OcupacoesBase b
            ) x
            WHERE x.rn = 1
        ),
        OcupacoesValidas AS (
            SELECT
                d.DataInicio,
                d.DataFim,
                SlotsConsumidos =
                    CASE
                      WHEN (SELECT EhDigital FROM Capacidade) = 1
                      THEN COALESCE(NULLIF(d.SpanQtd, 0), 1)
                      ELSE 1
                    END
            FROM OcupacoesDedup d
        ),
        UsoPorDia AS (
            SELECT
                c.[Data],
                SlotsOcupados =
                    CASE
                      WHEN cap.EhDigital = 1
                      THEN COALESCE(SUM(o.SlotsConsumidos), 0)
                      ELSE CASE WHEN COUNT(o.SlotsConsumidos) > 0 THEN 1 ELSE 0 END
                    END
            FROM [Integracao].[Silver].[DimCalendario] c
            CROSS JOIN Capacidade cap
            LEFT JOIN OcupacoesValidas o
                   ON c.[Data] >= o.DataInicio
                  AND c.[Data] <= o.DataFim
            WHERE c.[Data] >= @Inicio
              AND c.[Data] <  @Fim
            GROUP BY c.[Data], cap.EhDigital
        )
        SELECT
            Data = CONVERT(varchar(10), c.[Data], 23),
            CodPonto  = cap.CodPonto,
            TipoPainel  = cap.TipoPainel,
            EhDigital = cap.EhDigital,
            CapacidadeSlots= cap.CapacidadeSlots,
            SlotsOcupados  = u.SlotsOcupados,

            SlotsDisponiveis =
                CASE
                  WHEN cap.CapacidadeSlots - u.SlotsOcupados < 0 THEN 0
                  ELSE cap.CapacidadeSlots - u.SlotsOcupados
                END,

            OcupacaoPct =
                CASE
                  WHEN cap.CapacidadeSlots > 0
                  THEN CAST(u.SlotsOcupados * 100.0 / cap.CapacidadeSlots AS decimal(9,2))
                  ELSE NULL
                END,

            DiaDisponivel =
                CASE
                  WHEN cap.BitAtivo = 0 THEN 0
                  WHEN (cap.CapacidadeSlots - u.SlotsOcupados) > 0 THEN 1
                  ELSE 0
                END,

            StatusDia =
                CASE
                  WHEN cap.BitAtivo = 0 THEN 'INDISPONIVEL'
                  WHEN cap.EhDigital = 0 AND u.SlotsOcupados = 0 THEN 'DISPONIVEL'
                  WHEN cap.EhDigital = 0 AND u.SlotsOcupados = 1 THEN 'OCUPADO'
                  WHEN cap.EhDigital = 1 AND u.SlotsOcupados = 0 THEN 'LIVRE'
                  WHEN cap.EhDigital = 1 AND u.SlotsOcupados < cap.CapacidadeSlots THEN 'PARCIAL'
                  ELSE 'LOTADO'
                END
        FROM [Integracao].[Silver].[DimCalendario] c
        JOIN UsoPorDia u
          ON u.[Data] = c.[Data]
        CROSS JOIN Capacidade cap
        WHERE c.[Data] >= @Inicio
          AND c.[Data] <  @Fim
        ORDER BY c.[Data];
    """)

    rows = db.session.execute(sql, {
        "cod_face": cod_face,
        "mes_ref": mes_ref,
        "meses": meses
    }).all()

 
    cal = {}
    for r in rows:
        k = (r.Data or "").strip()
        if not k:
            continue
        cal[k] = {
            "status": (r.StatusDia or "").strip(),
            "disp": int(r.SlotsDisponiveis or 0),
            "cap": int(r.CapacidadeSlots or 0),
            "ocup": int(r.SlotsOcupados or 0),
            "pct": float(r.OcupacaoPct) if r.OcupacaoPct is not None else None,
            "dia_disponivel": int(r.DiaDisponivel or 0),
            "eh_digital": int(r.EhDigital or 0),
            "codponto": int(r.CodPonto) if r.CodPonto is not None else None,
            "tipo": (r.TipoPainel or "").strip(),
        }

    return jsonify({"ok": True, "cal": cal})





@paineis_bp.route("/api/ocupacao/reserva/dados-modal", methods=["GET"])
def api_ocupacao_reserva_dados_modal():

    cod_face = (request.args.get("cod_face") or request.args.get("codface") or "").strip()
    cod_ponto = (request.args.get("cod_ponto") or request.args.get("codponto") or "").strip()
    mes_ref = (request.args.get("mes_ref") or "").strip()
    cota = (request.args.get("cota") or "").strip()

    dt_ini = (request.args.get("dt_ini") or "").strip()
    dt_fim = (request.args.get("dt_fim") or "").strip()

    if not cod_face:
        return jsonify({"ok": False, "erro": "cod_face obrigatório"}), 400

    def executar_primeiro_sql_que_funciona(lista_sql, params):
        for sql_txt in lista_sql:
            try:
                rows = db.session.execute(text(sql_txt), params).all()
                return [dict(r._mapping) for r in rows]
            except Exception:
                continue
        return []

    sql_ocupacoes = """
        SELECT
            IDFatoOcupacaoPaineisEuromidia,
            CodFace,
            Status,
            CONVERT(varchar(10), DataInicio, 23) AS DataInicio,
            CONVERT(varchar(10), DataFim, 23) AS DataFim,
            ISNULL(Cota,0) AS Cota,
            MarcaExibida,
            Vendedor
        FROM [Integracao].[Silver].[FatoOcupacaoPaineisEuromidia]
        WHERE CodFace = :cod_face
          AND CanceladoEm IS NULL
          AND Status IN ('ATIVO','RESERVADO')
        ORDER BY DataInicio
    """
    ocupacoes = [dict(r._mapping) for r in db.session.execute(text(sql_ocupacoes), {"cod_face": cod_face})]

    sqls_empresas = [
        """
        SELECT DISTINCT IDEmpresa, RazaoSocial, CNPJ
        FROM [Integracao].[Silver].[DimEmpresas] WHERE IDempresaProprietaria = 3
        ORDER BY RazaoSocial
        """
    ]
    empresas = executar_primeiro_sql_que_funciona(sqls_empresas, {})

    sqls_vendedores = [
        """
        SELECT DISTINCT IDVendedor, NomeVendedor
        FROM [Integracao].[dbo].[Vendedores]
        ORDER BY NomeVendedor
        """
    ]
    vendedores = executar_primeiro_sql_que_funciona(sqls_vendedores, {})

    if not vendedores:
        sql_vendedores_fallback = """
            SELECT DISTINCT
                IDVendedor = NULLIF(LTRIM(RTRIM(CAST(IDVendedor AS varchar(50)))), ''),
                NomeVendedor = NULLIF(LTRIM(RTRIM(CAST(Vendedor AS varchar(200)))), '')
            FROM [Integracao].[Silver].[FatoOcupacaoPaineisEuromidia]
            WHERE CodFace = :cod_face
              AND CanceladoEm IS NULL
              AND Status IN ('ATIVO','RESERVADO')
              AND NULLIF(LTRIM(RTRIM(CAST(Vendedor AS varchar(200)))), '') IS NOT NULL
            ORDER BY NomeVendedor
        """
        vendedores = [dict(r._mapping) for r in db.session.execute(text(sql_vendedores_fallback), {"cod_face": cod_face})]

        for v in vendedores:
            if not (v.get("IDVendedor") or "").strip():
                v["IDVendedor"] = v.get("NomeVendedor") or ""

    params_contratos = {
        "cod_ponto": cod_ponto,
        "dt_ini": dt_ini,
        "dt_fim": dt_fim,
        "mes_ref": mes_ref,
        "cota": cota
    }

 

    sqls_contratos = [
        # ✅ 1) PRIORIDADE: pelo ITENS (tem CodPonto) e JOIN no HEADER pra pegar RazaoSocial correta
        """
        SELECT DISTINCT
            IDContrato = COALESCE(i.IDFatoControleContratoEuromidia, h.IDFatoControleContratosEuromidia),
            IDFatoControleContratosEuromidia = COALESCE(i.IDFatoControleContratoEuromidia, h.IDFatoControleContratosEuromidia),
            NumeroContrato = NULLIF(LTRIM(RTRIM(CAST(COALESCE(h.NumeroContrato, i.NumeroContrato) AS varchar(50)))), ''),
            NumeroPrevia   = NULLIF(LTRIM(RTRIM(CAST(COALESCE(h.NumeroPrevia,   i.NumeroPrevia)   AS varchar(50)))), ''),
            RazaoSocial    = NULLIF(LTRIM(RTRIM(COALESCE(h.RazaoSocial, i.RazaoSocial))), ''),
            CNPJ           = NULLIF(LTRIM(RTRIM(CAST(COALESCE(h.CNPJ, i.CNPJ) AS varchar(30)))), ''),
            IDEmpresa      = COALESCE(h.IDEmpresa, i.IDEmpresa)
        FROM [Integracao].[Silver].[FatoControleContratosItensEuromidia] i
        LEFT JOIN [Integracao].[Silver].[FatoControleContratosEuromidia] h
               ON h.IDFatoControleContratosEuromidia = i.IDFatoControleContratoEuromidia
        WHERE
            (NULLIF(:cod_ponto,'') IS NULL OR TRY_CONVERT(int, i.CodPonto) = TRY_CONVERT(int, :cod_ponto))
        ORDER BY
            NULLIF(LTRIM(RTRIM(CAST(COALESCE(h.NumeroContrato, i.NumeroContrato) AS varchar(50)))), ''),
            NULLIF(LTRIM(RTRIM(CAST(COALESCE(h.NumeroPrevia,   i.NumeroPrevia)   AS varchar(50)))), '')
        """,

        # ✅ 2) fallback: header puro (SEM filtro por cod_ponto porque não existe CodPonto no header)
        """
        SELECT DISTINCT
            IDContrato = IDFatoControleContratosEuromidia,
            IDFatoControleContratosEuromidia,
            NumeroContrato = NULLIF(LTRIM(RTRIM(CAST(NumeroContrato AS varchar(50)))), ''),
            NumeroPrevia   = NULLIF(LTRIM(RTRIM(CAST(NumeroPrevia   AS varchar(50)))), ''),
            RazaoSocial    = NULLIF(LTRIM(RTRIM(RazaoSocial)), ''),
            CNPJ           = NULLIF(LTRIM(RTRIM(CAST(CNPJ AS varchar(30)))), ''),
            IDEmpresa
        FROM [Integracao].[Silver].[FatoControleContratosEuromidia]
        ORDER BY
            NULLIF(LTRIM(RTRIM(CAST(NumeroContrato AS varchar(50)))), ''),
            NULLIF(LTRIM(RTRIM(CAST(NumeroPrevia   AS varchar(50)))), '')
        """,

        # ✅ 3) fallback: itens puro (se por algum motivo o JOIN falhar)
        """
        SELECT DISTINCT
            IDContrato = IDFatoControleContratosItensEuromidia,
            CAST(NULL AS int) AS IDFatoControleContratosEuromidia,
            NumeroContrato = NULLIF(LTRIM(RTRIM(CAST(NumeroContrato AS varchar(50)))), ''),
            NumeroPrevia   = NULLIF(LTRIM(RTRIM(CAST(NumeroPrevia   AS varchar(50)))), ''),
            RazaoSocial    = NULLIF(LTRIM(RTRIM(COALESCE(RazaoSocial,''))), ''),
            CNPJ           = NULLIF(LTRIM(RTRIM(CAST(CNPJ AS varchar(30)))), ''),
            IDEmpresa
        FROM [Integracao].[Silver].[FatoControleContratosItensEuromidia]
        WHERE (NULLIF(:cod_ponto,'') IS NULL OR TRY_CONVERT(int, CodPonto) = TRY_CONVERT(int, :cod_ponto))
        ORDER BY
            NULLIF(LTRIM(RTRIM(CAST(NumeroContrato AS varchar(50)))), ''),
            NULLIF(LTRIM(RTRIM(CAST(NumeroPrevia   AS varchar(50)))), '')
        """,

        # ✅ 4) fallback: tabela antiga FatoControleContratos (se existir/for usada)
        """
        SELECT DISTINCT
            IDContrato,
            CAST(NULL AS int) AS IDFatoControleContratosEuromidia,
            NumeroContrato = NULLIF(LTRIM(RTRIM(CAST(NumeroContrato AS varchar(50)))), ''),
            NumeroPrevia   = NULLIF(LTRIM(RTRIM(CAST(NumeroPrevia   AS varchar(50)))), ''),
            RazaoSocial    = NULLIF(LTRIM(RTRIM(COALESCE(RazaoSocial,''))), ''),
            CNPJ           = NULLIF(LTRIM(RTRIM(CAST(CNPJ AS varchar(30)))), ''),
            IDEmpresa
        FROM [Integracao].[Silver].[FatoControleContratos]
        WHERE (NULLIF(:cod_ponto,'') IS NULL OR TRY_CONVERT(int, CodPonto) = TRY_CONVERT(int, :cod_ponto))
        ORDER BY
            NULLIF(LTRIM(RTRIM(CAST(NumeroContrato AS varchar(50)))), ''),
            NULLIF(LTRIM(RTRIM(CAST(NumeroPrevia   AS varchar(50)))), '')
        """,
    ]

    contratos = executar_primeiro_sql_que_funciona(sqls_contratos, params_contratos)

    if not contratos:
        sql_contratos_fallback = """
            SELECT DISTINCT
                IDContrato = COALESCE(
                    NULLIF(LTRIM(RTRIM(CAST(IDFatoControleContratos AS varchar(50)))), ''),
                    NULLIF(LTRIM(RTRIM(CAST(NumeroContrato AS varchar(50)))), ''),
                    CONCAT('C:', CONVERT(varchar(10), MIN(DataInicio), 23), '|', CONVERT(varchar(10), MAX(DataFim), 23))
                ),
                IDFatoControleContratosEuromidia = TRY_CONVERT(int, IDFatoControleContratos),
                NumeroContrato = NULLIF(LTRIM(RTRIM(CAST(NumeroContrato AS varchar(50)))), ''),
                NumeroPrevia = NULLIF(LTRIM(RTRIM(CAST(NumeroPrevia AS varchar(50)))), ''),
                RazaoSocial = CAST('' AS varchar(200)),
                CNPJ = CAST('' AS varchar(30)),
                IDEmpresa = TRY_CONVERT(int, IDCliente)
            FROM [Integracao].[Silver].[FatoOcupacaoPaineisEuromidia]
            WHERE CodFace = :cod_face
              AND CanceladoEm IS NULL
              AND Status IN ('ATIVO','RESERVADO')
              AND (NumeroContrato IS NOT NULL OR NumeroPrevia IS NOT NULL OR IDFatoControleContratos IS NOT NULL)
            GROUP BY IDFatoControleContratos, NumeroContrato, NumeroPrevia, IDCliente
            ORDER BY NumeroContrato, NumeroPrevia
        """
        contratos = [dict(r._mapping) for r in db.session.execute(text(sql_contratos_fallback), {"cod_face": cod_face})]

        for c in contratos:
            if not (str(c.get("IDContrato") or "").strip()):
                nc = str(c.get("NumeroContrato") or "").strip()
                np = str(c.get("NumeroPrevia") or "").strip()
                c["IDContrato"] = (nc + "-" + np).strip("-") or "sem-id"

    return jsonify({
        "ok": True,
        "empresas": empresas,
        "vendedores": vendedores,
        "contratos": contratos,
        "ocupacoes": ocupacoes,
    })





@paineis_bp.route("/api/ocupacao/reserva/criar", methods=["POST"])
@login_required
def api_ocupacao_reserva_criar():

    payload = request.get_json(silent=True)
    if not isinstance(payload, dict):
        payload = {}

    csrf_token = (
        request.headers.get("X-CSRFToken")
        or request.headers.get("X-CSRF-Token")
        or payload.get("csrf_token")
        or request.form.get("csrf_token")
        or request.form.get("csrf")
    )

    try:
        validate_csrf(csrf_token)
    except Exception:
        return jsonify({"ok": False, "erro": "CSRF inválido/ausente"}), 400

    cod_ponto = (payload.get("cod_ponto") or request.form.get("cod_ponto") or "").strip()
    cod_face  = (payload.get("cod_face")  or request.form.get("cod_face")  or "").strip()

    cota_raw = payload.get("cota")
    if cota_raw is None:
        cota_raw = request.form.get("cota")

    data_inicio = (payload.get("data_inicio") or request.form.get("data_inicio") or "").strip()
    data_fim    = (payload.get("data_fim")    or request.form.get("data_fim")    or "").strip()

    id_cliente_raw = payload.get("id_cliente")
    if id_cliente_raw is None:
        id_cliente_raw = request.form.get("id_cliente")

    cnpj_cliente = (payload.get("cnpj_cliente") or request.form.get("cnpj_cliente") or "").strip()

    id_vendedor_raw = payload.get("id_vendedor")
    if id_vendedor_raw is None:
        id_vendedor_raw = request.form.get("id_vendedor")

    vendedor_nome_raw = (payload.get("vendedor") or request.form.get("vendedor") or "").strip()

    id_fato_controle_raw = payload.get("id_fato_controle_contratos")
    if id_fato_controle_raw is None:
        id_fato_controle_raw = request.form.get("id_fato_controle_contratos")

    numero_contrato = (payload.get("numero_contrato") or request.form.get("numero_contrato") or "").strip()
    numero_previa   = (payload.get("numero_previa")   or request.form.get("numero_previa")   or "").strip()

    dias_raw = payload.get("dias")
    if dias_raw is None:
        dias_raw = request.form.get("dias")

  
    marca_exibida = (payload.get("marca_exibida") or request.form.get("marca_exibida") or "").strip()

    if not cod_ponto:
        return jsonify({"ok": False, "erro": "cod_ponto obrigatório"}), 400
    if not cod_face:
        return jsonify({"ok": False, "erro": "cod_face obrigatório"}), 400
    if not data_inicio or not data_fim:
        return jsonify({"ok": False, "erro": "data_inicio e data_fim obrigatórios"}), 400

  
    if not marca_exibida:
        return jsonify({"ok": False, "erro": "marca_exibida obrigatória"}), 400
    if len(marca_exibida) > 200:
        return jsonify({"ok": False, "erro": "marca_exibida excede 200 caracteres"}), 400

    try:
        cod_ponto_int = int(cod_ponto)
    except Exception:
        return jsonify({"ok": False, "erro": "cod_ponto inválido"}), 400

    cota_int = None
    try:
        if cota_raw not in ("", None, "null", "None"):
            cota_int = int(cota_raw)
    except Exception:
        return jsonify({"ok": False, "erro": "cota inválida"}), 400

    id_cliente_int = None
    try:
        if id_cliente_raw not in ("", None, "null", "None"):
            id_cliente_int = int(id_cliente_raw)
    except Exception:
        return jsonify({"ok": False, "erro": "id_cliente inválido"}), 400

    id_vendedor_int = None
    try:
        if id_vendedor_raw not in ("", None, "null", "None"):
            id_vendedor_int = int(id_vendedor_raw)
    except Exception:
        return jsonify({"ok": False, "erro": "id_vendedor inválido"}), 400

    id_fato_controle_int = None
    try:
        if id_fato_controle_raw not in ("", None, "null", "None"):
            id_fato_controle_int = int(id_fato_controle_raw)
    except Exception:
        return jsonify({"ok": False, "erro": "id_fato_controle_contratos inválido"}), 400

    dias_int = None
    try:
        if dias_raw not in ("", None, "null", "None"):
            dias_int = int(dias_raw)
    except Exception:
        return jsonify({"ok": False, "erro": "dias inválido"}), 400

    if dias_int is None:
        dias_int = 7

    sql_painel = text("""
        SELECT TOP 1
            IDPainelEuromidia =  IDDimPaineisEuromidia
        FROM [Integracao].[Silver].[DimFacesPaineis]
        WHERE CodFace = :cod_face
        ORDER BY IDDimFacesPaineis DESC
    """)
    id_painel = db.session.execute(sql_painel, {"cod_face": cod_face}).scalar()
    try:
        id_painel_int = int(id_painel) if id_painel is not None else None
    except Exception:
        id_painel_int = None


    vendedor_nome = vendedor_nome_raw or None
    if id_vendedor_int is not None:
        sql_vend = text("""
            SELECT TOP 1 NomeVendedor
            FROM [Integracao].[dbo].[Vendedores]
            WHERE IDVendedor = :idv
        """)
        nome_db = db.session.execute(sql_vend, {"idv": id_vendedor_int}).scalar()
        if nome_db:
            vendedor_nome = str(nome_db).strip() or vendedor_nome

    if id_vendedor_int is None and vendedor_nome:
        sql_vend_id = text("""
            SELECT TOP 1 IDVendedor
            FROM [Integracao].[dbo].[Vendedores]
            WHERE NomeVendedor = :nome
        """)
        id_db = db.session.execute(sql_vend_id, {"nome": vendedor_nome}).scalar()
        try:
            id_vendedor_int = int(id_db) if id_db is not None else None
        except Exception:
            id_vendedor_int = None

    if id_cliente_int is None and cnpj_cliente:
        sql_cli = text("""
            SELECT TOP 1 IDEmpresa
            FROM [Integracao].[Silver].[DimEmpresas]
            WHERE REPLACE(REPLACE(REPLACE(REPLACE(COALESCE(CNPJ,''),'.',''),'-',''),'/',''),' ','')
                = REPLACE(REPLACE(REPLACE(REPLACE(:cnpj,'.',''),'-',''),'/',''),' ','')
        """)
        idc = db.session.execute(sql_cli, {"cnpj": cnpj_cliente}).scalar()
        try:
            id_cliente_int = int(idc) if idc is not None else None
        except Exception:
            id_cliente_int = None

    if id_cliente_int is None:
        return jsonify({"ok": False, "erro": "Não foi possível resolver IDEmpresa (IDCliente). Envie id_cliente (IDEmpresa) ou cnpj_cliente válido."}), 400


    sql_cap = text("""
        DECLARE @CodFace varchar(20) = :cod_face;

        DECLARE @CodPonto int =
        (
            SELECT TOP (1) fo.CodPonto
            FROM [Integracao].[Silver].[FatoOcupacaoPaineisEuromidia] fo
            WHERE fo.CodFace = @CodFace
              AND fo.CodPonto IS NOT NULL
            ORDER BY fo.DataAtualizacao DESC
        );

        ;WITH Painel AS (
            SELECT TOP (1)
                p.CodPonto,
                TipoPainel = UPPER(LTRIM(RTRIM(p.Tipo))),
                QuantidadeFaces = NULLIF(p.QuantidadeFaces, 0),
                BitAtivo = COALESCE(p.BitAtivo, 1)
            FROM [Integracao].[Silver].[DimPaineisEuromidia] p
            WHERE p.CodPonto = @CodPonto
        )
        SELECT TOP 1
            CodPonto = (SELECT CodPonto FROM Painel),
            TipoPainel = COALESCE((SELECT TipoPainel FROM Painel), 'DESCONHECIDO'),
            BitAtivo   = (SELECT BitAtivo FROM Painel),
            EhDigital =
                CASE
                    WHEN COALESCE((SELECT TipoPainel FROM Painel),'') LIKE '%DIGITAL%' THEN 1
                    ELSE 0
                END,
            CapacidadeSlots =
                CASE
                    WHEN COALESCE((SELECT TipoPainel FROM Painel),'') LIKE '%DIGITAL%'
                    THEN COALESCE((SELECT QuantidadeFaces FROM Painel), 16)
                    ELSE 1
                END;
    """)
    cap_row = db.session.execute(sql_cap, {"cod_face": cod_face}).mappings().first()

    if not cap_row:
        return jsonify({"ok": False, "erro": "Não consegui resolver capacidade do painel para essa CodFace"}), 400

    bit_ativo = int(cap_row.get("BitAtivo") or 0)
    eh_digital = int(cap_row.get("EhDigital") or 0)
    capacidade_slots = int(cap_row.get("CapacidadeSlots") or 0)

    if bit_ativo == 0:
        return jsonify({"ok": False, "erro": "Painel inativo (BitAtivo=0)"}), 409

    if capacidade_slots <= 0:
        return jsonify({"ok": False, "erro": "Capacidade inválida (<=0)"}), 400


    slots_novo = 1
    spanqtd_novo = 1 if eh_digital == 1 else None


    sem_capacidade = False

    if eh_digital == 0:
        sql_conflito = text("""
            SELECT TOP 1 1
            FROM [Integracao].[Silver].[FatoOcupacaoPaineisEuromidia]
            WHERE CodFace = :cod_face
              AND CanceladoEm IS NULL
              AND Status IN ('ATIVO','RESERVADO')
              AND NOT (
                    TRY_CONVERT(date, :data_fim) < DataInicio
                 OR TRY_CONVERT(date, :data_inicio) > DataFim
              )
        """)
        existe = db.session.execute(sql_conflito, {
            "cod_face": cod_face,
            "data_inicio": data_inicio,
            "data_fim": data_fim
        }).scalar()

        if existe:
            # ANTES: return 409
            # AGORA: entra em FILA
            sem_capacidade = True

    else:

        sql_conflito_digital = text("""
            DECLARE @CodFace varchar(20) = :cod_face;
            DECLARE @Ini date = TRY_CONVERT(date, :data_inicio);
            DECLARE @Fim date = TRY_CONVERT(date, :data_fim);

            IF @Ini IS NULL OR @Fim IS NULL
            BEGIN
                SELECT TOP 1
                    Conflito = 1,
                    Dia = CAST(NULL AS date),
                    SlotsOcupados = CAST(NULL AS int),
                    CapacidadeSlots = CAST(NULL AS int),
                    SlotsDisponiveis = CAST(NULL AS int);
                RETURN;
            END

            DECLARE @CodPonto int =
            (
                SELECT TOP (1) fo.CodPonto
                FROM [Integracao].[Silver].[FatoOcupacaoPaineisEuromidia] fo
                WHERE fo.CodFace = @CodFace
                  AND fo.CodPonto IS NOT NULL
                ORDER BY fo.DataAtualizacao DESC
            );

            ;WITH Painel AS (
                SELECT TOP (1)
                    p.CodPonto,
                    TipoPainel      = UPPER(LTRIM(RTRIM(p.Tipo))),
                    QuantidadeFaces = NULLIF(p.QuantidadeFaces, 0),
                    BitAtivo        = COALESCE(p.BitAtivo, 1)
                FROM [Integracao].[Silver].[DimPaineisEuromidia] p
                WHERE p.CodPonto = @CodPonto
            ),
            Capacidade AS (
                SELECT
                    BitAtivo   = (SELECT BitAtivo FROM Painel),
                    EhDigital =
                        CASE
                          WHEN COALESCE((SELECT TipoPainel FROM Painel),'') LIKE '%DIGITAL%' THEN 1
                          ELSE 0
                        END,
                    CapacidadeSlots =
                        CASE
                          WHEN COALESCE((SELECT TipoPainel FROM Painel),'') LIKE '%DIGITAL%'
                          THEN COALESCE((SELECT QuantidadeFaces FROM Painel), 16)
                          ELSE 1
                        END
            ),
            OcupacoesBase AS (
                SELECT
                    DataInicio = CAST(fo.DataInicio AS date),
                    DataFim    = CAST(fo.DataFim    AS date),
                    SpanQtd    = fo.SpanQtd,
                    NumeroContrato = fo.NumeroContrato,
                    NumeroPrevia   = fo.NumeroPrevia,
                    DataAtualizacao= fo.DataAtualizacao
                FROM [Integracao].[Silver].[FatoOcupacaoPaineisEuromidia] fo
                WHERE fo.CodFace = @CodFace
                  AND fo.DataInicio IS NOT NULL
                  AND fo.DataFim    IS NOT NULL
                  AND fo.CanceladoEm IS NULL
                  AND fo.Status IN ('ATIVO','RESERVADO')
            ),
            OcupacoesDedup AS (
                SELECT *
                FROM (
                    SELECT
                        b.*,
                        rn = ROW_NUMBER() OVER (
                            PARTITION BY
                                b.DataInicio, b.DataFim,
                                ISNULL(b.NumeroContrato,''), ISNULL(b.NumeroPrevia,'')
                            ORDER BY b.DataAtualizacao DESC
                        )
                    FROM OcupacoesBase b
                ) x
                WHERE x.rn = 1
            ),
            OcupacoesValidas AS (
                SELECT
                    DataInicio,
                    DataFim,
                    SlotsConsumidos = COALESCE(NULLIF(SpanQtd, 0), 1)
                FROM OcupacoesDedup
            ),
            UsoPorDia AS (
                SELECT
                    c.[Data] AS Dia,
                    SlotsOcupados = COALESCE(SUM(o.SlotsConsumidos), 0),
                    CapacidadeSlots = (SELECT CapacidadeSlots FROM Capacidade)
                FROM [Integracao].[Silver].[DimCalendario] c
                LEFT JOIN OcupacoesValidas o
                       ON c.[Data] >= o.DataInicio
                      AND c.[Data] <= o.DataFim
                WHERE c.[Data] >= @Ini
                  AND c.[Data] <= @Fim
                GROUP BY c.[Data]
            )
            SELECT TOP 1
                Conflito = 1,
                Dia,
                SlotsOcupados,
                CapacidadeSlots,
                SlotsDisponiveis =
                    CASE
                      WHEN CapacidadeSlots - SlotsOcupados < 0 THEN 0
                      ELSE CapacidadeSlots - SlotsOcupados
                    END
            FROM UsoPorDia
            WHERE SlotsOcupados + :slots_novo > CapacidadeSlots
            ORDER BY Dia;
        """)

        conflito = db.session.execute(sql_conflito_digital, {
            "cod_face": cod_face,
            "data_inicio": data_inicio,
            "data_fim": data_fim,
            "slots_novo": slots_novo
        }).mappings().first()

        if conflito and conflito.get("Conflito") == 1 and conflito.get("Dia") is not None:
            # ANTES: return 409
            # AGORA: entra em FILA
            sem_capacidade = True

    criado_por = int(getattr(current_user, "id", 0) or 0)



    reserva_ordem_prioridade_int = 1

    if sem_capacidade:
        sql_prioridade = text("""
            DECLARE @Ini date = TRY_CONVERT(date, :data_inicio);
            DECLARE @Fim date = TRY_CONVERT(date, :data_fim);

            IF @Ini IS NULL OR @Fim IS NULL
            BEGIN
                SELECT CAST(2 AS int) AS ProximaPrioridade;
                RETURN;
            END

            SELECT
                ProximaPrioridade = ISNULL(MAX(COALESCE(ReservaOrdemPrioridade, 0)), 0) + 1
            FROM [Integracao].[Silver].[FatoOcupacaoPaineisEuromidia] WITH (UPDLOCK, HOLDLOCK)
            WHERE CodFace = :cod_face
              AND CanceladoEm IS NULL
              AND Status IN ('ATIVO','RESERVADO')
              AND DataInicio = @Ini
              AND DataFim    = @Fim
              AND ( ( :cota IS NULL AND Cota IS NULL ) OR ( Cota = :cota ) )
              AND ( ( :spanqtd IS NULL AND SpanQtd IS NULL ) OR ( COALESCE(SpanQtd,0) = COALESCE(:spanqtd,0) ) );
        """)

        prox = db.session.execute(sql_prioridade, {
            "cod_face": cod_face,
            "data_inicio": data_inicio,
            "data_fim": data_fim,
            "cota": cota_int,
            "spanqtd": spanqtd_novo
        }).scalar()

        try:
            reserva_ordem_prioridade_int = int(prox) if prox is not None else 2
        except Exception:
            reserva_ordem_prioridade_int = 2


    sql_insert = text("""
        INSERT INTO [Integracao].[Silver].[FatoOcupacaoPaineisEuromidia] (
            DataAtualizacao,
            Referencia,
            CodPonto,
            CodFace,
            IDPainelEuromidia,
            Origem,
            Status,
            DataInicio,
            DataFim,
            SpanQtd,
            Cota,
            MarcaExibida,
            Vendedor,
            IDVendedor,
            IDCliente,
            IDFatoControleContratos,
            NumeroContrato,
            NumeroPrevia,
            Observacao,
            Dias,
            ExpiraEm,
            CriadoEm,
            CriadoPorIDUsuario,
            ReservaOrdemPrioridade
        )
        VALUES (
            SYSDATETIME(),
            CONVERT(varchar(64),
                HASHBYTES('SHA2_256',
                    CONCAT(
                        'RESERVA|',
                        COALESCE(CONVERT(varchar(30), :cod_ponto), ''), '|',
                        UPPER(LTRIM(RTRIM(COALESCE(:cod_face,'')))), '|',
                        COALESCE(CONVERT(varchar(10), TRY_CONVERT(date, :data_inicio), 23), ''), '|',
                        COALESCE(CONVERT(varchar(10), TRY_CONVERT(date, :data_fim), 23), ''), '|',
                        COALESCE(CONVERT(varchar(30), :spanqtd), ''), '|',
                        COALESCE(CONVERT(varchar(30), :cota), ''), '|',
                        COALESCE(CONVERT(varchar(30), :id_cliente), ''), '|',
                        COALESCE(CONVERT(varchar(30), :id_vendedor), ''), '|',
                        UPPER(LTRIM(RTRIM(COALESCE(:numero_contrato,'')))), '|',
                        UPPER(LTRIM(RTRIM(COALESCE(:numero_previa,''))))
                    )
                ), 2
            ),
            :cod_ponto,
            :cod_face,
            :id_painel,
            'RESERVA',
            'RESERVADO',
            TRY_CONVERT(date, :data_inicio),
            TRY_CONVERT(date, :data_fim),
            :spanqtd,
            :cota,
            :marca_exibida,
            :vendedor_nome,
            :id_vendedor,
            :id_cliente,
            :id_fato_controle,
            NULLIF(LTRIM(RTRIM(:numero_contrato)), ''),
            NULLIF(LTRIM(RTRIM(:numero_previa)), ''),
            NULLIF(LTRIM(RTRIM(:observacao)), ''),
            :dias,
            DATEADD(day, :dias, SYSDATETIME()),
            SYSDATETIME(),
            :criado_por,
            :reserva_ordem_prioridade
        )
    """)

    try:
        db.session.execute(sql_insert, {
            "cod_ponto": cod_ponto_int,
            "cod_face": cod_face,
            "id_painel": id_painel_int,
            "data_inicio": data_inicio,
            "data_fim": data_fim,
            "spanqtd": spanqtd_novo,
            "cota": cota_int,
            "marca_exibida": marca_exibida,
            "vendedor_nome": vendedor_nome,
            "id_vendedor": id_vendedor_int,
            "id_cliente": id_cliente_int,
            "id_fato_controle": id_fato_controle_int,
            "numero_contrato": numero_contrato,
            "numero_previa": numero_previa,
            "observacao": (payload.get("observacao") or request.form.get("observacao") or "").strip(),
            "dias": dias_int,
            "criado_por": criado_por,
            "reserva_ordem_prioridade": reserva_ordem_prioridade_int
        })
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"ok": False, "erro": "Erro ao inserir reserva na tabela"}), 500

    return jsonify({"ok": True})










@paineis_bp.route("/ocupacao", methods=["GET"])
@login_required
@limiter.limit("80 per minute", methods=["GET"])
@retry_get_view(db, attempts=6, base_delay=0.2, max_delay=1.5)
def lista_ocupacao():

    def _somente_digitos(s: str) -> str:
        try:
            if not s:
                return ""
            return "".join([c for c in str(s) if c.isdigit()])
        except:
            return ""

    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").strip()
    origem = (request.args.get("origem") or "").strip()

    try:
        page = int(request.args.get("page") or "1")
    except:
        page = 1

    try:
        per_page = int(request.args.get("per_page") or "20")
    except:
        per_page = 20

    per_page = max(5, min(per_page, 100))
    page = max(1, page)

    filtros_sql = []
    params = {}

  
    filtros_sql.append("ftcp.CodFace <> '0' AND ftcp.CodFace IS NOT NULL")

    if status:
        filtros_sql.append("ftcp.Status = :status")
        params["status"] = status

    if origem:
        filtros_sql.append("ftcp.Origem = :origem")
        params["origem"] = origem

    if q:
        filtros_sql.append("""
            (
                ftcp.CodFace LIKE :q_like
                OR emp.RazaoSocial LIKE :q_like
                OR ftcp.MarcaExibida LIKE :q_like
            )
        """)
        params["q_like"] = f"%{q}%"

    where_sql = " AND ".join([f"({x})" for x in filtros_sql]) if filtros_sql else "1=1"

   
    sql_total = text(f"""
        ;WITH dedupe AS (
            SELECT
                ftcp.IDFatoOcupacaoPaineisEuromidia,
                ftcp.CodFace,
                ftcp.Origem,
                ftcp.Status,
                ftcp.DataInicio,
                ftcp.DataFim,
                ftcp.MarcaExibida,
                ftcp.Cota,
                ftcp.CriadoEm,
                ftcp.DataAtualizacao,
                emp.RazaoSocial,

                ROW_NUMBER() OVER (
                    PARTITION BY
                        ISNULL(ftcp.CodFace,''),
                        ISNULL(ftcp.Origem,''),
                        ISNULL(ftcp.Status,''),
                        ISNULL(CAST(ftcp.DataInicio AS date),'1900-01-01'),
                        ISNULL(CAST(ftcp.DataFim AS date),'1900-01-01'),
                        ISNULL(ftcp.MarcaExibida,''),
                        ISNULL(CAST(ftcp.Cota AS varchar(50)),''),
                        ISNULL(CAST(ftcp.CriadoEm AS datetime2),'1900-01-01')
                    ORDER BY
                        ISNULL(ftcp.DataAtualizacao, ftcp.CriadoEm) DESC,
                        ftcp.IDFatoOcupacaoPaineisEuromidia DESC
                ) AS rn_key
            FROM [Integracao].[Silver].[FatoOcupacaoPaineisEuromidia] AS ftcp
            LEFT JOIN [Integracao].[Silver].[DimEmpresas] AS emp
                ON emp.IDEmpresa = ftcp.[IDCliente]
            WHERE {where_sql}
        )
        SELECT COUNT(1) AS total
        FROM dedupe
        WHERE rn_key = 1
    """)

    row_total = db.session.execute(sql_total, params).fetchone()
    total = int(row_total[0] or 0) if row_total else 0

    total_pages = max(1, (total + per_page - 1) // per_page)
    if page > total_pages:
        page = total_pages

    inicio = (page - 1) * per_page + 1 if total > 0 else 0
    fim = min(page * per_page, total)

    start_rn = (page - 1) * per_page + 1
    end_rn = page * per_page

    params_page = dict(params)
    params_page.update({"start_rn": start_rn, "end_rn": end_rn})


    sql_rows = text(f"""
        ;WITH dedupe AS (
            SELECT
                ftcp.IDFatoOcupacaoPaineisEuromidia,
                ftcp.CodFace,
                ftcp.Origem,
                ftcp.Status,
                ftcp.DataInicio,
                ftcp.DataFim,
                ftcp.MarcaExibida,
                ftcp.Cota,
                ftcp.CriadoEm,
                ftcp.DataAtualizacao,
                emp.RazaoSocial,

                ROW_NUMBER() OVER (
                    PARTITION BY
                        ISNULL(ftcp.CodFace,''),
                        ISNULL(ftcp.Origem,''),
                        ISNULL(ftcp.Status,''),
                        ISNULL(CAST(ftcp.DataInicio AS date),'1900-01-01'),
                        ISNULL(CAST(ftcp.DataFim AS date),'1900-01-01'),
                        ISNULL(ftcp.MarcaExibida,''),
                        ISNULL(CAST(ftcp.Cota AS varchar(50)),''),
                        ISNULL(CAST(ftcp.CriadoEm AS datetime2),'1900-01-01')
                    ORDER BY
                        ISNULL(ftcp.DataAtualizacao, ftcp.CriadoEm) DESC,
                        ftcp.IDFatoOcupacaoPaineisEuromidia DESC
                ) AS rn_key
            FROM [Integracao].[Silver].[FatoOcupacaoPaineisEuromidia] AS ftcp
            LEFT JOIN [Integracao].[Silver].[DimEmpresas] AS emp
                ON emp.IDEmpresa = ftcp.[IDCliente]
            WHERE {where_sql}
        ),
        filtrada AS (
            SELECT *
            FROM dedupe
            WHERE rn_key = 1
        ),
        ranked AS (
            SELECT
                *,
                ROW_NUMBER() OVER (
                    ORDER BY
                        ISNULL(CAST(DataInicio AS date), '1900-01-01') DESC,
                        ISNULL(CodFace, '') ASC
                ) AS rn
            FROM filtrada
        )
        SELECT
            IDFatoOcupacaoPaineisEuromidia,
            CodFace,
            Origem,
            Status,
            DataInicio,
            DataFim,
            MarcaExibida,
            Cota,
            CriadoEm,
            RazaoSocial
        FROM ranked
        WHERE rn BETWEEN :start_rn AND :end_rn
        ORDER BY rn
    """)

    rows = db.session.execute(sql_rows, params_page).fetchall()

    itens = []
    for r in rows:
        itens.append({
            "IDFatoOcupacaoPaineisEuromidia": int(r[0]) if r[0] is not None else None,
            "CodFace": (r[1] or "").strip(),
            "Origem": (r[2] or "").strip(),
            "Status": (r[3] or "").strip(),
            "DataInicio": r[4],
            "DataFim": r[5],
            "MarcaExibida": (r[6] or "").strip(),
            "Cota": r[7],
            "CriadoEm": r[8],
            "RazaoSocial": (r[9] or "").strip(),
        })

    
    sql_status = text("""
        SELECT DISTINCT Status
        FROM [Integracao].[Silver].[FatoOcupacaoPaineisEuromidia]
        WHERE Status IS NOT NULL AND LTRIM(RTRIM(Status)) <> ''
        ORDER BY Status ASC
    """)
    status_opcoes = [x[0] for x in db.session.execute(sql_status).fetchall()]

    sql_origem = text("""
        SELECT DISTINCT Origem
        FROM [Integracao].[Silver].[FatoOcupacaoPaineisEuromidia]
        WHERE Origem IS NOT NULL AND LTRIM(RTRIM(Origem)) <> ''
        ORDER BY Origem ASC
    """)
    origem_opcoes = [x[0] for x in db.session.execute(sql_origem).fetchall()]

    return render_template(
        "euromidia/ocupacao_lista.html",
        itens=itens,
        status_opcoes=status_opcoes,
        origem_opcoes=origem_opcoes,
        filtros={
            "q": q,
            "status": status,
            "origem": origem,
            "per_page": per_page,
        },
        paginacao={
            "page": page,
            "per_page": per_page,
            "total": total,
            "total_pages": total_pages,
            "inicio": inicio,
            "fim": fim,
        },
    )







@paineis_bp.route("/ocupacao/<int:id_ocupacao>", methods=["GET"])
@login_required
@limiter.limit("120 per minute", methods=["GET"])
@retry_get_view(db, attempts=6, base_delay=0.2, max_delay=1.5)
def ocupacao_detalhe(id_ocupacao: int):
    sql = text("""
        SELECT
            ftcp.IDFatoOcupacaoPaineisEuromidia,
            ftcp.DataAtualizacao,
            ftcp.Referencia,
            ftcp.CodPonto,
            ftcp.CodFace,
            ftcp.IDPainelEuromidia,
            ftcp.Origem,
            ftcp.Status,
            ftcp.DataInicio,
            ftcp.DataFim,
            ftcp.LoopInicio,
            ftcp.LoopFim,
            ftcp.SpanQtd,
            ftcp.Cota,
            ftcp.MarcaExibida,
            ftcp.Vendedor,
            ftcp.IDVendedor,
            ftcp.IDCliente,
            ftcp.IDFatoControleContratos,
            ftcp.NumeroContrato,
            ftcp.NumeroPrevia,
            ftcp.TextoOriginal,
            ftcp.CriadoEm,
            ftcp.CriadoPorIDUsuario,
            ftcp.ExpiraEm,
            ftcp.CanceladoEm,
            ftcp.CanceladoPorIDUsuario,
            ftcp.Observacao,
            ftcp.Dias,
            emp.RazaoSocial
        FROM [Integracao].[Silver].[FatoOcupacaoPaineisEuromidia] AS ftcp
        LEFT JOIN [Integracao].[Silver].[DimEmpresas] AS emp
            ON emp.IDEmpresa = ftcp.[IDCliente]
        WHERE ftcp.IDFatoOcupacaoPaineisEuromidia = :id
    """)

    r = db.session.execute(sql, {"id": id_ocupacao}).fetchone()
    if not r:
        abort(404)

    detalhe = {
        "IDFatoOcupacaoPaineisEuromidia": r[0],
        "DataAtualizacao": r[1],
        "Referencia": (r[2] or "").strip(),
        "CodPonto": (r[3] or ""),
        "CodFace": (r[4] or "").strip(),
        "IDPainelEuromidia": r[5],
        "Origem": (r[6] or "").strip(),
        "Status": (r[7] or "").strip(),
        "DataInicio": r[8],
        "DataFim": r[9],
        "LoopInicio": r[10],
        "LoopFim": r[11],
        "SpanQtd": r[12],
        "Cota": r[13],
        "MarcaExibida": (r[14] or "").strip(),
        "Vendedor": (r[15] or "").strip(),
        "IDVendedor": r[16],
        "IDCliente": r[17],
        "IDFatoControleContratos": r[18],
        "NumeroContrato": (r[19] or "").strip(),
        "NumeroPrevia": (r[20] or "").strip(),
        "TextoOriginal": r[21],
        "CriadoEm": r[22],
        "CriadoPorIDUsuario": r[23],
        "ExpiraEm": r[24],
        "CanceladoEm": r[25],
        "CanceladoPorIDUsuario": r[26],
        "Observacao": r[27],
        "Dias": r[28],
        "RazaoSocial": (r[29] or "").strip(),
    }

    return render_template(
        "euromidia/ocupacao_detalhe.html",
        detalhe=detalhe
    )






@paineis_bp.route("/ocupacao/<int:id_ocupacao>/editar", methods=["GET", "POST"])
@login_required
@limiter.limit("60 per minute", methods=["GET", "POST"])
@retry_get_view(db, attempts=6, base_delay=0.2, max_delay=1.5)
def ocupacao_editar(id_ocupacao: int):
    sql_get = text("""
        SELECT
            ftcp.IDFatoOcupacaoPaineisEuromidia,
            ftcp.CodFace,
            ftcp.Status,
            ftcp.Origem,
            ftcp.DataInicio,
            ftcp.DataFim,
            ftcp.LoopInicio,
            ftcp.LoopFim,
            ftcp.SpanQtd,
            ftcp.Cota,
            ftcp.MarcaExibida,
            ftcp.Vendedor,
            ftcp.NumeroContrato,
            ftcp.NumeroPrevia,
            ftcp.Observacao,
            ftcp.ExpiraEm,
            ftcp.CanceladoEm,
            emp.RazaoSocial
        FROM [Integracao].[Silver].[FatoOcupacaoPaineisEuromidia] AS ftcp
        LEFT JOIN [Integracao].[Silver].[DimEmpresas] AS emp
            ON emp.IDEmpresa = ftcp.[IDCliente]
        WHERE ftcp.IDFatoOcupacaoPaineisEuromidia = :id
    """)

    r = db.session.execute(sql_get, {"id": id_ocupacao}).fetchone()
    if not r:
        abort(404)

    detalhe = {
        "IDFatoOcupacaoPaineisEuromidia": r[0],
        "CodFace": (r[1] or "").strip(),
        "Status": (r[2] or "").strip(),
        "Origem": (r[3] or "").strip(),
        "DataInicio": r[4],
        "DataFim": r[5],
        "LoopInicio": r[6],
        "LoopFim": r[7],
        "SpanQtd": r[8],
        "Cota": r[9],
        "MarcaExibida": (r[10] or "").strip(),
        "Vendedor": (r[11] or "").strip(),
        "NumeroContrato": (r[12] or "").strip(),
        "NumeroPrevia": (r[13] or "").strip(),
        "Observacao": (r[14] or ""),
        "ExpiraEm": r[15],
        "CanceladoEm": r[16],
        "RazaoSocial": (r[17] or "").strip(),
    }


    ja_cancelado = detalhe["CanceladoEm"] is not None

    if request.method == "GET":
        return render_template(
            "euromidia/ocupacao_editar.html",
            detalhe=detalhe,
            ja_cancelado=ja_cancelado
        )

   
    if ja_cancelado:
        flash("Esta ocupação já está cancelada e não pode ser alterada.", "warning")
        return redirect(url_for("Paineis.ocupacao_detalhe", id_ocupacao=id_ocupacao))

 
    marca_exibida = (request.form.get("MarcaExibida") or "").strip()
    vendedor = (request.form.get("Vendedor") or "").strip()
    numero_contrato = (request.form.get("NumeroContrato") or "").strip()
    numero_previa = (request.form.get("NumeroPrevia") or "").strip()
    observacao = (request.form.get("Observacao") or "").strip()

    sql_upd = text("""
        UPDATE [Integracao].[Silver].[FatoOcupacaoPaineisEuromidia]
        SET
            MarcaExibida = :marca,
            Vendedor = :vendedor,
            NumeroContrato = :numero_contrato,
            NumeroPrevia = :numero_previa,
            Observacao = :observacao,
            DataAtualizacao = GETDATE()
        WHERE IDFatoOcupacaoPaineisEuromidia = :id
    """)

    db.session.execute(sql_upd, {
        "id": id_ocupacao,
        "marca": marca_exibida if marca_exibida else None,
        "vendedor": vendedor if vendedor else None,
        "numero_contrato": numero_contrato if numero_contrato else None,
        "numero_previa": numero_previa if numero_previa else None,
        "observacao": observacao if observacao else None,
    })
    db.session.commit()

    flash("Ocupação atualizada com sucesso.", "success")
    return redirect(url_for("Paineis.ocupacao_detalhe", id_ocupacao=id_ocupacao))







@paineis_bp.route("/ocupacao/<int:id_ocupacao>/cancelar", methods=["GET","POST"])
@login_required
@limiter.limit("40 per minute", methods=["POST"])
@retry_get_view(db, attempts=6, base_delay=0.2, max_delay=1.5)
def ocupacao_cancelar(id_ocupacao: int):
   
    sql_chk = text("""
        SELECT
            IDFatoOcupacaoPaineisEuromidia,
            CanceladoEm,
            Referencia,
            CodPonto,
            CodFace,
            DataInicio,
            DataFim
        FROM [Integracao].[Silver].[FatoOcupacaoPaineisEuromidia]
        WHERE IDFatoOcupacaoPaineisEuromidia = :id
    """)
    r = db.session.execute(sql_chk, {"id": id_ocupacao}).fetchone()
    if not r:
        abort(404)

    ja_cancelado = r[1] is not None
    if ja_cancelado:
        flash("Esta ocupação já estava cancelada.", "warning")
        return redirect(url_for("Paineis.ocupacao_detalhe", id_ocupacao=id_ocupacao))

    referencia = r[2]
    cod_ponto = r[3]
    cod_face = r[4]
    data_inicio = r[5]
    data_fim = r[6]

    motivo = (request.form.get("motivo_cancelamento") or "").strip()

    sql_cancel = text("""
        UPDATE [Integracao].[Silver].[FatoOcupacaoPaineisEuromidia]
        SET
            Status = 'CANCELADO',
            CanceladoEm = GETDATE(),
            CanceladoPorIDUsuario = :id_usuario,
            Observacao = CASE
                WHEN :motivo IS NULL OR LTRIM(RTRIM(:motivo)) = '' THEN Observacao
                WHEN Observacao IS NULL OR LTRIM(RTRIM(Observacao)) = '' THEN :motivo
                ELSE CONCAT(Observacao, CHAR(10), '[CANCELAMENTO] ', :motivo)
            END,
            DataAtualizacao = GETDATE()
        WHERE IDFatoOcupacaoPaineisEuromidia = :id
    """)


    sql_upd_item = text("""
        DECLARE @Ref varchar(64) = :referencia;

        DECLARE @IDContrato int =
        (
            SELECT TOP 1 c.IDFatoControleContratosEuromidia
            FROM [Integracao].[Silver].[FatoControleContratosEuromidia] c
            WHERE c.Referencia = @Ref
            ORDER BY c.DataAtualizacao DESC
        );

        IF @IDContrato IS NULL
        BEGIN
            SELECT 0 AS LinhasAfetadas;
            RETURN;
        END

        UPDATE i
        SET
            i.Status = 'CANCELADO',
            i.DataCancelamento = CAST(GETDATE() AS date)
        FROM [Integracao].[Silver].[FatoControleContratosItensEuromidia] i
        WHERE i.IDFatoControleContratoEuromidia = @IDContrato
          AND i.CodPonto = :cod_ponto
          AND LTRIM(RTRIM(COALESCE(i.CodFace,''))) = LTRIM(RTRIM(COALESCE(:cod_face,'')))
          AND i.DataInicioPrevisto = :data_inicio
          AND i.DataTerminoPrevisto = :data_fim;

        SELECT @@ROWCOUNT AS LinhasAfetadas;
    """)

    id_usuario = int(getattr(current_user, "id", 0) or 0)

    try:
        db.session.execute(sql_cancel, {
            "id": id_ocupacao,
            "id_usuario": id_usuario,
            "motivo": motivo if motivo else None
        })

       
        db.session.execute(sql_upd_item, {
            "referencia": str(referencia).strip() if referencia is not None else None,
            "cod_ponto": cod_ponto,
            "cod_face": cod_face,
            "data_inicio": data_inicio,
            "data_fim": data_fim
        })

        db.session.commit()
    except Exception:
        db.session.rollback()
        flash("Erro ao cancelar ocupação (ou ao cancelar item do contrato).", "danger")
        return redirect(url_for("Paineis.ocupacao_detalhe", id_ocupacao=id_ocupacao))

    flash("Ocupação cancelada com sucesso.", "success")
    return redirect(url_for("Paineis.lista_ocupacao"))





@paineis_bp.route("/ocupacao/<int:id_ocupacao>/status", methods=["POST", "GET"])
@login_required
@limiter.limit("120 per minute", methods=["POST"])
@retry_get_view(db, attempts=6, base_delay=0.2, max_delay=1.5)
def ocupacao_alterar_status(id_ocupacao: int):
    if request.method == "GET":
        return redirect(url_for("Paineis.ocupacao_detalhe", id_ocupacao=id_ocupacao))

    status_novo = (request.form.get("status_novo") or "").strip().upper()

    permitidos = {"CANCELADO", "RESERVADO", "ATIVO"}
    if status_novo not in permitidos:
        abort(400)

    id_usuario = None

    sql_chk = text("""
        SELECT
            oc.IDFatoOcupacaoPaineisEuromidia,
            LTRIM(RTRIM(ISNULL(oc.Status,'')))  AS StatusAtual,
            LTRIM(RTRIM(ISNULL(oc.Origem,'')))  AS OrigemAtual,
            oc.Referencia,
            oc.CodPonto,
            LTRIM(RTRIM(ISNULL(oc.CodFace,'')))   AS CodFace,
            oc.IDPainelEuromidia,
            oc.IDCliente,
            LTRIM(RTRIM(ISNULL(emp.RazaoSocial,''))) AS RazaoSocial,
            oc.MarcaExibida,
            oc.Vendedor,
            oc.IDVendedor,
            oc.NumeroContrato,
            oc.NumeroPrevia,
            oc.Cota,
            oc.DataInicio,
            oc.DataFim
        FROM [Integracao].[Silver].[FatoOcupacaoPaineisEuromidia] AS oc
        LEFT JOIN [Integracao].[Silver].[DimEmpresas] AS emp
            ON emp.IDEmpresa = oc.IDCliente
        WHERE oc.IDFatoOcupacaoPaineisEuromidia = :id
    """)
    r = db.session.execute(sql_chk, {"id": id_ocupacao}).fetchone()
    if not r:
        abort(404)

    status_atual = (r[1] or "").strip().upper()

    referencia = (r[3] or "").strip()
    codponto = int(r[4]) if r[4] is not None else None
    codface = (r[5] or "").strip()
    id_painel = r[6]
    id_cliente = r[7]
    razao_social = (r[8] or "").strip()
    marca_exibida = (r[9] or "").strip() if r[9] else None
    vendedor = (r[10] or "").strip() if r[10] else None
    id_vendedor = r[11]
    numero_contrato = (r[12] or "").strip() if r[12] else None
    numero_previa = (r[13] or "").strip() if r[13] else None
    cota = r[14]
    data_inicio = r[15]
    data_fim = r[16]

    if not referencia or codponto is None or not codface or data_inicio is None or data_fim is None:
        abort(400)

    origem_forcada = None
    if status_novo == "ATIVO":
        origem_forcada = "CONTRATO"
    elif status_novo == "RESERVADO":
        origem_forcada = "RESERVA"

    sql_upd_ocup = text("""
        UPDATE [Integracao].[Silver].[FatoOcupacaoPaineisEuromidia]
        SET
            Status = :status_novo,
            Origem = COALESCE(:origem_forcada, Origem),
            DataAtualizacao = GETDATE(),
            CanceladoEm = CASE
                WHEN :status_novo = 'CANCELADO' THEN ISNULL(CanceladoEm, GETDATE())
                ELSE NULL
            END,
            CanceladoPorIDUsuario = CASE
                WHEN :status_novo = 'CANCELADO' THEN :id_usuario
                ELSE NULL
            END
        WHERE IDFatoOcupacaoPaineisEuromidia = :id
    """)
    db.session.execute(sql_upd_ocup, {
        "id": id_ocupacao,
        "status_novo": status_novo,
        "origem_forcada": origem_forcada,
        "id_usuario": id_usuario
    })

    if status_novo == "ATIVO":

        sql_get_dimface = text("""
            SELECT TOP 1
                IDDimFacesPaineis
            FROM [Integracao].[Silver].[DimFacesPaineis]
            WHERE
                TRY_CONVERT(int, IDDimPaineisEuromidia) = TRY_CONVERT(int, :id_painel)
                AND LTRIM(RTRIM(ISNULL(CodFace,''))) = LTRIM(RTRIM(:codface))
            ORDER BY IDDimFacesPaineis DESC
        """)
        id_dim_faces = db.session.execute(sql_get_dimface, {
            "id_painel": id_painel,
            "codface": codface
        }).scalar()

        if id_dim_faces is None:
            sql_get_dimface_fallback = text("""
                SELECT TOP 1
                    IDDimFacesPaineis
                FROM [Integracao].[Silver].[DimFacesPaineis]
                WHERE
                    TRY_CONVERT(int, CodPonto) = TRY_CONVERT(int, :codponto)
                    AND LTRIM(RTRIM(ISNULL(CodFace,''))) = LTRIM(RTRIM(:codface))
                ORDER BY IDDimFacesPaineis DESC
            """)
            id_dim_faces = db.session.execute(sql_get_dimface_fallback, {
                "codponto": codponto,
                "codface": codface
            }).scalar()

        try:
            id_dim_faces_int = int(id_dim_faces) if id_dim_faces is not None else None
        except Exception:
            id_dim_faces_int = None

        sql_get_header = text("""
            SELECT TOP 1 IDFatoControleContratosEuromidia
            FROM [Integracao].[Silver].[FatoControleContratosEuromidia]
            WHERE Referencia = :ref
            ORDER BY IDFatoControleContratosEuromidia DESC
        """)
        row_header = db.session.execute(sql_get_header, {"ref": referencia}).fetchone()

        if row_header:
            id_contrato = int(row_header[0])
        else:
            sql_ins_header = text("""
                INSERT INTO [Integracao].[Silver].[FatoControleContratosEuromidia] (
                    DataAtualizacao,
                    Referencia,
                    NumeroContrato,
                    NumeroPrevia,
                    DataLancamento,
                    RazaoSocial,
                    MarcaExibida,
                    Vendedor,
                    Origem,
                    IDEmpresa
                )
                OUTPUT INSERTED.IDFatoControleContratosEuromidia AS IDNovo
                VALUES (
                    GETDATE(),
                    :ref,
                    :numero_contrato,
                    :numero_previa,
                    CAST(GETDATE() AS date),
                    :razao_social,
                    :marca_exibida,
                    :vendedor,
                    'CONTRATO',
                    :id_empresa
                );
            """)
            row_new = db.session.execute(sql_ins_header, {
                "ref": referencia,
                "numero_contrato": numero_contrato,
                "numero_previa": numero_previa,
                "razao_social": razao_social if razao_social else None,
                "marca_exibida": marca_exibida,
                "vendedor": vendedor,
                "id_empresa": id_cliente
            }).fetchone()

            if not row_new or row_new[0] is None:
                db.session.rollback()
                abort(500)

            id_contrato = int(row_new[0])

        sql_get_item = text("""
            SELECT TOP 1 IDFatoControleContratosItensEuromidia
            FROM [Integracao].[Silver].[FatoControleContratosItensEuromidia]
            WHERE
                Referencia = :ref
                AND ISNULL(CodPonto, -1) = :codponto
                AND LTRIM(RTRIM(ISNULL(CodFace,''))) = :codface
                AND DataInicioPrevisto = :dt_ini
                AND DataTerminoPrevisto = :dt_fim
            ORDER BY IDFatoControleContratosItensEuromidia DESC
        """)
        row_item = db.session.execute(sql_get_item, {
            "ref": referencia,
            "codponto": codponto,
            "codface": codface,
            "dt_ini": data_inicio,
            "dt_fim": data_fim
        }).fetchone()

        if row_item:

            id_item = int(row_item[0])
            sql_upd_item_ativo = text("""
                UPDATE [Integracao].[Silver].[FatoControleContratosItensEuromidia]
                SET
                    IDFatoControleContratoEuromidia = :id_contrato,
                    Status = 'ATIVO',
                    Origem = 'CONTRATO',
                    DataCancelamento = NULL,
                    AtivoCancelamento = NULL,
                    DataAtualizacao = GETDATE(),
                    IDPainelEuromidia = :id_painel,
                    IDDimFacesPaineis = :id_dim_faces
                WHERE IDFatoControleContratosItensEuromidia = :id_item
            """)
            db.session.execute(sql_upd_item_ativo, {
                "id_contrato": id_contrato,
                "id_item": id_item,
                "id_painel": id_painel,
                "id_dim_faces": id_dim_faces_int
            })
        else:

            sql_ins_item = text("""
                INSERT INTO [Integracao].[Silver].[FatoControleContratosItensEuromidia] (
                    IDFatoControleContratoEuromidia,
                    DataAtualizacao,
                    Referencia,
                    NumeroContrato,
                    NumeroPrevia,
                    CodPonto,
                    CodFace,
                    DataLancamento,
                    Cota,
                    Origem,
                    RazaoSocial,
                    MarcaExibida,
                    Vendedor,
                    IDVendedor,
                    IDPainelEuromidia,
                    IDDimFacesPaineis,
                    DataInicioPrevisto,
                    DataTerminoPrevisto,
                    Status
                )
                VALUES (
                    :id_contrato,
                    GETDATE(),
                    :ref,
                    :numero_contrato,
                    :numero_previa,
                    :codponto,
                    :codface,
                    CAST(GETDATE() AS date),
                    :cota,
                    'CONTRATO',
                    :razao_social,
                    :marca_exibida,
                    :vendedor,
                    :id_vendedor,
                    :id_painel,
                    :id_dim_faces,
                    :dt_ini,
                    :dt_fim,
                    'ATIVO'
                );
            """)
            db.session.execute(sql_ins_item, {
                "id_contrato": id_contrato,
                "ref": referencia,
                "numero_contrato": numero_contrato,
                "numero_previa": numero_previa,
                "codponto": codponto,
                "codface": codface,
                "cota": cota,
                "razao_social": razao_social if razao_social else None,
                "marca_exibida": marca_exibida,
                "vendedor": vendedor,
                "id_vendedor": id_vendedor,
                "id_painel": id_painel,
                "id_dim_faces": id_dim_faces_int,
                "dt_ini": data_inicio,
                "dt_fim": data_fim
            })

        sql_set_fk = text("""
            UPDATE [Integracao].[Silver].[FatoOcupacaoPaineisEuromidia]
            SET IDFatoControleContratos = :id_contrato
            WHERE IDFatoOcupacaoPaineisEuromidia = :id_ocup
        """)
        db.session.execute(sql_set_fk, {
            "id_contrato": id_contrato,
            "id_ocup": id_ocupacao
        })

    if status_novo == "CANCELADO":

        sql_get_header2 = text("""
            SELECT TOP 1 IDFatoControleContratosEuromidia
            FROM [Integracao].[Silver].[FatoControleContratosEuromidia]
            WHERE Referencia = :ref
            ORDER BY IDFatoControleContratosEuromidia DESC
        """)
        row_header2 = db.session.execute(sql_get_header2, {"ref": referencia}).fetchone()

        if row_header2:
            id_contrato2 = int(row_header2[0])

            sql_cancel_item = text("""
                UPDATE [Integracao].[Silver].[FatoControleContratosItensEuromidia]
                SET
                    Status = 'CANCELADO',
                    DataCancelamento = CAST(GETDATE() AS date),
                    AtivoCancelamento = 'S',
                    DataAtualizacao = GETDATE()
                WHERE
                    IDFatoControleContratoEuromidia = :id_contrato
                    AND ISNULL(CodPonto, -1) = :codponto
                    AND LTRIM(RTRIM(ISNULL(CodFace,''))) = :codface
                    AND DataInicioPrevisto = :dt_ini
                    AND DataTerminoPrevisto = :dt_fim
            """)
            db.session.execute(sql_cancel_item, {
                "id_contrato": id_contrato2,
                "codponto": codponto,
                "codface": codface,
                "dt_ini": data_inicio,
                "dt_fim": data_fim
            })

    db.session.commit()
    return redirect(url_for("Paineis.ocupacao_detalhe", id_ocupacao=id_ocupacao))














def _localizar_arquivo_imagem(pasta: Path) -> Path:
    extensoes = ("*.jpg", "*.jpeg", "*.png", "*.JPG", "*.JPEG", "*.PNG")

    arquivos = []
    for ext in extensoes:
        arquivos.extend(pasta.glob(ext))

    if not arquivos:
        raise FileNotFoundError(f"Nenhuma imagem encontrada em: {pasta}")

    arquivos = [arq for arq in arquivos if arq.is_file()]
    if not arquivos:
        raise FileNotFoundError(f"Nenhum arquivo de imagem válido encontrado em: {pasta}")

    return sorted(arquivos)[0]







def _normalizar_texto_checking(valor) -> str:
    try:
        return str(valor or "").strip()
    except Exception:
        return ""


def _somente_digitos(valor) -> str:
    return "".join(ch for ch in _normalizar_texto_checking(valor) if ch.isdigit())


def _garantir_pasta(caminho: Path) -> None:
    caminho.mkdir(parents=True, exist_ok=True)





def _obter_pasta_base_checking() -> Path:
    """Eu resolvo a pasta raiz do checking usando config quando existir e um padrão seguro quando não existir."""
    candidatos = [
        current_app.config.get("PASTA_BASE_CHECKING"),
        current_app.config.get("CHECKING_PASTA_BASE"),
        current_app.config.get("DIRETORIO_BASE_CHECKING"),
        "/home/guilherme_correa/PythonJobs/pipelines/FlaskApp/chekin/pontos",
    ]

    for candidato in candidatos:
        texto = str(candidato or "").strip()
        if texto:
            return Path(texto)

    raise RuntimeError("Não foi possível resolver a pasta base do checking.")


def _normalizar_extensao_arquivo(nome_arquivo: str) -> str:
    """Eu extraio a extensão do arquivo e devolvo sempre em minúsculo."""
    nome_arquivo = str(nome_arquivo or "").strip()
    if "." not in nome_arquivo:
        return ".jpg"
    return "." + nome_arquivo.rsplit(".", 1)[1].lower()


def _localizar_fundo_checking(cod_ponto: str, cod_face: str) -> Path:
    """Eu localizo o fundo oficial da face procurando primeiro na pasta fundo e depois na pasta da face."""
    pasta_base = _obter_pasta_base_checking()
    pasta_face = pasta_base / str(cod_ponto) / str(cod_face)

    if not pasta_face.exists():
        raise FileNotFoundError(f"Pasta da face não encontrada: {pasta_face}")

    pasta_fundo = pasta_face / "fundo"
    if pasta_fundo.exists() and pasta_fundo.is_dir():
        return _localizar_arquivo_imagem(pasta_fundo)

    return _localizar_arquivo_imagem(pasta_face)



def _salvar_upload_original_checking(
    *,
    arquivo,
    pasta_destino: Path,
    id_empresa: int,
    cod_face: str,
    data_checking: date,
) -> tuple[str, Path]:
    """
    Eu salvo a imagem importada dentro da pasta da empresa.

    Regra:
    - não crio pasta uploads
    - salvo em /pontos/<codponto>/<codface>/<IDEmpresa>/
    - o nome salvo fica no padrão:
      cod_face_data_idempresa_upload.ext
    """

    nome_original = _normalizar_texto_checking(getattr(arquivo, "filename", ""))
    nome_seguro = secure_filename(nome_original) if nome_original else ""
    extensao = Path(nome_seguro).suffix.lower()

    if not extensao:
        extensao = ".jpg"

    nome_base = f"{cod_face}_{data_checking.strftime('%Y%m%d')}_{int(id_empresa)}_upload"
    caminho_saida = pasta_destino / f"{nome_base}{extensao}"

    contador = 1
    while caminho_saida.exists():
        caminho_saida = pasta_destino / f"{nome_base}_{contador:02d}{extensao}"
        contador += 1

    arquivo.stream.seek(0)
    arquivo.save(caminho_saida)

    return nome_original, caminho_saida



def _montar_mockup_checking(
    *,
    caminho_fundo: Path,
    caminho_arte: Path,
    caminho_saida: Path,
) -> None:
    """Eu monto o mockup centralizando a arte sobre o fundo com redimensionamento proporcional."""
    fundo = Image.open(caminho_fundo).convert("RGBA")
    arte = Image.open(caminho_arte).convert("RGBA")

    largura_fundo, altura_fundo = fundo.size

    margem_lateral = int(largura_fundo * 0.08)
    margem_superior = int(altura_fundo * 0.08)

    largura_util = max(1, largura_fundo - (margem_lateral * 2))
    altura_util = max(1, altura_fundo - (margem_superior * 2))

    arte_ajustada = ImageOps.contain(arte, (largura_util, altura_util))

    posicao_x = int((largura_fundo - arte_ajustada.width) / 2)
    posicao_y = int((altura_fundo - arte_ajustada.height) / 2)

    composicao = fundo.copy()
    composicao.paste(arte_ajustada, (posicao_x, posicao_y), arte_ajustada)

    _garantir_pasta(caminho_saida.parent)
    composicao.convert("RGB").save(caminho_saida, quality=95)


def _buscar_razao_social_empresa(id_empresa: int | None) -> str:
    """Eu busco a razão social da empresa para gravar junto com o checking."""
    if not id_empresa:
        return ""

    sql = text("""
        SELECT TOP (1)
            RazaoSocial = NULLIF(LTRIM(RTRIM(COALESCE(RazaoSocial, NomeFantasia, ''))), '')
        FROM [Integracao].[Silver].[DimEmpresas]
        WHERE IDEmpresa = :id_empresa
    """)

    row = db.session.execute(sql, {"id_empresa": id_empresa}).mappings().first()
    if not row:
        return ""

    return _normalizar_texto_checking(row["RazaoSocial"])


def _processar_upload_checking(
    *,
    id_empresa: int | None,
    id_fato_controle_contratos: int,
    cod_ponto: str,
    cod_face: str,
    data_checking: date,
    arquivo,
    cnpj_digitado: str | None = None,
    observacao: str | None = None,
):
    cod_ponto_txt = _somente_digitos(cod_ponto)
    cod_face_txt = _normalizar_texto_checking(cod_face).upper()
    observacao_txt = _normalizar_texto_checking(observacao) or None
    cnpj_txt = _somente_digitos(cnpj_digitado) or None

    if not cod_ponto_txt:
        raise ValueError("CodPonto não informado.")

    if not cod_face_txt:
        raise ValueError("CodFace não informado.")

    if not id_empresa:
        raise ValueError("Empresa não informada.")

    if not id_fato_controle_contratos:
        raise ValueError("Contrato não informado.")

    cod_ponto_int = int(cod_ponto_txt)
    id_empresa_int = int(id_empresa)

    pasta_face, caminho_imagem_painel, caminho_fundo = _buscar_paths_mockup(
        cod_ponto_int,
        cod_face_txt,
    )

    pasta_empresa = pasta_face / str(id_empresa_int)
    _garantir_pasta(pasta_empresa)

    nome_original_cliente, caminho_imagem_upload = _salvar_upload_original_checking(
        arquivo=arquivo,
        pasta_destino=pasta_empresa,
        id_empresa=id_empresa_int,
        cod_face=cod_face_txt,
        data_checking=data_checking,
    )

    nome_base_gerado = f"{cod_face_txt}_{data_checking.strftime('%Y%m%d')}_gerado_{id_empresa_int}"
    caminho_imagem_gerada = pasta_empresa / f"{nome_base_gerado}.jpg"

    contador = 1
    while caminho_imagem_gerada.exists():
        caminho_imagem_gerada = pasta_empresa / f"{nome_base_gerado}_{contador:02d}.jpg"
        contador += 1

    _gerar_mockup_checking(
        caminho_fundo=caminho_fundo,
        caminho_imagem_upload=caminho_imagem_upload,
        caminho_saida=caminho_imagem_gerada,
    )

    razao_social = None
    if id_empresa:
        sql_empresa = text("""
            SELECT TOP (1)
                CNPJ = NULLIF(LTRIM(RTRIM(CAST(CNPJ AS varchar(30)))), ''),
                RazaoSocial = NULLIF(LTRIM(RTRIM(CAST(COALESCE(RazaoSocial, NomeFantasia, '') AS nvarchar(200)))), '')
            FROM [Integracao].[Silver].[DimEmpresas]
            WHERE IDEmpresa = :id_empresa
        """)
        row_empresa = db.session.execute(
            sql_empresa,
            {"id_empresa": id_empresa_int}
        ).mappings().first()

        if row_empresa:
            if not cnpj_txt:
                cnpj_txt = _somente_digitos(row_empresa["CNPJ"]) or None
            razao_social = _normalizar_texto_checking(row_empresa["RazaoSocial"]) or None

    tipo_painel = None
    tipo_face = None

    try:
        row_face = (
            db.session.query(DimFacesPaineis.TipoPainel)
            .filter(
                DimFacesPaineis.CodPonto == cod_ponto_int,
                DimFacesPaineis.CodFace == cod_face_txt,
            )
            .first()
        )

        if row_face:
            tipo_painel = _normalizar_texto_checking(row_face[0]) or None
    except Exception:
        tipo_painel = None

    id_usuario_criacao = None
    if current_user.is_authenticated:
        try:
            id_usuario_criacao = int(current_user.get_id())
        except Exception:
            id_usuario_criacao = None

    checking = DimCheckingHistorico()
    checking.DataAtualizacao = datetime.now()
    checking.DataChecking = data_checking
    checking.IDEmpresa = id_empresa_int
    checking.CNPJ = cnpj_txt
    checking.RazaoSocial = razao_social
    checking.IDFatoControleContratosEuromidia = int(id_fato_controle_contratos)
    checking.CodPonto = cod_ponto_int
    checking.CodFace = cod_face_txt
    checking.TipoPainel = tipo_painel
    checking.TipoFace = tipo_face
    checking.NomeArquivoOriginal = nome_original_cliente or None
    checking.NomeArquivoSalvo = caminho_imagem_upload.name
    checking.CaminhoImagemPainel = str(caminho_imagem_painel)
    checking.CaminhoImagemFundo = str(caminho_fundo)
    checking.CaminhoImagemUpload = str(caminho_imagem_upload)
    checking.CaminhoImagemGerada = str(caminho_imagem_gerada)
    checking.UrlImagemUpload = None
    checking.UrlImagemGerada = None
    checking.BitChekin = False
    checking.DataConfirmacao = None
    checking.IDUsuarioCriacao = id_usuario_criacao
    checking.IDUsuarioConfirmacao = None
    checking.Observacao = observacao_txt

    db.session.add(checking)
    db.session.flush()

    checking.UrlImagemGerada = url_for(
        "Paineis.checking_arquivo",
        id_checking=int(checking.IDDimCheckingHistorico),
    )

    db.session.commit()
    db.session.refresh(checking)

    return checking



def _buscar_dados_select_checking(id_empresa: int | None = None):
    sql_empresas = text("""
        SELECT
            TOP (500)
            e.IDEmpresa,
            CNPJ = NULLIF(LTRIM(RTRIM(CAST(e.CNPJ AS varchar(30)))), ''),
            RazaoSocial = NULLIF(LTRIM(RTRIM(COALESCE(e.RazaoSocial, e.NomeFantasia, ''))), ''),
            NomeFantasia = NULLIF(LTRIM(RTRIM(COALESCE(e.NomeFantasia, ''))), '')
        FROM [Integracao].[Silver].[DimEmpresas] e
        WHERE NULLIF(LTRIM(RTRIM(CAST(e.CNPJ AS varchar(30)))), '') IS NOT NULL
        ORDER BY
            NULLIF(LTRIM(RTRIM(COALESCE(e.RazaoSocial, e.NomeFantasia, ''))), ''),
            NULLIF(LTRIM(RTRIM(CAST(e.CNPJ AS varchar(30)))), '')
    """)

    sql_pontos = text("""
        SELECT
            TOP (1000)
            p.CodPonto,
            TipoPainel = NULLIF(LTRIM(RTRIM(COALESCE(p.Tipo, ''))), '')
        FROM [Integracao].[Silver].[DimPaineisEuromidia] p
        WHERE p.CodPonto IS NOT NULL
        ORDER BY p.CodPonto ASC
    """)

    sql_faces = text("""
        SELECT
            TOP (2000)
            f.IDDimFacesPaineis,
            f.CodPonto,
            f.CodFace,
            TipoFace = NULLIF(LTRIM(RTRIM(COALESCE(f.Tipo, ''))), ''),
            TipoPainel = NULLIF(LTRIM(RTRIM(COALESCE(p.Tipo, ''))), '')
        FROM [Integracao].[Silver].[DimFacesPaineis] f
        LEFT JOIN [Integracao].[Silver].[DimPaineisEuromidia] p
            ON p.IDDimPaineisEuromidia = f.IDDimPaineisEuromidia
        WHERE f.CodPonto IS NOT NULL
          AND NULLIF(LTRIM(RTRIM(COALESCE(f.CodFace, ''))), '') IS NOT NULL
        ORDER BY f.CodPonto ASC, f.CodFace ASC
    """)

    empresas_rows = db.session.execute(sql_empresas).mappings().all()
    pontos_rows = db.session.execute(sql_pontos).mappings().all()
    faces_rows = db.session.execute(sql_faces).mappings().all()

    empresas = []
    for row in empresas_rows:
        id_empresa_linha = row["IDEmpresa"]
        cnpj = _somente_digitos(row["CNPJ"])
        razao = _normalizar_texto_checking(row["RazaoSocial"])
        nome_fantasia = _normalizar_texto_checking(row["NomeFantasia"])

        nome_exibicao = razao or nome_fantasia or "Sem nome cadastrado"

        empresas.append({
            "IDEmpresa": id_empresa_linha,
            "CNPJ": cnpj,
            "RazaoSocial": razao,
            "NomeFantasia": nome_fantasia,
            "texto": f"{cnpj} | {nome_exibicao}",
        })

    pontos = []
    for row in pontos_rows:
        cod_ponto = row["CodPonto"]
        tipo_painel = _normalizar_texto_checking(row["TipoPainel"]) or "Sem tipo"

        pontos.append({
            "CodPonto": cod_ponto,
            "TipoPainel": tipo_painel,
            "texto": f"{cod_ponto} | {tipo_painel}",
        })

    faces = []
    for row in faces_rows:
        cod_ponto = row["CodPonto"]
        cod_face = _normalizar_texto_checking(row["CodFace"])
        tipo_face = _normalizar_texto_checking(row["TipoFace"]) or _normalizar_texto_checking(row["TipoPainel"]) or "Sem tipo"

        faces.append({
            "IDDimFacesPaineis": row["IDDimFacesPaineis"],
            "CodPonto": cod_ponto,
            "CodFace": cod_face,
            "TipoFace": tipo_face,
            "TipoPainel": _normalizar_texto_checking(row["TipoPainel"]),
            "texto_face": f"{cod_face} | {tipo_face}",
        })

    return empresas, [], pontos, faces




@paineis_bp.route("/checking/empresas/buscar", methods=["GET"])
@login_required
@retry_get_view(db, attempts=2, base_delay=0.2, max_delay=0.8)
def checking_empresas_buscar():
    q = _normalizar_texto_checking(request.args.get("q"))
    q_digitos = _somente_digitos(q)

    if not q:
        return jsonify([])

    def formatar_cnpj(cnpj: str) -> str:
        cnpj = _somente_digitos(cnpj)
        if len(cnpj) != 14:
            return cnpj
        return f"{cnpj[0:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"

    retorno = []

    if q_digitos:
        q_cnpj_puro = q_digitos
        q_cnpj_mascarado = formatar_cnpj(q_digitos)

        parametros = {
            "q": q,
            "q_like": f"%{q}%",
            "q_cnpj_puro": q_cnpj_puro,
            "q_cnpj_mascarado": q_cnpj_mascarado,
            "q_cnpj_puro_prefixo": f"{q_cnpj_puro}%",
            "q_cnpj_mascarado_prefixo": f"{q_cnpj_mascarado}%",
        }

        sql = text("""
            SELECT TOP (20)
                e.IDEmpresa,
                CNPJ = NULLIF(LTRIM(RTRIM(CAST(e.CNPJ AS varchar(30)))), ''),
                RazaoSocial = NULLIF(LTRIM(RTRIM(COALESCE(e.RazaoSocial, ''))), ''),
                NomeFantasia = NULLIF(LTRIM(RTRIM(COALESCE(e.NomeFantasia, ''))), '')
            FROM [Integracao].[Silver].[DimEmpresas] e
            WHERE
                NULLIF(LTRIM(RTRIM(CAST(e.CNPJ AS varchar(30)))), '') IS NOT NULL
                AND
                (
                    CAST(e.CNPJ AS varchar(30)) = :q_cnpj_puro
                    OR CAST(e.CNPJ AS varchar(30)) = :q_cnpj_mascarado
                    OR CAST(e.CNPJ AS varchar(30)) LIKE :q_cnpj_puro_prefixo
                    OR CAST(e.CNPJ AS varchar(30)) LIKE :q_cnpj_mascarado_prefixo
                )
            ORDER BY
                CASE
                    WHEN CAST(e.CNPJ AS varchar(30)) = :q_cnpj_puro THEN 0
                    WHEN CAST(e.CNPJ AS varchar(30)) = :q_cnpj_mascarado THEN 1
                    WHEN CAST(e.CNPJ AS varchar(30)) LIKE :q_cnpj_puro_prefixo THEN 2
                    WHEN CAST(e.CNPJ AS varchar(30)) LIKE :q_cnpj_mascarado_prefixo THEN 3
                    ELSE 4
                END,
                RazaoSocial,
                NomeFantasia,
                CNPJ
        """)
    else:
        parametros = {
            "q": q,
            "q_like": f"%{q}%",
        }

        sql = text("""
            SELECT TOP (20)
                e.IDEmpresa,
                CNPJ = NULLIF(LTRIM(RTRIM(CAST(e.CNPJ AS varchar(30)))), ''),
                RazaoSocial = NULLIF(LTRIM(RTRIM(COALESCE(e.RazaoSocial, ''))), ''),
                NomeFantasia = NULLIF(LTRIM(RTRIM(COALESCE(e.NomeFantasia, ''))), '')
            FROM [Integracao].[Silver].[DimEmpresas] e
            WHERE
                (
                    COALESCE(e.RazaoSocial, '') LIKE :q_like
                    OR COALESCE(e.NomeFantasia, '') LIKE :q_like
                )
            ORDER BY
                CASE
                    WHEN COALESCE(e.RazaoSocial, '') = :q THEN 0
                    WHEN COALESCE(e.NomeFantasia, '') = :q THEN 1
                    ELSE 2
                END,
                RazaoSocial,
                NomeFantasia,
                CNPJ
        """)

    rows = db.session.execute(sql, parametros).mappings().all()

    for row in rows:
        id_empresa = row["IDEmpresa"]
        cnpj = _normalizar_texto_checking(row["CNPJ"])
        cnpj_digitos = _somente_digitos(cnpj)
        razao = _normalizar_texto_checking(row["RazaoSocial"])
        nome_fantasia = _normalizar_texto_checking(row["NomeFantasia"])

        nome_exibicao = razao or nome_fantasia or "Sem nome cadastrado"

        retorno.append({
            "id": int(id_empresa),
            "cnpj": cnpj,
            "cnpj_digitos": cnpj_digitos,
            "razao_social": razao,
            "nome_fantasia": nome_fantasia,
            "texto": f"{cnpj} | {nome_exibicao}",
        })

    return jsonify(retorno)





def _buscar_paths_mockup(cod_ponto: int, cod_face: str) -> tuple[Path, Path, Path]:
    pasta_raiz = Path("/home/guilherme_correa/PythonJobs/pipelines/FlaskApp/chekin/pontos")
    pasta_ponto = pasta_raiz / str(cod_ponto)
    pasta_face = pasta_ponto / str(cod_face)
    pasta_fundo = pasta_ponto / "fundo"

    if not pasta_ponto.exists():
        raise FileNotFoundError(f"Pasta do ponto não encontrada: {pasta_ponto}")

    if not pasta_face.exists():
        raise FileNotFoundError(f"Pasta da face não encontrada: {pasta_face}")

    if not pasta_fundo.exists():
        raise FileNotFoundError(f"Pasta do fundo não encontrada: {pasta_fundo}")

    caminho_imagem_face = _localizar_arquivo_imagem(pasta_face)
    caminho_fundo = _localizar_arquivo_imagem(pasta_fundo)

    return pasta_face, caminho_imagem_face, caminho_fundo



def _gerar_mockup_checking(
    caminho_fundo: Path,
    caminho_imagem_upload: Path,
    caminho_saida: Path,
) -> None:
    """
    Eu gero o mockup final colocando a arte enviada em cima do fundo padrão.

    Lógica:
    - abro a imagem de fundo
    - abro a imagem enviada pelo usuário
    - ajusto a imagem exatamente na área útil definida
    - colo a imagem ajustada sobre o fundo
    - salvo o resultado final
    """

    fundo = Image.open(caminho_fundo).convert("RGB")
    imagem = Image.open(caminho_imagem_upload).convert("RGB")

    x1 = 430
    y1 = 35
    x2 = 1230
    y2 = 605

    largura_area = x2 - x1
    altura_area = y2 - y1

    imagem_ajustada = ImageOps.fit(
        imagem,
        (largura_area, altura_area),
        method=Image.Resampling.LANCZOS,
        centering=(0.5, 0.5),
    )

    fundo.paste(imagem_ajustada, (x1, y1))

    _garantir_pasta(caminho_saida.parent)
    fundo.save(caminho_saida, quality=95)










def _checking_item_pertence_ao_contrato(
    *,
    id_fato_controle_contratos: int,
    cod_ponto: str,
    cod_face: str,
) -> bool:
    sql = text("""
        SELECT TOP (1) 1 AS Existe
        FROM [Integracao].[Silver].[FatoControleContratosItensEuromidia] i
        WHERE
            i.IDFatoControleContratoEuromidia = :id_fato_controle_contratos
            AND TRY_CONVERT(int, i.CodPonto) = TRY_CONVERT(int, :cod_ponto)
            AND UPPER(LTRIM(RTRIM(CAST(i.CodFace AS varchar(50))))) = UPPER(LTRIM(RTRIM(:cod_face)))
    """)

    row = db.session.execute(
        sql,
        {
            "id_fato_controle_contratos": int(id_fato_controle_contratos),
            "cod_ponto": str(cod_ponto or "").strip(),
            "cod_face": str(cod_face or "").strip(),
        },
    ).mappings().first()

    return bool(row)














@paineis_bp.route("/checking/contratos/<int:id_fato_controle_contratos>/pontos", methods=["GET"])
@login_required
@retry_get_view(db, attempts=2, base_delay=0.2, max_delay=0.8)
def checking_pontos_do_contrato(id_fato_controle_contratos: int):
    sql = text("""
        SELECT TOP (300)
            CodPonto = TRY_CONVERT(int, i.CodPonto),
            TipoPainel = MAX(
                NULLIF(
                    LTRIM(RTRIM(COALESCE(p.Tipo, ''))),
                    ''
                )
            )
        FROM [Integracao].[Silver].[FatoControleContratosItensEuromidia] i
        LEFT JOIN [Integracao].[Silver].[DimPaineisEuromidia] p
            ON TRY_CONVERT(int, p.CodPonto) = TRY_CONVERT(int, i.CodPonto)
        WHERE
            i.IDFatoControleContratoEuromidia = :id_fato_controle_contratos
            AND TRY_CONVERT(int, i.CodPonto) IS NOT NULL
        GROUP BY TRY_CONVERT(int, i.CodPonto)
        ORDER BY TRY_CONVERT(int, i.CodPonto) ASC
    """)

    rows = db.session.execute(
        sql,
        {"id_fato_controle_contratos": int(id_fato_controle_contratos)},
    ).mappings().all()

    itens = []
    for row in rows:
        cod_ponto = row["CodPonto"]
        if cod_ponto is None:
            continue

        tipo_painel = _normalizar_texto_checking(row["TipoPainel"]) or "Sem tipo"

        itens.append(
            {
                "cod_ponto": int(cod_ponto),
                "tipo_painel": tipo_painel,
                "texto": f"{int(cod_ponto)} | {tipo_painel}",
            }
        )

    return jsonify({
        "id_fato_controle_contratos": int(id_fato_controle_contratos),
        "items": itens,
    })






@paineis_bp.route("/checking/contratos/<int:id_fato_controle_contratos>/pontos/<int:codponto>/faces", methods=["GET"])
@login_required
@retry_get_view(db, attempts=2, base_delay=0.2, max_delay=0.8)
def checking_faces_do_contrato(id_fato_controle_contratos: int, codponto: int):
    sql = text("""
        SELECT TOP (300)
            CodFace = UPPER(LTRIM(RTRIM(CAST(i.CodFace AS varchar(50))))),
            TipoFace = MAX(
                NULLIF(
                    LTRIM(RTRIM(COALESCE(f.Tipo, p.Tipo, ''))),
                    ''
                )
            )
        FROM [Integracao].[Silver].[FatoControleContratosItensEuromidia] i
        LEFT JOIN [Integracao].[Silver].[DimFacesPaineis] f
            ON TRY_CONVERT(int, f.CodPonto) = TRY_CONVERT(int, i.CodPonto)
            AND UPPER(LTRIM(RTRIM(CAST(f.CodFace AS varchar(50))))) = UPPER(LTRIM(RTRIM(CAST(i.CodFace AS varchar(50)))))
        LEFT JOIN [Integracao].[Silver].[DimPaineisEuromidia] p
            ON p.IDDimPaineisEuromidia = f.IDDimPaineisEuromidia
            OR TRY_CONVERT(int, p.CodPonto) = TRY_CONVERT(int, i.CodPonto)
        WHERE
            i.IDFatoControleContratoEuromidia = :id_fato_controle_contratos
            AND TRY_CONVERT(int, i.CodPonto) = :codponto
            AND NULLIF(LTRIM(RTRIM(CAST(i.CodFace AS varchar(50)))), '') IS NOT NULL
        GROUP BY UPPER(LTRIM(RTRIM(CAST(i.CodFace AS varchar(50)))))
        ORDER BY UPPER(LTRIM(RTRIM(CAST(i.CodFace AS varchar(50))))) ASC
    """)

    rows = db.session.execute(
        sql,
        {
            "id_fato_controle_contratos": int(id_fato_controle_contratos),
            "codponto": int(codponto),
        },
    ).mappings().all()

    itens = []
    for row in rows:
        cod_face = _normalizar_texto_checking(row["CodFace"])
        if not cod_face:
            continue

        tipo_face = _normalizar_texto_checking(row["TipoFace"]) or "Sem tipo"

        itens.append(
            {
                "cod_face": cod_face,
                "tipo_face": tipo_face,
                "texto": f"{cod_face} | {tipo_face}",
            }
        )

    return jsonify({
        "id_fato_controle_contratos": int(id_fato_controle_contratos),
        "codponto": int(codponto),
        "items": itens,
    })






@paineis_bp.route("/checking/novo", methods=["GET", "POST"])
@login_required
@limiter.limit("25 per minute", methods=["POST"])
@retry_get_view(db, attempts=2, base_delay=0.2, max_delay=0.8)
def checking_novo():
    extensoes_permitidas = current_app.config["EXTENSOES_PERMITIDAS_CHECKING"]
    largura_maxima = current_app.config["LARGURA_MAXIMA_IMAGEM"]
    altura_maxima = current_app.config["ALTURA_MAXIMA_IMAGEM"]

    hoje = datetime.now().date()

    if request.method == "GET":
        return render_template(
            "euromidia/checking_upload.html",
            hoje=hoje,
            checking=None,
        )

    id_empresa_txt = _normalizar_texto_checking(request.form.get("id_empresa"))

    try:
        id_empresa = int(id_empresa_txt)
    except Exception:
        flash("Selecione uma empresa válida.", "warning")
        return render_template(
            "euromidia/checking_upload.html",
            hoje=hoje,
            checking=None,
        )

    arquivo = request.files.get("imagem")
    if not arquivo or not arquivo.filename:
        flash("Selecione uma imagem para upload.", "warning")
        return render_template(
            "euromidia/checking_upload.html",
            hoje=hoje,
            checking=None,
        )

    nome_arquivo = str(arquivo.filename or "").lower().strip()
    if "." not in nome_arquivo or nome_arquivo.rsplit(".", 1)[1] not in extensoes_permitidas:
        flash("Extensão de arquivo inválida.", "danger")
        return render_template(
            "euromidia/checking_upload.html",
            hoje=hoje,
            checking=None,
        )

    try:
        id_fato_controle_contratos = int(request.form.get("id_fato_controle_contratos"))
    except Exception:
        flash("Selecione um contrato válido.", "warning")
        return render_template(
            "euromidia/checking_upload.html",
            hoje=hoje,
            checking=None,
        )

    cod_ponto = _normalizar_texto_checking(request.form.get("cod_ponto"))
    cod_face = _normalizar_texto_checking(request.form.get("cod_face"))
    observacao = _normalizar_texto_checking(request.form.get("observacao"))
    data_checking_txt = _normalizar_texto_checking(request.form.get("data_checking"))

    cnpj_digitado = None

    if not cod_ponto:
        flash("Selecione um ponto válido.", "warning")
        return render_template(
            "euromidia/checking_upload.html",
            hoje=hoje,
            checking=None,
        )

    if not cod_face:
        flash("Selecione uma face válida.", "warning")
        return render_template(
            "euromidia/checking_upload.html",
            hoje=hoje,
            checking=None,
        )

    if not _checking_item_pertence_ao_contrato(
        id_fato_controle_contratos=id_fato_controle_contratos,
        cod_ponto=cod_ponto,
        cod_face=cod_face,
    ):
        flash("O CodPonto e o CodFace selecionados não pertencem ao contrato informado.", "danger")
        return render_template(
            "euromidia/checking_upload.html",
            hoje=hoje,
            checking=None,
        )

    try:
        data_checking = datetime.strptime(data_checking_txt, "%Y-%m-%d").date()
    except Exception:
        flash("Data de checking inválida.", "warning")
        return render_template(
            "euromidia/checking_upload.html",
            hoje=hoje,
            checking=None,
        )

    try:
        imagem_bytes = arquivo.read()
        imagem = Image.open(BytesIO(imagem_bytes))
        largura, altura = imagem.size

        if largura > largura_maxima or altura > altura_maxima:
            flash(
                f"Imagem maior que o permitido. Máximo: {largura_maxima}x{altura_maxima}px.",
                "danger"
            )
            return render_template(
                "euromidia/checking_upload.html",
                hoje=hoje,
                checking=None,
            )

        arquivo.stream.seek(0)

        checking = _processar_upload_checking(
            id_empresa=id_empresa,
            id_fato_controle_contratos=id_fato_controle_contratos,
            cod_ponto=cod_ponto,
            cod_face=cod_face,
            data_checking=data_checking,
            arquivo=arquivo,
            cnpj_digitado=cnpj_digitado,
            observacao=observacao,
        )

        flash("Upload realizado e mockup gerado com sucesso.", "success")

        return render_template(
            "euromidia/checking_upload.html",
            hoje=hoje,
            checking=checking,
        )

    except Exception as e:
        db.session.rollback()
        current_app.logger.exception("[CHECKING] erro ao processar upload")
        flash(f"Falha ao processar checking: {str(e)}", "danger")

        return render_template(
            "euromidia/checking_upload.html",
            hoje=hoje,
            checking=None,
        )




@paineis_bp.route("/checking/contratos/buscar", methods=["GET"])
@login_required
def checking_contratos():
    id_empresa_txt = _normalizar_texto_checking(request.args.get("id_empresa"))
    cnpj_txt = _somente_digitos(request.args.get("cnpj"))
    q = _normalizar_texto_checking(request.args.get("q"))
    q_digitos = _somente_digitos(q)

    id_empresa = None
    try:
        if id_empresa_txt:
            id_empresa = int(id_empresa_txt)
    except Exception:
        id_empresa = None

    cnpj_empresa = cnpj_txt

    if not cnpj_empresa and id_empresa is not None:
        sql_empresa = text("""
            SELECT TOP (1)
                CNPJ = NULLIF(LTRIM(RTRIM(CAST(CNPJ AS varchar(30)))), '')
            FROM [Integracao].[Silver].[DimEmpresas]
            WHERE IDEmpresa = :id_empresa
        """)
        row_empresa = db.session.execute(
            sql_empresa,
            {"id_empresa": id_empresa}
        ).mappings().first()

        if row_empresa:
            cnpj_empresa = _somente_digitos(row_empresa["CNPJ"])

    contratos = []

    sql_cabecalho = text("""
        SELECT DISTINCT TOP (200)
            c.IDFatoControleContratosEuromidia,
            NumeroContrato = NULLIF(LTRIM(RTRIM(CAST(c.NumeroContrato AS varchar(50)))), ''),
            NumeroPrevia = NULLIF(LTRIM(RTRIM(CAST(c.NumeroPrevia AS varchar(50)))), ''),
            RazaoSocial = NULLIF(LTRIM(RTRIM(COALESCE(c.RazaoSocial, c.MarcaExibida, ''))), ''),
            CNPJ = NULLIF(LTRIM(RTRIM(CAST(c.CNPJ AS varchar(30)))), ''),
            c.IDEmpresa
        FROM [Integracao].[Silver].[FatoControleContratosEuromidia] c
        WHERE
            (
                (:id_empresa IS NOT NULL AND c.IDEmpresa = :id_empresa)
                OR
                (:cnpj IS NOT NULL AND REPLACE(REPLACE(REPLACE(CAST(c.CNPJ AS varchar(30)), '.', ''), '/', ''), '-', '') = :cnpj)
            )
            AND
            (
                :q = ''
                OR CAST(c.IDFatoControleContratosEuromidia AS varchar(50)) LIKE :q_like
                OR CAST(c.NumeroContrato AS varchar(50)) LIKE :q_like
                OR CAST(c.NumeroPrevia AS varchar(50)) LIKE :q_like
                OR COALESCE(c.RazaoSocial, c.MarcaExibida, '') LIKE :q_like
                OR REPLACE(REPLACE(REPLACE(CAST(c.CNPJ AS varchar(30)), '.', ''), '/', ''), '-', '') LIKE :q_digitos_like
            )
        ORDER BY c.IDFatoControleContratosEuromidia DESC
    """)

    contratos = db.session.execute(
        sql_cabecalho,
        {
            "id_empresa": id_empresa,
            "cnpj": cnpj_empresa or None,
            "q": q,
            "q_like": f"%{q}%",
            "q_digitos_like": f"%{q_digitos}%",
        }
    ).mappings().all()

    if not contratos and cnpj_empresa:
        sql_itens = text("""
            SELECT DISTINCT TOP (200)
                i.IDFatoControleContratoEuromidia AS IDFatoControleContratosEuromidia,
                NumeroContrato = NULLIF(LTRIM(RTRIM(CAST(i.NumeroContrato AS varchar(50)))), ''),
                NumeroPrevia = NULLIF(LTRIM(RTRIM(CAST(i.NumeroPrevia AS varchar(50)))), ''),
                RazaoSocial = NULLIF(LTRIM(RTRIM(COALESCE(i.RazaoSocial, i.MarcaExibida, ''))), ''),
                CNPJ = NULLIF(LTRIM(RTRIM(CAST(i.CNPJ AS varchar(30)))), '')
            FROM [Integracao].[Silver].[FatoControleContratosItensEuromidia] i
            WHERE
                REPLACE(REPLACE(REPLACE(CAST(i.CNPJ AS varchar(30)), '.', ''), '/', ''), '-', '') = :cnpj
                AND
                (
                    :q = ''
                    OR CAST(i.IDFatoControleContratoEuromidia AS varchar(50)) LIKE :q_like
                    OR CAST(i.NumeroContrato AS varchar(50)) LIKE :q_like
                    OR CAST(i.NumeroPrevia AS varchar(50)) LIKE :q_like
                    OR COALESCE(i.RazaoSocial, i.MarcaExibida, '') LIKE :q_like
                    OR REPLACE(REPLACE(REPLACE(CAST(i.CNPJ AS varchar(30)), '.', ''), '/', ''), '-', '') LIKE :q_digitos_like
                )
            ORDER BY i.IDFatoControleContratoEuromidia DESC
        """)

        contratos = db.session.execute(
            sql_itens,
            {
                "cnpj": cnpj_empresa,
                "q": q,
                "q_like": f"%{q}%",
                "q_digitos_like": f"%{q_digitos}%",
            }
        ).mappings().all()

    retorno = []
    ids_vistos = set()

    for row in contratos:
        id_contrato = row["IDFatoControleContratosEuromidia"]

        if not id_contrato or id_contrato in ids_vistos:
            continue

        ids_vistos.add(id_contrato)

        numero_contrato = _normalizar_texto_checking(row["NumeroContrato"])
        numero_previa = _normalizar_texto_checking(row["NumeroPrevia"])
        razao = _normalizar_texto_checking(row["RazaoSocial"]) or "Sem razão social"
        cnpj = _somente_digitos(row["CNPJ"])

        complemento = numero_contrato or numero_previa or cnpj or "Sem número"

        retorno.append({
            "id": int(id_contrato),
            "texto": f"{int(id_contrato)} | {razao} | {complemento}",
        })

    return jsonify(retorno)



@paineis_bp.route("/checking/arquivo/<int:id_checking>", methods=["GET"])
@login_required
@retry_get_view(db, attempts=6, base_delay=0.2, max_delay=1.5)
def checking_arquivo(id_checking: int):
    row = (
        db.session.query(DimCheckingHistorico)
        .filter(DimCheckingHistorico.IDDimCheckingHistorico == id_checking)
        .first()
    )

    if not row:
        abort(404)

    caminho = Path(str(row.CaminhoImagemGerada or "").strip())

    if not caminho.exists() or not caminho.is_file():
        abort(404)

    return send_file(caminho)



@paineis_bp.route("/checking/<int:id_checking>/confirmar", methods=["POST"])
@login_required
def checking_confirmar(id_checking: int):
    csrf_token = (
        request.form.get("csrf_token")
        or request.headers.get("X-CSRFToken")
        or request.headers.get("X-CSRF-Token")
    )

    try:
        validate_csrf(csrf_token)
    except Exception:
        flash("CSRF inválido ou ausente.", "danger")
        return redirect(url_for("Paineis.checking_novo"))

    row = (
        db.session.query(DimCheckingHistorico)
        .filter(DimCheckingHistorico.IDDimCheckingHistorico == id_checking)
        .first()
    )

    if not row:
        flash("Checking não encontrado.", "danger")
        return redirect(url_for("Paineis.checking_novo"))

    try:
        row.BitChekin = True
        row.DataConfirmacao = datetime.now()
        row.IDUsuarioConfirmacao = int(current_user.get_id()) if current_user.is_authenticated else None
        db.session.commit()
        flash("Checking confirmado com sucesso.", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Falha ao confirmar checking: {exc}", "danger")

    return redirect(url_for("Paineis.checking_novo"))





"""Lista Checkin"""




def _paginacao_basica(page: int, per_page: int, total: int):
    """Eu monto um objeto simples de paginação para o template."""
    if per_page <= 0:
        per_page = 10

    total_pages = max((total + per_page - 1) // per_page, 1)
    page = max(min(page, total_pages), 1)

    inicio = (page - 1) * per_page + 1 if total > 0 else 0
    fim = min(page * per_page, total) if total > 0 else 0

    return {
        "page": page,
        "per_page": per_page,
        "total": total,
        "total_pages": total_pages,
        "inicio": inicio,
        "fim": fim,
        "has_prev": page > 1,
        "has_next": page < total_pages,
        "prev_page": page - 1 if page > 1 else 1,
        "next_page": page + 1 if page < total_pages else total_pages,
    }




@paineis_bp.route("/checking/lista", methods=["GET"])
@login_required
@retry_get_view(db, attempts=2, base_delay=0.2, max_delay=0.8)
def lista_checkins():
    page = request.args.get("page", default=1, type=int) or 1
    per_page = 10

    sql_total = text("""
        SELECT COUNT(1) AS Total
        FROM [Integracao].[Silver].[DimCheckingHistorico] ch
        LEFT JOIN [Integracao].[Silver].[DimEmpresas] emp
            ON emp.IDEmpresa = ch.IDEmpresa
        LEFT JOIN [Integracao].[Silver].[DimFacesPaineis] fp
            ON fp.CodFace = ch.CodFace
           AND fp.CodPonto = ch.CodPonto
        LEFT JOIN [Integracao].[Silver].[DimPaineisEuromidia] p
            ON p.IDDimPaineisEuromidia = fp.IDDimPaineisEuromidia
    """)

    total = db.session.execute(sql_total).scalar() or 0
    paginacao = _paginacao_basica(page, per_page, total)
    offset = (paginacao["page"] - 1) * paginacao["per_page"]

    sql_lista = text("""
        SELECT
            ch.IDDimCheckingHistorico,
            ch.DataAtualizacao,
            CNPJ = COALESCE(emp.CNPJ, ch.CNPJ),
            RazaoSocial = COALESCE(emp.RazaoSocial, ch.RazaoSocial),
            ch.CodFace,
            Tipo = COALESCE(
                NULLIF(LTRIM(RTRIM(fp.Tipo)), ''),
                NULLIF(LTRIM(RTRIM(p.Tipo)), '')
            )
        FROM [Integracao].[Silver].[DimCheckingHistorico] ch
        LEFT JOIN [Integracao].[Silver].[DimEmpresas] emp
            ON emp.IDEmpresa = ch.IDEmpresa
        LEFT JOIN [Integracao].[Silver].[DimFacesPaineis] fp
            ON fp.CodFace = ch.CodFace
           AND fp.CodPonto = ch.CodPonto
        LEFT JOIN [Integracao].[Silver].[DimPaineisEuromidia] p
            ON p.IDDimPaineisEuromidia = fp.IDDimPaineisEuromidia
        ORDER BY ch.DataAtualizacao DESC
        OFFSET :offset ROWS
        FETCH NEXT :per_page ROWS ONLY
    """)

    resultado = db.session.execute(
        sql_lista,
        {
            "offset": offset,
            "per_page": paginacao["per_page"],
        },
    )

    itens = [dict(linha._mapping) for linha in resultado]

    return render_template(
        "euromidia/lista_checkins.html",
        itens=itens,
        paginacao=paginacao,
    )





@paineis_bp.route("/checking/<int:id_checking>", methods=["GET"])
@login_required
@retry_get_view(db, attempts=2, base_delay=0.2, max_delay=0.8)
def visualizar_checking(id_checking: int):
    sql = text("""
        SELECT TOP (1)
            ch.IDDimCheckingHistorico,
            ch.DataAtualizacao,
            ch.DataChecking,
            ch.IDEmpresa,
            CNPJ = COALESCE(NULLIF(LTRIM(RTRIM(emp.CNPJ)), ''), NULLIF(LTRIM(RTRIM(cc.CNPJ)), ''), NULLIF(LTRIM(RTRIM(ch.CNPJ)), '')),
            RazaoSocial = COALESCE(NULLIF(LTRIM(RTRIM(emp.RazaoSocial)), ''), NULLIF(LTRIM(RTRIM(cc.RazaoSocial)), ''), NULLIF(LTRIM(RTRIM(ch.RazaoSocial)), '')),
            MarcaExibida = NULLIF(LTRIM(RTRIM(cc.MarcaExibida)), ''),
            ch.IDFatoControleContratosEuromidia,
            ch.CodPonto,
            ch.CodFace,
            TipoPainel = COALESCE(NULLIF(LTRIM(RTRIM(p.Tipo)), ''), NULLIF(LTRIM(RTRIM(ch.TipoPainel)), '')),
            TipoFace = COALESCE(NULLIF(LTRIM(RTRIM(fp.Tipo)), ''), NULLIF(LTRIM(RTRIM(ch.TipoFace)), '')),
            fp.IDDimFacesPaineis,
            p.IDDimPaineisEuromidia,
            EnderecoPainel =
                LTRIM(RTRIM(
                    COALESCE(p.Logradouro, '') +
                    CASE WHEN p.Numero IS NOT NULL AND LTRIM(RTRIM(p.Numero)) <> '' THEN ', ' + p.Numero ELSE '' END +
                    CASE WHEN p.Bairro IS NOT NULL AND LTRIM(RTRIM(p.Bairro)) <> '' THEN ' - ' + p.Bairro ELSE '' END +
                    CASE WHEN p.Cidade IS NOT NULL AND LTRIM(RTRIM(p.Cidade)) <> '' THEN ' - ' + p.Cidade ELSE '' END +
                    CASE WHEN p.UF IS NOT NULL AND LTRIM(RTRIM(p.UF)) <> '' THEN '/' + p.UF ELSE '' END
                )),
            Logradouro = p.Logradouro,
            Numero = p.Numero,
            Bairro = p.Bairro,
            Cidade = p.Cidade,
            UF = p.UF,
            CEP = p.CEP,
            Referencia = p.Referencia,
            ch.NomeArquivoOriginal,
            ch.NomeArquivoSalvo,
            ch.CaminhoImagemPainel,
            ch.CaminhoImagemFundo,
            ch.CaminhoImagemUpload,
            ch.CaminhoImagemGerada,
            ch.UrlImagemUpload,
            ch.UrlImagemGerada,
            ch.BitChekin,
            ch.DataConfirmacao,
            ch.IDUsuarioCriacao,
            UsuarioCriacao = COALESCE(NULLIF(LTRIM(RTRIM(uc.NomeUsuario)), ''), CAST(ch.IDUsuarioCriacao AS VARCHAR(50))),
            ch.IDUsuarioConfirmacao,
            UsuarioConfirmacao = COALESCE(NULLIF(LTRIM(RTRIM(uf.NomeUsuario)), ''), CAST(ch.IDUsuarioConfirmacao AS VARCHAR(50))),
            ch.Observacao
        FROM [Integracao].[Silver].[DimCheckingHistorico] ch
        LEFT JOIN [Integracao].[Silver].[FatoControleContratosEuromidia] cc
               ON cc.IDFatoControleContratosEuromidia = ch.IDFatoControleContratosEuromidia
        LEFT JOIN [Integracao].[Silver].[DimEmpresas] emp
               ON emp.IDEmpresa = COALESCE(ch.IDEmpresa, cc.IDEmpresa)
        OUTER APPLY (
            SELECT TOP (1)
                fp2.IDDimFacesPaineis,
                fp2.CodPonto,
                fp2.CodFace,
                fp2.Tipo,
                fp2.IDDimPaineisEuromidia
            FROM [Integracao].[Silver].[DimFacesPaineis] fp2
            WHERE fp2.CodFace = ch.CodFace
               OR (fp2.CodPonto = ch.CodPonto AND fp2.Face = ch.CodFace)
            ORDER BY CASE WHEN fp2.CodFace = ch.CodFace THEN 0 ELSE 1 END,
                     fp2.IDDimFacesPaineis DESC
        ) fp
        LEFT JOIN [Integracao].[Silver].[DimPaineisEuromidia] p
               ON p.IDDimPaineisEuromidia = fp.IDDimPaineisEuromidia
        LEFT JOIN [Integracao].[Silver].[DimUsuarios] uc
               ON uc.IDDimUsuarios = ch.IDUsuarioCriacao
        LEFT JOIN [Integracao].[Silver].[DimUsuarios] uf
               ON uf.IDDimUsuarios = ch.IDUsuarioConfirmacao
        WHERE ch.IDDimCheckingHistorico = :id_checking
    """)

    item = db.session.execute(sql, {"id_checking": id_checking}).mappings().first()

    if not item:
        abort(404)

    return render_template("euromidia/visualizar_checking.html", item=item)